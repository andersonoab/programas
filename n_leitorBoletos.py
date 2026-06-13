# -*- coding: utf-8 -*-
"""
LEITOR DE BOLETOS - DP / Financeiro Sonova
Le boletos em PDF, extrai os dados de pagamento e exporta para planilha Excel.

Padrao Sonova: rodar direto via shortcut do python.exe (sem executavel compilado).
    Alvo do atalho:  C:\\caminho\\python.exe  "C:\\caminho\\leitor_boletos.py"

Dependencias:  pip install customtkinter pdfplumber openpyxl
Autor: Anderson Souza
"""

import os
import re
import glob
from datetime import datetime, date

import pdfplumber
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
# Paleta Sonova OFC
# ------------------------------------------------------------------
P   = "#0083CA"   # dominante
DK  = "#003C64"   # suporte escuro
MD  = "#005A64"   # suporte
LT  = "#6EB4DC"   # suporte claro
AC  = "#7D0041"   # acento (vence hoje)
AC2 = "#8C321E"   # acento 2 (vencido)
OK  = "#005A64"   # conferido
TX  = "#333333"   # texto
MU  = "#646464"   # texto secundario
BD  = "#CCCCCC"   # borda neutra
BG  = "#FFFFFF"   # fundo
BG2 = "#F4F6F9"   # fundo area

PLANILHA_PADRAO = "boletos.xlsx"

COLUNAS_XLS = [
    ("Conferido",        "conferido",         11),
    ("Valor (R$)",       "valor_num",         16),
    ("Vencimento",       "vencimento_dt",     14),
    ("Linha Digitavel",  "linha_digitavel",   46),
    ("Banco",            "banco",             8),
    ("Beneficiario",     "beneficiario",      34),
    ("CNPJ Benef.",      "beneficiario_cnpj", 20),
    ("Pagador",          "pagador",           38),
    ("CNPJ Pagador",     "pagador_cnpj",      20),
    ("Nosso Numero",     "nosso_numero",      16),
    ("N. Documento",     "num_documento",     16),
    ("Data Emissao",     "data_emissao",      14),
    ("Arquivo",          "arquivo",           28),
]


# ==================================================================
# EXTRACAO (logica validada)
# ==================================================================

def _buscar(padrao, texto, grupo=1, flags=0):
    m = re.search(padrao, texto, flags)
    return m.group(grupo).strip() if m else None


def _br_float(v):
    if not v:
        return None
    return float(v.replace(".", "").replace(",", "."))


def _br_data(v):
    if not v:
        return None
    try:
        return datetime.strptime(v, "%d/%m/%Y")
    except ValueError:
        return None


def extrair_dados(caminho):
    texto = ""
    with pdfplumber.open(caminho) as pdf:
        for pagina in pdf.pages:
            texto += (pagina.extract_text() or "") + "\n"

    d = {}
    d["linha_digitavel"] = _buscar(
        r"(\d{5}\.\d{5}\s+\d{5}\.\d{6}\s+\d{5}\.\d{6}\s+\d\s+\d{14})", texto)
    d["banco"] = _buscar(r"(\d{3}-\d)\s+\d{5}\.\d{5}", texto)
    d["vencimento"] = _buscar(
        r"PAG[AÁ]VEL.*?(\d{2}/\d{2}/\d{4})", texto, flags=re.IGNORECASE | re.DOTALL)

    valor = _buscar(r"\(\s*=\s*\)\s*Valor cobrado\s*R\$\s*([\d\.]+,\d{2})", texto,
                    flags=re.IGNORECASE | re.DOTALL)
    if not valor:
        valor = _buscar(r"Valor do documento\s*R\$\s*([\d\.]+,\d{2})", texto,
                        flags=re.IGNORECASE | re.DOTALL)
    d["valor"] = valor

    d["beneficiario"] = _buscar(
        r"Benefici[áa]rio.*?\n([A-ZÀ-Ú0-9 ].+?)\s*-?\s*CNPJ",
        texto, flags=re.IGNORECASE | re.DOTALL)
    d["beneficiario_cnpj"] = _buscar(
        r"Benefici[áa]rio.*?CNPJ:\s*([\d\./\-]+)",
        texto, flags=re.IGNORECASE | re.DOTALL)
    d["pagador"] = _buscar(
        r"Pagador\s*\n([A-ZÀ-Ú0-9 ].+?)\s+CNPJ", texto, flags=re.DOTALL)
    d["pagador_cnpj"] = _buscar(
        r"Pagador\s*\n.+?CNPJ:\s*([\d\./\-]+)", texto, flags=re.DOTALL)
    d["nosso_numero"] = _buscar(r"DM\s+N\s+\d{2}/\d{2}/\d{4}\s+(\d{10,})", texto)
    d["num_documento"] = _buscar(r"\d{2}/\d{2}/\d{4}\s+(\d{10,})\s+DM", texto)
    d["data_emissao"] = _buscar(r"Data de Emiss[ãa]o\s*:\s*(\d{2}/\d{2}/\d{4})", texto)

    d["valor_num"] = _br_float(d["valor"])
    d["vencimento_dt"] = _br_data(d["vencimento"])
    d["conferido"] = ""
    d["arquivo"] = os.path.basename(caminho)
    return d


# ==================================================================
# EXPORTACAO EXCEL
# ==================================================================

def exportar_excel(lista, caminho_xlsx):
    if os.path.exists(caminho_xlsx):
        wb = load_workbook(caminho_xlsx)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Boletos"
        borda = Side(style="thin", color=BD.lstrip("#"))
        for i, (titulo, _ch, larg) in enumerate(COLUNAS_XLS, start=1):
            c = ws.cell(row=1, column=i, value=titulo)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=P.lstrip("#"))
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(bottom=borda)
            ws.column_dimensions[get_column_letter(i)].width = larg
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    col_linha = [c[1] for c in COLUNAS_XLS].index("linha_digitavel") + 1
    existentes = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_linha).value
        if v:
            existentes.add(str(v).strip())

    novos = 0
    for d in lista:
        chave = (d.get("linha_digitavel") or "").strip()
        if chave and chave in existentes:
            continue
        linha = ws.max_row + 1
        for i, (_t, ch, _l) in enumerate(COLUNAS_XLS, start=1):
            ws.cell(row=linha, column=i, value=d.get(ch))
        cv = ws.cell(row=linha, column=2)
        cv.number_format = u'R$ #,##0.00'
        cv.font = Font(name="Arial", size=10, bold=True)
        cvenc = ws.cell(row=linha, column=3)
        cvenc.number_format = "DD/MM/YYYY"
        cvenc.alignment = Alignment(horizontal="center")
        ws.cell(row=linha, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=linha, column=5).alignment = Alignment(horizontal="center")
        if chave:
            existentes.add(chave)
        novos += 1

    wb.save(caminho_xlsx)
    return novos


# ==================================================================
# CARD DE BOLETO
# ==================================================================

class BoletoCard(ctk.CTkFrame):
    def __init__(self, master, dados, on_change):
        super().__init__(master, fg_color=BG, border_width=1,
                         border_color=BD, corner_radius=8)
        self.dados = dados
        self.on_change = on_change
        self.grid_columnconfigure(1, weight=1)

        # --- coluna 0: valor + vencimento ---
        bloco = ctk.CTkFrame(self, fg_color="transparent")
        bloco.grid(row=0, column=0, rowspan=3, padx=(16, 10), pady=14, sticky="nw")

        valor = dados.get("valor")
        ctk.CTkLabel(bloco, text=("R$ " + valor) if valor else "VALOR N/D",
                     font=ctk.CTkFont(size=22, weight="bold"),
                     text_color=DK).pack(anchor="w")

        venc_txt, venc_cor = self._status_vencimento()
        ctk.CTkLabel(bloco, text=venc_txt,
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color=venc_cor).pack(anchor="w", pady=(2, 0))

        # --- coluna 1: dados ---
        info = ctk.CTkFrame(self, fg_color="transparent")
        info.grid(row=0, column=1, padx=6, pady=(14, 4), sticky="nw")

        ctk.CTkLabel(info, text=dados.get("beneficiario") or "Beneficiario N/D",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=TX, anchor="w", justify="left").pack(anchor="w")
        ctk.CTkLabel(info, text="CNPJ benef.: " + (dados.get("beneficiario_cnpj") or "-"),
                     font=ctk.CTkFont(size=11), text_color=MU).pack(anchor="w")
        ctk.CTkLabel(info,
                     text="Pagador: " + (dados.get("pagador") or "-")
                          + "  -  " + (dados.get("pagador_cnpj") or ""),
                     font=ctk.CTkFont(size=11), text_color=MU,
                     anchor="w", justify="left").pack(anchor="w", pady=(2, 0))
        ctk.CTkLabel(info,
                     text="Nosso n.: " + (dados.get("nosso_numero") or "-")
                          + "   Doc.: " + (dados.get("num_documento") or "-")
                          + "   Arq.: " + (dados.get("arquivo") or "-"),
                     font=ctk.CTkFont(size=10), text_color=MU).pack(anchor="w", pady=(2, 0))

        # linha digitavel + copiar
        linha_box = ctk.CTkFrame(self, fg_color=BG2, corner_radius=6)
        linha_box.grid(row=1, column=1, padx=6, pady=(2, 4), sticky="we")
        linha_box.grid_columnconfigure(0, weight=1)
        ld = dados.get("linha_digitavel") or "Linha digitavel nao localizada"
        ctk.CTkLabel(linha_box, text=ld, font=ctk.CTkFont(size=12),
                     text_color=DK, anchor="w").grid(row=0, column=0,
                                                     padx=(10, 6), pady=6, sticky="w")
        ctk.CTkButton(linha_box, text="Copiar", width=70, height=28,
                      fg_color=P, hover_color=DK, text_color="white",
                      command=self._copiar).grid(row=0, column=1, padx=(0, 8), pady=6)

        # --- coluna 2: conferido ---
        self.var_conf = ctk.StringVar(value="off")
        self.chk = ctk.CTkCheckBox(self, text="Conferido", variable=self.var_conf,
                                   onvalue="on", offvalue="off",
                                   fg_color=OK, hover_color=DK, text_color=TX,
                                   font=ctk.CTkFont(size=12, weight="bold"),
                                   command=self._toggle_conf)
        self.chk.grid(row=0, column=2, rowspan=3, padx=(10, 16), pady=14, sticky="e")

    def _status_vencimento(self):
        venc = self.dados.get("vencimento_dt")
        if not venc:
            return ("Vencimento: N/D", MU)
        hoje = date.today()
        v = venc.date()
        txt = "Vencimento: " + venc.strftime("%d/%m/%Y")
        if v < hoje:
            return (txt + "  (VENCIDO)", AC2)
        if v == hoje:
            return (txt + "  (VENCE HOJE)", AC)
        dias = (v - hoje).days
        return (txt + f"  (em {dias} dia(s))", MD)

    def _copiar(self):
        ld = self.dados.get("linha_digitavel")
        if not ld:
            return
        self.clipboard_clear()
        self.clipboard_append(ld)
        messagebox.showinfo("Copiado", "Linha digitavel copiada para a area de transferencia.")

    def _toggle_conf(self):
        marcado = self.var_conf.get() == "on"
        self.dados["conferido"] = "X" if marcado else ""
        self.configure(border_color=OK if marcado else BD,
                       border_width=2 if marcado else 1)
        self.on_change()


# ==================================================================
# APLICACAO
# ==================================================================

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Leitor de Boletos - DP / Financeiro Sonova")
        self.geometry("1080x720")
        self.minsize(900, 600)
        self.configure(fg_color=BG2)
        self.boletos = []
        self.cards = []

        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # ---------- cabecalho (barra azul) ----------
        header = ctk.CTkFrame(self, fg_color=P, corner_radius=0, height=54)
        header.grid(row=0, column=0, sticky="we")
        header.grid_propagate(False)
        ctk.CTkLabel(header, text="Leitor de Boletos",
                     font=ctk.CTkFont(size=20, weight="bold"),
                     text_color="white").pack(side="left", padx=20)
        ctk.CTkLabel(header, text="DP / Financeiro Sonova",
                     font=ctk.CTkFont(size=13), text_color="white").pack(side="right", padx=20)

        # ---------- barra de acoes ----------
        toolbar = ctk.CTkFrame(self, fg_color=BG, corner_radius=0, height=58)
        toolbar.grid(row=1, column=0, sticky="we")
        toolbar.grid_propagate(False)

        def botao(texto, cor, cmd):
            return ctk.CTkButton(toolbar, text=texto, fg_color=cor, hover_color=DK,
                                 text_color="white", height=34, width=150,
                                 font=ctk.CTkFont(size=13, weight="bold"), command=cmd)

        botao("Adicionar PDF(s)", P, self.add_pdfs).pack(side="left", padx=(20, 8), pady=12)
        botao("Ler pasta", MD, self.ler_pasta).pack(side="left", padx=8, pady=12)
        botao("Exportar Excel", DK, self.exportar).pack(side="left", padx=8, pady=12)
        ctk.CTkButton(toolbar, text="Limpar", fg_color="white", hover_color=BG2,
                      text_color=TX, border_width=1, border_color=BD, height=34, width=110,
                      font=ctk.CTkFont(size=13), command=self.limpar).pack(side="left",
                                                                            padx=8, pady=12)

        # ---------- area de cards ----------
        self.area = ctk.CTkScrollableFrame(self, fg_color=BG2, corner_radius=0)
        self.area.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)
        self.area.grid_columnconfigure(0, weight=1)

        self.vazio = ctk.CTkLabel(self.area,
                                  text="Nenhum boleto carregado.\nUse 'Adicionar PDF(s)' ou 'Ler pasta'.",
                                  font=ctk.CTkFont(size=14), text_color=MU)
        self.vazio.grid(row=0, column=0, pady=60)

        # ---------- rodape (barra azul) ----------
        footer = ctk.CTkFrame(self, fg_color=P, corner_radius=0, height=46)
        footer.grid(row=3, column=0, sticky="we")
        footer.grid_propagate(False)
        self.lbl_total = ctk.CTkLabel(footer, text=self._texto_total(),
                                      font=ctk.CTkFont(size=14, weight="bold"),
                                      text_color="white")
        self.lbl_total.pack(side="left", padx=20)
        ctk.CTkLabel(footer, text="Anderson Souza",
                     font=ctk.CTkFont(size=12), text_color="white").pack(side="right", padx=20)

    # ------------------ acoes ------------------

    def add_pdfs(self):
        caminhos = filedialog.askopenfilenames(
            title="Selecione boleto(s) PDF",
            filetypes=[("PDF", "*.pdf")])
        if caminhos:
            self._processar(list(caminhos))

    def ler_pasta(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta com os boletos")
        if pasta:
            pdfs = glob.glob(os.path.join(pasta, "*.pdf"))
            if not pdfs:
                messagebox.showwarning("Pasta vazia", "Nenhum PDF encontrado na pasta.")
                return
            self._processar(pdfs)

    def _processar(self, caminhos):
        existentes = {b.get("linha_digitavel") for b in self.boletos}
        erros, novos = [], 0
        for c in caminhos:
            try:
                d = extrair_dados(c)
                if d.get("linha_digitavel") and d["linha_digitavel"] in existentes:
                    continue
                self.boletos.append(d)
                if d.get("linha_digitavel"):
                    existentes.add(d["linha_digitavel"])
                novos += 1
            except Exception as e:
                erros.append(os.path.basename(c) + ": " + str(e))
        self._render()
        if erros:
            messagebox.showerror("Erros de leitura", "\n".join(erros))

    def _render(self):
        for card in self.cards:
            card.destroy()
        self.cards.clear()
        self.vazio.grid_remove()

        if not self.boletos:
            self.vazio.grid(row=0, column=0, pady=60)
        else:
            # ordena por vencimento (mais proximo primeiro)
            self.boletos.sort(key=lambda b: b.get("vencimento_dt") or datetime.max)
            for i, d in enumerate(self.boletos):
                card = BoletoCard(self.area, d, self._atualizar_total)
                card.grid(row=i, column=0, sticky="we", padx=4, pady=5)
                self.cards.append(card)
        self._atualizar_total()

    def _texto_total(self):
        total = sum(b.get("valor_num") or 0 for b in self.boletos)
        conf = sum(1 for b in self.boletos if b.get("conferido") == "X")
        total_fmt = f"R$ {total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"Total a pagar: {total_fmt}    |    Boletos: {len(self.boletos)}    |    Conferidos: {conf}"

    def _atualizar_total(self):
        self.lbl_total.configure(text=self._texto_total())

    def exportar(self):
        if not self.boletos:
            messagebox.showwarning("Sem dados", "Carregue pelo menos um boleto antes de exportar.")
            return
        caminho = filedialog.asksaveasfilename(
            title="Salvar planilha", defaultextension=".xlsx",
            initialfile=PLANILHA_PADRAO, filetypes=[("Excel", "*.xlsx")])
        if not caminho:
            return
        try:
            novos = exportar_excel(self.boletos, caminho)
            messagebox.showinfo(
                "Exportado",
                f"{novos} boleto(s) adicionado(s).\n"
                f"{len(self.boletos) - novos} ja existente(s) ignorado(s).\n\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e))

    def limpar(self):
        if self.boletos and not messagebox.askyesno("Limpar", "Remover todos os boletos da tela?"):
            return
        self.boletos.clear()
        self._render()


if __name__ == "__main__":
    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("blue")
    App().mainloop()
