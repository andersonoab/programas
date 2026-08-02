"""
Preenchedor Scanner Admissao F8 v1.0
Anderson Marinho | Igarape Digital

Fluxo principal:
1. Selecione o colaborador no aplicativo.
2. No Windows Explorer, selecione um ou mais scanners.
3. Pressione F8.
4. Confirme o destino. Os arquivos sao COPIADOS, nunca movidos.

O indice de colaboradores e montado a partir:
- da coluna A de todas as abas dos arquivos Excel da pasta SRA;
- das pastas ja existentes em ADMISSOES\\ATIVOS.

Dependencias: customtkinter, openpyxl, pynput e pywin32.
Arquivos .xls antigos usam pandas + xlrd quando instalados.

& $Python -m cx_Freeze `
"C:\_RPA\aAppEnvios PDF\_dist_scanner\PreenchedorScannerAdmissao_F8_v1 (1).py" `
--base-name Win32GUI `
--target-dir "C:\_RPA\aAppEnvios PDF\build\ScannerAdmissao_v1_4_CF615" `
--target-name "ScannerAdmissao_v1_4.exe" `
--packages "customtkinter,openpyxl,pynput,win32com,win32gui,pythoncom,pywintypes,PIL,pymupdf,pandas,xlrd"
cx_Freeze 6.15.16
LIEF 0.12.3
Python 3.10


"""

from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Callable, Iterable

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from PIL import Image, ImageTk
except Exception as exc:  # pragma: no cover
    Image = None
    ImageTk = None
    PIL_ERROR = exc
else:
    PIL_ERROR = None

try:
    # Importacao oficial moderna. Evita conflito com o pacote antigo "fitz",
    # que falha no Windows com: No module named 'frontend'.
    import pymupdf
except Exception as exc:  # pragma: no cover
    pymupdf = None
    PYMUPDF_ERROR = exc
else:
    PYMUPDF_ERROR = None

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover - validacao em tempo de execucao
    load_workbook = None
    OPENPYXL_ERROR = exc
else:
    OPENPYXL_ERROR = None

try:
    from pynput import keyboard as pynput_keyboard
except Exception as exc:  # pragma: no cover
    pynput_keyboard = None
    PYNPUT_ERROR = exc
else:
    PYNPUT_ERROR = None

try:
    import pythoncom
    import win32gui
    import win32com.client
except Exception as exc:  # pragma: no cover
    pythoncom = None
    win32gui = None
    win32com = None
    PYWIN32_ERROR = exc
else:
    PYWIN32_ERROR = None

try:
    import pandas as pd
except Exception:
    pd = None


APP_NAME = "Preenchedor Scanner Admissao"
VERSION = "1.6"
DEFAULT_BASE = (
    r"C:\Users\99andsouza\Sonova\BR-SO-HR-RecursosHumanos - General"
    r"\ADMISSÕES\ATIVOS"
)
DEFAULT_SRA = (
    r"C:\Users\99andsouza\Sonova\BR-SO-HR-RecursosHumanos - General"
    r"\ACESSO ADP\SRA"
)
DESTINATIONS = ["E-mails e Anexos", "Documentos Admissionais", "Organizador", "(raiz do colaborador)"]
ROOT_DESTINATION = "(raiz do colaborador)"
PREVIEW_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff", ".webp"}


def app_data_dir() -> Path:
    root = os.getenv("APPDATA") or str(Path.home())
    return Path(root) / "IgarapeDigital" / "ScannerAdmissao"


CONFIG_PATH = app_data_dir() / "config.json"
HISTORY_PATH = app_data_dir() / "historico.csv"


def normalized(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().upper())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[.\-_/\\:;,()]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def safe_name(value: object, limit: int = 150) -> str:
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', " ", str(value or "").strip())
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text[:limit].rstrip(" .")


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for number in range(1, 10000):
        candidate = path.with_name(f"{path.stem} ({number}){path.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Nao foi possivel gerar um nome unico para {path.name}.")


def is_inside(child: Path, parent: Path) -> bool:
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except (ValueError, OSError):
        return False


@dataclass(frozen=True)
class Employee:
    name: str
    normalized_name: str
    source: str
    detail: str


def _employees_from_existing_folders(base: Path) -> list[Employee]:
    rows: list[Employee] = []
    if not base.is_dir():
        return rows
    try:
        folders = sorted((item for item in base.iterdir() if item.is_dir()), key=lambda x: normalized(x.name))
    except OSError:
        return rows
    for folder in folders:
        if normalized(folder.name):
            rows.append(Employee(folder.name, normalized(folder.name), "Pasta existente", str(folder)))
    return rows


def _employees_from_modern_excel(path: Path) -> Iterable[Employee]:
    if load_workbook is None:
        raise RuntimeError(f"openpyxl nao disponivel: {OPENPYXL_ERROR}")
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        for sheet in workbook.worksheets:
            for row_number, row in enumerate(
                sheet.iter_rows(min_col=1, max_col=1, values_only=True), start=1
            ):
                value = "" if not row else str(row[0] or "").strip()
                key = normalized(value)
                if key and key not in {"NOME", "COLABORADOR", "NOME COLABORADOR", "NOME DO COLABORADOR"}:
                    yield Employee(value, key, path.name, f"Aba {sheet.title} | linha {row_number}")
    finally:
        workbook.close()


def _employees_from_old_excel(path: Path) -> Iterable[Employee]:
    if pd is None:
        raise RuntimeError("Arquivo .xls encontrado. Instale pandas e xlrd para le-lo.")
    book = pd.ExcelFile(path, engine="xlrd")
    for sheet_name in book.sheet_names:
        frame = pd.read_excel(book, sheet_name=sheet_name, header=None, usecols=[0], dtype=str)
        for index, value in enumerate(frame.iloc[:, 0].fillna(""), start=1):
            text = str(value).strip()
            key = normalized(text)
            if key and key not in {"NOME", "COLABORADOR", "NOME COLABORADOR", "NOME DO COLABORADOR"}:
                yield Employee(text, key, path.name, f"Aba {sheet_name} | linha {index}")


def build_employee_index(
    base_dir: Path,
    excel_dir: Path,
    recursive: bool,
    progress: Callable[[str], None] | None = None,
) -> tuple[list[Employee], list[str], int]:
    """Retorna colaboradores unicos, avisos e quantidade de planilhas lidas."""
    found: dict[str, Employee] = {}
    warnings: list[str] = []

    for employee in _employees_from_existing_folders(base_dir):
        found[employee.normalized_name] = employee

    pattern = "**/*" if recursive else "*"
    files: list[Path] = []
    if excel_dir.is_file():
        if excel_dir.suffix.lower() in {".xlsx", ".xlsm", ".xls"} and not excel_dir.name.startswith("~$"):
            files = [excel_dir]
        else:
            warnings.append(f"O arquivo informado nao e uma planilha Excel valida: {excel_dir}")
    elif excel_dir.is_dir():
        try:
            files = sorted(
                path for path in excel_dir.glob(pattern)
                if path.is_file()
                and path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}
                and not path.name.startswith("~$")
            )
        except OSError as exc:
            warnings.append(f"Falha ao listar a pasta SRA: {exc}")
    else:
        warnings.append(f"Pasta SRA nao encontrada: {excel_dir}")

    read_count = 0
    for number, path in enumerate(files, start=1):
        if progress:
            progress(f"Lendo {number}/{len(files)}: {path.name}")
        try:
            iterator = (
                _employees_from_old_excel(path)
                if path.suffix.lower() == ".xls"
                else _employees_from_modern_excel(path)
            )
            for employee in iterator:
                # Pasta existente tem prioridade; depois fica a primeira ocorrencia no Excel.
                if employee.normalized_name not in found:
                    found[employee.normalized_name] = employee
            read_count += 1
        except Exception as exc:
            warnings.append(f"{path.name}: {exc}")

    result = sorted(found.values(), key=lambda item: item.normalized_name)
    return result, warnings, read_count


def search_employees(rows: list[Employee], term: str, limit: int = 200) -> list[Employee]:
    key = normalized(term)
    if not key:
        return rows[:limit]
    tokens = key.split()
    scored: list[tuple[float, Employee]] = []
    for employee in rows:
        name = employee.normalized_name
        if not all(token in name for token in tokens):
            continue
        score = SequenceMatcher(None, key, name).ratio()
        if name == key:
            score += 4
        elif name.startswith(key):
            score += 2
        elif f" {key}" in name:
            score += 1
        score -= max(0, len(name) - len(key)) / 1000
        scored.append((score, employee))
    scored.sort(key=lambda item: (-item[0], item[1].normalized_name))
    return [item[1] for item in scored[:limit]]


def _selected_files_via_shell() -> list[Path]:
    """Le a selecao do Windows Explorer sem alterar o clipboard."""
    if pythoncom is None or win32gui is None or win32com is None:
        return []
    pythoncom.CoInitialize()
    try:
        foreground = int(win32gui.GetForegroundWindow())
        shell = win32com.client.Dispatch("Shell.Application")
        for window in shell.Windows():
            try:
                if int(window.HWND) != foreground:
                    continue
                items = window.Document.SelectedItems()
                return [Path(str(item.Path)) for item in items if Path(str(item.Path)).is_file()]
            except Exception:
                continue
    finally:
        pythoncom.CoUninitialize()
    return []


def selected_explorer_files() -> list[Path]:
    # Nao le o clipboard como fallback: ele poderia conter uma selecao antiga.
    # Se o Explorer nao expuser a selecao, o usuario usa "Adicionar arquivos".
    rows = _selected_files_via_shell()
    unique: list[Path] = []
    seen: set[str] = set()
    for path in rows:
        key = os.path.normcase(os.path.abspath(path))
        if key not in seen:
            seen.add(key)
            unique.append(path)
    return unique


class AutoHideScrollbar(ttk.Scrollbar):
    """Scrollbar fina que some quando o conteudo cabe inteiro na area visivel.

    Precisa ser posicionada com grid(). Usada nas Treeviews para manter o
    padrao de barras finas (11px) com auto-hide.
    """

    def set(self, low, high) -> None:
        if float(low) <= 0.0 and float(high) >= 1.0:
            self.grid_remove()
        else:
            self.grid()
        ttk.Scrollbar.set(self, low, high)

    def pack(self, **_kwargs):  # pragma: no cover - protecao de uso indevido
        raise tk.TclError("AutoHideScrollbar so suporta grid().")

    def place(self, **_kwargs):  # pragma: no cover - protecao de uso indevido
        raise tk.TclError("AutoHideScrollbar so suporta grid().")


def attach_scrollbar(parent, tree: ttk.Treeview, row: int, column: int = 0) -> AutoHideScrollbar:
    """Encaixa uma Treeview + scrollbar fina auto-hide num grid existente."""
    tree.grid(row=row, column=column, padx=(14, 0), pady=4, sticky="nsew")
    scrollbar = AutoHideScrollbar(
        parent, orient="vertical", style="Fina.Vertical.TScrollbar", command=tree.yview,
    )
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.grid(row=row, column=column + 1, padx=(0, 14), pady=4, sticky="ns")
    return scrollbar


class SortableTreeview:
    """Adiciona ordenacao asc/desc por coluna (com setas) e zebra recalculada."""

    def __init__(
        self,
        tree: ttk.Treeview,
        headings: dict[str, str],
        key_funcs: dict[str, Callable[[str], object]] | None = None,
        priority_tags: frozenset[str] = frozenset(),
        zebra_tags: tuple[str, str] = ("zebra_par", "zebra_impar"),
    ) -> None:
        self.tree = tree
        self.headings = headings
        self.key_funcs = key_funcs or {}
        self.priority_tags = priority_tags
        self.zebra_tags = zebra_tags
        self.sort_column: str | None = None
        self.sort_reverse = False
        self.reset_headers()

    def _default_key(self, value: str):
        return value.strip().lower()

    def sort_by(self, column: str) -> None:
        reverse = self.sort_column == column and not self.sort_reverse
        key_func = self.key_funcs.get(column, self._default_key)
        rows = [(self.tree.set(iid, column), iid) for iid in self.tree.get_children("")]
        rows.sort(key=lambda pair: key_func(pair[0]), reverse=reverse)
        for position, (_value, iid) in enumerate(rows):
            self.tree.move(iid, "", position)
        self.sort_column = column
        self.sort_reverse = reverse
        self._paint_headers()
        self.apply_zebra()

    def _paint_headers(self) -> None:
        for column, label in self.headings.items():
            text = label
            if column == self.sort_column:
                text += " ↓" if self.sort_reverse else " ↑"
            self.tree.heading(column, text=text, command=lambda c=column: self.sort_by(c))

    def reset_headers(self) -> None:
        """Chamado apos recarregar os dados: limpa as setas (ordem natural)."""
        self.sort_column = None
        self.sort_reverse = False
        self._paint_headers()

    def apply_zebra(self) -> None:
        even_tag, odd_tag = self.zebra_tags
        for position, iid in enumerate(self.tree.get_children("")):
            current = set(self.tree.item(iid, "tags"))
            if current & self.priority_tags:
                continue
            self.tree.item(iid, tags=(even_tag if position % 2 == 0 else odd_tag,))


class PreviewWindow:
    """Visualizador interno de PDF e imagens, inclusive TIFF multipagina."""

    BG = "#20252B"
    PANEL = "#F4F6F9"
    BLUE = "#0083CA"
    BLUE_DARK = "#003C64"
    TEXT = "#333333"

    def __init__(
        self,
        parent: tk.Misc,
        files: list[Path],
        start_index: int,
        on_viewed: Callable[[Path], None],
    ) -> None:
        self.parent = parent
        self.files = list(files)
        self.file_index = max(0, min(start_index, len(self.files) - 1))
        self.page_index = 0
        self.page_count = 1
        self.zoom = 1.0
        self.rotation = 0
        self.on_viewed = on_viewed
        self.photo = None
        self._resize_job = None

        self.top = tk.Toplevel(parent)
        self.top.title("Visualizar scanner antes de copiar")
        self.top.geometry("1180x820")
        self.top.minsize(850, 600)
        self.top.configure(bg=self.PANEL)
        self.top.transient(parent)

        header = tk.Frame(self.top, bg=self.BLUE_DARK, height=66)
        header.pack(fill="x")
        header.pack_propagate(False)
        title_box = tk.Frame(header, bg=self.BLUE_DARK)
        title_box.pack(side="left", fill="both", expand=True, padx=18, pady=8)
        self.file_label = tk.Label(
            title_box, text="", bg=self.BLUE_DARK, fg="white",
            font=("Segoe UI", 15, "bold"), anchor="w",
        )
        self.file_label.pack(fill="x")
        self.path_label = tk.Label(
            title_box, text="", bg=self.BLUE_DARK, fg="#D9EDF8",
            font=("Segoe UI", 9), anchor="w",
        )
        self.path_label.pack(fill="x")
        tk.Button(
            header, text="Abrir no aplicativo padrao", command=self.open_external,
            bg=self.BLUE, fg="white", activebackground="#006BA8", activeforeground="white",
            relief="flat", font=("Segoe UI", 10, "bold"), padx=14, pady=7,
        ).pack(side="right", padx=18)

        file_bar = tk.Frame(self.top, bg=self.PANEL)
        file_bar.pack(fill="x", padx=14, pady=(10, 6))
        self.prev_file_button = self._button(file_bar, "◀ Arquivo anterior", self.prev_file)
        self.prev_file_button.pack(side="left")
        self.file_position_label = tk.Label(
            file_bar, text="", bg=self.PANEL, fg=self.BLUE_DARK,
            font=("Segoe UI", 10, "bold"),
        )
        self.file_position_label.pack(side="left", padx=14)
        self.next_file_button = self._button(file_bar, "Proximo arquivo ▶", self.next_file)
        self.next_file_button.pack(side="left")
        tk.Label(
            file_bar, text="Duplo clique na lista tambem abre esta tela.",
            bg=self.PANEL, fg="#646464", font=("Segoe UI", 9),
        ).pack(side="right")

        viewer = tk.Frame(self.top, bg=self.BG)
        viewer.pack(fill="both", expand=True, padx=14, pady=4)
        viewer.grid_rowconfigure(0, weight=1)
        viewer.grid_columnconfigure(0, weight=1)
        self.canvas = tk.Canvas(viewer, bg=self.BG, highlightthickness=0)
        self.scroll_y = ttk.Scrollbar(viewer, orient="vertical", command=self.canvas.yview)
        self.scroll_x = ttk.Scrollbar(viewer, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.scroll_y.set, xscrollcommand=self.scroll_x.set)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.scroll_y.grid(row=0, column=1, sticky="ns")
        self.scroll_x.grid(row=1, column=0, sticky="ew")
        self.canvas.bind("<Configure>", self._on_canvas_resize)

        controls = tk.Frame(self.top, bg=self.PANEL)
        controls.pack(fill="x", padx=14, pady=(7, 12))
        self.prev_page_button = self._button(controls, "◀ Pagina", self.prev_page)
        self.prev_page_button.pack(side="left")
        self.page_label = tk.Label(
            controls, text="Pagina 1/1", bg=self.PANEL, fg=self.TEXT,
            font=("Segoe UI", 10, "bold"), width=15,
        )
        self.page_label.pack(side="left", padx=5)
        self.next_page_button = self._button(controls, "Pagina ▶", self.next_page)
        self.next_page_button.pack(side="left")
        self._button(controls, "− Zoom", self.zoom_out).pack(side="left", padx=(22, 4))
        self.zoom_label = tk.Label(
            controls, text="Ajustar", bg=self.PANEL, fg=self.TEXT,
            font=("Segoe UI", 10, "bold"), width=10,
        )
        self.zoom_label.pack(side="left")
        self._button(controls, "+ Zoom", self.zoom_in).pack(side="left", padx=4)
        self._button(controls, "Ajustar", self.fit).pack(side="left", padx=(8, 4))
        self._button(controls, "↶ Girar", lambda: self.rotate(-90)).pack(side="left", padx=(22, 4))
        self._button(controls, "Girar ↷", lambda: self.rotate(90)).pack(side="left")
        self._button(controls, "Fechar", self.top.destroy, dark=True).pack(side="right")

        self.top.bind("<Escape>", lambda _event: self.top.destroy())
        self.top.bind("<Prior>", lambda _event: self.prev_page())
        self.top.bind("<Next>", lambda _event: self.next_page())
        self.top.bind("<Control-Left>", lambda _event: self.prev_file())
        self.top.bind("<Control-Right>", lambda _event: self.next_file())
        self.top.bind("<Control-plus>", lambda _event: self.zoom_in())
        self.top.bind("<Control-minus>", lambda _event: self.zoom_out())
        self.top.after(80, self.load_current_file)

    def _button(self, parent, text: str, command, dark: bool = False) -> tk.Button:
        color = self.BLUE_DARK if dark else self.BLUE
        return tk.Button(
            parent, text=text, command=command, bg=color, fg="white",
            activebackground="#002A47" if dark else "#006BA8", activeforeground="white",
            relief="flat", font=("Segoe UI", 9, "bold"), padx=11, pady=6,
        )

    @property
    def current_path(self) -> Path:
        return self.files[self.file_index]

    def _on_canvas_resize(self, _event=None) -> None:
        if self._resize_job is not None:
            try:
                self.top.after_cancel(self._resize_job)
            except Exception:
                pass
        self._resize_job = self.top.after(180, self.render)

    def _load_source_image(self) -> "Image.Image":
        path = self.current_path
        suffix = path.suffix.lower()
        if Image is None:
            raise RuntimeError(f"Pillow nao instalado: {PIL_ERROR}")
        if suffix == ".pdf":
            if pymupdf is None:
                raise RuntimeError(f"PyMuPDF nao instalado: {PYMUPDF_ERROR}")
            with pymupdf.open(str(path)) as document:
                self.page_count = max(1, document.page_count)
                self.page_index = min(self.page_index, self.page_count - 1)
                page = document.load_page(self.page_index)
                matrix = pymupdf.Matrix(max(1.8, 1.8 * self.zoom), max(1.8, 1.8 * self.zoom))
                pixmap = page.get_pixmap(matrix=matrix, alpha=False)
                return Image.frombytes("RGB", (pixmap.width, pixmap.height), pixmap.samples)
        if suffix in PREVIEW_IMAGE_EXTENSIONS:
            with Image.open(path) as source:
                self.page_count = max(1, int(getattr(source, "n_frames", 1)))
                self.page_index = min(self.page_index, self.page_count - 1)
                source.seek(self.page_index)
                return source.convert("RGB").copy()
        raise RuntimeError(
            f"Formato {suffix or '(sem extensao)'} sem visualizacao interna. "
            "Use 'Abrir no aplicativo padrao'."
        )

    def load_current_file(self) -> None:
        self.page_index = 0
        self.page_count = 1
        self.zoom = 1.0
        self.rotation = 0
        self.file_label.configure(text=self.current_path.name)
        self.path_label.configure(text=str(self.current_path))
        self.file_position_label.configure(
            text=f"Arquivo {self.file_index + 1} de {len(self.files)}"
        )
        self.prev_file_button.configure(state="normal" if self.file_index > 0 else "disabled")
        self.next_file_button.configure(
            state="normal" if self.file_index < len(self.files) - 1 else "disabled"
        )
        self.render()

    def render(self) -> None:
        if not self.top.winfo_exists():
            return
        self.canvas.delete("all")
        try:
            image = self._load_source_image()
            if self.rotation:
                image = image.rotate(-self.rotation, expand=True)
            canvas_width = max(300, self.canvas.winfo_width() - 28)
            canvas_height = max(300, self.canvas.winfo_height() - 28)
            fit_scale = min(canvas_width / image.width, canvas_height / image.height)
            scale = max(0.05, fit_scale * self.zoom)
            width = max(1, int(image.width * scale))
            height = max(1, int(image.height * scale))
            if (width, height) != image.size:
                image = image.resize((width, height), Image.Resampling.LANCZOS)
            self.photo = ImageTk.PhotoImage(image)
            x = max(12, (canvas_width - width) // 2)
            y = max(12, (canvas_height - height) // 2)
            self.canvas.create_image(x, y, anchor="nw", image=self.photo)
            self.canvas.configure(scrollregion=(0, 0, max(canvas_width, x + width + 12),
                                                max(canvas_height, y + height + 12)))
            self.on_viewed(self.current_path)
            self.page_label.configure(text=f"Pagina {self.page_index + 1}/{self.page_count}")
            self.prev_page_button.configure(state="normal" if self.page_index > 0 else "disabled")
            self.next_page_button.configure(
                state="normal" if self.page_index < self.page_count - 1 else "disabled"
            )
            self.zoom_label.configure(text="Ajustar" if self.zoom == 1.0 else f"{self.zoom:.0%}")
        except Exception as exc:
            self.page_count = 1
            self.page_label.configure(text="Sem previa")
            self.prev_page_button.configure(state="disabled")
            self.next_page_button.configure(state="disabled")
            self.canvas.create_text(
                max(250, self.canvas.winfo_width() // 2),
                max(180, self.canvas.winfo_height() // 2),
                text=f"Nao foi possivel visualizar este arquivo.\n\n{exc}",
                fill="white", font=("Segoe UI", 12), justify="center", width=650,
            )

    def prev_file(self) -> None:
        if self.file_index > 0:
            self.file_index -= 1
            self.load_current_file()

    def next_file(self) -> None:
        if self.file_index < len(self.files) - 1:
            self.file_index += 1
            self.load_current_file()

    def prev_page(self) -> None:
        if self.page_index > 0:
            self.page_index -= 1
            self.render()

    def next_page(self) -> None:
        if self.page_index < self.page_count - 1:
            self.page_index += 1
            self.render()

    def zoom_in(self) -> None:
        self.zoom = min(4.0, self.zoom * 1.25)
        self.render()

    def zoom_out(self) -> None:
        self.zoom = max(0.25, self.zoom / 1.25)
        self.render()

    def fit(self) -> None:
        self.zoom = 1.0
        self.render()

    def rotate(self, degrees: int) -> None:
        self.rotation = (self.rotation + degrees) % 360
        self.render()

    def open_external(self) -> None:
        try:
            if os.name == "nt":
                os.startfile(str(self.current_path))  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(self.current_path)])
            else:
                subprocess.Popen(["xdg-open", str(self.current_path)])
        except Exception as exc:
            messagebox.showerror("Abrir arquivo", f"Nao foi possivel abrir:\n\n{exc}", parent=self.top)


class ScannerAdmissionApp:
    AZUL = "#0083CA"
    AZUL_HOVER = "#006BA8"
    AZUL_ESCURO = "#003C64"
    AZUL_CLARO = "#6EB4DC"
    FUNDO = "#F4F6F9"
    CARD = "#FFFFFF"
    BORDA = "#CCCCCC"
    TEXTO = "#333333"
    TEXTO_SUAVE = "#646464"
    SUCESSO = "#1F7A3D"
    ALERTA = "#8C321E"
    AMBAR = "#A97C00"

    FONT_TITLE = ("Segoe UI", 18, "bold")
    FONT_SUB = ("Segoe UI", 11)
    FONT_LABEL = ("Segoe UI", 11, "bold")
    FONT_TEXT = ("Segoe UI", 11)
    FONT_BUTTON = ("Segoe UI", 11, "bold")

    def __init__(self) -> None:
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        self.app = ctk.CTk()
        self.app.title(f"{APP_NAME} | v{VERSION}")
        self.app.geometry("1260x820")
        self.app.minsize(1030, 700)
        self.app.configure(fg_color=self.FUNDO)

        self.config = self._load_config()
        self.employees: list[Employee] = []
        self.visible_employees: list[Employee] = []
        self.files: list[Path] = []
        self.previewed_files: set[str] = set()
        self.files_to_copy: set[str] = set()
        self.custom_names: dict[str, str] = {}
        self.selected_employee: Employee | None = None
        self.hotkey_listener = None
        self.hotkey_on = False
        self.indexing = False
        self.last_f8 = 0.0

        self._configure_ttk()
        self._build_ui()
        self.app.bind("<F8>", lambda _event: self.handle_f8())
        self.app.protocol("WM_DELETE_WINDOW", self.close)
        self.app.after(250, self.activate_hotkey)
        self.app.after(500, self.reindex)

    def _load_config(self) -> dict:
        defaults = {
            "base_dir": DEFAULT_BASE,
            "excel_dir": DEFAULT_SRA,
            "destination": DESTINATIONS[0],
            "recursive": True,
            "timestamp": True,
            "confirm": True,
        }
        try:
            if CONFIG_PATH.exists():
                data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
                defaults.update({key: data[key] for key in defaults if key in data})
        except Exception:
            pass
        return defaults

    def _save_config(self, quiet: bool = False) -> None:
        self.config = {
            "base_dir": self.base_entry.get().strip(),
            "excel_dir": self.excel_entry.get().strip(),
            "destination": self.destination_combo.get(),
            "recursive": bool(self.recursive_var.get()),
            "timestamp": bool(self.timestamp_var.get()),
            "confirm": bool(self.confirm_var.get()),
        }
        try:
            CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
            CONFIG_PATH.write_text(json.dumps(self.config, ensure_ascii=False, indent=2), encoding="utf-8")
            if not quiet:
                self.set_status("Configuracoes salvas.", self.SUCESSO)
        except Exception as exc:
            if not quiet:
                messagebox.showerror("Configuracoes", f"Nao foi possivel salvar:\n\n{exc}")

    def _configure_ttk(self) -> None:
        style = ttk.Style()
        style.theme_use("default")
        style.configure(
            "Sonova.Treeview", background="white", foreground=self.TEXTO,
            fieldbackground="white", rowheight=30, borderwidth=0,
            font=self.FONT_TEXT,
        )
        style.configure(
            "Sonova.Treeview.Heading", background="#E8F3FA", foreground=self.AZUL_ESCURO,
            relief="flat", padding=(8, 8), font=self.FONT_LABEL,
        )
        style.map(
            "Sonova.Treeview",
            background=[("selected", "#BFE0F2")],
            foreground=[("selected", self.AZUL_ESCURO)],
        )
        style.configure(
            "Fina.Vertical.TScrollbar", gripcount=0, width=11, arrowsize=11,
            troughcolor=self.FUNDO, background=self.AZUL_CLARO,
            bordercolor=self.FUNDO, lightcolor=self.AZUL_CLARO, darkcolor=self.AZUL_CLARO,
            relief="flat",
        )
        style.map(
            "Fina.Vertical.TScrollbar",
            background=[("active", self.AZUL)],
        )

    def _button(self, parent, text: str, command, width: int = 145, secondary: bool = False):
        return ctk.CTkButton(
            parent, text=text, command=command, width=width, height=36,
            corner_radius=8, font=self.FONT_BUTTON,
            fg_color=self.AZUL_ESCURO if secondary else self.AZUL,
            hover_color="#002A47" if secondary else self.AZUL_HOVER,
        )

    def _entry(self, parent, **kwargs):
        return ctk.CTkEntry(
            parent, height=35, corner_radius=8, fg_color="white",
            border_color=self.BORDA, text_color=self.TEXTO, font=self.FONT_TEXT, **kwargs,
        )

    def _card(self, parent):
        return ctk.CTkFrame(
            parent, fg_color=self.CARD, corner_radius=12,
            border_width=1, border_color=self.BORDA,
        )

    def _build_ui(self) -> None:
        header = ctk.CTkFrame(self.app, height=82, corner_radius=0, fg_color=self.AZUL)
        header.pack(fill="x")
        header.pack_propagate(False)
        title_box = ctk.CTkFrame(header, fg_color="transparent")
        title_box.pack(side="left", fill="y", padx=24, pady=13)
        ctk.CTkLabel(
            title_box, text="Scanner de Admissao", font=("Segoe UI", 22, "bold"),
            text_color="white",
        ).pack(anchor="w")
        ctk.CTkLabel(
            title_box, text="Selecione o colaborador > selecione os arquivos no Explorer > F8",
            font=self.FONT_SUB, text_color="white",
        ).pack(anchor="w")
        self.hotkey_label = ctk.CTkLabel(
            header, text="F8 local", font=self.FONT_LABEL, text_color="white",
            fg_color="#646464", corner_radius=8, width=150, height=32,
        )
        self.hotkey_label.pack(side="right", padx=24)

        container = ctk.CTkFrame(self.app, fg_color=self.FUNDO)
        container.pack(fill="both", expand=True, padx=18, pady=14)

        config_card = self._card(container)
        config_card.pack(fill="x", pady=(0, 12))
        config_card.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(config_card, text="Pasta de admissoes", font=self.FONT_LABEL,
                     text_color=self.AZUL_ESCURO).grid(row=0, column=0, padx=(14, 8), pady=(12, 6), sticky="w")
        self.base_entry = self._entry(config_card)
        self.base_entry.insert(0, self.config["base_dir"])
        self.base_entry.grid(row=0, column=1, padx=4, pady=(12, 6), sticky="ew")
        self._button(config_card, "Escolher", self.choose_base, 90, True).grid(
            row=0, column=2, padx=(6, 14), pady=(12, 6)
        )

        ctk.CTkLabel(config_card, text="Planilhas SRA", font=self.FONT_LABEL,
                     text_color=self.AZUL_ESCURO).grid(row=1, column=0, padx=(14, 8), pady=6, sticky="w")
        self.excel_entry = self._entry(config_card)
        self.excel_entry.insert(0, self.config["excel_dir"])
        self.excel_entry.grid(row=1, column=1, padx=4, pady=6, sticky="ew")
        self._button(config_card, "Escolher", self.choose_excel, 90, True).grid(
            row=1, column=2, padx=(6, 14), pady=6
        )

        options = ctk.CTkFrame(config_card, fg_color="transparent")
        options.grid(row=2, column=0, columnspan=3, sticky="ew", padx=14, pady=(5, 12))
        ctk.CTkLabel(options, text="Subpasta:", font=self.FONT_LABEL,
                     text_color=self.TEXTO).pack(side="left")
        self.destination_combo = ctk.CTkComboBox(
            options, values=DESTINATIONS, width=220, height=34, state="readonly",
            fg_color="white", border_color=self.BORDA, button_color=self.AZUL,
            button_hover_color=self.AZUL_HOVER, text_color=self.TEXTO,
        )
        self.destination_combo.set(self.config["destination"])
        self.destination_combo.pack(side="left", padx=(7, 18))
        self.recursive_var = tk.BooleanVar(value=self.config["recursive"])
        self.timestamp_var = tk.BooleanVar(value=self.config["timestamp"])
        self.confirm_var = tk.BooleanVar(value=self.config["confirm"])
        for text, variable in (
            ("Ler subpastas do SRA", self.recursive_var),
            ("Prefixar data e hora", self.timestamp_var),
            ("Confirmar antes de copiar", self.confirm_var),
        ):
            ctk.CTkCheckBox(
                options, text=text, variable=variable, font=self.FONT_TEXT,
                text_color=self.TEXTO, fg_color=self.AZUL, hover_color=self.AZUL_HOVER,
            ).pack(side="left", padx=(0, 18))
        self._button(options, "Salvar", self._save_config, 85, True).pack(side="right")
        self._button(options, "Reindexar", self.reindex, 105).pack(side="right", padx=(0, 7))

        work = ctk.CTkFrame(container, fg_color="transparent")
        work.pack(fill="both", expand=True)
        work.grid_columnconfigure(0, weight=45)
        work.grid_columnconfigure(1, weight=55)
        work.grid_rowconfigure(0, weight=1)

        files_card = self._card(work)
        files_card.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        files_card.grid_rowconfigure(2, weight=1)
        files_card.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(files_card, text="1. Arquivos selecionados", font=self.FONT_TITLE,
                     text_color=self.AZUL_ESCURO).grid(row=0, column=0, columnspan=2, padx=14, pady=(13, 2), sticky="w")
        self.files_info = ctk.CTkLabel(
            files_card, text="Use F8 no Windows Explorer ou adicione manualmente.",
            font=self.FONT_SUB, text_color=self.TEXTO_SUAVE,
        )
        self.files_info.grid(row=1, column=0, columnspan=2, padx=14, pady=(0, 8), sticky="w")
        self.files_tree = ttk.Treeview(
            files_card, columns=("check", "file", "type", "size", "viewed", "status"),
            show="headings", style="Sonova.Treeview",
        )
        files_headings = {
            "check": "Copiar", "file": "Arquivo", "type": "Tipo",
            "size": "Tamanho", "viewed": "Visualizado", "status": "Status",
        }
        self.files_tree.column("check", width=52, anchor="center", stretch=False)
        self.files_tree.column("file", width=180, anchor="w", stretch=True)
        self.files_tree.column("type", width=42, anchor="center", stretch=False)
        self.files_tree.column("size", width=62, anchor="e", stretch=False)
        self.files_tree.column("viewed", width=67, anchor="center", stretch=False)
        self.files_tree.column("status", width=67, anchor="center", stretch=False)
        attach_scrollbar(files_card, self.files_tree, row=2, column=0)
        self.files_tree.bind("<Button-1>", self.on_file_tree_click)
        self.files_tree.bind("<Double-1>", self.on_file_tree_double_click)
        self.files_tree.bind("<space>", self.toggle_focused_files)
        self.files_tree.bind("<F2>", lambda _event: self.rename_selected_file())
        self.files_sorter = SortableTreeview(
            self.files_tree, files_headings,
            key_funcs={
                "check": lambda v: v == "☑",
                "size": self._size_sort_key,
                "viewed": lambda v: v == "Sim",
                "status": lambda v: v != "Disponivel",
            },
            priority_tags=frozenset({"missing", "checked", "viewed", "renamed"}),
        )

        selection_buttons = ctk.CTkFrame(files_card, fg_color="transparent")
        selection_buttons.grid(row=3, column=0, columnspan=2, padx=14, pady=(6, 0), sticky="ew")
        self._button(
            selection_buttons, "Marcar disponiveis", self.mark_all_available, 145
        ).pack(side="left")
        self._button(
            selection_buttons, "Desmarcar todos", self.unmark_all_files, 135, True
        ).pack(side="left", padx=6)
        self.copy_count_label = ctk.CTkLabel(
            selection_buttons, text="0 marcado(s)", font=self.FONT_LABEL,
            text_color=self.ALERTA,
        )
        self.copy_count_label.pack(side="right")

        file_buttons = ctk.CTkFrame(files_card, fg_color="transparent")
        file_buttons.grid(row=4, column=0, columnspan=2, padx=14, pady=(7, 13), sticky="ew")
        self._button(file_buttons, "Adicionar arquivos", self.choose_files, 150).pack(side="left")
        self._button(file_buttons, "Visualizar", self.preview_selected_file, 105).pack(side="left", padx=(6, 0))
        self._button(file_buttons, "Renomear", self.rename_selected_file, 100, True).pack(side="left", padx=6)
        self._button(file_buttons, "Remover", self.remove_selected_files, 95, True).pack(side="left", padx=6)
        self._button(file_buttons, "Limpar", self.clear_files, 85, True).pack(side="left")

        employee_card = self._card(work)
        employee_card.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        employee_card.grid_rowconfigure(3, weight=1)
        employee_card.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(employee_card, text="2. Colaborador de destino", font=self.FONT_TITLE,
                     text_color=self.AZUL_ESCURO).grid(row=0, column=0, columnspan=2, padx=14, pady=(13, 2), sticky="w")
        self.index_info = ctk.CTkLabel(
            employee_card, text="Preparando indice...", font=self.FONT_SUB,
            text_color=self.TEXTO_SUAVE,
        )
        self.index_info.grid(row=1, column=0, columnspan=2, padx=14, pady=(0, 8), sticky="w")
        self.search_entry = self._entry(employee_card, placeholder_text="Digite parte do nome: LETICIA, ALICE...")
        self.search_entry.grid(row=2, column=0, columnspan=2, padx=14, pady=(0, 8), sticky="ew")
        self.search_entry.bind("<KeyRelease>", lambda _event: self.render_search())
        self.search_entry.bind("<Return>", lambda _event: self.select_first_employee())
        self.employee_tree = ttk.Treeview(
            employee_card, columns=("name", "source", "detail"), show="headings",
            style="Sonova.Treeview",
        )
        employee_headings = {"name": "Nome", "source": "Origem", "detail": "Local"}
        self.employee_tree.column("name", width=250, anchor="w", stretch=True)
        self.employee_tree.column("source", width=135, anchor="w", stretch=False)
        self.employee_tree.column("detail", width=180, anchor="w", stretch=True)
        attach_scrollbar(employee_card, self.employee_tree, row=3, column=0)
        self.employee_tree.bind("<<TreeviewSelect>>", self.on_employee_selection)
        self.employee_sorter = SortableTreeview(self.employee_tree, employee_headings)
        self.employee_tree.bind("<Double-1>", lambda _event: self.copy_files())

        manual = ctk.CTkFrame(employee_card, fg_color="transparent")
        manual.grid(row=4, column=0, columnspan=2, padx=14, pady=(7, 6), sticky="ew")
        self.manual_entry = self._entry(manual, width=300, placeholder_text="Nome manual, se nao estiver na base")
        self.manual_entry.pack(side="left", fill="x", expand=True)
        self._button(manual, "Usar nome manual", self.use_manual_name, 145, True).pack(side="left", padx=(7, 0))

        self.selected_label = ctk.CTkLabel(
            employee_card, text="Nenhum colaborador selecionado", font=self.FONT_LABEL,
            text_color=self.ALERTA, anchor="w",
        )
        self.selected_label.grid(row=5, column=0, columnspan=2, padx=14, pady=(4, 4), sticky="ew")
        self.destination_label = ctk.CTkLabel(
            employee_card, text="Destino: —", font=self.FONT_SUB,
            text_color=self.TEXTO_SUAVE, anchor="w", wraplength=600, justify="left",
        )
        self.destination_label.grid(row=6, column=0, columnspan=2, padx=14, pady=(0, 8), sticky="ew")
        self.copy_button = ctk.CTkButton(
            employee_card, text="Copiar scanners para a admissao", command=self.copy_files,
            height=44, corner_radius=8, font=("Segoe UI", 12, "bold"),
            fg_color=self.AZUL, hover_color=self.AZUL_HOVER,
        )
        self.copy_button.grid(row=7, column=0, columnspan=2, padx=14, pady=(2, 14), sticky="ew")

        footer = ctk.CTkFrame(container, fg_color="transparent")
        footer.pack(fill="x", pady=(10, 0))
        self.status_label = ctk.CTkLabel(
            footer, text="Status: iniciando...", font=self.FONT_LABEL,
            text_color=self.AZUL_ESCURO, anchor="w",
        )
        self.status_label.pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(
            footer, text="Igarape Digital | copia segura + historico CSV",
            font=self.FONT_SUB, text_color=self.TEXTO_SUAVE,
        ).pack(side="right")

    def choose_base(self) -> None:
        path = filedialog.askdirectory(title="Escolha a pasta ADMISSOES\\ATIVOS")
        if path:
            self.base_entry.delete(0, "end")
            self.base_entry.insert(0, path)
            self._save_config(quiet=True)
            self.reindex()

    def choose_excel(self) -> None:
        path = filedialog.askdirectory(title="Escolha a pasta das planilhas SRA")
        if path:
            self.excel_entry.delete(0, "end")
            self.excel_entry.insert(0, path)
            self._save_config(quiet=True)
            self.reindex()

    def choose_files(self) -> None:
        names = filedialog.askopenfilenames(title="Selecione os scanners")
        if names:
            self.set_files([Path(name) for name in names])

    def set_files(self, paths: Iterable[Path]) -> None:
        unique: list[Path] = []
        seen: set[str] = set()
        for path in paths:
            item = Path(path)
            if not item.is_file():
                continue
            key = os.path.normcase(os.path.abspath(item))
            if key not in seen:
                seen.add(key)
                unique.append(item)
        self.files = unique
        valid_keys = {self._path_key(path) for path in self.files}
        self.previewed_files.intersection_update(valid_keys)
        self.files_to_copy.intersection_update(valid_keys)
        self.custom_names = {k: v for k, v in self.custom_names.items() if k in valid_keys}
        self.render_files()

    @staticmethod
    def _human_size(size: int) -> str:
        value = float(size)
        for unit in ("B", "KB", "MB", "GB"):
            if value < 1024 or unit == "GB":
                return f"{value:.0f} {unit}" if unit == "B" else f"{value:.1f} {unit}"
            value /= 1024
        return f"{size} B"

    @staticmethod
    def _size_sort_key(text: str) -> float:
        match = re.match(r"([\d.,]+)\s*(B|KB|MB|GB)", text.strip())
        if not match:
            return -1.0
        amount = float(match.group(1).replace(",", "."))
        multiplier = {"B": 1, "KB": 1024, "MB": 1024**2, "GB": 1024**3}[match.group(2)]
        return amount * multiplier

    def render_files(self) -> None:
        self.files_tree.delete(*self.files_tree.get_children())
        available_count = 0
        missing_count = 0
        for index, path in enumerate(self.files):
            exists = path.is_file()
            if exists:
                available_count += 1
            else:
                missing_count += 1
            try:
                size = self._human_size(path.stat().st_size)
            except OSError:
                size = "—"
            suffix = path.suffix.upper().lstrip(".") or "—"
            key = self._path_key(path)
            checked = key in self.files_to_copy
            renamed = key in self.custom_names
            display_name = self.custom_names.get(key, path.name)
            check_box = "☑" if checked else "☐"
            viewed = "Sim" if key in self.previewed_files else "Nao"
            status = "Disponivel" if exists else "Ausente"
            if not exists:
                tag = "missing"
            elif checked:
                tag = "checked"
            elif renamed:
                tag = "renamed"
            elif viewed == "Sim":
                tag = "viewed"
            else:
                tag = ""
            self.files_tree.insert(
                "", "end", iid=str(index),
                values=(check_box, display_name, suffix, size, viewed, status),
                tags=(tag,) if tag else (),
            )
        self.files_tree.tag_configure("viewed", foreground=self.SUCESSO)
        self.files_tree.tag_configure("checked", foreground=self.AZUL_ESCURO, background="#DCEEF8")
        self.files_tree.tag_configure("missing", foreground=self.ALERTA, background="#FCECE8")
        self.files_tree.tag_configure("renamed", foreground=self.AZUL_ESCURO, background="#EFF7FC")
        self.files_tree.tag_configure("zebra_par", background=self.CARD, foreground=self.TEXTO)
        self.files_tree.tag_configure("zebra_impar", background=self.FUNDO, foreground=self.TEXTO)
        self.files_sorter.reset_headers()
        self.files_sorter.apply_zebra()
        marked_count = sum(
            1 for path in self.files if self._path_key(path) in self.files_to_copy
        )
        if hasattr(self, "copy_count_label"):
            self.copy_count_label.configure(
                text=f"{marked_count} marcado(s)",
                text_color=self.SUCESSO if marked_count else self.ALERTA,
            )
        if missing_count:
            info_text = (
                f"{marked_count} marcado(s) | {available_count} disponivel(is) | "
                f"{missing_count} ausente(s). "
                "Os ausentes nao impedirao a copia dos demais."
            )
            info_color = self.AMBAR
        elif self.files:
            info_text = (
                f"{marked_count} marcado(s) para copiar | {available_count} disponivel(is)."
            )
            info_color = self.SUCESSO if marked_count else self.AMBAR
        else:
            info_text = "Use F8 no Windows Explorer ou adicione manualmente."
            info_color = self.TEXTO_SUAVE
        self.files_info.configure(
            text=info_text,
            text_color=info_color,
        )
        self.refresh_destination()

    @staticmethod
    def _path_key(path: Path) -> str:
        return os.path.normcase(os.path.abspath(path))

    def toggle_file_check(self, index: int) -> None:
        if not 0 <= index < len(self.files):
            return
        path = self.files[index]
        key = self._path_key(path)
        if key in self.files_to_copy:
            self.files_to_copy.remove(key)
        else:
            self.files_to_copy.add(key)
        self.render_files()
        if str(index) in self.files_tree.get_children():
            self.files_tree.selection_set(str(index))
            self.files_tree.focus(str(index))

    def on_file_tree_click(self, event) -> str | None:
        row = self.files_tree.identify_row(event.y)
        column = self.files_tree.identify_column(event.x)
        if row and column == "#1":
            self.toggle_file_check(int(row))
            return "break"
        return None

    def on_file_tree_double_click(self, event) -> str | None:
        column = self.files_tree.identify_column(event.x)
        if column == "#1":
            return "break"
        if column == "#2":
            self.rename_selected_file()
            return "break"
        self.preview_selected_file()
        return None

    def toggle_focused_files(self, _event=None) -> str:
        selected = self.files_tree.selection()
        if not selected:
            focused = self.files_tree.focus()
            selected = (focused,) if focused else ()
        for item in selected:
            index = int(item)
            if 0 <= index < len(self.files):
                key = self._path_key(self.files[index])
                if key in self.files_to_copy:
                    self.files_to_copy.remove(key)
                else:
                    self.files_to_copy.add(key)
        self.render_files()
        return "break"

    def mark_all_available(self) -> None:
        self.files_to_copy.clear()
        self.files_to_copy.update(
            self._path_key(path) for path in self.files if path.is_file()
        )
        self.render_files()

    def unmark_all_files(self) -> None:
        self.files_to_copy.clear()
        self.render_files()

    def preview_selected_file(self) -> None:
        if not self.files:
            messagebox.showwarning("Visualizar", "Adicione ou capture os scanners primeiro.")
            return
        selection = self.files_tree.selection()
        index = int(selection[0]) if selection else 0
        if not self.files[index].is_file():
            self.render_files()
            messagebox.showwarning(
                "Arquivo ausente",
                f"O arquivo nao esta mais disponivel na origem:\n\n{self.files[index]}\n\n"
                "Reconecte a unidade ou capture os scanners novamente.",
            )
            return
        suffix = self.files[index].suffix.lower()
        if suffix == ".pdf" and pymupdf is None:
            messagebox.showerror(
                "Visualizar PDF",
                f"Instale PyMuPDF para visualizar PDFs dentro do programa.\n\n{PYMUPDF_ERROR}",
            )
            return
        if Image is None:
            messagebox.showerror(
                "Visualizar",
                f"Instale Pillow para visualizar PDFs e imagens.\n\n{PIL_ERROR}",
            )
            return
        PreviewWindow(self.app, self.files, index, self.mark_previewed)

    def mark_previewed(self, path: Path) -> None:
        key = self._path_key(path)
        if key in self.previewed_files:
            return
        self.previewed_files.add(key)
        self.render_files()
        self.set_status(f"Visualizado: {path.name}", self.SUCESSO)

    def rename_selected_file(self) -> None:
        if not self.files:
            messagebox.showwarning("Renomear", "Adicione ou capture os scanners primeiro.")
            return
        selection = self.files_tree.selection()
        if selection:
            index = int(selection[0])
        else:
            focused = self.files_tree.focus()
            index = int(focused) if focused else 0
        if not 0 <= index < len(self.files):
            messagebox.showwarning("Renomear", "Selecione um arquivo na lista.")
            return
        self._open_rename_dialog(index)

    def _open_rename_dialog(self, index: int) -> None:
        path = self.files[index]
        key = self._path_key(path)
        suffix = path.suffix
        current_display = self.custom_names.get(key, path.name)
        current_stem = (
            current_display[: -len(suffix)]
            if suffix and current_display.lower().endswith(suffix.lower())
            else current_display
        )

        dialog = tk.Toplevel(self.app)
        dialog.title("Renomear arquivo")
        dialog.configure(bg=self.FUNDO)
        dialog.transient(self.app)
        dialog.resizable(False, False)
        dialog.geometry("+{}+{}".format(
            self.app.winfo_rootx() + 140, self.app.winfo_rooty() + 140
        ))

        tk.Label(
            dialog, text="Nome de destino (a extensao e mantida)", bg=self.FUNDO,
            fg=self.TEXTO, font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w", padx=16, pady=(16, 6))

        row = tk.Frame(dialog, bg=self.FUNDO)
        row.pack(fill="x", padx=16)
        entry = tk.Entry(
            row, font=("Segoe UI", 11), relief="flat", highlightthickness=1,
            highlightbackground=self.BORDA, highlightcolor=self.AZUL,
        )
        entry.insert(0, current_stem)
        entry.pack(side="left", fill="x", expand=True, ipady=6)
        tk.Label(
            row, text=suffix, bg=self.FUNDO, fg=self.TEXTO_SUAVE,
            font=("Segoe UI", 11, "bold"),
        ).pack(side="left", padx=(6, 0))

        tk.Label(
            dialog, text=f"Original: {path.name}", bg=self.FUNDO, fg=self.TEXTO_SUAVE,
            font=("Segoe UI", 9),
        ).pack(anchor="w", padx=16, pady=(8, 0))

        buttons = tk.Frame(dialog, bg=self.FUNDO)
        buttons.pack(fill="x", padx=16, pady=16)

        def confirm(_event=None) -> None:
            new_stem = safe_name(entry.get())
            if not new_stem:
                messagebox.showwarning("Renomear", "Digite um nome valido.", parent=dialog)
                return
            new_name = f"{new_stem}{suffix}"
            if new_name == path.name:
                self.custom_names.pop(key, None)
            else:
                self.custom_names[key] = new_name
            dialog.destroy()
            self.render_files()
            self.set_status(f"Renomeado para: {new_name}", self.SUCESSO)

        def clear_rename() -> None:
            self.custom_names.pop(key, None)
            dialog.destroy()
            self.render_files()

        def cancel() -> None:
            dialog.destroy()

        tk.Button(
            buttons, text="Restaurar original", command=clear_rename,
            bg="white", fg=self.TEXTO, relief="flat", font=("Segoe UI", 10),
            padx=10, pady=6, activebackground="#EDEDED",
        ).pack(side="left")
        tk.Button(
            buttons, text="Cancelar", command=cancel,
            bg="white", fg=self.TEXTO, relief="flat", font=("Segoe UI", 10),
            padx=10, pady=6, activebackground="#EDEDED",
        ).pack(side="right", padx=(0, 6))
        tk.Button(
            buttons, text="Renomear", command=confirm,
            bg=self.AZUL, fg="white", relief="flat", font=("Segoe UI", 10, "bold"),
            padx=14, pady=6, activebackground=self.AZUL_HOVER, activeforeground="white",
        ).pack(side="right")

        entry.bind("<Return>", confirm)
        dialog.bind("<Escape>", lambda _event: cancel())
        dialog.grab_set()
        entry.focus_set()
        entry.select_range(0, "end")
        entry.icursor("end")

    def remove_selected_files(self) -> None:
        indexes = sorted((int(item) for item in self.files_tree.selection()), reverse=True)
        for index in indexes:
            if 0 <= index < len(self.files):
                key = self._path_key(self.files[index])
                self.files_to_copy.discard(key)
                self.previewed_files.discard(key)
                self.custom_names.pop(key, None)
                self.files.pop(index)
        self.render_files()

    def clear_files(self) -> None:
        self.files = []
        self.previewed_files.clear()
        self.files_to_copy.clear()
        self.custom_names.clear()
        self.render_files()

    def reindex(self) -> None:
        if self.indexing:
            return
        self._save_config(quiet=True)
        self.indexing = True
        self.index_info.configure(text="Lendo planilhas e pastas...", text_color=self.AMBAR)
        self.set_status("Montando indice de colaboradores...", self.AMBAR)
        base = Path(self.base_entry.get().strip())
        excel = Path(self.excel_entry.get().strip())
        recursive = bool(self.recursive_var.get())

        def progress(text: str) -> None:
            self.app.after(0, lambda value=text: self.index_info.configure(text=value))

        def worker() -> None:
            try:
                rows, warnings, read_count = build_employee_index(base, excel, recursive, progress)
                self.app.after(0, lambda: self.finish_reindex(rows, warnings, read_count))
            except Exception as exc:
                self.app.after(0, lambda err=exc: self.fail_reindex(err))

        threading.Thread(target=worker, daemon=True, name="employee-index").start()

    def finish_reindex(self, rows: list[Employee], warnings: list[str], read_count: int) -> None:
        self.indexing = False
        self.employees = rows
        self.render_search()
        self.index_info.configure(
            text=f"{len(rows)} colaborador(es) | {read_count} planilha(s) lida(s)",
            text_color=self.SUCESSO if rows else self.ALERTA,
        )
        if warnings:
            self.set_status(f"Indice pronto com {len(warnings)} aviso(s).", self.AMBAR)
        else:
            self.set_status("Indice de colaboradores pronto.", self.SUCESSO)

    def fail_reindex(self, exc: Exception) -> None:
        self.indexing = False
        self.index_info.configure(text="Falha ao montar o indice.", text_color=self.ALERTA)
        self.set_status(f"Falha no indice: {exc}", self.ALERTA)

    def render_search(self) -> None:
        term = self.search_entry.get() if hasattr(self, "search_entry") else ""
        self.visible_employees = search_employees(self.employees, term)
        self.employee_tree.delete(*self.employee_tree.get_children())
        for index, employee in enumerate(self.visible_employees):
            self.employee_tree.insert(
                "", "end", iid=str(index),
                values=(employee.name, employee.source, employee.detail),
            )
        self.employee_tree.tag_configure("zebra_par", background=self.CARD, foreground=self.TEXTO)
        self.employee_tree.tag_configure("zebra_impar", background=self.FUNDO, foreground=self.TEXTO)
        self.employee_sorter.reset_headers()
        self.employee_sorter.apply_zebra()

    def select_first_employee(self) -> None:
        children = self.employee_tree.get_children()
        if children:
            self.employee_tree.selection_set(children[0])
            self.employee_tree.focus(children[0])
            self.on_employee_selection()

    def on_employee_selection(self, _event=None) -> None:
        selection = self.employee_tree.selection()
        if not selection:
            return
        index = int(selection[0])
        if 0 <= index < len(self.visible_employees):
            self.selected_employee = self.visible_employees[index]
            self.selected_label.configure(
                text=f"Selecionado: {self.selected_employee.name}", text_color=self.SUCESSO
            )
            self.refresh_destination()

    def use_manual_name(self) -> None:
        name = safe_name(self.manual_entry.get())
        if not name:
            messagebox.showwarning("Nome manual", "Digite o nome completo do colaborador.")
            return
        if not messagebox.askyesno(
            "Confirmar nome manual",
            f"Usar o nome abaixo e criar a pasta, se necessario?\n\n{name}",
        ):
            return
        self.selected_employee = Employee(name, normalized(name), "Digitado manualmente", "Nova pasta")
        self.selected_label.configure(text=f"Selecionado: {name}", text_color=self.SUCESSO)
        self.refresh_destination()

    def destination_path(self) -> Path | None:
        if self.selected_employee is None:
            return None
        base = Path(self.base_entry.get().strip())
        person = base / safe_name(self.selected_employee.name)
        destination = self.destination_combo.get()
        return person if destination == ROOT_DESTINATION else person / safe_name(destination)

    def refresh_destination(self) -> None:
        destination = self.destination_path()
        self.destination_label.configure(text=f"Destino: {destination}" if destination else "Destino: —")

    def activate_hotkey(self) -> None:
        if self.hotkey_on:
            return
        if pynput_keyboard is None:
            self.hotkey_label.configure(text="F8 local", fg_color="#646464")
            self.set_status(f"Hotkey global indisponivel: {PYNPUT_ERROR}", self.AMBAR)
            return
        try:
            self.hotkey_listener = pynput_keyboard.GlobalHotKeys(
                {"<f8>": lambda: self.app.after(0, self.handle_f8)}
            )
            self.hotkey_listener.start()
            self.hotkey_on = True
            self.hotkey_label.configure(text="F8 global ativo", fg_color=self.SUCESSO)
            self.set_status("F8 global ativo. Pronto para capturar scanners.", self.SUCESSO)
        except Exception as exc:
            self.hotkey_label.configure(text="F8 local", fg_color="#646464")
            self.set_status(f"Nao foi possivel ativar F8 global: {exc}", self.ALERTA)

    def bring_to_front(self) -> None:
        self.app.deiconify()
        self.app.lift()
        self.app.attributes("-topmost", True)
        self.app.after(700, lambda: self.app.attributes("-topmost", False))

    def handle_f8(self) -> None:
        now = time.monotonic()
        if now - self.last_f8 < 0.7:
            return
        self.last_f8 = now
        try:
            paths = selected_explorer_files()
        except Exception as exc:
            paths = []
            self.set_status(f"Falha ao ler a selecao do Explorer: {exc}", self.ALERTA)
        if not paths:
            self.bring_to_front()
            messagebox.showwarning(
                "Nenhum scanner selecionado",
                "Selecione um ou mais arquivos no Windows Explorer e pressione F8.\n\n"
                "Alternativa: use o botao 'Adicionar arquivos'.",
            )
            return
        self.set_files(paths)
        self.bring_to_front()
        if self.selected_employee is None:
            self.set_status("Scanners capturados. Pesquise e selecione o colaborador.", self.AMBAR)
            self.search_entry.focus_set()
            return
        self.copy_files()

    def _copy_plan(
        self, destination: Path, sources: Iterable[Path] | None = None
    ) -> list[tuple[Path, Path]]:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        plan: list[tuple[Path, Path]] = []
        reserved: set[str] = set()
        for source in (list(sources) if sources is not None else self.files):
            display_name = self.custom_names.get(self._path_key(source), source.name)
            original = safe_name(display_name) or "Scanner"
            filename = f"{stamp} - {original}" if self.timestamp_var.get() else original
            target = destination / filename
            base_target = target
            counter = 1
            while target.exists() or os.path.normcase(str(target)) in reserved:
                target = base_target.with_name(f"{base_target.stem} ({counter}){base_target.suffix}")
                counter += 1
            reserved.add(os.path.normcase(str(target)))
            plan.append((source, target))
        return plan

    def copy_files(self) -> None:
        if not self.files:
            messagebox.showwarning("Arquivos", "Selecione os scanners no Explorer e pressione F8.")
            return
        if self.selected_employee is None:
            messagebox.showwarning("Colaborador", "Pesquise e selecione o colaborador de destino.")
            self.search_entry.focus_set()
            return
        chosen = [
            path for path in self.files
            if self._path_key(path) in self.files_to_copy
        ]
        if not chosen:
            messagebox.showwarning(
                "Nenhum arquivo marcado",
                "Marque com ☑ os arquivos que deseja copiar.\n\n"
                "A linha azul serve apenas para selecionar e visualizar.",
            )
            return
        self._save_config(quiet=True)
        base = Path(self.base_entry.get().strip())
        destination = self.destination_path()
        if destination is None:
            return
        if not base.is_dir():
            messagebox.showerror("Pasta base", f"A pasta de admissoes nao foi encontrada:\n\n{base}")
            return
        if not is_inside(destination, base):
            messagebox.showerror("Destino invalido", "O destino calculado ficou fora da pasta de admissoes.")
            return
        available = [path for path in chosen if path.is_file()]
        missing = [path for path in chosen if not path.is_file()]
        if not available:
            self.render_files()
            messagebox.showerror(
                "Arquivos indisponiveis",
                "Nenhum dos arquivos marcados continua disponivel na origem.\n\n"
                "Reconecte a unidade ou capture os scanners novamente.",
            )
            return
        if missing:
            missing_preview = "\n".join(f"• {path.name}" for path in missing[:10])
            if len(missing) > 10:
                missing_preview += f"\n• ... e mais {len(missing) - 10}"
            if not messagebox.askyesno(
                "Arquivos ausentes",
                f"{len(missing)} arquivo(s) nao existe(m) mais na origem:\n\n"
                f"{missing_preview}\n\n"
                f"Deseja continuar com os {len(available)} arquivo(s) disponivel(is)?",
            ):
                self.render_files()
                self.set_status("Copia cancelada para revisar arquivos ausentes.", self.AMBAR)
                return
            for path in missing:
                self._append_history("ERRO", path, destination, "Arquivo nao existe mais na origem")

        plan = self._copy_plan(destination, available)
        preview = "\n".join(f"• {target.name}" for _source, target in plan[:10])
        if len(plan) > 10:
            preview += f"\n• ... e mais {len(plan) - 10}"
        if self.confirm_var.get() and not messagebox.askyesno(
            "Confirmar copia",
            f"Colaborador: {self.selected_employee.name}\n\n"
            f"Destino:\n{destination}\n\n"
            f"Arquivos ({len(plan)}):\n{preview}\n\n"
            "Os originais serao mantidos. Confirmar?",
        ):
            self.set_status("Copia cancelada pelo usuario.", self.AMBAR)
            return

        copied: list[tuple[Path, Path]] = []
        failed: list[tuple[Path, str]] = []
        try:
            destination.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            messagebox.showerror("Pasta de destino", f"Nao foi possivel criar o destino:\n\n{exc}")
            self.set_status(f"Falha ao criar o destino: {exc}", self.ALERTA)
            return

        for source, target in plan:
            try:
                shutil.copy2(source, target)
                copied.append((source, target))
                self._append_history("OK", source, target, "")
            except Exception as exc:
                failed.append((source, str(exc)))
                self._append_history("ERRO", source, target, str(exc))

        unresolved: list[Path] = []
        unresolved_keys: set[str] = set()
        for path in missing + [source for source, _error in failed]:
            key = self._path_key(path)
            if key not in unresolved_keys:
                unresolved_keys.add(key)
                unresolved.append(path)

        copied_keys = {self._path_key(source) for source, _target in copied}
        self.files = [
            path for path in self.files if self._path_key(path) not in copied_keys
        ]
        remaining_keys = {self._path_key(path) for path in self.files}
        self.files_to_copy.intersection_update(remaining_keys)
        self.previewed_files.intersection_update(remaining_keys)
        self.custom_names = {k: v for k, v in self.custom_names.items() if k in remaining_keys}

        if unresolved:
            self.render_files()
            failure_preview = "\n".join(f"• {path.name}" for path in unresolved[:10])
            if len(unresolved) > 10:
                failure_preview += f"\n• ... e mais {len(unresolved) - 10}"
            self.set_status(
                f"Copia parcial: {len(copied)} copiado(s) | {len(unresolved)} pendente(s).",
                self.AMBAR,
            )
            messagebox.showwarning(
                "Copia concluida com pendencias",
                f"{len(copied)} arquivo(s) copiado(s) com sucesso.\n\n"
                f"{len(unresolved)} arquivo(s) permaneceu(ram) na lista para nova tentativa:\n\n"
                f"{failure_preview}",
            )
            return

        self.set_status(
            f"Concluido: {len(copied)} scanner(s) copiado(s) para {self.selected_employee.name}.",
            self.SUCESSO,
        )
        messagebox.showinfo(
            "Concluido",
            f"{len(copied)} arquivo(s) copiado(s) com sucesso.\n\n"
            f"Colaborador: {self.selected_employee.name}\n\nDestino:\n{destination}",
        )
        self.render_files()

    def _append_history(self, status: str, source: Path, target: Path, error: str) -> None:
        try:
            HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
            new_file = not HISTORY_PATH.exists()
            with HISTORY_PATH.open("a", newline="", encoding="utf-8-sig") as handle:
                writer = csv.writer(handle, delimiter=";")
                if new_file:
                    writer.writerow([
                        "DATA_HORA", "STATUS", "COLABORADOR", "ORIGEM",
                        "DESTINO", "ERRO",
                    ])
                writer.writerow([
                    datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    status,
                    self.selected_employee.name if self.selected_employee else "",
                    str(source),
                    str(target),
                    error,
                ])
        except Exception:
            pass

    def set_status(self, text: str, color: str | None = None) -> None:
        if hasattr(self, "status_label"):
            self.status_label.configure(text=f"Status: {text}", text_color=color or self.AZUL_ESCURO)

    def close(self) -> None:
        self._save_config(quiet=True)
        try:
            if self.hotkey_listener is not None:
                self.hotkey_listener.stop()
        except Exception:
            pass
        self.app.destroy()

    def run(self) -> None:
        self.app.mainloop()


def validate_runtime() -> None:
    missing: list[str] = []
    if load_workbook is None:
        missing.append(f"openpyxl ({OPENPYXL_ERROR})")
    if os.name == "nt" and pythoncom is None:
        missing.append(f"pywin32 ({PYWIN32_ERROR})")
    if missing:
        raise RuntimeError("Dependencias obrigatorias ausentes: " + ", ".join(missing))


if __name__ == "__main__":
    validate_runtime()
    ScannerAdmissionApp().run()
