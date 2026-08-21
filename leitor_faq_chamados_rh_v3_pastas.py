import os
import re
import json
import time
import difflib
import threading
import traceback
from pathlib import Path
from datetime import datetime, date

import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk

import pythoncom
import win32com.client as win32

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


APP_NAME = "Leitor Outlook RH - FAQ de Chamados"
APP_VERSION = "2.1.0"

BASE_DIR = Path(r"C:\_RPA\AppEmailRH")
SAIDA_DIR = BASE_DIR / "saida"
LOGS_DIR = BASE_DIR / "logs"
CONFIG_DIR = BASE_DIR / "config"
CONFIG_FILE = CONFIG_DIR / "config_email_rh.json"

DEFAULT_EXCEL_NAME = "faq_chamados_gente_gestao.xlsx"
DEFAULT_HTML_NAME = "faq_chamados_gente_gestao.html"
DEFAULT_PASTA_OUTLOOK = "aCHAMADOS"

CORES = {
    "navy": "#121C4E",
    "azul": "#0083CA",
    "azul_apoio": "#003C64",
    "fundo": "#F5F7FA",
    "card": "#FFFFFF",
    "borda": "#D9DDE3",
    "texto": "#1F2937",
    "texto_sec": "#6B7280",
    "sucesso": "#2E7D32",
    "erro": "#B91C1C",
    "alerta": "#B7791F",
    "branco": "#FFFFFF",
}


COLUNAS_BASE = [
    "EntryID",
    "Data Recebimento",
    "Hora Recebimento",
    "Dias Desde Recebimento",
    "Remetente Nome",
    "Remetente E-mail",
    "Para",
    "CC",
    "Assunto",
    "Corpo do E-mail",
    "Corpo Limpo",
    "Texto Truncado",
    "Tem Anexo",
    "Quantidade de Anexos",
    "Nomes dos Anexos",
    "Status Outlook",
    "Lido",
    "Importância Outlook",
    "Tamanho do Texto",
    "Data Extração",
    "Análise ChatGPT",
    "Urgência Sugerida",
    "Motivo da Urgência",
    "Categoria Sugerida",
    "Ação Recomendada",
    "Responsável Sugerido",
    "Observação RH",
    "Status Atendimento",
]


COLUNAS_GLOSSARIO = [
    "Palavra-chave",
    "Categoria",
    "Prioridade",
    "Ação sugerida",
    "Responsável",
    "Observação",
    "Ativo",
]


# ---------------------------------------------------------------------------
# Glossário de temas de RH usado para classificar chamados e organizar o FAQ.
# Cada tema tem uma lista de palavras-chave (busca por substring, sem acento,
# minúsculas). A ordem importa: o primeiro tema que casar vence, então os
# temas mais específicos vêm antes dos mais genéricos.
# ---------------------------------------------------------------------------
TEMAS_RH = [
    ("Férias", [
        "ferias", "abono pecuniario", "venda de ferias", "programacao de ferias",
        "aviso de ferias", "retorno de ferias", "13 na ferias", "1/3",
    ]),
    ("Holerite / Contracheque", [
        "holerite", "contracheque", "contra cheque", "demonstrativo de pagamento",
        "recibo de pagamento", "espelho de pagamento",
    ]),
    ("Ponto / Frequência", [
        "ponto", "batida", "espelho de ponto", "banco de horas", "hora extra",
        "horas extras", "frequencia", "abono de falta", "marcacao",
    ]),
    ("Vale-Transporte", [
        "vale transporte", "vale-transporte", "vt ", " vt", "bilhete unico",
        "recarga de transporte", "passagem",
    ]),
    ("Vale-Refeição / Alimentação", [
        "vale refeicao", "vale-refeicao", "vale alimentacao", "vale-alimentacao",
        "vr ", " vr", "va ", " va", "flash", "alelo", "ticket", "sodexo", "cartao refeicao",
    ]),
    ("Plano de Saúde / Odonto", [
        "plano de saude", "convenio", "sulamerica", "sul america", "amil",
        "unimed", "reembolso medico", "carteirinha", "odonto", "dependente no plano",
    ]),
    ("Admissão", [
        "admissao", "documentos de admissao", "contratacao", "primeiro dia",
        "exame admissional", "aso admissional", "integracao",
    ]),
    ("Desligamento / Rescisão", [
        "desligamento", "rescisao", "demissao", "aviso previo", "homologacao",
        "termo de rescisao", "trct", "acerto de contas", "seguro desemprego",
    ]),
    ("Atestado / Afastamento", [
        "atestado", "afastamento", "inss", "auxilio doenca", "licenca medica",
        "cid", "pericia", "acidente de trabalho", "cat ",
    ]),
    ("Licença-Maternidade / Paternidade", [
        "licenca maternidade", "licenca-maternidade", "salario maternidade",
        "licenca paternidade", "licenca-paternidade",
    ]),
    ("Adiantamento / Empréstimo", [
        "adiantamento", "vale", "emprestimo consignado", "consignado", "antecipacao",
    ]),
    ("Informe de Rendimentos / IRPF", [
        "informe de rendimentos", "imposto de renda", "irpf", "declaracao de imposto",
    ]),
    ("Benefícios (geral)", [
        "beneficio", "beneficios", "gympass", "totalpass", "seguro de vida",
        "auxilio creche", "cesta",
    ]),
    ("Documentos / Declarações", [
        "declaracao", "carta", "comprovante de vinculo", "carteira de trabalho",
        "ctps", "pis", "documento", "2 via", "segunda via",
    ]),
    ("Dados Cadastrais", [
        "atualizacao cadastral", "mudanca de endereco", "dados bancarios",
        "conta salario", "alteracao de dados", "novo telefone",
    ]),
    ("PLR / Remuneração Variável", [
        "plr", "participacao nos lucros", "remuneracao variavel", "comissao",
        "premiacao", "bonus",
    ]),
    ("Treinamento / Desenvolvimento", [
        "treinamento", "curso", "capacitacao", "trilha", "certificado",
    ]),
    ("Sistema / Acesso", [
        "acesso", "senha", "login", "portal do colaborador", "aplicativo",
        "sistema fora do ar", "nao consigo acessar", "adp",
    ]),
]

TEMA_PADRAO = "Outros / Não classificado"


COLUNAS_CHAMADOS = [
    "ID Chamado",
    "Tema",
    "Assunto Original",
    "Pergunta (colaborador)",
    "Resposta (RH)",
    "Status Resposta",
    "Colaborador",
    "E-mail Colaborador",
    "Respondido por",
    "Data Abertura",
    "Data Resposta",
    "Tempo de Resposta (h)",
    "Qtd. Mensagens na Thread",
    "Tem Anexo",
]


COLUNAS_FAQ = [
    "Nº",
    "Tema",
    "Pergunta Frequente",
    "Resposta Sugerida",
    "Frequência",
    "% do Total",
    "Status Resposta",
    "Exemplos de Assunto",
    "Primeira Ocorrência",
    "Última Ocorrência",
]


COLUNAS_TEMAS = [
    "Tema",
    "Qtd. Chamados",
    "% do Total",
    "Respondidos",
    "Sem Resposta",
    "Tempo Médio de Resposta (h)",
]


def garantir_pastas():
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    SAIDA_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)


def agora_texto():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def hoje_data():
    return datetime.now().date()


def limpar_texto(texto):
    if texto is None:
        return ""

    texto = str(texto)
    texto = texto.replace("\r", "\n")
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    texto = re.sub(r"[ \t]+", " ", texto)
    texto = texto.replace("\x00", "")
    return texto.strip()


def limpar_texto_uma_linha(texto):
    texto = limpar_texto(texto)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def limitar_excel(texto, limite=32000):
    texto = "" if texto is None else str(texto)
    if len(texto) <= limite:
        return texto, "Não"
    return texto[:limite] + "\n\n[TEXTO TRUNCADO PARA COMPATIBILIDADE COM EXCEL]", "Sim"


def normalizar_data_outlook(valor):
    if valor is None:
        return None

    try:
        if hasattr(valor, "date"):
            return valor
    except Exception:
        pass

    try:
        return datetime.fromtimestamp(time.mktime(valor.timetuple()))
    except Exception:
        return None


def dias_desde_recebimento(recebido):
    if not recebido:
        return ""

    try:
        data_email = recebido.date()
        return (hoje_data() - data_email).days
    except Exception:
        return ""


def traduzir_importancia(valor):
    try:
        valor = int(valor)
    except Exception:
        return ""

    mapa = {
        0: "Baixa",
        1: "Normal",
        2: "Alta",
    }
    return mapa.get(valor, str(valor))


# ---------------------------------------------------------------------------
# Motor de FAQ: classificação por tema, normalização de assunto,
# isolamento da mensagem do topo (sem histórico citado) e identificação
# de quem é o RH dentro da thread.
# ---------------------------------------------------------------------------

_ACENTOS = str.maketrans(
    "áàâãäéèêëíìîïóòôõöúùûüçÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇ",
    "aaaaaeeeeiiiiooooouuuucAAAAAEEEEIIIIOOOOOUUUUC",
)


def remover_acentos(texto):
    if texto is None:
        return ""
    return str(texto).translate(_ACENTOS)


def texto_normalizado(texto):
    """Minúsculas, sem acento, espaços colapsados. Base para comparação."""
    t = remover_acentos(texto).lower()
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def classificar_tema(texto):
    """Retorna o primeiro tema de RH cujas palavras-chave aparecem no texto."""
    base = texto_normalizado(texto)
    if not base:
        return TEMA_PADRAO

    for tema, chaves in TEMAS_RH:
        for chave in chaves:
            if chave in base:
                return tema
    return TEMA_PADRAO


# Prefixos de resposta/encaminhamento em PT e EN, para normalizar o assunto.
_PREFIXO_ASSUNTO = re.compile(
    r"^\s*(re|res|enc|ens|fw|fwd|encaminhar|encaminhada|encaminhado|"
    r"em resposta a|automatic reply|resposta automatica)\s*:\s*",
    re.IGNORECASE,
)

# Padrões de protocolo/chamado que poluem o agrupamento (ex.: [#12345], (Ticket 987)).
_PROTOCOLO = re.compile(r"[\[\(]?\s*(#|n[ºo°.]?\s*|ticket|chamado|protocolo)\s*[:#]?\s*\d+\s*[\]\)]?",
                        re.IGNORECASE)


def normalizar_assunto(assunto):
    """
    Remove RE:/ENC:/FW: repetidos, números de protocolo e espaços,
    devolvendo o "assunto-raiz" usado para agrupar chamados iguais.
    """
    if not assunto:
        return ""

    txt = str(assunto).strip()

    # Remove prefixos repetidos (RE: RES: ENC: ...)
    anterior = None
    while anterior != txt:
        anterior = txt
        txt = _PREFIXO_ASSUNTO.sub("", txt).strip()

    txt = _PROTOCOLO.sub(" ", txt)
    txt = re.sub(r"\s+", " ", txt).strip(" -–—:")
    return txt


# Marcadores que indicam o início do histórico citado dentro de um e-mail.
_MARCADORES_CITACAO = [
    re.compile(r"^\s*De:\s", re.IGNORECASE | re.MULTILINE),
    re.compile(r"^\s*From:\s", re.IGNORECASE | re.MULTILINE),
    re.compile(r"^\s*Enviad[ao]( em)?:\s", re.IGNORECASE | re.MULTILINE),
    re.compile(r"^\s*Sent:\s", re.IGNORECASE | re.MULTILINE),
    re.compile(r"-{3,}\s*Mensagem original\s*-{3,}", re.IGNORECASE),
    re.compile(r"-{3,}\s*Original Message\s*-{3,}", re.IGNORECASE),
    re.compile(r"^\s*Em\s.+escreveu:\s*$", re.IGNORECASE | re.MULTILINE),
    re.compile(r"^\s*On\s.+wrote:\s*$", re.IGNORECASE | re.MULTILINE),
    re.compile(r"_{10,}"),
]


def isolar_mensagem_topo(corpo):
    """
    Devolve apenas o texto escrito na mensagem atual, cortando o histórico
    citado (o bloco 'De:/From:/Enviada em:' e afins que o Outlook empilha).
    """
    if not corpo:
        return ""

    texto = str(corpo)
    corte = len(texto)

    for padrao in _MARCADORES_CITACAO:
        m = padrao.search(texto)
        if m and m.start() < corte:
            corte = m.start()

    topo = texto[:corte].strip()
    # Se o corte deixou algo muito curto (ex.: só uma saudação), mantém tudo.
    if len(topo) < 3:
        return limpar_texto(texto)
    return limpar_texto(topo)


def pertence_ao_rh(email, nome, dominios_rh, emails_rh):
    """
    Decide se um remetente faz parte do time de RH (quem responde os chamados),
    com base nos e-mails e domínios informados na tela. Sem lista, retorna None
    (indefinido) e o pareamento cai para a heurística de ordem/remetente.
    """
    if not dominios_rh and not emails_rh:
        return None

    email_n = (email or "").strip().lower()
    nome_n = texto_normalizado(nome)

    for e in emails_rh:
        if e and e in email_n:
            return True

    for d in dominios_rh:
        if d and email_n.endswith(d):
            return True

    for e in emails_rh:
        # Permite informar um nome em vez de e-mail (ex.: "Jair", "Sabrina").
        if e and "@" not in e and e in nome_n:
            return True

    return False


def parse_dominios_emails(texto):
    """
    Interpreta o campo 'quem responde os chamados' da tela.
    Aceita e-mails, domínios (@empresa.com) e nomes, separados por ; , ou espaço.
    Retorna (dominios, emails_ou_nomes).
    """
    if not texto:
        return [], []

    partes = re.split(r"[;,\s]+", texto.strip())
    dominios = []
    emails = []
    for p in partes:
        p = p.strip().lower()
        if not p:
            continue
        if p.startswith("@"):
            dominios.append(p)
        elif p.startswith("*@"):
            dominios.append(p[1:])
        else:
            emails.append(p)
    return dominios, emails


def carregar_config():
    garantir_pastas()

    config_padrao = {
        "modo_caixa": "principal",
        "caixa_compartilhada": "",
        "pasta_outlook": DEFAULT_PASTA_OUTLOOK,
        "quantidade": 500,
        "ultimos_dias": 90,
        "somente_nao_lidos": False,
        "salvar_msg": False,
        "emails_rh": "",
        "pasta_saida": str(SAIDA_DIR),
        "nome_excel": DEFAULT_EXCEL_NAME,
        "nome_html": DEFAULT_HTML_NAME,
    }

    if not CONFIG_FILE.exists():
        salvar_config(config_padrao)
        return config_padrao

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            config_lido = json.load(f)

        for chave, valor in config_padrao.items():
            config_lido.setdefault(chave, valor)

        return config_lido

    except Exception:
        return config_padrao


def salvar_config(config):
    garantir_pastas()
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=4)


def registrar_log_arquivo(mensagem):
    garantir_pastas()
    caminho_log = LOGS_DIR / "log_execucao.txt"
    with open(caminho_log, "a", encoding="utf-8") as f:
        f.write(f"[{agora_texto()}] {mensagem}\n")


def carregar_glossario(caminho_excel):
    """
    Função reservada para próxima iteração.

    Objetivo futuro:
    Ler a aba GLOSSARIO do Excel e usar as palavras-chave para sugerir
    categoria, prioridade, responsável e ação recomendada.

    Nesta primeira versão, a função não aplica classificação.
    """
    return []


def classificar_por_glossario(texto_email, glossario):
    """
    Função reservada para próxima iteração.

    Futuramente:
    - Receberá o texto do e-mail
    - Comparará com o glossário
    - Retornará categoria, prioridade, palavras encontradas,
      responsável e ação sugerida

    Nesta primeira versão:
    Retorna valores em branco ou Não classificado.
    """
    return {
        "analise_chatgpt": "",
        "urgencia": "",
        "motivo_urgencia": "",
        "categoria": "Não classificado",
        "acao_recomendada": "",
        "responsavel": "",
        "observacao": "",
        "status_atendimento": "Pendente",
    }


def get_smtp_address(mail):
    """
    Tenta obter o e-mail SMTP real do remetente.
    Funciona melhor em contas Exchange/Outlook corporativas.
    """
    try:
        if getattr(mail, "SenderEmailType", "") == "EX":
            try:
                exchange_user = mail.Sender.GetExchangeUser()
                if exchange_user:
                    smtp = exchange_user.PrimarySmtpAddress
                    if smtp:
                        return smtp
            except Exception:
                pass

            try:
                PR_SMTP_ADDRESS = "http://schemas.microsoft.com/mapi/proptag/0x39FE001E"
                smtp = mail.Sender.PropertyAccessor.GetProperty(PR_SMTP_ADDRESS)
                if smtp:
                    return smtp
            except Exception:
                pass

        return getattr(mail, "SenderEmailAddress", "") or ""

    except Exception:
        return ""


NOMES_INBOX = {"", "caixa de entrada", "inbox", "entrada"}


def _iter_subpastas(pasta):
    try:
        for sub in pasta.Folders:
            yield sub
    except Exception:
        return


def buscar_pasta_recursiva(pasta_raiz, nome_pasta, parcial=False):
    """
    Busca uma pasta pelo nome em toda a árvore a partir de pasta_raiz.
    Retorna a pasta encontrada ou None (não levanta exceção).
    Com parcial=True aceita correspondência por 'contém' (ex.: CHAMADOS em aCHAMADOS).
    Prioriza sempre a correspondência exata.
    """
    alvo = (nome_pasta or "").strip().lower()
    if not alvo:
        return None

    exatos = []
    parciais = []

    def caminhar(pasta):
        try:
            nome = pasta.Name.strip().lower()
        except Exception:
            nome = ""
        if nome == alvo:
            exatos.append(pasta)
        elif parcial and alvo in nome:
            parciais.append(pasta)
        for sub in _iter_subpastas(pasta):
            caminhar(sub)

    caminhar(pasta_raiz)

    if exatos:
        return exatos[0]
    if len(parciais) == 1:
        return parciais[0]
    return None


def caminho_pasta(pasta):
    """Caminho legível da pasta: Conta \\ Pasta \\ Subpasta."""
    partes = []
    atual = pasta
    for _ in range(12):
        try:
            partes.append(atual.Name)
            atual = atual.Parent
        except Exception:
            break
    return " \\ ".join(reversed(partes))


def listar_pastas_nivel1(base):
    nomes = []
    for sub in _iter_subpastas(base):
        try:
            nomes.append(sub.Name)
        except Exception:
            pass
    return nomes


def _nome_pasta_outlook(pasta):
    try:
        return str(pasta.Name or "").strip()
    except Exception:
        return ""


def _buscar_subpasta_exata(pasta, nome):
    """Busca uma subpasta direta pelo nome, sem varrer toda a árvore."""
    alvo = (nome or "").strip().casefold()
    if not alvo:
        return None

    try:
        return pasta.Folders.Item(nome)
    except Exception:
        pass

    for sub in _iter_subpastas(pasta):
        try:
            if _nome_pasta_outlook(sub).casefold() == alvo:
                return sub
        except Exception:
            continue
    return None


def resolver_pasta_por_caminho(outlook, caminho, raizes_extras=None):
    """
    Resolve um caminho completo selecionado na interface, por exemplo:
        Caixa Sonova \\ Caixa de Entrada \\ aCHAMADOS

    Retorna o objeto COM da pasta ou None.
    """
    caminho = (caminho or "").replace("/", "\\").strip(" \\")
    partes = [p.strip() for p in caminho.split("\\") if p.strip()]
    if not partes:
        return None

    raizes = []
    vistos = set()

    try:
        for raiz in outlook.Folders:
            nome = _nome_pasta_outlook(raiz)
            chave = nome.casefold()
            if nome and chave not in vistos:
                raizes.append(raiz)
                vistos.add(chave)
    except Exception:
        pass

    for raiz in (raizes_extras or []):
        nome = _nome_pasta_outlook(raiz)
        chave = nome.casefold()
        if nome and chave not in vistos:
            raizes.append(raiz)
            vistos.add(chave)

    raiz_escolhida = None
    for raiz in raizes:
        if _nome_pasta_outlook(raiz).casefold() == partes[0].casefold():
            raiz_escolhida = raiz
            break

    if raiz_escolhida is None:
        return None

    atual = raiz_escolhida
    for nome in partes[1:]:
        atual = _buscar_subpasta_exata(atual, nome)
        if atual is None:
            return None

    return atual


def listar_pastas_outlook_disponiveis(modo_caixa="principal", caixa_compartilhada="", log_callback=None):
    """
    Retorna a árvore de pastas visíveis no Outlook sem devolver objetos COM.
    Cada item contém nome, caminho completo, profundidade e raiz/store.

    A função inicializa COM internamente para poder ser usada em thread.
    """
    pythoncom.CoInitialize()
    try:
        outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
        modo_caixa = (modo_caixa or "principal").strip().lower()
        caixa_compartilhada = (caixa_compartilhada or "").strip()

        raizes = []
        vistos = set()

        try:
            for raiz in outlook.Folders:
                nome = _nome_pasta_outlook(raiz)
                chave = nome.casefold()
                if nome and chave not in vistos:
                    raizes.append(raiz)
                    vistos.add(chave)
        except Exception:
            pass

        if modo_caixa == "compartilhada" and caixa_compartilhada:
            try:
                recipient = outlook.CreateRecipient(caixa_compartilhada)
                recipient.Resolve()
                if recipient.Resolved:
                    inbox_comp = outlook.GetSharedDefaultFolder(recipient, 6)
                    raiz_comp = inbox_comp.Parent
                    nome = _nome_pasta_outlook(raiz_comp)
                    chave = nome.casefold()
                    if nome and chave not in vistos:
                        raizes.append(raiz_comp)
                        vistos.add(chave)
            except Exception as e:
                if log_callback:
                    log_callback(f"Não foi possível incluir a caixa compartilhada na lista: {e}")

        registros = []

        def caminhar(pasta, caminho_pai="", profundidade=0, store_nome=""):
            nome = _nome_pasta_outlook(pasta)
            if not nome:
                return

            caminho = f"{caminho_pai} \\ {nome}" if caminho_pai else nome
            registros.append({
                "nome": nome,
                "caminho": caminho,
                "profundidade": profundidade,
                "store": store_nome or nome,
            })

            for sub in _iter_subpastas(pasta):
                caminhar(sub, caminho, profundidade + 1, store_nome or nome)

        for raiz in raizes:
            caminhar(raiz)

        registros.sort(key=lambda x: (x["store"].casefold(), x["caminho"].casefold()))

        if log_callback:
            log_callback(f"Pastas do Outlook encontradas: {len(registros)}")

        return registros

    finally:
        pythoncom.CoUninitialize()


def obter_pasta_outlook(modo_caixa, caixa_compartilhada, pasta_outlook, log_callback=None):
    outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")

    modo_caixa = (modo_caixa or "principal").strip().lower()
    caixa_compartilhada = (caixa_compartilhada or "").strip()
    pasta_outlook = (pasta_outlook or "Caixa de Entrada").strip()

    raiz_store = None
    raizes_extras = []

    if modo_caixa == "compartilhada":
        if not caixa_compartilhada:
            raise Exception("Informe o nome ou e-mail da caixa compartilhada.")

        if log_callback:
            log_callback(f"Tentando acessar caixa compartilhada: {caixa_compartilhada}")

        recipient = outlook.CreateRecipient(caixa_compartilhada)
        recipient.Resolve()

        if not recipient.Resolved:
            raise Exception(f"Não foi possível resolver a caixa compartilhada: {caixa_compartilhada}")

        inbox = outlook.GetSharedDefaultFolder(recipient, 6)
        try:
            raiz_store = inbox.Parent
            raizes_extras.append(raiz_store)
        except Exception:
            raiz_store = None
    else:
        if log_callback:
            log_callback("Acessando as contas disponíveis no Outlook.")
        inbox = outlook.GetDefaultFolder(6)
        try:
            raiz_store = inbox.Parent
        except Exception:
            raiz_store = None

    if "\\" in pasta_outlook or "/" in pasta_outlook:
        pasta_final = resolver_pasta_por_caminho(outlook, pasta_outlook, raizes_extras=raizes_extras)
        if pasta_final is not None:
            if log_callback:
                try:
                    total = pasta_final.Items.Count
                    log_callback(f"Pasta selecionada pelo caminho: {pasta_outlook} ({total} itens).")
                except Exception:
                    log_callback(f"Pasta selecionada pelo caminho: {pasta_outlook}")
            return pasta_final

        raise Exception(
            f"O caminho selecionado não foi encontrado no Outlook: '{pasta_outlook}'. "
            "Clique em 'Ver pastas' e selecione novamente."
        )

    if pasta_outlook.lower() in NOMES_INBOX:
        if log_callback:
            log_callback("Pasta selecionada: Caixa de Entrada.")
        return inbox

    pasta_final = None

    for origem in [raiz_store, inbox]:
        if origem is None:
            continue
        pasta_final = buscar_pasta_recursiva(origem, pasta_outlook)
        if pasta_final is not None:
            break

    if pasta_final is None:
        try:
            for store_folder in outlook.Folders:
                pasta_final = buscar_pasta_recursiva(store_folder, pasta_outlook)
                if pasta_final is not None:
                    break
        except Exception:
            pass

    if pasta_final is None:
        base = raiz_store or inbox
        pasta_final = buscar_pasta_recursiva(base, pasta_outlook, parcial=True)

    if pasta_final is None:
        base = raiz_store or inbox
        disponiveis = listar_pastas_nivel1(base)
        if log_callback and disponiveis:
            log_callback("Pastas neste nível: " + " | ".join(disponiveis))
        raise Exception(
            f"Pasta não encontrada no Outlook: '{pasta_outlook}'. "
            "Use o botão 'Ver pastas' para escolher a pasta exata."
        )

    if log_callback:
        try:
            total = pasta_final.Items.Count
            log_callback(f"Pasta selecionada: {caminho_pasta(pasta_final)} ({total} itens).")
        except Exception:
            log_callback("Pasta selecionada com sucesso.")

    return pasta_final


def extrair_dados_email(mail):
    try:
        entry_id = getattr(mail, "EntryID", "") or ""
    except Exception:
        entry_id = ""

    try:
        recebido = normalizar_data_outlook(getattr(mail, "ReceivedTime", None))
    except Exception:
        recebido = None

    try:
        conversation_id = getattr(mail, "ConversationID", "") or ""
    except Exception:
        conversation_id = ""

    try:
        conversation_topic = getattr(mail, "ConversationTopic", "") or ""
    except Exception:
        conversation_topic = ""

    data_recebimento = ""
    hora_recebimento = ""

    if recebido:
        try:
            data_recebimento = recebido.strftime("%d/%m/%Y")
            hora_recebimento = recebido.strftime("%H:%M:%S")
        except Exception:
            pass

    try:
        remetente_nome = getattr(mail, "SenderName", "") or ""
    except Exception:
        remetente_nome = ""

    remetente_email = get_smtp_address(mail)

    try:
        para = getattr(mail, "To", "") or ""
    except Exception:
        para = ""

    try:
        cc = getattr(mail, "CC", "") or ""
    except Exception:
        cc = ""

    try:
        assunto = getattr(mail, "Subject", "") or ""
    except Exception:
        assunto = ""

    try:
        corpo = getattr(mail, "Body", "") or ""
    except Exception:
        corpo = ""

    corpo = limpar_texto(corpo)
    corpo_limpo = limpar_texto_uma_linha(corpo)

    corpo_excel, truncado = limitar_excel(corpo)
    corpo_limpo_excel, truncado_limpo = limitar_excel(corpo_limpo)

    texto_truncado = "Sim" if truncado == "Sim" or truncado_limpo == "Sim" else "Não"

    nomes_anexos = []
    qtd_anexos = 0

    try:
        qtd_anexos = int(mail.Attachments.Count)
        for i in range(1, qtd_anexos + 1):
            try:
                nomes_anexos.append(mail.Attachments.Item(i).FileName)
            except Exception:
                nomes_anexos.append("Anexo não identificado")
    except Exception:
        qtd_anexos = 0

    tem_anexo = "Sim" if qtd_anexos > 0 else "Não"

    try:
        nao_lido = bool(getattr(mail, "UnRead", False))
        lido = "Não" if nao_lido else "Sim"
    except Exception:
        lido = ""

    try:
        importancia = traduzir_importancia(getattr(mail, "Importance", ""))
    except Exception:
        importancia = ""

    tamanho_texto = len(corpo)

    glossario = []
    classificacao = classificar_por_glossario(
        f"{assunto}\n{corpo}",
        glossario
    )

    linha = {
        "EntryID": entry_id,
        "Data Recebimento": data_recebimento,
        "Hora Recebimento": hora_recebimento,
        "Dias Desde Recebimento": dias_desde_recebimento(recebido),
        "Remetente Nome": remetente_nome,
        "Remetente E-mail": remetente_email,
        "Para": para,
        "CC": cc,
        "Assunto": assunto,
        "Corpo do E-mail": corpo_excel,
        "Corpo Limpo": corpo_limpo_excel,
        "Texto Truncado": texto_truncado,
        "Tem Anexo": tem_anexo,
        "Quantidade de Anexos": qtd_anexos,
        "Nomes dos Anexos": " | ".join(nomes_anexos),
        "Status Outlook": "Não lido" if lido == "Não" else "Lido",
        "Lido": lido,
        "Importância Outlook": importancia,
        "Tamanho do Texto": tamanho_texto,
        "Data Extração": agora_texto(),
        "Análise ChatGPT": classificacao["analise_chatgpt"],
        "Urgência Sugerida": classificacao["urgencia"],
        "Motivo da Urgência": classificacao["motivo_urgencia"],
        "Categoria Sugerida": classificacao["categoria"],
        "Ação Recomendada": classificacao["acao_recomendada"],
        "Responsável Sugerido": classificacao["responsavel"],
        "Observação RH": classificacao["observacao"],
        "Status Atendimento": classificacao["status_atendimento"],
    }

    # Campos técnicos usados pelo motor de FAQ (não vão para a aba BASE_EMAILS,
    # pois não estão em COLUNAS_BASE). Prefixo "_" para deixar claro.
    linha["_conversation_id"] = conversation_id
    linha["_conversation_topic"] = conversation_topic
    linha["_assunto_normalizado"] = normalizar_assunto(assunto)
    linha["_dt"] = recebido
    linha["_corpo_original"] = corpo
    linha["_mensagem_topo"] = isolar_mensagem_topo(corpo)
    linha["_tema"] = classificar_tema(f"{assunto}\n{corpo}")

    return linha


# ---------------------------------------------------------------------------
# Salvar cópias dos e-mails em disco (.msg com anexos) e ler de volta de uma
# pasta local, para reprocessar o FAQ sem depender do Outlook/pasta original.
# ---------------------------------------------------------------------------

OL_SAVE_MSG_UNICODE = 9  # olMSGUnicode: .msg nativo, preserva anexos e formatação
OL_DISCARD = 1           # olDiscard: fecha o item aberto sem salvar alterações

_INVALIDOS_ARQ = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def sanitizar_nome(texto, limite=60):
    """Transforma texto em nome de arquivo seguro para Windows."""
    texto = (texto or "").strip()
    texto = _INVALIDOS_ARQ.sub(" ", texto)
    texto = re.sub(r"\s+", " ", texto).strip(" .")
    if not texto:
        texto = "sem_assunto"
    return texto[:limite].strip()


def salvar_email_msg(mail, destino_msg, destino_anexos, idx, salvar_anexos=True):
    """
    Salva o e-mail como .msg (formato nativo, com anexos embutidos) e,
    opcionalmente, extrai os anexos soltos numa subpasta. Retorna o caminho.
    """
    try:
        recebido = normalizar_data_outlook(getattr(mail, "ReceivedTime", None))
        data_txt = recebido.strftime("%Y-%m-%d_%H%M") if recebido else "sem_data"
    except Exception:
        data_txt = "sem_data"

    try:
        rem = getattr(mail, "SenderName", "") or ""
    except Exception:
        rem = ""

    try:
        assunto = getattr(mail, "Subject", "") or ""
    except Exception:
        assunto = ""

    nome = f"{idx:04d}_{data_txt}_{sanitizar_nome(rem, 30)}_{sanitizar_nome(assunto, 50)}.msg"
    caminho = os.path.join(destino_msg, nome)

    mail.SaveAs(caminho, OL_SAVE_MSG_UNICODE)

    if salvar_anexos and destino_anexos:
        try:
            qtd = int(mail.Attachments.Count)
            for i in range(1, qtd + 1):
                try:
                    att = mail.Attachments.Item(i)
                    nome_anexo = sanitizar_nome(att.FileName, 80) or f"anexo_{i}"
                    att.SaveAsFile(os.path.join(destino_anexos, f"{idx:04d}_{nome_anexo}"))
                except Exception:
                    continue
        except Exception:
            pass

    return caminho


def ler_emails_de_pasta_msg(
    caminho_pasta,
    quantidade,
    ultimos_dias,
    somente_nao_lidos,
    progress_callback=None,
    log_callback=None,
):
    """
    Lê os arquivos .msg de uma pasta local (reabrindo cada um via Outlook) e
    roda o mesmo pipeline de extração. Não toca na caixa/pasta original.
    """
    pythoncom.CoInitialize()

    try:
        outlook = win32.Dispatch("Outlook.Application")

        pasta = Path(caminho_pasta)
        arquivos = sorted(pasta.glob("*.msg"))

        if log_callback:
            log_callback(f"Arquivos .msg encontrados: {len(arquivos)}")

        registros = []
        erros = []

        quantidade = int(quantidade)
        ultimos_dias = int(ultimos_dias)
        coletados = 0
        total_alvo = min(quantidade, len(arquivos)) or 1

        for arq in arquivos:
            if coletados >= quantidade:
                break

            try:
                mail = outlook.Session.OpenSharedItem(str(arq))
            except Exception as e:
                erros.append(f"{arq.name}: {e}")
                continue

            try:
                try:
                    recebido = normalizar_data_outlook(getattr(mail, "ReceivedTime", None))
                except Exception:
                    recebido = None

                if ultimos_dias > 0 and recebido:
                    dias = dias_desde_recebimento(recebido)
                    if isinstance(dias, int) and dias > ultimos_dias:
                        continue

                if somente_nao_lidos:
                    try:
                        if not bool(getattr(mail, "UnRead", False)):
                            continue
                    except Exception:
                        continue

                dados = extrair_dados_email(mail)
                registros.append(dados)
                coletados += 1

                if progress_callback:
                    progress_callback(coletados, total_alvo)

                if log_callback and coletados % 10 == 0:
                    log_callback(f"E-mails lidos da pasta: {coletados}")

            except Exception as e:
                erros.append(str(e))
            finally:
                try:
                    mail.Close(OL_DISCARD)
                except Exception:
                    pass

        if log_callback:
            log_callback(f"Leitura da pasta local finalizada. E-mails lidos: {coletados}.")

        return registros, erros

    finally:
        pythoncom.CoUninitialize()


def ler_emails_outlook(
    modo_caixa,
    caixa_compartilhada,
    pasta_outlook,
    quantidade,
    ultimos_dias,
    somente_nao_lidos,
    progress_callback=None,
    log_callback=None,
    salvar_msg=False,
    destino_msg=None,
    destino_anexos=None,
    salvar_anexos=True,
):
    pythoncom.CoInitialize()

    try:
        pasta = obter_pasta_outlook(
            modo_caixa=modo_caixa,
            caixa_compartilhada=caixa_compartilhada,
            pasta_outlook=pasta_outlook,
            log_callback=log_callback,
        )

        if salvar_msg and destino_msg:
            Path(destino_msg).mkdir(parents=True, exist_ok=True)
            if destino_anexos:
                Path(destino_anexos).mkdir(parents=True, exist_ok=True)
            if log_callback:
                log_callback(f"Salvando cópias .msg em: {destino_msg}")

        items = pasta.Items
        items.Sort("[ReceivedTime]", True)

        total_outlook = items.Count

        if log_callback:
            log_callback(f"Total aparente de itens na pasta: {total_outlook}")

        registros = []
        erros = []

        quantidade = int(quantidade)
        ultimos_dias = int(ultimos_dias)

        limite_data = None
        if ultimos_dias > 0:
            limite_data = hoje_data()
            if log_callback:
                log_callback(f"Filtro de período ativado: últimos {ultimos_dias} dias.")

        analisados = 0
        coletados = 0

        for item in items:
            if coletados >= quantidade:
                break

            analisados += 1

            try:
                item_class = getattr(item, "Class", None)
                if item_class != 43:
                    continue

                try:
                    recebido = normalizar_data_outlook(getattr(item, "ReceivedTime", None))
                except Exception:
                    recebido = None

                if ultimos_dias > 0 and recebido:
                    dias = dias_desde_recebimento(recebido)
                    if isinstance(dias, int) and dias > ultimos_dias:
                        continue

                if somente_nao_lidos:
                    try:
                        if not bool(getattr(item, "UnRead", False)):
                            continue
                    except Exception:
                        continue

                dados = extrair_dados_email(item)
                registros.append(dados)
                coletados += 1

                if salvar_msg and destino_msg:
                    try:
                        salvar_email_msg(item, destino_msg, destino_anexos, coletados, salvar_anexos)
                    except Exception as e:
                        erros.append(f"Falha ao salvar .msg do item {coletados}: {e}")

                if progress_callback:
                    progress_callback(coletados, quantidade)

                if log_callback and coletados % 10 == 0:
                    log_callback(f"E-mails coletados: {coletados}")

            except Exception as e:
                erros.append(str(e))
                continue

        if log_callback:
            log_callback(f"Leitura finalizada. Itens analisados: {analisados}. E-mails coletados: {coletados}.")

        return registros, erros

    finally:
        pythoncom.CoUninitialize()


# ---------------------------------------------------------------------------
# Construção dos chamados (threads pareadas) e do FAQ consolidado.
# ---------------------------------------------------------------------------

def _chave_thread(registro):
    """Chave de agrupamento da thread: ConversationID, senão tópico/assunto normalizado."""
    cid = (registro.get("_conversation_id") or "").strip()
    if cid:
        return cid

    topico = registro.get("_conversation_topic") or ""
    if topico.strip():
        return "TOPICO::" + texto_normalizado(normalizar_assunto(topico))

    return "ASSUNTO::" + texto_normalizado(registro.get("_assunto_normalizado") or "")


def _dt_ordenacao(registro):
    dt = registro.get("_dt")
    return dt if dt is not None else datetime.min


def montar_chamados(registros, dominios_rh=None, emails_rh=None):
    """
    Agrupa os e-mails em threads e, dentro de cada thread, identifica a
    pergunta (primeira mensagem do colaborador) e a resposta (primeira
    mensagem do RH depois dela).
    """
    dominios_rh = dominios_rh or []
    emails_rh = emails_rh or []

    grupos = {}
    for r in registros:
        grupos.setdefault(_chave_thread(r), []).append(r)

    chamados = []
    contador = 0

    for chave, itens in grupos.items():
        itens.sort(key=_dt_ordenacao)

        # Marca cada mensagem como RH ou não (quando há lista informada).
        for it in itens:
            it["_eh_rh"] = pertence_ao_rh(
                it.get("Remetente E-mail", ""),
                it.get("Remetente Nome", ""),
                dominios_rh,
                emails_rh,
            )

        # Escolha da PERGUNTA.
        pergunta = None
        for it in itens:
            if it.get("_eh_rh") is True:
                continue  # começa por alguém que não seja do RH, se possível
            pergunta = it
            break
        if pergunta is None:
            pergunta = itens[0]

        idx_pergunta = itens.index(pergunta)
        remetente_pergunta = (pergunta.get("Remetente E-mail") or pergunta.get("Remetente Nome") or "").lower()

        # Escolha da RESPOSTA.
        resposta = None
        # 1) Preferência: primeira mensagem do RH depois da pergunta.
        for it in itens[idx_pergunta + 1:]:
            if it.get("_eh_rh") is True:
                resposta = it
                break
        # 2) Fallback: primeira mensagem de remetente diferente (troca de
        #    remetente). Cobre o caso sem lista de RH ou lista incompleta.
        if resposta is None:
            for it in itens[idx_pergunta + 1:]:
                rem = (it.get("Remetente E-mail") or it.get("Remetente Nome") or "").lower()
                if rem and rem != remetente_pergunta:
                    resposta = it
                    break

        contador += 1
        tema = pergunta.get("_tema") or TEMA_PADRAO

        dt_abertura = pergunta.get("_dt")
        dt_resposta = resposta.get("_dt") if resposta else None

        tempo_h = ""
        if dt_abertura and dt_resposta:
            try:
                delta = dt_resposta - dt_abertura
                tempo_h = round(delta.total_seconds() / 3600.0, 1)
            except Exception:
                tempo_h = ""

        tem_anexo = "Sim" if any(i.get("Tem Anexo") == "Sim" for i in itens) else "Não"

        chamado = {
            "ID Chamado": f"CH-{contador:04d}",
            "Tema": tema,
            "Assunto Original": pergunta.get("Assunto", ""),
            "Assunto Normalizado": pergunta.get("_assunto_normalizado", ""),
            "Pergunta (colaborador)": pergunta.get("_mensagem_topo", "") or pergunta.get("Corpo Limpo", ""),
            "Resposta (RH)": (resposta.get("_mensagem_topo", "") if resposta else ""),
            "Status Resposta": "Respondido" if resposta else "Sem resposta",
            "Colaborador": pergunta.get("Remetente Nome", ""),
            "E-mail Colaborador": pergunta.get("Remetente E-mail", ""),
            "Respondido por": (resposta.get("Remetente Nome", "") if resposta else ""),
            "Data Abertura": dt_abertura.strftime("%d/%m/%Y %H:%M") if dt_abertura else pergunta.get("Data Recebimento", ""),
            "Data Resposta": dt_resposta.strftime("%d/%m/%Y %H:%M") if dt_resposta else "",
            "Tempo de Resposta (h)": tempo_h,
            "Qtd. Mensagens na Thread": len(itens),
            "Tem Anexo": tem_anexo,
            "_dt_abertura": dt_abertura,
        }
        chamados.append(chamado)

    chamados.sort(key=lambda c: c.get("_dt_abertura") or datetime.min, reverse=True)
    return chamados


_STOPWORDS_FAQ = {
    "de", "do", "da", "dos", "das", "o", "a", "os", "as", "e", "em", "para",
    "pra", "por", "um", "uma", "no", "na", "meu", "minha", "como", "faco",
    "sobre", "referente", "solicitacao", "solicito", "gostaria", "preciso",
    "duvida", "ajuda", "favor", "prezados", "bom", "dia", "tarde", "noite",
}

_SINONIMOS_FAQ = [
    (re.compile(r"\b(segunda|2a|2\.?)\s*via\b"), "2via"),
    (re.compile(r"\bprimeira\s*via\b"), "1via"),
    (re.compile(r"\bcontra\s*cheque\b"), "contracheque"),
    (re.compile(r"\bvale\s*transporte\b"), "vt"),
    (re.compile(r"\bvale\s*(refeicao|alimentacao)\b"), "vr"),
    (re.compile(r"\bplano\s*de\s*saude\b"), "plano"),
]


def _chave_faq(assunto):
    """Chave agressiva para agrupar perguntas equivalentes: sem acento,
    sinônimos colapsados, sem stopwords, palavras ordenadas."""
    t = texto_normalizado(assunto)
    for padrao, repl in _SINONIMOS_FAQ:
        t = padrao.sub(repl, t)
    palavras = [p for p in re.split(r"[^a-z0-9]+", t) if p and p not in _STOPWORDS_FAQ]
    return " ".join(sorted(palavras))


def montar_faq(chamados, limiar=0.78):
    """
    Consolida os chamados em perguntas frequentes. Agrupa por tema e, dentro
    do tema, funde perguntas equivalentes por similaridade (difflib), para que
    variações do mesmo pedido virem uma única pergunta com a frequência somada.
    A resposta sugerida é a mais recente não vazia do grupo.
    """
    total = len(chamados) or 1

    # 1) separa por tema
    por_tema = {}
    for c in chamados:
        por_tema.setdefault(c.get("Tema", TEMA_PADRAO), []).append(c)

    # 2) dentro de cada tema, clustering guloso por similaridade da chave
    grupos = {}
    for tema, itens in por_tema.items():
        clusters = []  # lista de (chave_representante, [chamados])
        for c in itens:
            chave = _chave_faq(c.get("Assunto Normalizado", "") or c.get("Assunto Original", ""))
            destino = None
            melhor = limiar
            for i, (chave_rep, _lista) in enumerate(clusters):
                r = difflib.SequenceMatcher(None, chave, chave_rep).ratio()
                if r >= melhor:
                    melhor = r
                    destino = i
            if destino is None:
                clusters.append((chave, [c]))
            else:
                clusters[destino][1].append(c)

        for idx, (_chave_rep, lista) in enumerate(clusters):
            grupos[(tema, idx)] = lista

    faq = []

    for (tema, _idx), itens in grupos.items():
        itens.sort(key=lambda x: x.get("_dt_abertura") or datetime.min)

        # Pergunta representativa: o assunto original mais comprido do grupo.
        assunto_repr = max((i.get("Assunto Original", "") for i in itens), key=len, default="")

        # Resposta sugerida: a mais recente não vazia.
        resposta_repr = ""
        for i in reversed(itens):
            r = (i.get("Resposta (RH)") or "").strip()
            if r:
                resposta_repr = r
                break

        respondidos = sum(1 for i in itens if i.get("Status Resposta") == "Respondido")
        if respondidos == len(itens):
            status = "Todos respondidos"
        elif respondidos == 0:
            status = "Nenhum respondido"
        else:
            status = f"{respondidos}/{len(itens)} respondidos"

        datas = [i.get("_dt_abertura") for i in itens if i.get("_dt_abertura")]
        primeira = min(datas).strftime("%d/%m/%Y") if datas else ""
        ultima = max(datas).strftime("%d/%m/%Y") if datas else ""

        exemplos = []
        vistos = set()
        for i in itens:
            a = (i.get("Assunto Original") or "").strip()
            if a and a.lower() not in vistos:
                exemplos.append(a)
                vistos.add(a.lower())
            if len(exemplos) >= 3:
                break

        faq.append({
            "Tema": tema,
            "Pergunta Frequente": assunto_repr,
            "Resposta Sugerida": resposta_repr,
            "Frequência": len(itens),
            "% do Total": round(100.0 * len(itens) / total, 1),
            "Status Resposta": status,
            "Exemplos de Assunto": " | ".join(exemplos),
            "Primeira Ocorrência": primeira,
            "Última Ocorrência": ultima,
        })

    faq.sort(key=lambda x: x["Frequência"], reverse=True)
    for i, item in enumerate(faq, start=1):
        item["Nº"] = i
    return faq


def montar_temas(chamados):
    """Resumo por tema: volume, % e SLA médio de resposta."""
    grupos = {}
    for c in chamados:
        grupos.setdefault(c.get("Tema", TEMA_PADRAO), []).append(c)

    total = len(chamados) or 1
    linhas = []

    for tema, itens in grupos.items():
        respondidos = [i for i in itens if i.get("Status Resposta") == "Respondido"]
        tempos = [i.get("Tempo de Resposta (h)") for i in respondidos
                  if isinstance(i.get("Tempo de Resposta (h)"), (int, float))]
        media = round(sum(tempos) / len(tempos), 1) if tempos else ""

        linhas.append({
            "Tema": tema,
            "Qtd. Chamados": len(itens),
            "% do Total": round(100.0 * len(itens) / total, 1),
            "Respondidos": len(respondidos),
            "Sem Resposta": len(itens) - len(respondidos),
            "Tempo Médio de Resposta (h)": media,
        })

    linhas.sort(key=lambda x: x["Qtd. Chamados"], reverse=True)
    return linhas


def aplicar_estilo_planilha(ws, tipo="base"):
    azul = CORES["azul"].replace("#", "")
    navy = CORES["navy"].replace("#", "")
    branco = "FFFFFF"
    cinza_claro = "F2F2F2"
    borda_cor = "D9DDE3"

    fill_header = PatternFill("solid", fgColor=navy)
    fill_sub = PatternFill("solid", fgColor=cinza_claro)
    font_header = Font(color=branco, bold=True)
    font_titulo = Font(color=navy, bold=True, size=14)
    border = Border(
        left=Side(style="thin", color=borda_cor),
        right=Side(style="thin", color=borda_cor),
        top=Side(style="thin", color=borda_cor),
        bottom=Side(style="thin", color=borda_cor),
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    larguras = {}

    if tipo == "base":
        larguras = {
            "A": 28,
            "B": 16,
            "C": 14,
            "D": 18,
            "E": 28,
            "F": 32,
            "G": 36,
            "H": 28,
            "I": 45,
            "J": 80,
            "K": 80,
            "L": 16,
            "M": 14,
            "N": 18,
            "O": 40,
            "P": 18,
            "Q": 12,
            "R": 18,
            "S": 18,
            "T": 22,
            "U": 35,
            "V": 18,
            "W": 35,
            "X": 22,
            "Y": 35,
            "Z": 24,
            "AA": 35,
            "AB": 18,
        }
    elif tipo == "glossario":
        larguras = {
            "A": 28,
            "B": 24,
            "C": 18,
            "D": 40,
            "E": 24,
            "F": 40,
            "G": 12,
        }
    elif tipo == "resumo":
        larguras = {
            "A": 35,
            "B": 35,
        }
    elif tipo == "chamados":
        larguras = {
            "A": 12, "B": 26, "C": 42, "D": 60, "E": 60, "F": 16,
            "G": 26, "H": 32, "I": 24, "J": 18, "K": 18, "L": 16,
            "M": 14, "N": 12,
        }
    elif tipo == "faq":
        larguras = {
            "A": 6, "B": 26, "C": 55, "D": 70, "E": 12, "F": 12,
            "G": 20, "H": 50, "I": 18, "J": 18,
        }
    elif tipo == "temas":
        larguras = {
            "A": 30, "B": 14, "C": 12, "D": 14, "E": 14, "F": 24,
        }

    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    ws.row_dimensions[1].height = 28

    if tipo == "base":
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 80
    elif tipo in ("chamados", "faq"):
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 60


def criar_aba_base(wb, registros):
    ws = wb.active
    ws.title = "BASE_EMAILS"

    ws.append(COLUNAS_BASE)

    for registro in registros:
        ws.append([registro.get(coluna, "") for coluna in COLUNAS_BASE])

    aplicar_estilo_planilha(ws, tipo="base")


def aplicar_fonte_calibri_light(ws):
    """Padrão de export Anderson Souza: Calibri Light 8 em todas as células
    (cabeçalho pode ser negrito, mas sempre Calibri Light 8)."""
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            atual = cell.font
            cell.font = Font(
                name="Calibri Light",
                size=8,
                bold=bool(atual.bold),
                color=atual.color,
            )


def criar_aba_glossario(wb):
    ws = wb.create_sheet("GLOSSARIO")
    ws.append(COLUNAS_GLOSSARIO)

    # Pré-carrega os temas de RH já usados na classificação, com as
    # palavras-chave de cada um. Editável pela operação.
    for tema, chaves in TEMAS_RH:
        ws.append([
            ", ".join(chaves),
            tema,
            "",
            "Preencher resposta-padrão do tema (opcional).",
            "",
            "Termos usados na classificação automática. Ajuste conforme a operação.",
            "Sim",
        ])

    aplicar_estilo_planilha(ws, tipo="glossario")


def criar_aba_chamados(wb, chamados):
    ws = wb.create_sheet("CHAMADOS")
    ws.append(COLUNAS_CHAMADOS)

    for c in chamados:
        ws.append([c.get(coluna, "") for coluna in COLUNAS_CHAMADOS])

    aplicar_estilo_planilha(ws, tipo="chamados")
    aplicar_fonte_calibri_light(ws)


def criar_aba_faq(wb, faq):
    ws = wb.create_sheet("FAQ")
    ws.append(COLUNAS_FAQ)

    for f in faq:
        ws.append([f.get(coluna, "") for coluna in COLUNAS_FAQ])

    aplicar_estilo_planilha(ws, tipo="faq")
    aplicar_fonte_calibri_light(ws)


def criar_aba_temas(wb, temas):
    ws = wb.create_sheet("TEMAS")
    ws.append(COLUNAS_TEMAS)

    for t in temas:
        ws.append([t.get(coluna, "") for coluna in COLUNAS_TEMAS])

    aplicar_estilo_planilha(ws, tipo="temas")
    aplicar_fonte_calibri_light(ws)


def criar_aba_resumo(wb, registros):
    ws = wb.create_sheet("RESUMO")

    total = len(registros)
    total_anexo = sum(1 for r in registros if r.get("Tem Anexo") == "Sim")
    total_sem_anexo = total - total_anexo
    total_nao_lidos = sum(1 for r in registros if r.get("Lido") == "Não")
    total_lidos = sum(1 for r in registros if r.get("Lido") == "Sim")

    datas = []
    for r in registros:
        data_txt = r.get("Data Recebimento", "")
        try:
            datas.append(datetime.strptime(data_txt, "%d/%m/%Y").date())
        except Exception:
            pass

    data_inicial = min(datas).strftime("%d/%m/%Y") if datas else ""
    data_final = max(datas).strftime("%d/%m/%Y") if datas else ""

    linhas = [
        ["Indicador", "Valor"],
        ["Total de e-mails extraídos", total],
        ["Total com anexo", total_anexo],
        ["Total sem anexo", total_sem_anexo],
        ["Total não lidos", total_nao_lidos],
        ["Total lidos", total_lidos],
        ["Data inicial", data_inicial],
        ["Data final", data_final],
        ["Data da extração", agora_texto()],
        ["Aplicativo", f"{APP_NAME} v{APP_VERSION}"],
        ["Observação", "A classificação será feita posteriormente com apoio do ChatGPT e/ou glossário."],
    ]

    for linha in linhas:
        ws.append(linha)

    aplicar_estilo_planilha(ws, tipo="resumo")

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=CORES["navy"].replace("#", ""))
        cell.font = Font(color="FFFFFF", bold=True)


def gerar_excel(registros, caminho_saida, dominios_rh=None, emails_rh=None):
    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    if caminho_saida.exists():
        try:
            os.rename(caminho_saida, caminho_saida)
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            novo_nome = caminho_saida.with_name(f"{caminho_saida.stem}_{timestamp}{caminho_saida.suffix}")
            caminho_saida = novo_nome

    # Motor de FAQ: threads pareadas -> chamados -> FAQ -> temas.
    chamados = montar_chamados(registros, dominios_rh=dominios_rh, emails_rh=emails_rh)
    faq = montar_faq(chamados)
    temas = montar_temas(chamados)

    wb = Workbook()
    criar_aba_base(wb, registros)
    criar_aba_faq(wb, faq)
    criar_aba_chamados(wb, chamados)
    criar_aba_temas(wb, temas)
    criar_aba_glossario(wb)
    criar_aba_resumo(wb, registros)

    wb.save(caminho_saida)

    return {
        "caminho": caminho_saida,
        "chamados": chamados,
        "faq": faq,
        "temas": temas,
    }


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>FAQ - Chamados Gente & Gestao</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Rufina:wght@400;700&family=Roboto:wght@300;400;500&display=swap" rel="stylesheet">
<style>
:root{
  --azul:#0083CA; --navy:#003C64; --apoio:#005A64; --claro:#6EB4DC;
  --texto:#333333; --texto2:#646464; --borda:#CCCCCC; --fundo:#F5F7FA;
}
*{box-sizing:border-box}
body{margin:0;background:var(--fundo);color:var(--texto);
  font-family:'Roboto',Arial,sans-serif;font-weight:300}
header{background:var(--azul);color:#fff;padding:22px 28px}
header h1{font-family:'Rufina',Georgia,serif;font-weight:700;margin:0;font-size:26px}
header p{margin:6px 0 0;font-size:14px;opacity:.92}
.wrap{max-width:1100px;margin:0 auto;padding:20px 22px 60px}
.kpis{display:flex;gap:14px;flex-wrap:wrap;margin:18px 0}
.kpi{flex:1;min-width:150px;background:#fff;border:1px solid var(--borda);
  border-radius:12px;padding:14px 16px}
.kpi .n{font-family:'Rufina',serif;font-size:26px;color:var(--navy);font-weight:700}
.kpi .l{font-size:12px;color:var(--texto2);margin-top:2px}
.controls{display:flex;gap:12px;flex-wrap:wrap;align-items:center;margin:14px 0}
#busca{flex:1;min-width:240px;padding:11px 14px;border:1px solid var(--borda);
  border-radius:10px;font-size:14px;font-family:inherit}
#tema{padding:11px 14px;border:1px solid var(--borda);border-radius:10px;
  font-size:14px;font-family:inherit;background:#fff}
.item{background:#fff;border:1px solid var(--borda);border-radius:12px;
  margin-bottom:10px;overflow:hidden}
.q{display:flex;align-items:flex-start;gap:12px;padding:14px 16px;cursor:pointer}
.q:hover{background:#f2f8fc}
.q .txt{flex:1}
.q .perg{font-weight:500;color:var(--navy);font-size:15px}
.badges{display:flex;gap:8px;flex-wrap:wrap;margin-top:6px}
.badge{font-size:11px;padding:3px 9px;border-radius:20px;background:#eaf4fb;
  color:var(--navy);border:1px solid #d4e8f5}
.badge.freq{background:var(--azul);color:#fff;border-color:var(--azul)}
.badge.warn{background:#fbeee9;color:#8C321E;border-color:#f0d6cc}
.chev{color:var(--azul);font-size:18px;transition:transform .2s;margin-top:2px}
.item.open .chev{transform:rotate(90deg)}
.a{display:none;padding:0 16px 16px 44px;font-size:14px;color:var(--texto)}
.item.open .a{display:block}
.a .rot{font-size:11px;text-transform:uppercase;letter-spacing:.5px;
  color:var(--texto2);margin:12px 0 4px}
.a .resp{white-space:pre-wrap;background:#f7fafc;border-left:3px solid var(--azul);
  padding:10px 12px;border-radius:6px}
.a .semresp{color:#8C321E;font-style:italic}
.a .ex{color:var(--texto2);font-size:12px;margin-top:10px}
.vazio{text-align:center;color:var(--texto2);padding:40px}
footer{max-width:1100px;margin:0 auto;padding:16px 22px;color:var(--texto2);
  font-size:12px;text-align:right;border-top:1px solid var(--borda)}
</style>
</head>
<body>
<header>
  <h1>FAQ - Chamados Gente &amp; Gestao</h1>
  <p>Perguntas frequentes extraidas dos chamados. __PERIODO__</p>
</header>
<div class="wrap">
  <div class="kpis">
    <div class="kpi"><div class="n">__TOTAL_PERG__</div><div class="l">Perguntas frequentes</div></div>
    <div class="kpi"><div class="n">__TOTAL_CH__</div><div class="l">Chamados analisados</div></div>
    <div class="kpi"><div class="n">__PCT_RESP__%</div><div class="l">Chamados respondidos</div></div>
    <div class="kpi"><div class="n">__SLA__</div><div class="l">Tempo medio de resposta (h)</div></div>
  </div>
  <div class="controls">
    <input id="busca" type="text" placeholder="Buscar pergunta, tema ou palavra...">
    <select id="tema"><option value="">Todos os temas</option></select>
  </div>
  <div id="lista"></div>
  <div id="vazio" class="vazio" style="display:none">Nenhuma pergunta encontrada com esse filtro.</div>
</div>
<footer>Anderson Souza &middot; Gerado em __GERADO__ &middot; __APP__</footer>
<script>
const FAQ = __DADOS_JSON__;
function esc(s){return (s||"").replace(/[&<>]/g,c=>({"&":"&amp;","<":"&lt;",">":"&gt;"}[c]));}
const sel=document.getElementById("tema");
[...new Set(FAQ.map(f=>f.Tema))].sort().forEach(t=>{
  const o=document.createElement("option");o.value=t;o.textContent=t;sel.appendChild(o);
});
function render(){
  const q=(document.getElementById("busca").value||"").toLowerCase();
  const t=sel.value;
  const lista=document.getElementById("lista");
  lista.innerHTML="";
  let n=0;
  FAQ.filter(f=>{
    const alvo=((f["Pergunta Frequente"]||"")+" "+(f.Tema||"")+" "+(f["Resposta Sugerida"]||"")+" "+(f["Exemplos de Assunto"]||"")).toLowerCase();
    return (!t||f.Tema===t) && (!q||alvo.includes(q));
  }).forEach(f=>{
    n++;
    const div=document.createElement("div");div.className="item";
    const temResp=(f["Resposta Sugerida"]||"").trim().length>0;
    div.innerHTML=
      '<div class="q"><div class="txt"><div class="perg">'+esc(f["Pergunta Frequente"])+'</div>'+
      '<div class="badges"><span class="badge">'+esc(f.Tema)+'</span>'+
      '<span class="badge freq">'+f["Frequência"]+'x</span>'+
      (temResp?'':'<span class="badge warn">Sem resposta padrao</span>')+
      '</div></div><div class="chev">&#9656;</div></div>'+
      '<div class="a"><div class="rot">Resposta sugerida</div>'+
      (temResp?'<div class="resp">'+esc(f["Resposta Sugerida"])+'</div>':'<div class="semresp">Ainda nao ha resposta registrada. Preencher.</div>')+
      (f["Exemplos de Assunto"]?'<div class="ex">Exemplos: '+esc(f["Exemplos de Assunto"])+'</div>':'')+
      '<div class="ex">Ocorrencias: '+esc(f["Primeira Ocorrência"])+' a '+esc(f["Última Ocorrência"])+'</div></div>';
    div.querySelector(".q").onclick=()=>div.classList.toggle("open");
    lista.appendChild(div);
  });
  document.getElementById("vazio").style.display=n?"none":"block";
}
document.getElementById("busca").addEventListener("input",render);
sel.addEventListener("change",render);
render();
</script>
</body>
</html>"""


def gerar_html_faq(faq, chamados, temas, caminho_saida, meta=None):
    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    meta = meta or {}
    total_ch = len(chamados)
    respondidos = sum(1 for c in chamados if c.get("Status Resposta") == "Respondido")
    pct_resp = round(100.0 * respondidos / total_ch) if total_ch else 0

    tempos = [c.get("Tempo de Resposta (h)") for c in chamados
              if isinstance(c.get("Tempo de Resposta (h)"), (int, float))]
    sla = round(sum(tempos) / len(tempos), 1) if tempos else "-"

    # Só os campos usados pelo HTML, para não vazar dados técnicos.
    faq_html = [{
        "Tema": f.get("Tema", ""),
        "Pergunta Frequente": f.get("Pergunta Frequente", ""),
        "Resposta Sugerida": f.get("Resposta Sugerida", ""),
        "Frequência": f.get("Frequência", 0),
        "Exemplos de Assunto": f.get("Exemplos de Assunto", ""),
        "Primeira Ocorrência": f.get("Primeira Ocorrência", ""),
        "Última Ocorrência": f.get("Última Ocorrência", ""),
    } for f in faq]

    dados_json = json.dumps(faq_html, ensure_ascii=False).replace("</", "<\\/")

    periodo = meta.get("periodo", "")
    html = (
        HTML_TEMPLATE
        .replace("__DADOS_JSON__", dados_json)
        .replace("__TOTAL_PERG__", str(len(faq)))
        .replace("__TOTAL_CH__", str(total_ch))
        .replace("__PCT_RESP__", str(pct_resp))
        .replace("__SLA__", str(sla))
        .replace("__PERIODO__", periodo)
        .replace("__GERADO__", agora_texto())
        .replace("__APP__", f"{APP_NAME} v{APP_VERSION}")
    )

    with open(caminho_saida, "w", encoding="utf-8") as f:
        f.write(html)

    return caminho_saida


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        garantir_pastas()

        self.config_app = carregar_config()
        self.processo_rodando = False
        self.caminho_excel_gerado = None
        self.caminho_html_gerado = None

        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.title(APP_NAME)
        self.geometry("1180x760")
        self.minsize(1050, 680)
        self.configure(fg_color=CORES["fundo"])

        self.criar_variaveis()
        self.criar_layout()
        self.carregar_variaveis_config()

    def criar_variaveis(self):
        self.var_modo_caixa = ctk.StringVar(value="principal")
        self.var_caixa_compartilhada = ctk.StringVar(value="")
        self.var_pasta_outlook = ctk.StringVar(value=DEFAULT_PASTA_OUTLOOK)
        self.var_quantidade = ctk.StringVar(value="500")
        self.var_ultimos_dias = ctk.StringVar(value="90")
        self.var_somente_nao_lidos = ctk.BooleanVar(value=False)
        self.var_salvar_msg = ctk.BooleanVar(value=False)
        self.var_emails_rh = ctk.StringVar(value="")
        self.var_pasta_saida = ctk.StringVar(value=str(SAIDA_DIR))
        self.var_nome_excel = ctk.StringVar(value=DEFAULT_EXCEL_NAME)
        self.var_nome_html = ctk.StringVar(value=DEFAULT_HTML_NAME)

    def carregar_variaveis_config(self):
        self.var_modo_caixa.set(self.config_app.get("modo_caixa", "principal"))
        self.var_caixa_compartilhada.set(self.config_app.get("caixa_compartilhada", ""))
        self.var_pasta_outlook.set(self.config_app.get("pasta_outlook", DEFAULT_PASTA_OUTLOOK))
        self.var_quantidade.set(str(self.config_app.get("quantidade", 500)))
        self.var_ultimos_dias.set(str(self.config_app.get("ultimos_dias", 90)))
        self.var_somente_nao_lidos.set(bool(self.config_app.get("somente_nao_lidos", False)))
        self.var_salvar_msg.set(bool(self.config_app.get("salvar_msg", False)))
        self.var_emails_rh.set(self.config_app.get("emails_rh", ""))
        self.var_pasta_saida.set(self.config_app.get("pasta_saida", str(SAIDA_DIR)))
        self.var_nome_excel.set(self.config_app.get("nome_excel", DEFAULT_EXCEL_NAME))
        self.var_nome_html.set(self.config_app.get("nome_html", DEFAULT_HTML_NAME))

    def obter_config_tela(self):
        return {
            "modo_caixa": self.var_modo_caixa.get(),
            "caixa_compartilhada": self.var_caixa_compartilhada.get().strip(),
            "pasta_outlook": self.var_pasta_outlook.get().strip(),
            "quantidade": int(self.var_quantidade.get()),
            "ultimos_dias": int(self.var_ultimos_dias.get()),
            "somente_nao_lidos": bool(self.var_somente_nao_lidos.get()),
            "salvar_msg": bool(self.var_salvar_msg.get()),
            "emails_rh": self.var_emails_rh.get().strip(),
            "pasta_saida": self.var_pasta_saida.get().strip(),
            "nome_excel": self.var_nome_excel.get().strip(),
            "nome_html": self.var_nome_html.get().strip(),
        }

    def criar_layout(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        self.criar_header()
        self.criar_cards_config()
        self.criar_area_log()
        self.criar_rodape()

    def criar_header(self):
        header = ctk.CTkFrame(self, fg_color=CORES["navy"], corner_radius=0, height=92)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        titulo = ctk.CTkLabel(
            header,
            text="Leitor Outlook RH",
            font=("Segoe UI", 26, "bold"),
            text_color=CORES["branco"],
            anchor="w",
        )
        titulo.grid(row=0, column=0, padx=28, pady=(18, 0), sticky="ew")

        subtitulo = ctk.CTkLabel(
            header,
            text="Lê a pasta de chamados no Outlook, pareia pergunta e resposta nas threads e gera um FAQ em Excel e HTML. Modo seguro: nada é movido, respondido ou excluído.",
            font=("Segoe UI", 14),
            text_color=CORES["branco"],
            anchor="w",
        )
        subtitulo.grid(row=1, column=0, padx=28, pady=(2, 18), sticky="ew")

    def criar_cards_config(self):
        area = ctk.CTkFrame(self, fg_color=CORES["fundo"], corner_radius=0)
        area.grid(row=1, column=0, sticky="ew", padx=22, pady=18)
        area.grid_columnconfigure(0, weight=1)
        area.grid_columnconfigure(1, weight=1)

        card_origem = ctk.CTkFrame(
            area,
            fg_color=CORES["card"],
            border_color=CORES["borda"],
            border_width=1,
            corner_radius=14,
        )
        card_origem.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        card_origem.grid_columnconfigure(1, weight=1)

        card_saida = ctk.CTkFrame(
            area,
            fg_color=CORES["card"],
            border_color=CORES["borda"],
            border_width=1,
            corner_radius=14,
        )
        card_saida.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        card_saida.grid_columnconfigure(1, weight=1)

        self.criar_card_origem(card_origem)
        self.criar_card_saida(card_saida)

    def criar_titulo_card(self, parent, texto):
        label = ctk.CTkLabel(
            parent,
            text=texto,
            font=("Segoe UI", 17, "bold"),
            text_color=CORES["navy"],
            anchor="w",
        )
        label.grid(row=0, column=0, columnspan=3, padx=18, pady=(16, 10), sticky="ew")

    def criar_label(self, parent, texto, row):
        label = ctk.CTkLabel(
            parent,
            text=texto,
            font=("Segoe UI", 13),
            text_color=CORES["texto"],
            anchor="w",
        )
        label.grid(row=row, column=0, padx=(18, 8), pady=7, sticky="w")
        return label

    def criar_entry(self, parent, variable, row, placeholder=""):
        entry = ctk.CTkEntry(
            parent,
            textvariable=variable,
            placeholder_text=placeholder,
            height=34,
            fg_color=CORES["branco"],
            border_color=CORES["borda"],
            text_color=CORES["texto"],
        )
        entry.grid(row=row, column=1, padx=(4, 18), pady=7, sticky="ew")
        return entry

    def criar_card_origem(self, parent):
        self.criar_titulo_card(parent, "Origem dos e-mails")

        radio_principal = ctk.CTkRadioButton(
            parent,
            text="Caixa principal",
            variable=self.var_modo_caixa,
            value="principal",
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        radio_principal.grid(row=1, column=0, padx=18, pady=5, sticky="w")

        radio_compartilhada = ctk.CTkRadioButton(
            parent,
            text="Caixa compartilhada",
            variable=self.var_modo_caixa,
            value="compartilhada",
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        radio_compartilhada.grid(row=1, column=1, padx=8, pady=5, sticky="w")

        self.criar_label(parent, "Caixa compartilhada", 2)
        self.criar_entry(
            parent,
            self.var_caixa_compartilhada,
            2,
            "Exemplo: genteegestao@empresa.com.br",
        )

        self.criar_label(parent, "Pasta Outlook", 3)

        frame_pasta_outlook = ctk.CTkFrame(parent, fg_color="transparent")
        frame_pasta_outlook.grid(row=3, column=1, padx=(4, 18), pady=7, sticky="ew")
        frame_pasta_outlook.grid_columnconfigure(0, weight=1)

        self.entry_pasta_outlook = ctk.CTkEntry(
            frame_pasta_outlook,
            textvariable=self.var_pasta_outlook,
            placeholder_text="Selecione na árvore de pastas do Outlook",
            height=34,
            fg_color=CORES["branco"],
            border_color=CORES["borda"],
            text_color=CORES["texto"],
        )
        self.entry_pasta_outlook.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        self.btn_pastas_outlook = ctk.CTkButton(
            frame_pasta_outlook,
            text="Ver pastas",
            command=self.iniciar_listagem_pastas_outlook,
            width=105,
            height=34,
            fg_color=CORES["azul_apoio"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
        )
        self.btn_pastas_outlook.grid(row=0, column=1, sticky="e")

        self.criar_label(parent, "Quantidade máxima", 4)
        self.criar_entry(parent, self.var_quantidade, 4, "200")

        self.criar_label(parent, "Últimos dias", 5)
        self.criar_entry(parent, self.var_ultimos_dias, 5, "90. Use 0 para não filtrar")

        self.criar_label(parent, "Quem responde (RH)", 6)
        self.criar_entry(
            parent,
            self.var_emails_rh,
            6,
            "E-mails, @dominio ou nomes de quem responde. Ex: @sonova.com; Jair; Sabrina",
        )

        chk = ctk.CTkCheckBox(
            parent,
            text="Ler somente e-mails não lidos",
            variable=self.var_somente_nao_lidos,
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        chk.grid(row=7, column=0, columnspan=2, padx=18, pady=(8, 4), sticky="w")

        chk_msg = ctk.CTkCheckBox(
            parent,
            text="Salvar cópias .msg dos e-mails (com anexos) na pasta de saída",
            variable=self.var_salvar_msg,
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        chk_msg.grid(row=8, column=0, columnspan=2, padx=18, pady=(0, 16), sticky="w")

    def criar_card_saida(self, parent):
        self.criar_titulo_card(parent, "Saída em Excel")

        self.criar_label(parent, "Pasta de saída", 1)

        frame_saida = ctk.CTkFrame(parent, fg_color="transparent")
        frame_saida.grid(row=1, column=1, padx=(4, 18), pady=7, sticky="ew")
        frame_saida.grid_columnconfigure(0, weight=1)

        entry_saida = ctk.CTkEntry(
            frame_saida,
            textvariable=self.var_pasta_saida,
            height=34,
            fg_color=CORES["branco"],
            border_color=CORES["borda"],
            text_color=CORES["texto"],
        )
        entry_saida.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        btn_pasta = ctk.CTkButton(
            frame_saida,
            text="Selecionar",
            command=self.selecionar_pasta_saida,
            width=95,
            height=34,
            fg_color=CORES["azul_apoio"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
        )
        btn_pasta.grid(row=0, column=1)

        self.criar_label(parent, "Nome do Excel", 2)
        self.criar_entry(parent, self.var_nome_excel, 2, DEFAULT_EXCEL_NAME)

        self.criar_label(parent, "Nome do HTML", 3)
        self.criar_entry(parent, self.var_nome_html, 3, DEFAULT_HTML_NAME)

        info = ctk.CTkLabel(
            parent,
            text="Gera um Excel (abas FAQ, CHAMADOS, TEMAS, BASE_EMAILS, GLOSSARIO, RESUMO) e um FAQ em HTML consultável. As threads são agrupadas e a pergunta do colaborador é pareada com a resposta do RH. Modo seguro: nada é movido, respondido ou excluído.",
            font=("Segoe UI", 13),
            text_color=CORES["texto_sec"],
            justify="left",
            wraplength=480,
            anchor="w",
        )
        info.grid(row=4, column=0, columnspan=2, padx=18, pady=(8, 12), sticky="ew")

        botoes = ctk.CTkFrame(parent, fg_color="transparent")
        botoes.grid(row=5, column=0, columnspan=2, padx=18, pady=(4, 16), sticky="ew")
        botoes.grid_columnconfigure(0, weight=1)
        botoes.grid_columnconfigure(1, weight=1)

        self.btn_extrair = ctk.CTkButton(
            botoes,
            text="Gerar FAQ (Excel + HTML)",
            command=self.iniciar_extracao,
            height=42,
            fg_color=CORES["azul"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
            font=("Segoe UI", 14, "bold"),
        )
        self.btn_extrair.grid(row=0, column=0, columnspan=2, padx=0, pady=(0, 8), sticky="ew")

        self.btn_faq = ctk.CTkButton(
            botoes,
            text="Abrir FAQ (HTML)",
            command=self.abrir_faq_html,
            height=40,
            fg_color=CORES["azul_apoio"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
            font=("Segoe UI", 13, "bold"),
        )
        self.btn_faq.grid(row=1, column=0, padx=(0, 8), sticky="ew")

        self.btn_abrir = ctk.CTkButton(
            botoes,
            text="Abrir pasta de saída",
            command=self.abrir_pasta_saida,
            height=40,
            fg_color=CORES["azul_apoio"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
            font=("Segoe UI", 13, "bold"),
        )
        self.btn_abrir.grid(row=1, column=1, padx=(8, 0), sticky="ew")

        self.btn_ler_pasta = ctk.CTkButton(
            botoes,
            text="Ler de pasta .msg (offline, sem Outlook)…",
            command=self.iniciar_leitura_pasta_msg,
            height=40,
            fg_color=CORES["azul_apoio"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
            font=("Segoe UI", 13, "bold"),
        )
        self.btn_ler_pasta.grid(row=2, column=0, columnspan=2, padx=0, pady=(8, 0), sticky="ew")

    def criar_area_log(self):
        area = ctk.CTkFrame(self, fg_color=CORES["fundo"], corner_radius=0)
        area.grid(row=2, column=0, sticky="nsew", padx=22, pady=(0, 12))
        area.grid_columnconfigure(0, weight=1)
        area.grid_rowconfigure(2, weight=1)

        card = ctk.CTkFrame(
            area,
            fg_color=CORES["card"],
            border_color=CORES["borda"],
            border_width=1,
            corner_radius=14,
        )
        card.grid(row=0, column=0, sticky="nsew")
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(2, weight=1)

        titulo = ctk.CTkLabel(
            card,
            text="Execução",
            font=("Segoe UI", 17, "bold"),
            text_color=CORES["navy"],
            anchor="w",
        )
        titulo.grid(row=0, column=0, padx=18, pady=(16, 8), sticky="ew")

        self.progress = ctk.CTkProgressBar(
            card,
            height=14,
            progress_color=CORES["azul"],
            fg_color=CORES["borda"],
        )
        self.progress.grid(row=1, column=0, padx=18, pady=(0, 10), sticky="ew")
        self.progress.set(0)

        self.txt_log = ctk.CTkTextbox(
            card,
            height=280,
            fg_color="#0F172A",
            text_color="#E5E7EB",
            border_color=CORES["borda"],
            border_width=1,
            font=("Consolas", 12),
        )
        self.txt_log.grid(row=2, column=0, padx=18, pady=(0, 18), sticky="nsew")

        self.log("Aplicativo iniciado.")
        self.log("Modo seguro: leitura e exportação. Nenhum e-mail será movido, excluído, respondido ou marcado.")

    def criar_rodape(self):
        rodape = ctk.CTkFrame(self, fg_color=CORES["fundo"], corner_radius=0, height=36)
        rodape.grid(row=3, column=0, sticky="ew", padx=22, pady=(0, 8))
        rodape.grid_columnconfigure(0, weight=1)

        label = ctk.CTkLabel(
            rodape,
            text="Anderson Souza",
            font=("Segoe UI", 12),
            text_color=CORES["texto_sec"],
            anchor="e",
        )
        label.grid(row=0, column=0, sticky="e")

    def log(self, mensagem):
        mensagem = str(mensagem)
        linha = f"[{agora_texto()}] {mensagem}\n"

        try:
            self.txt_log.insert("end", linha)
            self.txt_log.see("end")
        except Exception:
            pass

        try:
            registrar_log_arquivo(mensagem)
        except Exception:
            pass

    def iniciar_listagem_pastas_outlook(self):
        """Carrega a árvore completa de pastas sem congelar a interface."""
        if self.processo_rodando:
            messagebox.showinfo("Outlook", "Aguarde o processamento atual terminar.")
            return

        modo = self.var_modo_caixa.get()
        compartilhada = self.var_caixa_compartilhada.get().strip()

        if modo == "compartilhada" and not compartilhada:
            messagebox.showerror(
                "Outlook",
                "Para consultar uma caixa compartilhada não montada no Outlook, informe o nome ou e-mail da caixa.\n\n"
                "Se a caixa já aparece na barra lateral do Outlook, você também pode usar 'Caixa principal' e selecionar a pasta na árvore."
            )
            return

        try:
            self.btn_pastas_outlook.configure(state="disabled", text="Carregando...")
        except Exception:
            pass

        self.log("Consultando a árvore de pastas do Outlook...")
        thread = threading.Thread(
            target=self._thread_listar_pastas_outlook,
            args=(modo, compartilhada),
            daemon=True,
        )
        thread.start()

    def _thread_listar_pastas_outlook(self, modo, compartilhada):
        try:
            pastas = listar_pastas_outlook_disponiveis(
                modo_caixa=modo,
                caixa_compartilhada=compartilhada,
                log_callback=lambda msg: self.after(0, self.log, msg),
            )
            self.after(0, self._mostrar_seletor_pastas_outlook, pastas)
        except Exception as e:
            erro = traceback.format_exc()
            registrar_log_arquivo(erro)
            self.after(0, self.log, f"Erro ao listar pastas do Outlook: {e}")
            self.after(0, messagebox.showerror, "Outlook", f"Não foi possível listar as pastas do Outlook:\n\n{e}")
        finally:
            self.after(0, self._restaurar_botao_pastas_outlook)

    def _restaurar_botao_pastas_outlook(self):
        try:
            self.btn_pastas_outlook.configure(state="normal", text="Ver pastas")
        except Exception:
            pass

    def _mostrar_seletor_pastas_outlook(self, pastas):
        if not pastas:
            messagebox.showinfo("Outlook", "Nenhuma pasta foi localizada nas contas abertas do Outlook.")
            return

        janela = ctk.CTkToplevel(self)
        janela.title("Escolher pasta do Outlook")
        janela.geometry("900x620")
        janela.minsize(720, 480)
        janela.configure(fg_color=CORES["fundo"])
        janela.transient(self)
        janela.grab_set()

        janela.grid_columnconfigure(0, weight=1)
        janela.grid_rowconfigure(2, weight=1)

        titulo = ctk.CTkLabel(
            janela,
            text="Pastas disponíveis no Outlook",
            font=("Segoe UI", 20, "bold"),
            text_color=CORES["navy"],
            anchor="w",
        )
        titulo.grid(row=0, column=0, padx=20, pady=(18, 4), sticky="ew")

        info = ctk.CTkLabel(
            janela,
            text="Selecione a conta e a pasta que deverá ser lida. O caminho completo será salvo para evitar ambiguidades.",
            font=("Segoe UI", 13),
            text_color=CORES["texto_sec"],
            anchor="w",
        )
        info.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="ew")

        frame_tree = ctk.CTkFrame(
            janela,
            fg_color=CORES["card"],
            border_color=CORES["borda"],
            border_width=1,
            corner_radius=10,
        )
        frame_tree.grid(row=2, column=0, padx=20, pady=(0, 12), sticky="nsew")
        frame_tree.grid_columnconfigure(0, weight=1)
        frame_tree.grid_rowconfigure(0, weight=1)

        tree = ttk.Treeview(frame_tree, columns=("caminho",), show="tree headings", selectmode="browse")
        tree.heading("#0", text="Pasta")
        tree.heading("caminho", text="Caminho completo")
        tree.column("#0", width=310, minwidth=180, stretch=True)
        tree.column("caminho", width=520, minwidth=280, stretch=True)

        scroll_y = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        scroll_x = ttk.Scrollbar(frame_tree, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        tree.grid(row=0, column=0, sticky="nsew", padx=(10, 0), pady=(10, 0))
        scroll_y.grid(row=0, column=1, sticky="ns", pady=(10, 0))
        scroll_x.grid(row=1, column=0, sticky="ew", padx=(10, 0), pady=(0, 10))

        iid_por_caminho = {}
        caminho_por_iid = {}
        for idx, item in enumerate(pastas, start=1):
            caminho = item["caminho"]
            nome = item["nome"]
            partes = [p.strip() for p in caminho.split("\\") if p.strip()]
            caminho_pai = " \\ ".join(partes[:-1]) if len(partes) > 1 else ""
            parent_iid = iid_por_caminho.get(caminho_pai, "")
            iid = f"p{idx}"
            tree.insert(parent_iid, "end", iid=iid, text=nome, values=(caminho,), open=(item["profundidade"] == 0))
            iid_por_caminho[caminho] = iid
            caminho_por_iid[iid] = caminho

        frame_botoes = ctk.CTkFrame(janela, fg_color="transparent")
        frame_botoes.grid(row=3, column=0, padx=20, pady=(0, 18), sticky="ew")
        frame_botoes.grid_columnconfigure(0, weight=1)

        lbl_total = ctk.CTkLabel(
            frame_botoes,
            text=f"{len(pastas)} pastas encontradas",
            font=("Segoe UI", 12),
            text_color=CORES["texto_sec"],
        )
        lbl_total.grid(row=0, column=0, sticky="w")

        def selecionar():
            iid = tree.focus() or (tree.selection()[0] if tree.selection() else "")
            caminho = caminho_por_iid.get(iid)
            if not caminho:
                messagebox.showinfo("Outlook", "Selecione uma pasta na árvore.", parent=janela)
                return
            self.var_pasta_outlook.set(caminho)
            self.log(f"Pasta escolhida para leitura: {caminho}")
            janela.destroy()

        btn_cancelar = ctk.CTkButton(
            frame_botoes,
            text="Cancelar",
            command=janela.destroy,
            width=110,
            fg_color=CORES["borda"],
            hover_color="#C7CDD4",
            text_color=CORES["texto"],
        )
        btn_cancelar.grid(row=0, column=1, padx=(8, 8))

        btn_selecionar = ctk.CTkButton(
            frame_botoes,
            text="Selecionar pasta",
            command=selecionar,
            width=150,
            fg_color=CORES["azul"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
        )
        btn_selecionar.grid(row=0, column=2)

        tree.bind("<Double-1>", lambda _event: selecionar())

        atual = self.var_pasta_outlook.get().strip()
        if atual in iid_por_caminho:
            iid = iid_por_caminho[atual]
            tree.selection_set(iid)
            tree.focus(iid)
            tree.see(iid)

    def selecionar_pasta_saida(self):
        pasta = filedialog.askdirectory(
            title="Selecione a pasta de saída",
            initialdir=self.var_pasta_saida.get() or str(SAIDA_DIR),
        )
        if pasta:
            self.var_pasta_saida.set(pasta)

    def abrir_pasta_saida(self):
        pasta = self.var_pasta_saida.get().strip()
        if not pasta:
            pasta = str(SAIDA_DIR)

        Path(pasta).mkdir(parents=True, exist_ok=True)
        os.startfile(pasta)

    def abrir_faq_html(self):
        caminho = self.caminho_html_gerado
        if not caminho:
            # Tenta o último HTML pela config, senão avisa.
            candidato = Path(self.var_pasta_saida.get().strip() or str(SAIDA_DIR)) / self.var_nome_html.get().strip()
            if candidato.exists():
                caminho = candidato

        if not caminho or not Path(caminho).exists():
            messagebox.showinfo(
                "FAQ",
                "Nenhum FAQ HTML foi gerado ainda. Clique em 'Gerar FAQ (Excel + HTML)' primeiro.",
            )
            return

        os.startfile(str(caminho))

    def validar_campos(self):
        try:
            quantidade = int(self.var_quantidade.get())
            if quantidade <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Validação", "Informe uma quantidade válida maior que zero.")
            return False

        try:
            dias = int(self.var_ultimos_dias.get())
            if dias < 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Validação", "Informe os últimos dias como número. Use 0 para não filtrar.")
            return False

        if self.var_modo_caixa.get() == "compartilhada" and not self.var_caixa_compartilhada.get().strip():
            messagebox.showerror("Validação", "Informe o nome ou e-mail da caixa compartilhada.")
            return False

        pasta_saida = self.var_pasta_saida.get().strip()
        if not pasta_saida:
            messagebox.showerror("Validação", "Informe a pasta de saída.")
            return False

        nome_excel = self.var_nome_excel.get().strip()
        if not nome_excel:
            messagebox.showerror("Validação", "Informe o nome do arquivo Excel.")
            return False

        if not nome_excel.lower().endswith(".xlsx"):
            self.var_nome_excel.set(nome_excel + ".xlsx")

        nome_html = self.var_nome_html.get().strip()
        if not nome_html:
            self.var_nome_html.set(DEFAULT_HTML_NAME)
        elif not nome_html.lower().endswith(".html"):
            self.var_nome_html.set(nome_html + ".html")

        return True

    def travar_interface(self, travar=True):
        estado = "disabled" if travar else "normal"

        for nome in ("btn_extrair", "btn_ler_pasta", "btn_pastas_outlook"):
            try:
                getattr(self, nome).configure(state=estado)
            except Exception:
                pass

        self.processo_rodando = travar

    def atualizar_progresso(self, atual, total):
        try:
            valor = min(max(atual / total, 0), 1)
            self.progress.set(valor)
        except Exception:
            pass

    def iniciar_extracao(self):
        if self.processo_rodando:
            messagebox.showinfo("Execução", "Já existe uma extração em andamento.")
            return

        if not self.validar_campos():
            return

        config = self.obter_config_tela()
        salvar_config(config)

        self.travar_interface(True)
        self.progress.set(0)
        self.log("Iniciando leitura dos e-mails no Outlook.")

        thread = threading.Thread(target=self.executar_extracao_thread, args=(config,), daemon=True)
        thread.start()

    def executar_extracao_thread(self, config):
        try:
            pasta_saida = Path(config["pasta_saida"])
            salvar_msg = bool(config.get("salvar_msg", False))
            destino_msg = str(pasta_saida / "emails_msg") if salvar_msg else None
            destino_anexos = str(pasta_saida / "emails_msg" / "anexos") if salvar_msg else None

            registros, erros = ler_emails_outlook(
                modo_caixa=config["modo_caixa"],
                caixa_compartilhada=config["caixa_compartilhada"],
                pasta_outlook=config["pasta_outlook"],
                quantidade=config["quantidade"],
                ultimos_dias=config["ultimos_dias"],
                somente_nao_lidos=config["somente_nao_lidos"],
                progress_callback=lambda atual, total: self.after(0, self.atualizar_progresso, atual, total),
                log_callback=lambda msg: self.after(0, self.log, msg),
                salvar_msg=salvar_msg,
                destino_msg=destino_msg,
                destino_anexos=destino_anexos,
                salvar_anexos=True,
            )

            if salvar_msg and registros:
                self.after(0, self.log, f"Cópias .msg salvas em: {destino_msg}")

            self._finalizar_processamento(registros, erros, config)

        except Exception as e:
            erro = traceback.format_exc()
            self.after(0, self.log, f"Erro na execução: {e}")
            registrar_log_arquivo(erro)
            self.after(0, messagebox.showerror, "Erro", f"Ocorreu um erro na execução:\n\n{e}")

        finally:
            self.after(0, self.travar_interface, False)

    def iniciar_leitura_pasta_msg(self):
        if self.processo_rodando:
            messagebox.showinfo("Execução", "Já existe um processamento em andamento.")
            return

        pasta = filedialog.askdirectory(
            title="Selecione a pasta com os arquivos .msg",
            initialdir=self.var_pasta_saida.get() or str(SAIDA_DIR),
        )
        if not pasta:
            return

        if not self.validar_campos():
            return

        config = self.obter_config_tela()
        salvar_config(config)

        self.travar_interface(True)
        self.progress.set(0)
        self.log(f"Lendo e-mails da pasta local: {pasta}")

        thread = threading.Thread(
            target=self.executar_leitura_pasta_thread, args=(config, pasta), daemon=True
        )
        thread.start()

    def executar_leitura_pasta_thread(self, config, caminho_pasta):
        try:
            registros, erros = ler_emails_de_pasta_msg(
                caminho_pasta=caminho_pasta,
                quantidade=config["quantidade"],
                ultimos_dias=config["ultimos_dias"],
                somente_nao_lidos=config["somente_nao_lidos"],
                progress_callback=lambda atual, total: self.after(0, self.atualizar_progresso, atual, total),
                log_callback=lambda msg: self.after(0, self.log, msg),
            )

            self._finalizar_processamento(registros, erros, config)

        except Exception as e:
            erro = traceback.format_exc()
            self.after(0, self.log, f"Erro na leitura da pasta: {e}")
            registrar_log_arquivo(erro)
            self.after(0, messagebox.showerror, "Erro", f"Ocorreu um erro na leitura da pasta:\n\n{e}")

        finally:
            self.after(0, self.travar_interface, False)

    def _finalizar_processamento(self, registros, erros, config):
        """Etapa comum: a partir dos registros lidos, gera Excel e FAQ HTML."""
        if not registros:
            self.after(0, self.log, "Nenhum e-mail foi lido com os filtros informados.")
            self.after(0, messagebox.showwarning, "Resultado", "Nenhum e-mail foi lido com os filtros informados.")
            return

        pasta_saida = Path(config["pasta_saida"])
        nome_excel = config["nome_excel"]
        nome_html = config.get("nome_html") or DEFAULT_HTML_NAME

        if not nome_excel.lower().endswith(".xlsx"):
            nome_excel += ".xlsx"
        if not nome_html.lower().endswith(".html"):
            nome_html += ".html"

        caminho_saida = pasta_saida / nome_excel
        caminho_html = pasta_saida / nome_html

        dominios_rh, emails_rh = parse_dominios_emails(config.get("emails_rh", ""))
        if dominios_rh or emails_rh:
            self.after(0, self.log, f"Pareamento RH ativado. Domínios: {dominios_rh or '-'} | E-mails/nomes: {emails_rh or '-'}")
        else:
            self.after(0, self.log, "Sem lista de RH informada. Usando heurística de ordem e troca de remetente na thread.")

        self.after(0, self.log, "Agrupando threads e montando o FAQ.")
        resultado = gerar_excel(registros, caminho_saida, dominios_rh=dominios_rh, emails_rh=emails_rh)

        caminho_gerado = resultado["caminho"]
        chamados = resultado["chamados"]
        faq = resultado["faq"]
        temas = resultado["temas"]

        self.caminho_excel_gerado = caminho_gerado

        datas = [r.get("_dt") for r in registros if r.get("_dt")]
        periodo = ""
        if datas:
            periodo = f"Período: {min(datas).strftime('%d/%m/%Y')} a {max(datas).strftime('%d/%m/%Y')}."

        self.after(0, self.log, "Gerando FAQ HTML consultável.")
        caminho_html_gerado = gerar_html_faq(
            faq, chamados, temas, caminho_html, meta={"periodo": periodo}
        )
        self.caminho_html_gerado = caminho_html_gerado

        respondidos = sum(1 for c in chamados if c.get("Status Resposta") == "Respondido")

        self.after(0, self.progress.set, 1)
        self.after(0, self.log, f"Excel gerado: {caminho_gerado}")
        self.after(0, self.log, f"FAQ HTML gerado: {caminho_html_gerado}")
        self.after(0, self.log, f"Chamados identificados: {len(chamados)} | Respondidos: {respondidos} | Perguntas no FAQ: {len(faq)} | Temas: {len(temas)}")

        if erros:
            self.after(0, self.log, f"Alguns itens não puderam ser lidos. Total de erros ignorados: {len(erros)}")

        self.after(
            0,
            messagebox.showinfo,
            "Concluído",
            (
                "FAQ gerado com sucesso.\n\n"
                f"E-mails lidos: {len(registros)}\n"
                f"Chamados (threads): {len(chamados)}\n"
                f"Respondidos: {respondidos}\n"
                f"Perguntas no FAQ: {len(faq)}\n\n"
                f"Excel:\n{caminho_gerado}\n\n"
                f"FAQ HTML:\n{caminho_html_gerado}"
            ),
        )


def main():
    garantir_pastas()
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()