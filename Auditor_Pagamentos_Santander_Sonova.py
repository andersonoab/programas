"""
Auditor de Pagamentos Santander - Sonova

Aplicativo para ler relatórios PDF do Internet Banking Empresarial Santander,
consolidar pagamentos e localizar possíveis duplicidades dentro de uma empresa
ou entre empresas do grupo.

Compatibilidade: Python 3.10+
Dependências: pypdf e openpyxl

Instalação:
    python -m pip install pypdf openpyxl

Execução gráfica:
    python Auditor_Pagamentos_Santander_Sonova.py

Execução em linha de comando:
    python Auditor_Pagamentos_Santander_Sonova.py --cli pasta_ou_pdf --export resultado.xlsx
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
import threading
import unicodedata
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Iterable, Optional

try:
    from pypdf import PdfReader
except ImportError as exc:  # pragma: no cover - mensagem amigavel ao usuario
    raise SystemExit(
        "Biblioteca 'pypdf' não encontrada. Execute: python -m pip install pypdf openpyxl"
    ) from exc

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Biblioteca 'openpyxl' não encontrada. Execute: python -m pip install pypdf openpyxl"
    ) from exc


APP_NAME = "Auditor de Pagamentos Santander"
APP_VERSION = "1.0.0"

# Identidade visual Sonova
SONOVA_BLUE = "0083CA"
SONOVA_DARK = "003C64"
SONOVA_LIGHT = "6EB4DC"
SONOVA_ACCENT = "7D0041"
WHITE = "FFFFFF"
TEXT = "333333"
SOFT_GRAY = "F2F5F7"
WARNING = "F59E0B"
DANGER = "C62828"
SUCCESS = "2E7D32"

DATE_RE = r"\d{2}/\d{2}/\d{4}"
MONEY_RE = r"[\d.]+,\d{2}"
ROW_RE = re.compile(
    rf"(?ms)^\s*(\d{{9}})\s+(\d{{20}})\s+(.*?)\s+"
    rf"({DATE_RE})\s+([A-Z]{{2,5}})\s+({MONEY_RE})\s*$"
)


@dataclass(frozen=True)
class Payment:
    company: str
    agency: str
    account: str
    payment_number: str
    client_number: str
    cpf: str
    employee: str
    payment_date: str
    payment_type: str
    amount: Decimal
    source_file: str
    source_path: str
    page: int


@dataclass(frozen=True)
class FileSummary:
    company: str
    source_file: str
    source_path: str
    extracted_count: int
    extracted_total: Decimal
    informed_count: Optional[int]
    informed_total: Optional[Decimal]
    status: str
    observation: str


@dataclass(frozen=True)
class AlertGroup:
    group_id: str
    risk: str
    rule: str
    explanation: str
    payment_indexes: tuple[int, ...]
    potential_exposure: Decimal


@dataclass
class AnalysisResult:
    payments: list[Payment]
    files: list[FileSummary]
    alerts: list[AlertGroup]
    errors: list[str]

    @property
    def total_amount(self) -> Decimal:
        return sum((p.amount for p in self.payments), Decimal("0"))

    @property
    def potential_exposure(self) -> Decimal:
        return sum((a.potential_exposure for a in self.alerts), Decimal("0"))


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def normalize_text(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return normalize_spaces(value).upper()


def digits_only(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def normalize_cpf(client_number: str) -> str:
    digits = digits_only(client_number)
    return digits[-11:].zfill(11)


def format_cpf(cpf: str) -> str:
    cpf = normalize_cpf(cpf)
    return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"


def parse_money(value: str) -> Decimal:
    cleaned = (value or "").replace("R$", "").replace(" ", "")
    cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except InvalidOperation as exc:
        raise ValueError(f"Valor monetário inválido: {value!r}") from exc


def brl(value: Decimal) -> str:
    value = value.quantize(Decimal("0.01"))
    formatted = f"{value:,.2f}"
    return "R$ " + formatted.replace(",", "X").replace(".", ",").replace("X", ".")


def safe_filename_company(pdf_path: Path) -> str:
    name = normalize_text(pdf_path.stem)
    name = re.sub(r"\bGERARPDF\b", "", name)
    name = re.sub(r"\b\d{14}\b", "", name)
    name = normalize_spaces(name.replace("_", " ").replace("-", " "))
    return name or pdf_path.stem.upper()


def detect_company(text: str, pdf_path: Path) -> str:
    patterns = [
        r"Internet Banking Empresarial\s*\n\s*(.+?)\s+Ag[eê]ncia:\s*\d+",
        r"^\s*(.+?)\s+Ag[eê]ncia:\s*\d+\s+Conta:",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I | re.M | re.S)
        if match:
            candidate = normalize_spaces(match.group(1))
            if 2 <= len(candidate) <= 150:
                return candidate.upper()
    return safe_filename_company(pdf_path)


def detect_account(text: str) -> tuple[str, str]:
    match = re.search(
        r"Ag[eê]ncia:\s*([\d.-]+)\s+Conta:\s*([\d.-]+)", text, flags=re.I
    )
    if not match:
        return "", ""
    return digits_only(match.group(1)), digits_only(match.group(2))


def extract_informed_totals(text: str) -> tuple[Optional[int], Optional[Decimal]]:
    count_matches = re.findall(r"Total Compromissos:\s*(\d+)", text, flags=re.I)
    total_matches = re.findall(
        rf"Valor Total:\s*(?:R\$)?\s*({MONEY_RE})", text, flags=re.I
    )
    informed_count = int(count_matches[-1]) if count_matches else None
    informed_total = parse_money(total_matches[-1]) if total_matches else None
    return informed_count, informed_total


def extract_pdf(pdf_path: Path) -> tuple[list[Payment], FileSummary]:
    reader = PdfReader(str(pdf_path))
    if reader.is_encrypted:
        try:
            reader.decrypt("")
        except Exception as exc:
            raise ValueError("PDF protegido por senha") from exc

    page_texts: list[str] = []
    for page in reader.pages:
        page_texts.append(page.extract_text() or "")

    full_text = "\n".join(page_texts)
    company = detect_company(page_texts[0] if page_texts else full_text, pdf_path)
    agency, account = detect_account(page_texts[0] if page_texts else full_text)
    informed_count, informed_total = extract_informed_totals(full_text)

    payments: list[Payment] = []
    for page_number, page_text in enumerate(page_texts, start=1):
        for match in ROW_RE.finditer(page_text):
            payment_number, client_number, employee, date, payment_type, amount = match.groups()
            employee = normalize_spaces(employee)
            payments.append(
                Payment(
                    company=company,
                    agency=agency,
                    account=account,
                    payment_number=payment_number,
                    client_number=client_number,
                    cpf=normalize_cpf(client_number),
                    employee=employee,
                    payment_date=date,
                    payment_type=payment_type,
                    amount=parse_money(amount),
                    source_file=pdf_path.name,
                    source_path=str(pdf_path.resolve()),
                    page=page_number,
                )
            )

    extracted_total = sum((p.amount for p in payments), Decimal("0"))
    observations: list[str] = []
    status = "OK"

    if not payments:
        status = "ERRO"
        observations.append("Nenhum pagamento reconhecido")
    if informed_count is not None and informed_count != len(payments):
        status = "REVISAR"
        observations.append(
            f"PDF informa {informed_count} registros; foram extraídos {len(payments)}"
        )
    if informed_total is not None and informed_total != extracted_total:
        status = "REVISAR"
        observations.append(
            f"PDF informa {brl(informed_total)}; extração somou {brl(extracted_total)}"
        )
    if not observations:
        observations.append("Quantidade e total conferidos")

    summary = FileSummary(
        company=company,
        source_file=pdf_path.name,
        source_path=str(pdf_path.resolve()),
        extracted_count=len(payments),
        extracted_total=extracted_total,
        informed_count=informed_count,
        informed_total=informed_total,
        status=status,
        observation="; ".join(observations),
    )
    return payments, summary


def _group_indexes(payments: list[Payment], key_func: Callable[[Payment], tuple]) -> dict:
    grouped: dict[tuple, list[int]] = defaultdict(list)
    for index, payment in enumerate(payments):
        grouped[key_func(payment)].append(index)
    return grouped


def detect_alerts(payments: list[Payment]) -> list[AlertGroup]:
    alerts: list[AlertGroup] = []
    serial = 1

    def add_alert(
        risk: str,
        rule: str,
        explanation: str,
        indexes: Iterable[int],
        exposure: Decimal = Decimal("0"),
    ) -> None:
        nonlocal serial
        indexes_tuple = tuple(sorted(set(indexes)))
        if len(indexes_tuple) < 2:
            return
        alerts.append(
            AlertGroup(
                group_id=f"ALT-{serial:04d}",
                risk=risk,
                rule=rule,
                explanation=explanation,
                payment_indexes=indexes_tuple,
                potential_exposure=exposure,
            )
        )
        serial += 1

    # 1. Mesmo número de pagamento dentro da mesma conta/empresa.
    number_groups = _group_indexes(
        payments,
        lambda p: (normalize_text(p.company), p.agency, p.account, p.payment_number),
    )
    for indexes in number_groups.values():
        if len(indexes) > 1:
            values = [payments[i].amount for i in indexes]
            exposure = sum(values, Decimal("0")) - min(values)
            add_alert(
                "CRÍTICO",
                "Número de pagamento repetido",
                "O mesmo número de pagamento aparece mais de uma vez na mesma empresa/conta.",
                indexes,
                exposure,
            )

    # 2. Mesmo CPF, data e valor: duplicidade exata, ainda que os números mudem.
    exact_groups = _group_indexes(
        payments,
        lambda p: (p.cpf, p.payment_date, p.amount),
    )
    exact_members: set[tuple[int, ...]] = set()
    for indexes in exact_groups.values():
        if len(indexes) > 1:
            exact_members.add(tuple(sorted(indexes)))
            value = payments[indexes[0]].amount
            add_alert(
                "CRÍTICO",
                "CPF + data + valor repetidos",
                "Pagamentos com o mesmo CPF, mesma data e mesmo valor.",
                indexes,
                value * (len(indexes) - 1),
            )

    # 3. Mesmo CPF e valor, porém em datas diferentes: provável reprocessamento.
    cpf_value_groups = _group_indexes(payments, lambda p: (p.cpf, p.amount))
    for indexes in cpf_value_groups.values():
        dates = {payments[i].payment_date for i in indexes}
        if len(indexes) > 1 and len(dates) > 1:
            value = payments[indexes[0]].amount
            add_alert(
                "ALTO",
                "CPF + valor repetidos em datas diferentes",
                "Possível reprocessamento ou pagamento repetido; validar a natureza da verba.",
                indexes,
                value * (len(indexes) - 1),
            )

    # 4. Mesmo CPF e mesma data com valores diferentes: pode ser complemento.
    cpf_date_groups = _group_indexes(payments, lambda p: (p.cpf, p.payment_date))
    for indexes in cpf_date_groups.values():
        values = {payments[i].amount for i in indexes}
        if len(indexes) > 1 and len(values) > 1:
            add_alert(
                "ATENÇÃO",
                "CPF repetido no mesmo dia com valores diferentes",
                "Pode representar folha, complemento, férias ou rescisão; validar as rubricas.",
                indexes,
            )

    # 5. Mesmo CPF pago por empresas diferentes.
    cpf_groups = _group_indexes(payments, lambda p: (p.cpf,))
    for indexes in cpf_groups.values():
        companies = {normalize_text(payments[i].company) for i in indexes}
        if len(companies) > 1:
            add_alert(
                "ATENÇÃO",
                "CPF presente em empresas diferentes",
                "Validar vínculo, transferência, sucessão, férias ou rescisão entre empresas do grupo.",
                indexes,
            )

    risk_order = {"CRÍTICO": 0, "ALTO": 1, "ATENÇÃO": 2}
    alerts.sort(key=lambda a: (risk_order.get(a.risk, 99), a.group_id))
    return alerts


def collect_pdf_paths(inputs: Iterable[str | Path]) -> list[Path]:
    paths: list[Path] = []
    seen: set[str] = set()
    for raw in inputs:
        path = Path(raw).expanduser()
        candidates = sorted(path.rglob("*.pdf")) if path.is_dir() else [path]
        for candidate in candidates:
            if not candidate.is_file() or candidate.suffix.lower() != ".pdf":
                continue
            resolved = str(candidate.resolve()).casefold()
            if resolved not in seen:
                seen.add(resolved)
                paths.append(candidate.resolve())
    return paths


def analyze_files(
    pdf_paths: Iterable[str | Path],
    progress_callback: Optional[Callable[[int, int, Path], None]] = None,
) -> AnalysisResult:
    paths = collect_pdf_paths(pdf_paths)
    payments: list[Payment] = []
    summaries: list[FileSummary] = []
    errors: list[str] = []

    total = len(paths)
    for position, path in enumerate(paths, start=1):
        if progress_callback:
            progress_callback(position, total, path)
        try:
            file_payments, summary = extract_pdf(path)
            payments.extend(file_payments)
            summaries.append(summary)
        except Exception as exc:
            errors.append(f"{path.name}: {exc}")
            summaries.append(
                FileSummary(
                    company=safe_filename_company(path),
                    source_file=path.name,
                    source_path=str(path),
                    extracted_count=0,
                    extracted_total=Decimal("0"),
                    informed_count=None,
                    informed_total=None,
                    status="ERRO",
                    observation=str(exc),
                )
            )

    alerts = detect_alerts(payments)
    return AnalysisResult(payments=payments, files=summaries, alerts=alerts, errors=errors)


def _apply_worksheet_style(ws, table_name: Optional[str] = None) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor=SONOVA_BLUE)
    header_font = Font(color=WHITE, bold=True)
    thin_gray = Side(style="thin", color="D9E2E8")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin_gray)
    ws.row_dimensions[1].height = 28

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
            cell.border = Border(bottom=Side(style="hair", color="E6EDF1"))

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 11), 48)

    if table_name and ws.max_row >= 2 and ws.max_column >= 1:
        safe_name = re.sub(r"\W", "", table_name)
        table = Table(displayName=safe_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def export_excel(result: AnalysisResult, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumo"

    summary_headers = [
        "Indicador",
        "Valor",
        "Observação",
    ]
    ws_summary.append(summary_headers)
    ws_summary.append(["Arquivos processados", len(result.files), "Relatórios PDF selecionados"])
    ws_summary.append(["Pagamentos extraídos", len(result.payments), "Linhas consolidadas"])
    ws_summary.append(["Valor total", float(result.total_amount), "Soma dos pagamentos"])
    ws_summary.append(["Grupos de alerta", len(result.alerts), "Ocorrências para validação"])
    ws_summary.append(
        ["Exposição potencial", float(result.potential_exposure), "Estimativa das duplicidades de valor"]
    )
    ws_summary.append(
        ["Alertas críticos", sum(a.risk == "CRÍTICO" for a in result.alerts), "Maior prioridade"]
    )
    ws_summary.append(["Alertas altos", sum(a.risk == "ALTO" for a in result.alerts), "Provável duplicidade"])
    ws_summary.append(
        ["Alertas de atenção", sum(a.risk == "ATENÇÃO" for a in result.alerts), "Requer conciliação"]
    )
    ws_summary.append(["Versão do aplicativo", APP_VERSION, APP_NAME])
    for row in range(2, ws_summary.max_row + 1):
        if ws_summary.cell(row, 1).value in {"Valor total", "Exposição potencial"}:
            ws_summary.cell(row, 2).number_format = 'R$ #,##0.00'
    _apply_worksheet_style(ws_summary, "TabelaResumo")

    ws_files = wb.create_sheet("Arquivos")
    ws_files.append(
        [
            "Empresa",
            "Arquivo",
            "Caminho",
            "Qtd. extraída",
            "Total extraído",
            "Qtd. informada",
            "Total informado",
            "Status",
            "Observação",
        ]
    )
    for item in result.files:
        ws_files.append(
            [
                item.company,
                item.source_file,
                item.source_path,
                item.extracted_count,
                float(item.extracted_total),
                item.informed_count,
                float(item.informed_total) if item.informed_total is not None else None,
                item.status,
                item.observation,
            ]
        )
    for row in range(2, ws_files.max_row + 1):
        ws_files.cell(row, 5).number_format = 'R$ #,##0.00'
        ws_files.cell(row, 7).number_format = 'R$ #,##0.00'
    _apply_worksheet_style(ws_files, "TabelaArquivos")

    ws_alerts = wb.create_sheet("Alertas")
    ws_alerts.append(
        [
            "Grupo",
            "Risco",
            "Regra",
            "Explicação",
            "Exposição potencial",
            "Empresa",
            "CPF",
            "CPF sem máscara",
            "Funcionário",
            "Data",
            "Número do pagamento",
            "Valor",
            "Tipo",
            "Arquivo",
            "Página",
        ]
    )
    for alert in result.alerts:
        for index in alert.payment_indexes:
            payment = result.payments[index]
            ws_alerts.append(
                [
                    alert.group_id,
                    alert.risk,
                    alert.rule,
                    alert.explanation,
                    float(alert.potential_exposure),
                    payment.company,
                    format_cpf(payment.cpf),
                    payment.cpf,
                    payment.employee,
                    datetime.strptime(payment.payment_date, "%d/%m/%Y").date(),
                    payment.payment_number,
                    float(payment.amount),
                    payment.payment_type,
                    payment.source_file,
                    payment.page,
                ]
            )
    for row in range(2, ws_alerts.max_row + 1):
        ws_alerts.cell(row, 5).number_format = 'R$ #,##0.00'
        ws_alerts.cell(row, 8).number_format = "@"
        ws_alerts.cell(row, 10).number_format = "dd/mm/yyyy"
        ws_alerts.cell(row, 11).number_format = "@"
        ws_alerts.cell(row, 12).number_format = 'R$ #,##0.00'
        risk = ws_alerts.cell(row, 2).value
        color = DANGER if risk == "CRÍTICO" else WARNING if risk == "ALTO" else SONOVA_LIGHT
        ws_alerts.cell(row, 2).fill = PatternFill("solid", fgColor=color)
        ws_alerts.cell(row, 2).font = Font(color=WHITE if risk != "ATENÇÃO" else TEXT, bold=True)
    _apply_worksheet_style(ws_alerts, "TabelaAlertas")

    ws_payments = wb.create_sheet("Pagamentos")
    ws_payments.append(
        [
            "Empresa",
            "Agência",
            "Conta",
            "Número do pagamento",
            "Número do cliente",
            "CPF",
            "CPF sem máscara",
            "Funcionário",
            "Data",
            "Tipo",
            "Valor",
            "Arquivo",
            "Caminho",
            "Página",
        ]
    )
    for payment in result.payments:
        ws_payments.append(
            [
                payment.company,
                payment.agency,
                payment.account,
                payment.payment_number,
                payment.client_number,
                format_cpf(payment.cpf),
                payment.cpf,
                payment.employee,
                datetime.strptime(payment.payment_date, "%d/%m/%Y").date(),
                payment.payment_type,
                float(payment.amount),
                payment.source_file,
                payment.source_path,
                payment.page,
            ]
        )
    for row in range(2, ws_payments.max_row + 1):
        for col in (2, 3, 4, 5, 7):
            ws_payments.cell(row, col).number_format = "@"
        ws_payments.cell(row, 9).number_format = "dd/mm/yyyy"
        ws_payments.cell(row, 11).number_format = 'R$ #,##0.00'
    _apply_worksheet_style(ws_payments, "TabelaPagamentos")

    ws_errors = wb.create_sheet("Logs")
    ws_errors.append(["Tipo", "Mensagem"])
    if result.errors:
        for error in result.errors:
            ws_errors.append(["ERRO", error])
    else:
        ws_errors.append(["INFO", "Processamento concluído sem erros de leitura."])
    _apply_worksheet_style(ws_errors, "TabelaLogs")

    wb.save(output_path)
    return output_path.resolve()


def open_local_path(path: str | Path) -> None:
    path = str(Path(path).resolve())
    if sys.platform.startswith("win"):
        os.startfile(path)  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])


class AuditorApp:
    def __init__(self) -> None:
        import tkinter as tk
        from tkinter import ttk

        self.tk = tk
        self.ttk = ttk
        self.root = tk.Tk()
        self.root.title(f"{APP_NAME} | Sonova")
        self.root.geometry("1380x860")
        self.root.minsize(1080, 680)

        self.pdf_paths: list[Path] = []
        self.result: Optional[AnalysisResult] = None
        self.alert_rows: list[tuple[AlertGroup, Payment]] = []
        self.settings_path = Path(__file__).with_name("auditor_pagamentos_config.json")
        self.last_directory = str(Path.home())

        self._load_settings()
        self._configure_styles()
        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _load_settings(self) -> None:
        try:
            data = json.loads(self.settings_path.read_text(encoding="utf-8"))
            self.last_directory = data.get("last_directory", self.last_directory)
            geometry = data.get("geometry")
            if geometry:
                self.root.geometry(geometry)
        except Exception:
            pass

    def _save_settings(self) -> None:
        data = {
            "last_directory": self.last_directory,
            "geometry": self.root.geometry(),
            "version": APP_VERSION,
        }
        try:
            self.settings_path.write_text(
                json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
            )
        except Exception:
            pass

    def _configure_styles(self) -> None:
        style = self.ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("TFrame", background="#FFFFFF")
        style.configure("TLabel", background="#FFFFFF", foreground="#333333", font=("Segoe UI", 10))
        style.configure("Title.TLabel", font=("Georgia", 22, "bold"), foreground="#003C64")
        style.configure("Section.TLabel", font=("Segoe UI", 11, "bold"), foreground="#003C64")
        style.configure("Kpi.TLabel", font=("Segoe UI", 19, "bold"), foreground="#003C64")
        style.configure("KpiCaption.TLabel", font=("Segoe UI", 9), foreground="#5B6770")
        style.configure("Primary.TButton", background="#0083CA", foreground="#FFFFFF", font=("Segoe UI", 10, "bold"), padding=(14, 8))
        style.map("Primary.TButton", background=[("active", "#003C64"), ("disabled", "#AAB7C0")])
        style.configure("Accent.TButton", background="#7D0041", foreground="#FFFFFF", font=("Segoe UI", 10, "bold"), padding=(14, 8))
        style.map("Accent.TButton", background=[("active", "#59002E"), ("disabled", "#B7A0AC")])
        style.configure("Secondary.TButton", background="#EAF4FA", foreground="#003C64", padding=(12, 7))
        style.map("Secondary.TButton", background=[("active", "#CDE7F5")])
        style.configure("Treeview", rowheight=27, font=("Segoe UI", 9), background="#FFFFFF", fieldbackground="#FFFFFF")
        style.configure("Treeview.Heading", background="#003C64", foreground="#FFFFFF", font=("Segoe UI", 9, "bold"), padding=6)
        style.map("Treeview.Heading", background=[("active", "#0083CA")])
        style.configure("TNotebook", background="#FFFFFF", borderwidth=0)
        style.configure("TNotebook.Tab", padding=(16, 8), font=("Segoe UI", 10, "bold"))
        style.map("TNotebook.Tab", background=[("selected", "#0083CA")], foreground=[("selected", "#FFFFFF")])
        style.configure("Sonova.Horizontal.TProgressbar", troughcolor="#E6EDF1", background="#0083CA")

    def _build_ui(self) -> None:
        tk, ttk = self.tk, self.ttk

        header = tk.Frame(self.root, bg="#0083CA", height=88)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="SONOVA", bg="#0083CA", fg="white", font=("Segoe UI", 20, "bold")).pack(side="left", padx=(28, 18))
        tk.Frame(header, bg="#6EB4DC", width=1, height=48).pack(side="left", pady=20)
        title_box = tk.Frame(header, bg="#0083CA")
        title_box.pack(side="left", padx=18, pady=14)
        tk.Label(title_box, text="Auditor de Pagamentos Santander", bg="#0083CA", fg="white", font=("Georgia", 19, "bold")).pack(anchor="w")
        tk.Label(title_box, text="Cruzamento de duplicidades | Payroll Brasil", bg="#0083CA", fg="white", font=("Segoe UI", 10)).pack(anchor="w", pady=(3, 0))
        tk.Label(header, text=f"v{APP_VERSION}", bg="#0083CA", fg="#D9F1FF", font=("Segoe UI", 9)).pack(side="right", padx=24)

        body = ttk.Frame(self.root, padding=(22, 16, 22, 10))
        body.pack(fill="both", expand=True)

        toolbar = ttk.Frame(body)
        toolbar.pack(fill="x")
        ttk.Button(toolbar, text="+ Selecionar PDFs", style="Primary.TButton", command=self.select_files).pack(side="left")
        ttk.Button(toolbar, text="Selecionar pasta", style="Secondary.TButton", command=self.select_folder).pack(side="left", padx=(8, 0))
        ttk.Button(toolbar, text="Remover selecionados", style="Secondary.TButton", command=self.remove_selected_files).pack(side="left", padx=(8, 0))
        ttk.Button(toolbar, text="Limpar", style="Secondary.TButton", command=self.clear_all).pack(side="left", padx=(8, 0))
        self.export_button = ttk.Button(toolbar, text="Exportar Excel", style="Accent.TButton", command=self.export_result, state="disabled")
        self.export_button.pack(side="right")
        self.analyze_button = ttk.Button(toolbar, text="Analisar cruzamentos", style="Primary.TButton", command=self.start_analysis)
        self.analyze_button.pack(side="right", padx=(0, 8))

        file_frame = ttk.Frame(body)
        file_frame.pack(fill="x", pady=(14, 10))
        ttk.Label(file_frame, text="Arquivos selecionados", style="Section.TLabel").pack(anchor="w", pady=(0, 5))
        file_table_frame = ttk.Frame(file_frame)
        file_table_frame.pack(fill="x")
        self.file_tree = ttk.Treeview(file_table_frame, columns=("arquivo", "empresa", "registros", "total", "status"), show="headings", height=4)
        headings = {
            "arquivo": ("Arquivo", 380),
            "empresa": ("Empresa", 340),
            "registros": ("Registros", 90),
            "total": ("Total", 120),
            "status": ("Status", 120),
        }
        for column, (title, width) in headings.items():
            self.file_tree.heading(column, text=title)
            self.file_tree.column(column, width=width, anchor="w" if column in {"arquivo", "empresa"} else "center")
        file_scroll = ttk.Scrollbar(file_table_frame, orient="vertical", command=self.file_tree.yview)
        self.file_tree.configure(yscrollcommand=file_scroll.set)
        self.file_tree.pack(side="left", fill="x", expand=True)
        file_scroll.pack(side="right", fill="y")
        self.file_tree.bind("<Double-1>", self.open_selected_file)

        kpi_frame = ttk.Frame(body)
        kpi_frame.pack(fill="x", pady=(0, 10))
        self.kpi_values: dict[str, object] = {}
        kpis = [
            ("payments", "0", "Pagamentos"),
            ("total", "R$ 0,00", "Valor processado"),
            ("alerts", "0", "Grupos de alerta"),
            ("exposure", "R$ 0,00", "Exposição potencial"),
        ]
        for key, initial, caption in kpis:
            card = tk.Frame(kpi_frame, bg="#F2F5F7", highlightbackground="#DCE5EA", highlightthickness=1)
            card.pack(side="left", fill="x", expand=True, padx=(0, 9) if key != "exposure" else 0)
            value_label = tk.Label(card, text=initial, bg="#F2F5F7", fg="#003C64", font=("Segoe UI", 18, "bold"))
            value_label.pack(anchor="w", padx=14, pady=(9, 0))
            tk.Label(card, text=caption, bg="#F2F5F7", fg="#5B6770", font=("Segoe UI", 9)).pack(anchor="w", padx=14, pady=(0, 9))
            self.kpi_values[key] = value_label

        filter_frame = ttk.Frame(body)
        filter_frame.pack(fill="x", pady=(0, 8))
        ttk.Label(filter_frame, text="Filtro:").pack(side="left")
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(filter_frame, textvariable=self.search_var, width=42)
        search_entry.pack(side="left", padx=(6, 12))
        search_entry.bind("<KeyRelease>", lambda _event: self.refresh_views())
        ttk.Label(filter_frame, text="Risco:").pack(side="left")
        self.risk_var = tk.StringVar(value="Todos")
        risk_combo = ttk.Combobox(filter_frame, textvariable=self.risk_var, values=["Todos", "CRÍTICO", "ALTO", "ATENÇÃO"], state="readonly", width=13)
        risk_combo.pack(side="left", padx=(6, 0))
        risk_combo.bind("<<ComboboxSelected>>", lambda _event: self.refresh_views())

        notebook = ttk.Notebook(body)
        notebook.pack(fill="both", expand=True)
        self.alert_tab = ttk.Frame(notebook)
        self.payment_tab = ttk.Frame(notebook)
        self.company_tab = ttk.Frame(notebook)
        notebook.add(self.alert_tab, text="Alertas")
        notebook.add(self.payment_tab, text="Todos os pagamentos")
        notebook.add(self.company_tab, text="Resumo por empresa")

        self.alert_tree = self._make_tree(
            self.alert_tab,
            [
                ("grupo", "Grupo", 78), ("risco", "Risco", 82), ("regra", "Regra", 270),
                ("cpf", "CPF", 115), ("nome", "Funcionário", 230), ("empresa", "Empresa", 210),
                ("data", "Data", 90), ("pagamento", "Nº pagamento", 105), ("valor", "Valor", 100),
                ("arquivo", "Arquivo", 210), ("pagina", "Pág.", 50),
            ],
        )
        self.alert_tree.tag_configure("CRÍTICO", background="#FDECEC", foreground="#8A1010")
        self.alert_tree.tag_configure("ALTO", background="#FFF4D9", foreground="#7A4B00")
        self.alert_tree.tag_configure("ATENÇÃO", background="#EAF4FA", foreground="#003C64")
        self.alert_tree.bind("<Double-1>", self.open_alert_file)

        self.payment_tree = self._make_tree(
            self.payment_tab,
            [
                ("empresa", "Empresa", 230), ("cpf", "CPF", 115), ("nome", "Funcionário", 250),
                ("data", "Data", 90), ("pagamento", "Nº pagamento", 110), ("tipo", "Tipo", 60),
                ("valor", "Valor", 110), ("arquivo", "Arquivo", 250), ("pagina", "Pág.", 50),
            ],
        )
        self.payment_tree.bind("<Double-1>", self.open_payment_file)

        self.company_tree = self._make_tree(
            self.company_tab,
            [
                ("empresa", "Empresa", 390), ("arquivos", "Arquivos", 80),
                ("pagamentos", "Pagamentos", 100), ("cpfs", "CPFs únicos", 100),
                ("total", "Valor total", 140), ("alertas", "Grupos de alerta", 130),
            ],
        )

        footer = ttk.Frame(self.root, padding=(22, 5, 22, 10))
        footer.pack(fill="x")
        self.status_var = tk.StringVar(value="Selecione os relatórios PDF para iniciar.")
        ttk.Label(footer, textvariable=self.status_var).pack(side="left")
        self.progress = ttk.Progressbar(footer, mode="determinate", length=240, style="Sonova.Horizontal.TProgressbar")
        self.progress.pack(side="right")

    def _make_tree(self, parent, columns: list[tuple[str, str, int]]):
        frame = self.ttk.Frame(parent, padding=8)
        frame.pack(fill="both", expand=True)
        tree = self.ttk.Treeview(frame, columns=[c[0] for c in columns], show="headings")
        for key, title, width in columns:
            tree.heading(key, text=title)
            anchor = "e" if key in {"valor", "total"} else "center" if key in {"data", "pagina", "pagamento", "tipo", "risco", "grupo", "arquivos", "pagamentos", "cpfs", "alertas"} else "w"
            tree.column(key, width=width, anchor=anchor, stretch=key in {"empresa", "nome", "arquivo", "regra"})
        y_scroll = self.ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        x_scroll = self.ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        return tree

    def select_files(self) -> None:
        from tkinter import filedialog

        files = filedialog.askopenfilenames(
            title="Selecionar relatórios PDF do Santander",
            initialdir=self.last_directory,
            filetypes=[("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*")],
        )
        if files:
            self.last_directory = str(Path(files[0]).parent)
            self._add_paths(files)

    def select_folder(self) -> None:
        from tkinter import filedialog

        folder = filedialog.askdirectory(title="Selecionar pasta com PDFs", initialdir=self.last_directory)
        if folder:
            self.last_directory = folder
            self._add_paths([folder])

    def _add_paths(self, raw_paths: Iterable[str | Path]) -> None:
        existing = {str(path).casefold() for path in self.pdf_paths}
        for path in collect_pdf_paths(raw_paths):
            if str(path).casefold() not in existing:
                self.pdf_paths.append(path)
                existing.add(str(path).casefold())
        self.pdf_paths.sort(key=lambda p: p.name.casefold())
        self.result = None
        self.export_button.configure(state="disabled")
        self.refresh_file_tree()
        self.status_var.set(f"{len(self.pdf_paths)} arquivo(s) selecionado(s).")

    def refresh_file_tree(self) -> None:
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        summary_by_path = {s.source_path.casefold(): s for s in self.result.files} if self.result else {}
        for index, path in enumerate(self.pdf_paths):
            summary = summary_by_path.get(str(path.resolve()).casefold())
            values = (
                path.name,
                summary.company if summary else "Aguardando análise",
                summary.extracted_count if summary else "-",
                brl(summary.extracted_total) if summary else "-",
                summary.status if summary else "PENDENTE",
            )
            self.file_tree.insert("", "end", iid=str(index), values=values)

    def remove_selected_files(self) -> None:
        selected = sorted((int(iid) for iid in self.file_tree.selection()), reverse=True)
        for index in selected:
            if 0 <= index < len(self.pdf_paths):
                self.pdf_paths.pop(index)
        self.result = None
        self.export_button.configure(state="disabled")
        self.refresh_file_tree()
        self.clear_result_views()
        self.status_var.set(f"{len(self.pdf_paths)} arquivo(s) selecionado(s).")

    def clear_all(self) -> None:
        self.pdf_paths.clear()
        self.result = None
        self.export_button.configure(state="disabled")
        self.refresh_file_tree()
        self.clear_result_views()
        self.status_var.set("Selecione os relatórios PDF para iniciar.")

    def clear_result_views(self) -> None:
        for tree in (self.alert_tree, self.payment_tree, self.company_tree):
            for item in tree.get_children():
                tree.delete(item)
        self.alert_rows.clear()
        for key, value in {"payments": "0", "total": "R$ 0,00", "alerts": "0", "exposure": "R$ 0,00"}.items():
            self.kpi_values[key].configure(text=value)

    def start_analysis(self) -> None:
        from tkinter import messagebox

        if not self.pdf_paths:
            messagebox.showwarning(APP_NAME, "Selecione pelo menos um arquivo PDF.")
            return
        self.analyze_button.configure(state="disabled")
        self.export_button.configure(state="disabled")
        self.progress.configure(value=0, maximum=max(len(self.pdf_paths), 1))
        self.status_var.set("Iniciando leitura dos PDFs...")
        thread = threading.Thread(target=self._analysis_worker, daemon=True)
        thread.start()

    def _analysis_worker(self) -> None:
        def progress(current: int, total: int, path: Path) -> None:
            self.root.after(0, lambda: self._update_progress(current, total, path.name))

        result = analyze_files(self.pdf_paths, progress_callback=progress)
        self.root.after(0, lambda: self._finish_analysis(result))

    def _update_progress(self, current: int, total: int, filename: str) -> None:
        self.progress.configure(maximum=max(total, 1), value=current - 1)
        self.status_var.set(f"Lendo {current}/{total}: {filename}")

    def _finish_analysis(self, result: AnalysisResult) -> None:
        from tkinter import messagebox

        self.result = result
        self.progress.configure(value=len(self.pdf_paths))
        self.analyze_button.configure(state="normal")
        self.export_button.configure(state="normal" if result.payments else "disabled")
        self.refresh_file_tree()
        self.refresh_views()
        self.kpi_values["payments"].configure(text=f"{len(result.payments):,}".replace(",", "."))
        self.kpi_values["total"].configure(text=brl(result.total_amount))
        self.kpi_values["alerts"].configure(text=str(len(result.alerts)))
        self.kpi_values["exposure"].configure(text=brl(result.potential_exposure))
        if result.errors:
            self.status_var.set(f"Concluído com {len(result.errors)} erro(s). Consulte o Excel para detalhes.")
            messagebox.showwarning(APP_NAME, "A análise terminou, mas alguns arquivos não puderam ser lidos:\n\n" + "\n".join(result.errors[:8]))
        else:
            self.status_var.set(
                f"Análise concluída: {len(result.payments)} pagamentos e {len(result.alerts)} grupo(s) de alerta."
            )

    def refresh_views(self) -> None:
        if not self.result:
            return
        search = normalize_text(self.search_var.get())
        selected_risk = self.risk_var.get()

        for item in self.alert_tree.get_children():
            self.alert_tree.delete(item)
        self.alert_rows = []
        for alert in self.result.alerts:
            if selected_risk != "Todos" and alert.risk != selected_risk:
                continue
            for index in alert.payment_indexes:
                payment = self.result.payments[index]
                searchable = normalize_text(
                    " ".join(
                        [
                            alert.group_id, alert.risk, alert.rule, payment.cpf,
                            payment.employee, payment.company, payment.payment_date,
                            payment.payment_number, payment.source_file,
                        ]
                    )
                )
                if search and search not in searchable:
                    continue
                iid = str(len(self.alert_rows))
                self.alert_rows.append((alert, payment))
                self.alert_tree.insert(
                    "", "end", iid=iid,
                    values=(
                        alert.group_id, alert.risk, alert.rule, format_cpf(payment.cpf),
                        payment.employee, payment.company, payment.payment_date,
                        payment.payment_number, brl(payment.amount), payment.source_file, payment.page,
                    ),
                    tags=(alert.risk,),
                )

        for item in self.payment_tree.get_children():
            self.payment_tree.delete(item)
        for index, payment in enumerate(self.result.payments):
            searchable = normalize_text(
                " ".join(
                    [payment.company, payment.cpf, payment.employee, payment.payment_date,
                     payment.payment_number, payment.source_file]
                )
            )
            if search and search not in searchable:
                continue
            self.payment_tree.insert(
                "", "end", iid=str(index),
                values=(
                    payment.company, format_cpf(payment.cpf), payment.employee,
                    payment.payment_date, payment.payment_number, payment.payment_type,
                    brl(payment.amount), payment.source_file, payment.page,
                ),
            )

        for item in self.company_tree.get_children():
            self.company_tree.delete(item)
        grouped: dict[str, list[int]] = defaultdict(list)
        for index, payment in enumerate(self.result.payments):
            grouped[payment.company].append(index)
        for company, indexes in sorted(grouped.items()):
            company_payments = [self.result.payments[i] for i in indexes]
            files = {p.source_path for p in company_payments}
            cpfs = {p.cpf for p in company_payments}
            alert_count = sum(
                1
                for alert in self.result.alerts
                if any(i in indexes for i in alert.payment_indexes)
            )
            self.company_tree.insert(
                "", "end",
                values=(
                    company, len(files), len(company_payments), len(cpfs),
                    brl(sum((p.amount for p in company_payments), Decimal("0"))), alert_count,
                ),
            )

    def export_result(self) -> None:
        from tkinter import filedialog, messagebox

        if not self.result:
            return
        default_name = f"Auditoria_Pagamentos_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        output = filedialog.asksaveasfilename(
            title="Salvar relatório de auditoria",
            initialdir=self.last_directory,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")],
        )
        if not output:
            return
        try:
            saved = export_excel(self.result, output)
            self.last_directory = str(saved.parent)
            self.status_var.set(f"Excel salvo: {saved.name}")
            if messagebox.askyesno(APP_NAME, "Relatório exportado com sucesso.\n\nDeseja abrir o arquivo agora?"):
                open_local_path(saved)
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"Não foi possível exportar o Excel:\n\n{exc}")

    def open_selected_file(self, _event=None) -> None:
        selection = self.file_tree.selection()
        if selection:
            index = int(selection[0])
            if 0 <= index < len(self.pdf_paths):
                open_local_path(self.pdf_paths[index])

    def open_alert_file(self, _event=None) -> None:
        selection = self.alert_tree.selection()
        if selection:
            index = int(selection[0])
            if 0 <= index < len(self.alert_rows):
                open_local_path(self.alert_rows[index][1].source_path)

    def open_payment_file(self, _event=None) -> None:
        if not self.result:
            return
        selection = self.payment_tree.selection()
        if selection:
            index = int(selection[0])
            if 0 <= index < len(self.result.payments):
                open_local_path(self.result.payments[index].source_path)

    def _on_close(self) -> None:
        self._save_settings()
        self.root.destroy()

    def run(self) -> None:
        self.root.mainloop()


def run_cli(inputs: list[str], output: Optional[str]) -> int:
    paths = collect_pdf_paths(inputs)
    if not paths:
        print("Nenhum arquivo PDF encontrado.", file=sys.stderr)
        return 2

    def progress(current: int, total: int, path: Path) -> None:
        print(f"[{current}/{total}] {path.name}")

    result = analyze_files(paths, progress_callback=progress)
    print(f"Pagamentos: {len(result.payments)}")
    print(f"Valor total: {brl(result.total_amount)}")
    print(f"Grupos de alerta: {len(result.alerts)}")
    print(f"Exposição potencial: {brl(result.potential_exposure)}")
    for alert in result.alerts:
        print(
            f"{alert.group_id} | {alert.risk} | {alert.rule} | "
            f"{len(alert.payment_indexes)} pagamentos | {brl(alert.potential_exposure)}"
        )
    if result.errors:
        print("Erros:", file=sys.stderr)
        for error in result.errors:
            print(f"- {error}", file=sys.stderr)

    if output:
        saved = export_excel(result, output)
        print(f"Excel: {saved}")
    return 1 if result.errors else 0


def main() -> int:
    parser = argparse.ArgumentParser(description=APP_NAME)
    parser.add_argument("inputs", nargs="*", help="Arquivos PDF ou pastas")
    parser.add_argument("--cli", action="store_true", help="Executa sem interface gráfica")
    parser.add_argument("--export", help="Caminho do Excel de saída")
    args = parser.parse_args()

    if args.cli or args.inputs:
        return run_cli(args.inputs, args.export)

    app = AuditorApp()
    app.run()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
