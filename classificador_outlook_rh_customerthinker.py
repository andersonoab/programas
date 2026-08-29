import os
import re
import json
import time
import threading
import traceback
from pathlib import Path
from datetime import datetime, date

import customtkinter as ctk
from tkinter import filedialog, messagebox

import pythoncom
import win32com.client as win32

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


APP_NAME = "Leitor Outlook RH - CustomerThinker"
APP_VERSION = "1.1.0"

# ---------------------------------------------------------------------------
# IA (Claude / Anthropic)
# ---------------------------------------------------------------------------
# Modelos disponiveis no seletor. Rotulo amigavel -> id tecnico do modelo.
IA_MODELOS = {
    "Claude Haiku 4.5 (rapido e economico)": "claude-haiku-4-5-20251001",
    "Claude Sonnet 5 (respostas mais elaboradas)": "claude-sonnet-5",
}
IA_MODELO_PADRAO = "claude-haiku-4-5-20251001"

# Limite de caracteres do corpo enviado a IA (minimizacao de dados / custo).
IA_LIMITE_CORPO = 4000

BASE_DIR = Path(r"C:\_RPA\AppEmailRH")
SAIDA_DIR = BASE_DIR / "saida"
LOGS_DIR = BASE_DIR / "logs"
CONFIG_DIR = BASE_DIR / "config"
CONFIG_FILE = CONFIG_DIR / "config_email_rh.json"

DEFAULT_EXCEL_NAME = "base_atendimento_gente_gestao.xlsx"

CORES = {
    # Paleta Sonova OFC
    "azul": "#0083CA",         # dominante
    "azul_apoio": "#003C64",   # apoio (barra lateral, textos de titulo)
    "verde_apoio": "#005A64",  # apoio
    "azul_claro": "#6EB4DC",   # apoio / botao secundario
    "vinho": "#7D0041",        # acento
    "telha": "#8C321E",        # acento
    "navy": "#003C64",         # compat.: antigo "navy" agora aponta para o apoio Sonova
    "fundo": "#FFFFFF",        # fundo branco
    "fundo_lateral": "#003C64",# barra lateral
    "card": "#FFFFFF",
    "borda": "#CCCCCC",        # borda de card (nunca azul claro)
    "texto": "#333333",
    "texto_sec": "#646464",
    "sucesso": "#2E7D32",      # verde = avancar
    "erro": "#B91C1C",         # vermelho = destrutivo
    "alerta": "#B7791F",
    "auxiliar": "#E0E0E0",     # cinza = auxiliar
    "auxiliar_hover": "#CFCFCF",
    "branco": "#FFFFFF",
}


COLUNAS_BASE = [
    "EntryID",
    "Data Recebimento",
    "Hora Recebimento",
    "Dias Desde Recebimento",
    "Remetente Nome",
    "Remetente E-mail",
    "Para",
    "CC",
    "Assunto",
    "Corpo do E-mail",
    "Corpo Limpo",
    "Texto Truncado",
    "Tem Anexo",
    "Quantidade de Anexos",
    "Nomes dos Anexos",
    "Status Outlook",
    "Lido",
    "Importância Outlook",
    "Tamanho do Texto",
    "Data Extração",
    "Análise ChatGPT",
    "Urgência Sugerida",
    "Motivo da Urgência",
    "Categoria Sugerida",
    "Ação Recomendada",
    "Responsável Sugerido",
    "Observação RH",
    "Status Atendimento",
    "Resposta Sugerida (IA)",
    "Rascunho Outlook Criado",
]


COLUNAS_GLOSSARIO = [
    "Palavra-chave",
    "Categoria",
    "Prioridade",
    "Ação sugerida",
    "Responsável",
    "Observação",
    "Ativo",
]


def garantir_pastas():
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    SAIDA_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)


def agora_texto():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def hoje_data():
    return datetime.now().date()


def limpar_texto(texto):
    if texto is None:
        return ""

    texto = str(texto)
    texto = texto.replace("\r", "\n")
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    texto = re.sub(r"[ \t]+", " ", texto)
    texto = texto.replace("\x00", "")
    return texto.strip()


def limpar_texto_uma_linha(texto):
    texto = limpar_texto(texto)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def limitar_excel(texto, limite=32000):
    texto = "" if texto is None else str(texto)
    if len(texto) <= limite:
        return texto, "Não"
    return texto[:limite] + "\n\n[TEXTO TRUNCADO PARA COMPATIBILIDADE COM EXCEL]", "Sim"


def normalizar_data_outlook(valor):
    if valor is None:
        return None

    try:
        if hasattr(valor, "date"):
            return valor
    except Exception:
        pass

    try:
        return datetime.fromtimestamp(time.mktime(valor.timetuple()))
    except Exception:
        return None


def dias_desde_recebimento(recebido):
    if not recebido:
        return ""

    try:
        data_email = recebido.date()
        return (hoje_data() - data_email).days
    except Exception:
        return ""


def traduzir_importancia(valor):
    try:
        valor = int(valor)
    except Exception:
        return ""

    mapa = {
        0: "Baixa",
        1: "Normal",
        2: "Alta",
    }
    return mapa.get(valor, str(valor))


def carregar_config():
    garantir_pastas()

    config_padrao = {
        "modo_caixa": "principal",
        "caixa_compartilhada": "",
        "pasta_outlook": "Caixa de Entrada",
        "pasta_entry_id": "",
        "pasta_store_id": "",
        "pasta_display": "",
        "quantidade": 200,
        "ultimos_dias": 30,
        "somente_nao_lidos": False,
        "pasta_saida": str(SAIDA_DIR),
        "nome_excel": DEFAULT_EXCEL_NAME,
        "ia_ativa": False,
        "criar_rascunho_outlook": False,
        "ia_modelo": IA_MODELO_PADRAO,
        "ia_api_key": "",
        "ia_workspace_id": "",
        "ia_assinatura": "Atenciosamente,\nGente e Gestão – Sonova",
        "ia_contexto_rh": (
            "Você atende a caixa de Gente e Gestão (RH/DP) da Sonova Brasil. "
            "Os remetentes costumam ser colaboradores e gestores com dúvidas "
            "sobre folha de pagamento, benefícios, férias, admissão, "
            "desligamento, atestados e afastamentos."
        ),
    }

    if not CONFIG_FILE.exists():
        salvar_config(config_padrao)
        return config_padrao

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            config_lido = json.load(f)

        for chave, valor in config_padrao.items():
            config_lido.setdefault(chave, valor)

        return config_lido

    except Exception:
        return config_padrao


def salvar_config(config):
    garantir_pastas()
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=4)


def registrar_log_arquivo(mensagem):
    garantir_pastas()
    caminho_log = LOGS_DIR / "log_execucao.txt"
    with open(caminho_log, "a", encoding="utf-8") as f:
        f.write(f"[{agora_texto()}] {mensagem}\n")


def carregar_glossario(caminho_excel):
    """
    Função reservada para próxima iteração.

    Objetivo futuro:
    Ler a aba GLOSSARIO do Excel e usar as palavras-chave para sugerir
    categoria, prioridade, responsável e ação recomendada.

    Nesta primeira versão, a função não aplica classificação.
    """
    return []


def classificar_por_glossario(texto_email, glossario):
    """
    Função reservada para próxima iteração.

    Futuramente:
    - Receberá o texto do e-mail
    - Comparará com o glossário
    - Retornará categoria, prioridade, palavras encontradas,
      responsável e ação sugerida

    Nesta primeira versão:
    Retorna valores em branco ou Não classificado.
    """
    return {
        "analise_chatgpt": "",
        "urgencia": "",
        "motivo_urgencia": "",
        "categoria": "Não classificado",
        "acao_recomendada": "",
        "responsavel": "",
        "observacao": "",
        "status_atendimento": "Pendente",
    }


# ===========================================================================
# IA (Claude / Anthropic) - analise e geracao de resposta
# ===========================================================================

def _extrair_json(texto):
    """
    Extrai o primeiro objeto JSON de um texto, tolerando cercas ```json
    e eventual preambulo do modelo. Retorna dict (ou {} em falha).
    """
    if not texto:
        return {}

    limpo = str(texto).strip()
    limpo = limpo.replace("```json", "").replace("```", "").strip()

    try:
        return json.loads(limpo)
    except Exception:
        pass

    inicio = limpo.find("{")
    fim = limpo.rfind("}")
    if inicio != -1 and fim != -1 and fim > inicio:
        trecho = limpo[inicio:fim + 1]
        try:
            return json.loads(trecho)
        except Exception:
            return {}

    return {}


def analisar_email_com_ia(assunto, corpo, remetente_nome, remetente_email, config_ia, log_callback=None):
    """
    Envia o e-mail ao Claude e retorna classificacao + resposta sugerida.

    config_ia: dict com ia_modelo, ia_api_key, ia_contexto_rh.
    Em qualquer falha, retorna um dicionario de fallback (nada trava a extracao).
    A biblioteca 'anthropic' e importada aqui dentro, entao o app roda
    normalmente mesmo sem ela instalada quando a IA esta desligada.
    """
    fallback = {
        "analise_chatgpt": "",
        "urgencia": "",
        "motivo_urgencia": "",
        "categoria": "Não classificado",
        "acao_recomendada": "",
        "responsavel": "",
        "observacao": "",
        "status_atendimento": "Pendente",
        "resposta_sugerida": "",
    }

    api_key = (config_ia.get("ia_api_key") or "").strip() or os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        fallback["analise_chatgpt"] = "[IA ligada, mas nenhuma chave da API foi informada]"
        return fallback

    modelo = (config_ia.get("ia_modelo") or IA_MODELO_PADRAO).strip()
    contexto_rh = (config_ia.get("ia_contexto_rh") or "").strip()
    workspace_id = (config_ia.get("ia_workspace_id") or "").strip() or os.environ.get("ANTHROPIC_WORKSPACE_ID", "").strip()

    corpo_curto = (corpo or "")[:IA_LIMITE_CORPO]

    system_prompt = (
        "Você é um assistente de atendimento de RH (Gente e Gestão) que trabalha "
        "em português do Brasil. " + contexto_rh + " "
        "Sua tarefa é ler um e-mail recebido e devolver uma análise objetiva mais "
        "um rascunho de resposta ao remetente. Regras: seja profissional e "
        "empático; NÃO invente políticas, prazos, valores ou regras internas; se "
        "faltar informação para responder com segurança, escreva uma resposta que "
        "acuse o recebimento e peça os dados que faltam de forma cordial; marque "
        "precisa_revisao_humana como true sempre que houver risco jurídico, dado "
        "sensível, reclamação, rescisão, ponto/hora extra ou qualquer assunto que "
        "exija validação de um humano. "
        "Responda EXCLUSIVAMENTE com um objeto JSON válido, sem texto fora dele, "
        "sem comentários e sem cercas de código, exatamente com estas chaves: "
        "analise, urgencia, motivo_urgencia, categoria, acao_recomendada, "
        "responsavel, observacao, status_atendimento, resposta_sugerida, "
        "precisa_revisao_humana. "
        "urgencia deve ser Alta, Média ou Baixa. "
        "resposta_sugerida é o texto pronto para o remetente (sem assinatura, "
        "ela é adicionada depois)."
    )

    user_content = (
        f"Remetente: {remetente_nome} <{remetente_email}>\n"
        f"Assunto: {assunto}\n\n"
        f"Corpo do e-mail:\n{corpo_curto}"
    )

    try:
        from anthropic import Anthropic
    except Exception:
        fallback["analise_chatgpt"] = "[Biblioteca 'anthropic' não instalada. Rode: pip install anthropic]"
        if log_callback:
            log_callback("Biblioteca 'anthropic' não encontrada. Instale com: pip install anthropic")
        return fallback

    try:
        default_headers = {}
        if workspace_id:
            default_headers["anthropic-workspace-id"] = workspace_id

        if default_headers:
            client = Anthropic(api_key=api_key, default_headers=default_headers)
        else:
            client = Anthropic(api_key=api_key)

        resposta = client.messages.create(
            model=modelo,
            max_tokens=1200,
            system=system_prompt,
            messages=[{"role": "user", "content": user_content}],
        )

        partes = []
        for bloco in resposta.content:
            texto_bloco = getattr(bloco, "text", None)
            if texto_bloco:
                partes.append(texto_bloco)
        texto_final = "\n".join(partes).strip()

        dados = _extrair_json(texto_final)

        if not dados:
            fallback["analise_chatgpt"] = "[IA respondeu em formato inesperado]"
            fallback["resposta_sugerida"] = texto_final
            return fallback

        observacao = str(dados.get("observacao", "") or "")
        if bool(dados.get("precisa_revisao_humana", False)):
            marca = "REVISÃO HUMANA RECOMENDADA."
            observacao = (marca + " " + observacao).strip()

        return {
            "analise_chatgpt": str(dados.get("analise", "") or ""),
            "urgencia": str(dados.get("urgencia", "") or ""),
            "motivo_urgencia": str(dados.get("motivo_urgencia", "") or ""),
            "categoria": str(dados.get("categoria", "") or "Não classificado"),
            "acao_recomendada": str(dados.get("acao_recomendada", "") or ""),
            "responsavel": str(dados.get("responsavel", "") or ""),
            "observacao": observacao,
            "status_atendimento": str(dados.get("status_atendimento", "") or "Pendente"),
            "resposta_sugerida": str(dados.get("resposta_sugerida", "") or ""),
        }

    except Exception as e:
        if log_callback:
            log_callback(f"Falha na chamada à IA: {e}")
        fallback["analise_chatgpt"] = f"[Erro na IA: {e}]"
        return fallback


def criar_rascunho_reply(mail, corpo_resposta, assinatura, log_callback=None):
    """
    Cria um rascunho de RESPOSTA no Outlook e o salva na pasta Rascunhos.
    NUNCA envia. O usuário revisa e clica em enviar manualmente.
    Retorna True se o rascunho foi salvo, False caso contrário.
    """
    corpo_resposta = (corpo_resposta or "").strip()
    if not corpo_resposta:
        return False

    try:
        reply = mail.Reply()  # cria a resposta (somente remetente), sem enviar

        try:
            corpo_original = reply.Body or ""
        except Exception:
            corpo_original = ""

        novo_corpo = corpo_resposta
        if assinatura:
            novo_corpo += "\n\n" + assinatura
        novo_corpo += "\n\n" + corpo_original

        reply.Body = novo_corpo
        reply.Save()  # salva em Rascunhos; nao dispara envio
        return True

    except Exception as e:
        if log_callback:
            log_callback(f"Falha ao criar rascunho no Outlook: {e}")
        return False


def get_smtp_address(mail):
    """
    Tenta obter o e-mail SMTP real do remetente.
    Funciona melhor em contas Exchange/Outlook corporativas.
    """
    try:
        if getattr(mail, "SenderEmailType", "") == "EX":
            try:
                exchange_user = mail.Sender.GetExchangeUser()
                if exchange_user:
                    smtp = exchange_user.PrimarySmtpAddress
                    if smtp:
                        return smtp
            except Exception:
                pass

            try:
                PR_SMTP_ADDRESS = "http://schemas.microsoft.com/mapi/proptag/0x39FE001E"
                smtp = mail.Sender.PropertyAccessor.GetProperty(PR_SMTP_ADDRESS)
                if smtp:
                    return smtp
            except Exception:
                pass

        return getattr(mail, "SenderEmailAddress", "") or ""

    except Exception:
        return ""


def obter_pasta_por_nome(pasta_raiz, nome_pasta):
    """
    Procura uma pasta pelo nome dentro da árvore do Outlook.
    Se nome_pasta for vazio ou Caixa de Entrada, usa a própria pasta recebida.
    """
    nome_pasta = (nome_pasta or "").strip()

    if nome_pasta == "":
        return pasta_raiz

    # Tolerancia: se o usuario digitou um caminho (ex.: "conta\Caixa de Entrada\Sub"),
    # usa apenas o ultimo segmento como nome da pasta a procurar.
    if "\\" in nome_pasta or "/" in nome_pasta:
        partes = [p.strip() for p in re.split(r"[\\/]", nome_pasta) if p.strip()]
        if partes:
            nome_pasta = partes[-1]

    nomes_padrao_inbox = {
        "caixa de entrada",
        "inbox",
        "entrada",
    }

    if nome_pasta.lower() in nomes_padrao_inbox:
        return pasta_raiz

    def buscar_recursivo(pasta):
        try:
            if pasta.Name.strip().lower() == nome_pasta.lower():
                return pasta
        except Exception:
            pass

        try:
            for subpasta in pasta.Folders:
                achou = buscar_recursivo(subpasta)
                if achou is not None:
                    return achou
        except Exception:
            pass

        return None

    encontrado = buscar_recursivo(pasta_raiz)

    if encontrado is None:
        raise Exception(f"Pasta não encontrada no Outlook: {nome_pasta}")

    return encontrado


def obter_pasta_outlook(modo_caixa, caixa_compartilhada, pasta_outlook,
                        pasta_entry_id="", pasta_store_id="", log_callback=None):
    outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")

    # 1) Se o usuario escolheu uma pasta pelo seletor do Outlook, ela tem prioridade.
    pasta_entry_id = (pasta_entry_id or "").strip()
    pasta_store_id = (pasta_store_id or "").strip()
    if pasta_entry_id and pasta_store_id:
        try:
            pasta_final = outlook.GetFolderFromID(pasta_entry_id, pasta_store_id)
            if log_callback:
                try:
                    log_callback(f"Pasta selecionada: {pasta_final.FolderPath}")
                except Exception:
                    log_callback("Pasta selecionada pelo seletor do Outlook.")
            return pasta_final
        except Exception as e:
            if log_callback:
                log_callback(f"Não foi possível abrir a pasta escolhida ({e}). Usando os campos de origem.")

    modo_caixa = (modo_caixa or "principal").strip().lower()
    caixa_compartilhada = (caixa_compartilhada or "").strip()
    pasta_outlook = (pasta_outlook or "Caixa de Entrada").strip()

    if modo_caixa == "compartilhada":
        if not caixa_compartilhada:
            raise Exception("Informe o nome ou e-mail da caixa compartilhada.")

        if log_callback:
            log_callback(f"Tentando acessar caixa compartilhada: {caixa_compartilhada}")

        recipient = outlook.CreateRecipient(caixa_compartilhada)
        recipient.Resolve()

        if not recipient.Resolved:
            raise Exception(f"Não foi possível resolver a caixa compartilhada: {caixa_compartilhada}")

        inbox = outlook.GetSharedDefaultFolder(recipient, 6)

    else:
        if log_callback:
            log_callback("Acessando Caixa de Entrada principal do Outlook.")
        inbox = outlook.GetDefaultFolder(6)

    pasta_final = obter_pasta_por_nome(inbox, pasta_outlook)

    if log_callback:
        try:
            log_callback(f"Pasta selecionada: {pasta_final.Name}")
        except Exception:
            log_callback("Pasta selecionada com sucesso.")

    return pasta_final


def extrair_dados_email(mail, config_ia=None, log_callback=None):
    try:
        entry_id = getattr(mail, "EntryID", "") or ""
    except Exception:
        entry_id = ""

    try:
        recebido = normalizar_data_outlook(getattr(mail, "ReceivedTime", None))
    except Exception:
        recebido = None

    data_recebimento = ""
    hora_recebimento = ""

    if recebido:
        try:
            data_recebimento = recebido.strftime("%d/%m/%Y")
            hora_recebimento = recebido.strftime("%H:%M:%S")
        except Exception:
            pass

    try:
        remetente_nome = getattr(mail, "SenderName", "") or ""
    except Exception:
        remetente_nome = ""

    remetente_email = get_smtp_address(mail)

    try:
        para = getattr(mail, "To", "") or ""
    except Exception:
        para = ""

    try:
        cc = getattr(mail, "CC", "") or ""
    except Exception:
        cc = ""

    try:
        assunto = getattr(mail, "Subject", "") or ""
    except Exception:
        assunto = ""

    try:
        corpo = getattr(mail, "Body", "") or ""
    except Exception:
        corpo = ""

    corpo = limpar_texto(corpo)
    corpo_limpo = limpar_texto_uma_linha(corpo)

    corpo_excel, truncado = limitar_excel(corpo)
    corpo_limpo_excel, truncado_limpo = limitar_excel(corpo_limpo)

    texto_truncado = "Sim" if truncado == "Sim" or truncado_limpo == "Sim" else "Não"

    nomes_anexos = []
    qtd_anexos = 0

    try:
        qtd_anexos = int(mail.Attachments.Count)
        for i in range(1, qtd_anexos + 1):
            try:
                nomes_anexos.append(mail.Attachments.Item(i).FileName)
            except Exception:
                nomes_anexos.append("Anexo não identificado")
    except Exception:
        qtd_anexos = 0

    tem_anexo = "Sim" if qtd_anexos > 0 else "Não"

    try:
        nao_lido = bool(getattr(mail, "UnRead", False))
        lido = "Não" if nao_lido else "Sim"
    except Exception:
        lido = ""

    try:
        importancia = traduzir_importancia(getattr(mail, "Importance", ""))
    except Exception:
        importancia = ""

    tamanho_texto = len(corpo)

    if config_ia and config_ia.get("ia_ativa"):
        classificacao = analisar_email_com_ia(
            assunto=assunto,
            corpo=corpo,
            remetente_nome=remetente_nome,
            remetente_email=remetente_email,
            config_ia=config_ia,
            log_callback=log_callback,
        )
    else:
        glossario = []
        classificacao = classificar_por_glossario(
            f"{assunto}\n{corpo}",
            glossario
        )
        classificacao.setdefault("resposta_sugerida", "")

    linha = {
        "EntryID": entry_id,
        "Data Recebimento": data_recebimento,
        "Hora Recebimento": hora_recebimento,
        "Dias Desde Recebimento": dias_desde_recebimento(recebido),
        "Remetente Nome": remetente_nome,
        "Remetente E-mail": remetente_email,
        "Para": para,
        "CC": cc,
        "Assunto": assunto,
        "Corpo do E-mail": corpo_excel,
        "Corpo Limpo": corpo_limpo_excel,
        "Texto Truncado": texto_truncado,
        "Tem Anexo": tem_anexo,
        "Quantidade de Anexos": qtd_anexos,
        "Nomes dos Anexos": " | ".join(nomes_anexos),
        "Status Outlook": "Não lido" if lido == "Não" else "Lido",
        "Lido": lido,
        "Importância Outlook": importancia,
        "Tamanho do Texto": tamanho_texto,
        "Data Extração": agora_texto(),
        "Análise ChatGPT": classificacao["analise_chatgpt"],
        "Urgência Sugerida": classificacao["urgencia"],
        "Motivo da Urgência": classificacao["motivo_urgencia"],
        "Categoria Sugerida": classificacao["categoria"],
        "Ação Recomendada": classificacao["acao_recomendada"],
        "Responsável Sugerido": classificacao["responsavel"],
        "Observação RH": classificacao["observacao"],
        "Status Atendimento": classificacao["status_atendimento"],
        "Resposta Sugerida (IA)": classificacao.get("resposta_sugerida", ""),
        "Rascunho Outlook Criado": "Não",
    }

    return linha


def ler_emails_outlook(
    modo_caixa,
    caixa_compartilhada,
    pasta_outlook,
    quantidade,
    ultimos_dias,
    somente_nao_lidos,
    config_ia=None,
    pasta_entry_id="",
    pasta_store_id="",
    progress_callback=None,
    log_callback=None,
):
    pythoncom.CoInitialize()

    try:
        pasta = obter_pasta_outlook(
            modo_caixa=modo_caixa,
            caixa_compartilhada=caixa_compartilhada,
            pasta_outlook=pasta_outlook,
            pasta_entry_id=pasta_entry_id,
            pasta_store_id=pasta_store_id,
            log_callback=log_callback,
        )

        items = pasta.Items
        items.Sort("[ReceivedTime]", True)

        total_outlook = items.Count

        if log_callback:
            log_callback(f"Total aparente de itens na pasta: {total_outlook}")

        registros = []
        erros = []

        quantidade = int(quantidade)
        ultimos_dias = int(ultimos_dias)

        limite_data = None
        if ultimos_dias > 0:
            limite_data = hoje_data()
            if log_callback:
                log_callback(f"Filtro de período ativado: últimos {ultimos_dias} dias.")

        analisados = 0
        coletados = 0

        for item in items:
            if coletados >= quantidade:
                break

            analisados += 1

            try:
                item_class = getattr(item, "Class", None)
                if item_class != 43:
                    continue

                try:
                    recebido = normalizar_data_outlook(getattr(item, "ReceivedTime", None))
                except Exception:
                    recebido = None

                if ultimos_dias > 0 and recebido:
                    dias = dias_desde_recebimento(recebido)
                    if isinstance(dias, int) and dias > ultimos_dias:
                        continue

                if somente_nao_lidos:
                    try:
                        if not bool(getattr(item, "UnRead", False)):
                            continue
                    except Exception:
                        continue

                dados = extrair_dados_email(item, config_ia=config_ia, log_callback=log_callback)

                if (
                    config_ia
                    and config_ia.get("ia_ativa")
                    and config_ia.get("criar_rascunho_outlook")
                    and dados.get("Resposta Sugerida (IA)")
                ):
                    ok_rascunho = criar_rascunho_reply(
                        mail=item,
                        corpo_resposta=dados.get("Resposta Sugerida (IA)", ""),
                        assinatura=config_ia.get("ia_assinatura", ""),
                        log_callback=log_callback,
                    )
                    dados["Rascunho Outlook Criado"] = "Sim" if ok_rascunho else "Falha"

                registros.append(dados)
                coletados += 1

                if progress_callback:
                    progress_callback(coletados, quantidade)

                if log_callback and coletados % 10 == 0:
                    log_callback(f"E-mails coletados: {coletados}")

            except Exception as e:
                erros.append(str(e))
                continue

        if log_callback:
            log_callback(f"Leitura finalizada. Itens analisados: {analisados}. E-mails coletados: {coletados}.")

        return registros, erros

    finally:
        pythoncom.CoUninitialize()


def aplicar_estilo_planilha(ws, tipo="base"):
    azul = CORES["azul"].replace("#", "")
    navy = CORES["navy"].replace("#", "")
    branco = "FFFFFF"
    cinza_claro = "F2F2F2"
    borda_cor = "D9DDE3"

    fill_header = PatternFill("solid", fgColor=navy)
    fill_sub = PatternFill("solid", fgColor=cinza_claro)
    font_header = Font(color=branco, bold=True)
    font_titulo = Font(color=navy, bold=True, size=14)
    border = Border(
        left=Side(style="thin", color=borda_cor),
        right=Side(style="thin", color=borda_cor),
        top=Side(style="thin", color=borda_cor),
        bottom=Side(style="thin", color=borda_cor),
    )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    larguras = {}

    if tipo == "base":
        larguras = {
            "A": 28,
            "B": 16,
            "C": 14,
            "D": 18,
            "E": 28,
            "F": 32,
            "G": 36,
            "H": 28,
            "I": 45,
            "J": 80,
            "K": 80,
            "L": 16,
            "M": 14,
            "N": 18,
            "O": 40,
            "P": 18,
            "Q": 12,
            "R": 18,
            "S": 18,
            "T": 22,
            "U": 35,
            "V": 18,
            "W": 35,
            "X": 22,
            "Y": 35,
            "Z": 24,
            "AA": 35,
            "AB": 18,
            "AC": 90,
            "AD": 22,
        }
    elif tipo == "glossario":
        larguras = {
            "A": 28,
            "B": 24,
            "C": 18,
            "D": 40,
            "E": 24,
            "F": 40,
            "G": 12,
        }
    elif tipo == "resumo":
        larguras = {
            "A": 35,
            "B": 35,
        }

    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    ws.row_dimensions[1].height = 28

    if tipo == "base":
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 80


def criar_aba_base(wb, registros):
    ws = wb.active
    ws.title = "BASE_EMAILS"

    ws.append(COLUNAS_BASE)

    for registro in registros:
        ws.append([registro.get(coluna, "") for coluna in COLUNAS_BASE])

    aplicar_estilo_planilha(ws, tipo="base")


def criar_aba_glossario(wb):
    ws = wb.create_sheet("GLOSSARIO")
    ws.append(COLUNAS_GLOSSARIO)

    exemplos = [
        ["", "", "", "", "", "Preencher futuramente com termos validados pela operação.", "Sim"],
    ]

    for linha in exemplos:
        ws.append(linha)

    aplicar_estilo_planilha(ws, tipo="glossario")


def criar_aba_resumo(wb, registros):
    ws = wb.create_sheet("RESUMO")

    total = len(registros)
    total_anexo = sum(1 for r in registros if r.get("Tem Anexo") == "Sim")
    total_sem_anexo = total - total_anexo
    total_nao_lidos = sum(1 for r in registros if r.get("Lido") == "Não")
    total_lidos = sum(1 for r in registros if r.get("Lido") == "Sim")
    total_com_resposta = sum(1 for r in registros if str(r.get("Resposta Sugerida (IA)", "")).strip())
    total_rascunhos = sum(1 for r in registros if r.get("Rascunho Outlook Criado") == "Sim")

    datas = []
    for r in registros:
        data_txt = r.get("Data Recebimento", "")
        try:
            datas.append(datetime.strptime(data_txt, "%d/%m/%Y").date())
        except Exception:
            pass

    data_inicial = min(datas).strftime("%d/%m/%Y") if datas else ""
    data_final = max(datas).strftime("%d/%m/%Y") if datas else ""

    linhas = [
        ["Indicador", "Valor"],
        ["Total de e-mails extraídos", total],
        ["Total com anexo", total_anexo],
        ["Total sem anexo", total_sem_anexo],
        ["Total não lidos", total_nao_lidos],
        ["Total lidos", total_lidos],
        ["Data inicial", data_inicial],
        ["Data final", data_final],
        ["Respostas sugeridas pela IA", total_com_resposta],
        ["Rascunhos criados no Outlook", total_rascunhos],
        ["Data da extração", agora_texto()],
        ["Aplicativo", f"{APP_NAME} v{APP_VERSION}"],
        ["Observação", "Classificação e resposta geradas pela IA (Claude) quando ativada; caso contrário, ficam pendentes para o glossário."],
    ]

    for linha in linhas:
        ws.append(linha)

    aplicar_estilo_planilha(ws, tipo="resumo")

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=CORES["navy"].replace("#", ""))
        cell.font = Font(color="FFFFFF", bold=True)


def gerar_excel(registros, caminho_saida):
    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    if caminho_saida.exists():
        try:
            os.rename(caminho_saida, caminho_saida)
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            novo_nome = caminho_saida.with_name(f"{caminho_saida.stem}_{timestamp}{caminho_saida.suffix}")
            caminho_saida = novo_nome

    wb = Workbook()
    criar_aba_base(wb, registros)
    criar_aba_glossario(wb)
    criar_aba_resumo(wb, registros)

    wb.save(caminho_saida)
    return caminho_saida


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        garantir_pastas()

        self.config_app = carregar_config()
        self.processo_rodando = False
        self.caminho_excel_gerado = None

        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.title(APP_NAME)
        self.geometry("1240x880")
        self.minsize(1120, 760)
        self.configure(fg_color=CORES["fundo"])

        self.criar_variaveis()
        self.criar_layout()
        self.carregar_variaveis_config()

    def criar_variaveis(self):
        self.var_modo_caixa = ctk.StringVar(value="principal")
        self.var_caixa_compartilhada = ctk.StringVar(value="")
        self.var_pasta_outlook = ctk.StringVar(value="Caixa de Entrada")
        self.var_pasta_entry_id = ctk.StringVar(value="")
        self.var_pasta_store_id = ctk.StringVar(value="")
        self.var_pasta_display = ctk.StringVar(value="")
        self.var_quantidade = ctk.StringVar(value="200")
        self.var_ultimos_dias = ctk.StringVar(value="30")
        self.var_somente_nao_lidos = ctk.BooleanVar(value=False)
        self.var_pasta_saida = ctk.StringVar(value=str(SAIDA_DIR))
        self.var_nome_excel = ctk.StringVar(value=DEFAULT_EXCEL_NAME)

        self.var_ia_ativa = ctk.BooleanVar(value=False)
        self.var_criar_rascunho = ctk.BooleanVar(value=False)
        self.var_ia_api_key = ctk.StringVar(value="")
        self.var_ia_workspace_id = ctk.StringVar(value="")
        self.var_ia_modelo_rotulo = ctk.StringVar(value=self._rotulo_por_modelo(IA_MODELO_PADRAO))

    def _rotulo_por_modelo(self, modelo_id):
        for rotulo, mid in IA_MODELOS.items():
            if mid == modelo_id:
                return rotulo
        return next(iter(IA_MODELOS))

    def _modelo_por_rotulo(self, rotulo):
        return IA_MODELOS.get(rotulo, IA_MODELO_PADRAO)

    def carregar_variaveis_config(self):
        self.var_modo_caixa.set(self.config_app.get("modo_caixa", "principal"))
        self.var_caixa_compartilhada.set(self.config_app.get("caixa_compartilhada", ""))
        self.var_pasta_outlook.set(self.config_app.get("pasta_outlook", "Caixa de Entrada"))
        self.var_pasta_entry_id.set(self.config_app.get("pasta_entry_id", ""))
        self.var_pasta_store_id.set(self.config_app.get("pasta_store_id", ""))
        self.var_pasta_display.set(self.config_app.get("pasta_display", ""))
        self.var_quantidade.set(str(self.config_app.get("quantidade", 200)))
        self.var_ultimos_dias.set(str(self.config_app.get("ultimos_dias", 30)))
        self.var_somente_nao_lidos.set(bool(self.config_app.get("somente_nao_lidos", False)))
        self.var_pasta_saida.set(self.config_app.get("pasta_saida", str(SAIDA_DIR)))
        self.var_nome_excel.set(self.config_app.get("nome_excel", DEFAULT_EXCEL_NAME))

        self.var_ia_ativa.set(bool(self.config_app.get("ia_ativa", False)))
        self.var_criar_rascunho.set(bool(self.config_app.get("criar_rascunho_outlook", False)))
        self.var_ia_api_key.set(self.config_app.get("ia_api_key", ""))
        self.var_ia_workspace_id.set(self.config_app.get("ia_workspace_id", ""))
        self.var_ia_modelo_rotulo.set(self._rotulo_por_modelo(self.config_app.get("ia_modelo", IA_MODELO_PADRAO)))

    def obter_config_tela(self):
        return {
            "modo_caixa": self.var_modo_caixa.get(),
            "caixa_compartilhada": self.var_caixa_compartilhada.get().strip(),
            "pasta_outlook": self.var_pasta_outlook.get().strip(),
            "pasta_entry_id": self.var_pasta_entry_id.get().strip(),
            "pasta_store_id": self.var_pasta_store_id.get().strip(),
            "pasta_display": self.var_pasta_display.get().strip(),
            "quantidade": int(self.var_quantidade.get()),
            "ultimos_dias": int(self.var_ultimos_dias.get()),
            "somente_nao_lidos": bool(self.var_somente_nao_lidos.get()),
            "pasta_saida": self.var_pasta_saida.get().strip(),
            "nome_excel": self.var_nome_excel.get().strip(),
            "ia_ativa": bool(self.var_ia_ativa.get()),
            "criar_rascunho_outlook": bool(self.var_criar_rascunho.get()),
            "ia_modelo": self._modelo_por_rotulo(self.var_ia_modelo_rotulo.get()),
            "ia_api_key": self.var_ia_api_key.get().strip(),
            "ia_workspace_id": self.var_ia_workspace_id.get().strip(),
            "ia_assinatura": self.config_app.get("ia_assinatura", "Atenciosamente,\nGente e Gestão – Sonova"),
            "ia_contexto_rh": self.config_app.get("ia_contexto_rh", ""),
        }

    def criar_layout(self):
        self.grid_columnconfigure(0, weight=0)   # barra lateral (largura fixa)
        self.grid_columnconfigure(1, weight=1)   # conteudo
        self.grid_rowconfigure(0, weight=1)

        self.views = {}
        self.nav_botoes = {}

        self.criar_sidebar()

        direita = ctk.CTkFrame(self, fg_color=CORES["fundo"], corner_radius=0)
        direita.grid(row=0, column=1, sticky="nsew")
        direita.grid_columnconfigure(0, weight=1)
        direita.grid_rowconfigure(1, weight=1)

        self.criar_header(direita)
        self.criar_area_conteudo(direita)
        self.criar_rodape(direita)

        self.mostrar_view("config")

    def criar_sidebar(self):
        barra = ctk.CTkFrame(self, fg_color=CORES["fundo_lateral"], corner_radius=0, width=232)
        barra.grid(row=0, column=0, sticky="nsew")
        barra.grid_propagate(False)
        barra.grid_columnconfigure(0, weight=1)

        marca = ctk.CTkLabel(
            barra,
            text="Sonova",
            font=("Segoe UI", 22, "bold"),
            text_color=CORES["branco"],
            anchor="w",
        )
        marca.grid(row=0, column=0, padx=22, pady=(26, 0), sticky="ew")

        sub = ctk.CTkLabel(
            barra,
            text="Gente e Gestão",
            font=("Segoe UI", 12),
            text_color=CORES["azul_claro"],
            anchor="w",
        )
        sub.grid(row=1, column=0, padx=22, pady=(0, 26), sticky="ew")

        itens = [
            ("config", "Origem e Saída"),
            ("ia", "Assistente de IA"),
            ("exec", "Execução"),
        ]

        linha = 2
        for chave, texto in itens:
            btn = ctk.CTkButton(
                barra,
                text=texto,
                anchor="w",
                command=lambda c=chave: self.mostrar_view(c),
                height=44,
                corner_radius=8,
                fg_color="transparent",
                hover_color=CORES["azul"],
                text_color=CORES["branco"],
                font=("Segoe UI", 14),
            )
            btn.grid(row=linha, column=0, padx=12, pady=4, sticky="ew")
            self.nav_botoes[chave] = btn
            linha += 1

        barra.grid_rowconfigure(linha, weight=1)

        versao = ctk.CTkLabel(
            barra,
            text=f"Leitor Outlook RH · v{APP_VERSION}",
            font=("Segoe UI", 11),
            text_color=CORES["azul_claro"],
            anchor="w",
            justify="left",
            wraplength=190,
        )
        versao.grid(row=linha + 1, column=0, padx=22, pady=(0, 16), sticky="sw")

    def mostrar_view(self, chave):
        for k, frame in self.views.items():
            if k == chave:
                frame.grid()
            else:
                frame.grid_remove()

        for k, btn in self.nav_botoes.items():
            if k == chave:
                btn.configure(fg_color=CORES["azul"])
            else:
                btn.configure(fg_color="transparent")

    def criar_header(self, parent):
        header = ctk.CTkFrame(parent, fg_color=CORES["azul"], corner_radius=0, height=92)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        titulo = ctk.CTkLabel(
            header,
            text="Leitor Outlook RH",
            font=("Segoe UI", 26, "bold"),
            text_color=CORES["branco"],
            anchor="w",
        )
        titulo.grid(row=0, column=0, padx=28, pady=(18, 0), sticky="ew")

        subtitulo = ctk.CTkLabel(
            header,
            text="Leitura e exportação segura dos e-mails de Gente e Gestão. Nada é enviado; com IA, gera rascunhos de resposta para sua revisão.",
            font=("Segoe UI", 14),
            text_color=CORES["branco"],
            anchor="w",
        )
        subtitulo.grid(row=1, column=0, padx=28, pady=(2, 18), sticky="ew")

    def _nova_view_scroll(self, container):
        view = ctk.CTkScrollableFrame(
            container,
            fg_color=CORES["fundo"],
            scrollbar_button_color=CORES["azul_claro"],
            scrollbar_button_hover_color=CORES["azul"],
        )
        view.grid(row=0, column=0, sticky="nsew", padx=18, pady=18)
        return view

    def _novo_card(self, parent, coluna=0, colspan=1, padx=0):
        card = ctk.CTkFrame(
            parent,
            fg_color=CORES["card"],
            border_color=CORES["borda"],
            border_width=1,
            corner_radius=14,
        )
        card.grid(row=0, column=coluna, columnspan=colspan, sticky="nsew", padx=padx)
        card.grid_columnconfigure(1, weight=1)
        return card

    def criar_area_conteudo(self, parent):
        container = ctk.CTkFrame(parent, fg_color=CORES["fundo"], corner_radius=0)
        container.grid(row=1, column=0, sticky="nsew")
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(0, weight=1)

        # View: Origem e Saida
        view_cfg = self._nova_view_scroll(container)
        view_cfg.grid_columnconfigure(0, weight=1)
        view_cfg.grid_columnconfigure(1, weight=1)
        self.views["config"] = view_cfg
        card_origem = self._novo_card(view_cfg, coluna=0, padx=(0, 10))
        card_saida = self._novo_card(view_cfg, coluna=1, padx=(10, 0))
        self.criar_card_origem(card_origem)
        self.criar_card_saida(card_saida)

        # View: Assistente de IA
        view_ia = self._nova_view_scroll(container)
        view_ia.grid_columnconfigure(0, weight=1)
        self.views["ia"] = view_ia
        card_ia = self._novo_card(view_ia, coluna=0)
        self.criar_card_ia(card_ia)

        # View: Execucao
        view_exec = ctk.CTkFrame(container, fg_color=CORES["fundo"], corner_radius=0)
        view_exec.grid(row=0, column=0, sticky="nsew", padx=18, pady=18)
        view_exec.grid_columnconfigure(0, weight=1)
        view_exec.grid_rowconfigure(0, weight=1)
        self.views["exec"] = view_exec
        self.criar_area_log(view_exec)

    def criar_card_ia(self, parent):
        self.criar_titulo_card(parent, "Assistente de IA (Claude)")

        chk_ia = ctk.CTkCheckBox(
            parent,
            text="Gerar análise e resposta com IA (Claude)",
            variable=self.var_ia_ativa,
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        chk_ia.grid(row=1, column=0, columnspan=3, padx=18, pady=(4, 4), sticky="w")

        chk_rascunho = ctk.CTkCheckBox(
            parent,
            text="Criar rascunho de resposta no Outlook (pasta Rascunhos, sem enviar)",
            variable=self.var_criar_rascunho,
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        chk_rascunho.grid(row=2, column=0, columnspan=3, padx=18, pady=(0, 8), sticky="w")

        self.criar_label(parent, "Modelo", 3)
        menu_modelo = ctk.CTkOptionMenu(
            parent,
            variable=self.var_ia_modelo_rotulo,
            values=list(IA_MODELOS.keys()),
            fg_color=CORES["azul"],
            button_color=CORES["azul_apoio"],
            button_hover_color=CORES["navy"],
            text_color=CORES["branco"],
        )
        menu_modelo.grid(row=3, column=1, columnspan=2, padx=(4, 18), pady=7, sticky="ew")

        self.criar_label(parent, "Chave da API", 4)
        entry_key = ctk.CTkEntry(
            parent,
            textvariable=self.var_ia_api_key,
            placeholder_text="ANTHROPIC_API_KEY (deixe vazio para usar a variável de ambiente)",
            height=34,
            show="•",
            fg_color=CORES["branco"],
            border_color=CORES["borda"],
            text_color=CORES["texto"],
        )
        entry_key.grid(row=4, column=1, columnspan=2, padx=(4, 18), pady=7, sticky="ew")

        self.criar_label(parent, "Workspace ID", 5)
        entry_ws = ctk.CTkEntry(
            parent,
            textvariable=self.var_ia_workspace_id,
            placeholder_text="wrkspc_... (só se a chave for multi-workspace / vinculada à identidade)",
            height=34,
            fg_color=CORES["branco"],
            border_color=CORES["borda"],
            text_color=CORES["texto"],
        )
        entry_ws.grid(row=5, column=1, columnspan=2, padx=(4, 18), pady=7, sticky="ew")

        info = ctk.CTkLabel(
            parent,
            text=(
                "A IA lê assunto e corpo do e-mail e sugere classificação + resposta. "
                "O corpo é enviado à API da Anthropic — envie apenas o necessário (LGPD) e "
                "restrinja o acesso à pasta de configuração, pois a chave é salva nela. "
                "Preencha o Workspace ID apenas se sua chave for multi-workspace (erro "
                "\"anthropic-workspace-id is required\"). "
                "Nenhum e-mail é enviado: o rascunho fica em Rascunhos para sua revisão."
            ),
            font=("Segoe UI", 12),
            text_color=CORES["texto_sec"],
            justify="left",
            wraplength=980,
            anchor="w",
        )
        info.grid(row=6, column=0, columnspan=3, padx=18, pady=(6, 16), sticky="ew")

    def criar_titulo_card(self, parent, texto):
        label = ctk.CTkLabel(
            parent,
            text=texto,
            font=("Segoe UI", 17, "bold"),
            text_color=CORES["navy"],
            anchor="w",
        )
        label.grid(row=0, column=0, columnspan=3, padx=18, pady=(16, 10), sticky="ew")

    def criar_label(self, parent, texto, row):
        label = ctk.CTkLabel(
            parent,
            text=texto,
            font=("Segoe UI", 13),
            text_color=CORES["texto"],
            anchor="w",
        )
        label.grid(row=row, column=0, padx=(18, 8), pady=7, sticky="w")
        return label

    def criar_entry(self, parent, variable, row, placeholder=""):
        entry = ctk.CTkEntry(
            parent,
            textvariable=variable,
            placeholder_text=placeholder,
            height=34,
            fg_color=CORES["branco"],
            border_color=CORES["borda"],
            text_color=CORES["texto"],
        )
        entry.grid(row=row, column=1, padx=(4, 18), pady=7, sticky="ew")
        return entry

    def criar_card_origem(self, parent):
        self.criar_titulo_card(parent, "Origem dos e-mails")

        radio_principal = ctk.CTkRadioButton(
            parent,
            text="Caixa principal",
            variable=self.var_modo_caixa,
            value="principal",
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        radio_principal.grid(row=1, column=0, padx=18, pady=5, sticky="w")

        radio_compartilhada = ctk.CTkRadioButton(
            parent,
            text="Caixa compartilhada",
            variable=self.var_modo_caixa,
            value="compartilhada",
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        radio_compartilhada.grid(row=1, column=1, padx=8, pady=5, sticky="w")

        self.criar_label(parent, "Caixa compartilhada", 2)
        self.criar_entry(
            parent,
            self.var_caixa_compartilhada,
            2,
            "Exemplo: genteegestao@empresa.com.br",
        )

        self.criar_label(parent, "Pasta Outlook", 3)
        self.criar_entry(
            parent,
            self.var_pasta_outlook,
            3,
            "Caixa de Entrada ou nome de subpasta",
        )

        frame_sel = ctk.CTkFrame(parent, fg_color="transparent")
        frame_sel.grid(row=4, column=0, columnspan=3, padx=18, pady=(2, 2), sticky="ew")

        btn_escolher = ctk.CTkButton(
            frame_sel,
            text="Escolher pasta do Outlook...",
            command=self.escolher_pasta_outlook,
            height=34,
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["branco"],
            font=("Segoe UI", 13, "bold"),
        )
        btn_escolher.grid(row=0, column=0, padx=(0, 8), sticky="w")

        btn_limpar_sel = ctk.CTkButton(
            frame_sel,
            text="Limpar seleção",
            command=self.limpar_pasta_escolhida,
            height=34,
            width=120,
            fg_color=CORES["borda"],
            hover_color="#C4C9D1",
            text_color=CORES["texto"],
            font=("Segoe UI", 13),
        )
        btn_limpar_sel.grid(row=0, column=1, sticky="w")

        self.lbl_pasta_escolhida = ctk.CTkLabel(
            parent,
            text=self._texto_pasta_escolhida(),
            font=("Segoe UI", 12),
            text_color=CORES["texto_sec"],
            justify="left",
            wraplength=480,
            anchor="w",
        )
        self.lbl_pasta_escolhida.grid(row=5, column=0, columnspan=3, padx=18, pady=(0, 6), sticky="ew")

        self.criar_label(parent, "Quantidade máxima", 6)
        self.criar_entry(parent, self.var_quantidade, 6, "200")

        self.criar_label(parent, "Últimos dias", 7)
        self.criar_entry(parent, self.var_ultimos_dias, 7, "30. Use 0 para não filtrar")

        chk = ctk.CTkCheckBox(
            parent,
            text="Ler somente e-mails não lidos",
            variable=self.var_somente_nao_lidos,
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        chk.grid(row=8, column=0, columnspan=2, padx=18, pady=(8, 16), sticky="w")

    def _texto_pasta_escolhida(self):
        display = self.var_pasta_display.get().strip()
        if display:
            return f"Pasta selecionada: {display}  (tem prioridade sobre os campos acima)"
        return "Nenhuma pasta selecionada pelo seletor — usando os campos de origem acima."

    def escolher_pasta_outlook(self):
        try:
            try:
                pythoncom.CoInitialize()
            except Exception:
                pass

            outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
            folder = outlook.PickFolder()

            if folder is None:
                return  # usuário cancelou

            entry_id = getattr(folder, "EntryID", "") or ""
            store_id = getattr(folder, "StoreID", "") or ""
            try:
                display = folder.FolderPath
            except Exception:
                display = getattr(folder, "Name", "Pasta selecionada")

            self.var_pasta_entry_id.set(entry_id)
            self.var_pasta_store_id.set(store_id)
            self.var_pasta_display.set(display)
            self.lbl_pasta_escolhida.configure(text=self._texto_pasta_escolhida())
            self.log(f"Pasta escolhida no seletor: {display}")

        except Exception as e:
            messagebox.showerror(
                "Escolher pasta",
                f"Não foi possível abrir o seletor de pastas do Outlook.\n\n{e}",
            )

    def limpar_pasta_escolhida(self):
        self.var_pasta_entry_id.set("")
        self.var_pasta_store_id.set("")
        self.var_pasta_display.set("")
        self.lbl_pasta_escolhida.configure(text=self._texto_pasta_escolhida())
        self.log("Seleção de pasta limpa. Usando os campos de origem.")

    def criar_card_saida(self, parent):
        self.criar_titulo_card(parent, "Saída em Excel")

        self.criar_label(parent, "Pasta de saída", 1)

        frame_saida = ctk.CTkFrame(parent, fg_color="transparent")
        frame_saida.grid(row=1, column=1, padx=(4, 18), pady=7, sticky="ew")
        frame_saida.grid_columnconfigure(0, weight=1)

        entry_saida = ctk.CTkEntry(
            frame_saida,
            textvariable=self.var_pasta_saida,
            height=34,
            fg_color=CORES["branco"],
            border_color=CORES["borda"],
            text_color=CORES["texto"],
        )
        entry_saida.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        btn_pasta = ctk.CTkButton(
            frame_saida,
            text="Selecionar",
            command=self.selecionar_pasta_saida,
            width=95,
            height=34,
            fg_color=CORES["azul_apoio"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
        )
        btn_pasta.grid(row=0, column=1)

        self.criar_label(parent, "Nome do Excel", 2)
        self.criar_entry(parent, self.var_nome_excel, 2, DEFAULT_EXCEL_NAME)

        info = ctk.CTkLabel(
            parent,
            text="O Excel será gerado com as abas BASE_EMAILS, GLOSSARIO e RESUMO. A classificação será feita depois, com análise da planilha.",
            font=("Segoe UI", 13),
            text_color=CORES["texto_sec"],
            justify="left",
            wraplength=480,
            anchor="w",
        )
        info.grid(row=3, column=0, columnspan=2, padx=18, pady=(8, 12), sticky="ew")

        botoes = ctk.CTkFrame(parent, fg_color="transparent")
        botoes.grid(row=4, column=0, columnspan=2, padx=18, pady=(4, 16), sticky="ew")
        botoes.grid_columnconfigure(0, weight=1)
        botoes.grid_columnconfigure(1, weight=1)

        self.btn_extrair = ctk.CTkButton(
            botoes,
            text="Extrair e-mails para Excel",
            command=self.iniciar_extracao,
            height=42,
            fg_color=CORES["azul"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
            font=("Segoe UI", 14, "bold"),
        )
        self.btn_extrair.grid(row=0, column=0, padx=(0, 8), sticky="ew")

        self.btn_abrir = ctk.CTkButton(
            botoes,
            text="Abrir pasta de saída",
            command=self.abrir_pasta_saida,
            height=42,
            fg_color=CORES["azul_apoio"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
            font=("Segoe UI", 14, "bold"),
        )
        self.btn_abrir.grid(row=0, column=1, padx=(8, 0), sticky="ew")

    def criar_area_log(self, parent):
        card = ctk.CTkFrame(
            parent,
            fg_color=CORES["card"],
            border_color=CORES["borda"],
            border_width=1,
            corner_radius=14,
        )
        card.grid(row=0, column=0, sticky="nsew")
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(2, weight=1)

        header_exec = ctk.CTkFrame(card, fg_color="transparent")
        header_exec.grid(row=0, column=0, padx=18, pady=(16, 8), sticky="ew")
        header_exec.grid_columnconfigure(0, weight=1)

        titulo = ctk.CTkLabel(
            header_exec,
            text="Execução",
            font=("Segoe UI", 17, "bold"),
            text_color=CORES["navy"],
            anchor="w",
        )
        titulo.grid(row=0, column=0, sticky="w")

        self.btn_executar = ctk.CTkButton(
            header_exec,
            text="Executar extração",
            command=self.iniciar_extracao,
            height=44,
            width=230,
            fg_color=CORES["sucesso"],
            hover_color="#256628",
            text_color=CORES["branco"],
            font=("Segoe UI", 15, "bold"),
        )
        self.btn_executar.grid(row=0, column=1, sticky="e")

        self.progress = ctk.CTkProgressBar(
            card,
            height=14,
            progress_color=CORES["azul"],
            fg_color=CORES["borda"],
        )
        self.progress.grid(row=1, column=0, padx=18, pady=(0, 10), sticky="ew")
        self.progress.set(0)

        self.txt_log = ctk.CTkTextbox(
            card,
            height=280,
            fg_color="#0F172A",
            text_color="#E5E7EB",
            border_color=CORES["borda"],
            border_width=1,
            font=("Consolas", 12),
        )
        self.txt_log.grid(row=2, column=0, padx=18, pady=(0, 18), sticky="nsew")

        self.log("Aplicativo iniciado.")
        self.log("Modo seguro: nenhum e-mail é movido, excluído, marcado ou enviado.")
        self.log("Com a IA ativada, respostas são apenas rascunhos salvos em Rascunhos para sua revisão.")

    def criar_rodape(self, parent):
        rodape = ctk.CTkFrame(parent, fg_color=CORES["fundo"], corner_radius=0, height=36)
        rodape.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 8))
        rodape.grid_columnconfigure(0, weight=1)

        label = ctk.CTkLabel(
            rodape,
            text="Anderson Marinho | Igarapé Digital",
            font=("Segoe UI", 12),
            text_color=CORES["texto_sec"],
            anchor="e",
        )
        label.grid(row=0, column=0, sticky="e")

    def log(self, mensagem):
        mensagem = str(mensagem)
        linha = f"[{agora_texto()}] {mensagem}\n"

        try:
            self.txt_log.insert("end", linha)
            self.txt_log.see("end")
        except Exception:
            pass

        try:
            registrar_log_arquivo(mensagem)
        except Exception:
            pass

    def selecionar_pasta_saida(self):
        pasta = filedialog.askdirectory(
            title="Selecione a pasta de saída",
            initialdir=self.var_pasta_saida.get() or str(SAIDA_DIR),
        )
        if pasta:
            self.var_pasta_saida.set(pasta)

    def abrir_pasta_saida(self):
        pasta = self.var_pasta_saida.get().strip()
        if not pasta:
            pasta = str(SAIDA_DIR)

        Path(pasta).mkdir(parents=True, exist_ok=True)
        os.startfile(pasta)

    def validar_campos(self):
        try:
            quantidade = int(self.var_quantidade.get())
            if quantidade <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Validação", "Informe uma quantidade válida maior que zero.")
            return False

        try:
            dias = int(self.var_ultimos_dias.get())
            if dias < 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Validação", "Informe os últimos dias como número. Use 0 para não filtrar.")
            return False

        if self.var_modo_caixa.get() == "compartilhada" and not self.var_caixa_compartilhada.get().strip():
            messagebox.showerror("Validação", "Informe o nome ou e-mail da caixa compartilhada.")
            return False

        pasta_saida = self.var_pasta_saida.get().strip()
        if not pasta_saida:
            messagebox.showerror("Validação", "Informe a pasta de saída.")
            return False

        nome_excel = self.var_nome_excel.get().strip()
        if not nome_excel:
            messagebox.showerror("Validação", "Informe o nome do arquivo Excel.")
            return False

        if not nome_excel.lower().endswith(".xlsx"):
            self.var_nome_excel.set(nome_excel + ".xlsx")

        if self.var_criar_rascunho.get() and not self.var_ia_ativa.get():
            messagebox.showinfo(
                "Assistente de IA",
                "Para criar rascunhos no Outlook, ative também "
                "'Gerar análise e resposta com IA (Claude)'.",
            )
            return False

        if self.var_ia_ativa.get():
            chave = self.var_ia_api_key.get().strip() or os.environ.get("ANTHROPIC_API_KEY", "").strip()
            if not chave:
                messagebox.showerror(
                    "Assistente de IA",
                    "A IA está ativada, mas nenhuma chave da API foi encontrada.\n\n"
                    "Informe a chave no campo 'Chave da API' ou defina a variável "
                    "de ambiente ANTHROPIC_API_KEY.",
                )
                return False

        return True

    def travar_interface(self, travar=True):
        estado = "disabled" if travar else "normal"

        try:
            self.btn_extrair.configure(state=estado)
        except Exception:
            pass

        try:
            self.btn_executar.configure(state=estado)
        except Exception:
            pass

        self.processo_rodando = travar

    def atualizar_progresso(self, atual, total):
        try:
            valor = min(max(atual / total, 0), 1)
            self.progress.set(valor)
        except Exception:
            pass

    def iniciar_extracao(self):
        if self.processo_rodando:
            messagebox.showinfo("Execução", "Já existe uma extração em andamento.")
            return

        if not self.validar_campos():
            return

        config = self.obter_config_tela()
        salvar_config(config)

        self.travar_interface(True)
        self.progress.set(0)
        self.log("Iniciando extração dos e-mails.")

        thread = threading.Thread(target=self.executar_extracao_thread, args=(config,), daemon=True)
        thread.start()

    def executar_extracao_thread(self, config):
        try:
            config_ia = {
                "ia_ativa": config.get("ia_ativa", False),
                "criar_rascunho_outlook": config.get("criar_rascunho_outlook", False),
                "ia_modelo": config.get("ia_modelo", IA_MODELO_PADRAO),
                "ia_api_key": config.get("ia_api_key", ""),
                "ia_workspace_id": config.get("ia_workspace_id", ""),
                "ia_assinatura": config.get("ia_assinatura", ""),
                "ia_contexto_rh": config.get("ia_contexto_rh", ""),
            }

            if config_ia["ia_ativa"]:
                self.after(0, self.log, f"IA ativada. Modelo: {config_ia['ia_modelo']}.")
                if config_ia["criar_rascunho_outlook"]:
                    self.after(0, self.log, "Rascunhos de resposta serão salvos em Rascunhos do Outlook (não enviados).")
            else:
                self.after(0, self.log, "IA desligada. Classificação ficará pendente (glossário).")

            registros, erros = ler_emails_outlook(
                modo_caixa=config["modo_caixa"],
                caixa_compartilhada=config["caixa_compartilhada"],
                pasta_outlook=config["pasta_outlook"],
                quantidade=config["quantidade"],
                ultimos_dias=config["ultimos_dias"],
                somente_nao_lidos=config["somente_nao_lidos"],
                config_ia=config_ia,
                pasta_entry_id=config.get("pasta_entry_id", ""),
                pasta_store_id=config.get("pasta_store_id", ""),
                progress_callback=lambda atual, total: self.after(0, self.atualizar_progresso, atual, total),
                log_callback=lambda msg: self.after(0, self.log, msg),
            )

            if not registros:
                self.after(0, self.log, "Nenhum e-mail foi extraído com os filtros informados.")
                self.after(0, messagebox.showwarning, "Resultado", "Nenhum e-mail foi extraído com os filtros informados.")
                return

            pasta_saida = Path(config["pasta_saida"])
            nome_excel = config["nome_excel"]

            if not nome_excel.lower().endswith(".xlsx"):
                nome_excel += ".xlsx"

            caminho_saida = pasta_saida / nome_excel

            self.after(0, self.log, "Gerando arquivo Excel.")
            caminho_gerado = gerar_excel(registros, caminho_saida)

            self.caminho_excel_gerado = caminho_gerado

            self.after(0, self.progress.set, 1)
            self.after(0, self.log, f"Excel gerado com sucesso: {caminho_gerado}")

            if erros:
                self.after(0, self.log, f"Alguns itens não puderam ser lidos. Total de erros ignorados: {len(erros)}")

            self.after(
                0,
                messagebox.showinfo,
                "Concluído",
                f"Extração concluída com sucesso.\n\nE-mails extraídos: {len(registros)}\nArquivo:\n{caminho_gerado}",
            )

        except Exception as e:
            erro = traceback.format_exc()
            self.after(0, self.log, f"Erro na execução: {e}")
            registrar_log_arquivo(erro)
            self.after(0, messagebox.showerror, "Erro", f"Ocorreu um erro na execução:\n\n{e}")

        finally:
            self.after(0, self.travar_interface, False)


def main():
    garantir_pastas()
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()