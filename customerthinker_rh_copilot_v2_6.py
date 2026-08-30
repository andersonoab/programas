import os
import sys
import re
import json
import time
import threading
import traceback
import unicodedata
import hashlib
from collections import Counter, defaultdict
from pathlib import Path
from datetime import datetime, date

import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk

import pythoncom
import win32com.client as win32

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


APP_NAME = "Agente IA Outlook RH Copilot"
APP_VERSION = "2.6.0"

# ---------------------------------------------------------------------------
# IA (Claude / Anthropic)
# ---------------------------------------------------------------------------
# Modelos disponiveis no seletor. Rotulo amigavel -> id tecnico do modelo.
IA_MODELOS = {
    "Claude Haiku 4.5 (rapido e economico)": "claude-haiku-4-5-20251001",
    "Claude Sonnet 5 (respostas mais elaboradas)": "claude-sonnet-5",
}
IA_MODELO_PADRAO = "claude-haiku-4-5-20251001"

# Limite de caracteres do corpo enviado a IA (minimizacao de dados / custo).
IA_LIMITE_CORPO = 4000

# O Copilot separa duas dimensões que antes estavam misturadas:
# 1) MODO OPERACIONAL = como analisar o trabalho.
# 2) DESTINATÁRIO/TOM = como escrever a comunicação.
MODO_PADRAO = "Geral RH"
MODOS_OPERACIONAIS = {
    "Geral RH": {
        "descricao": "Triagem geral de RH/DP: identifica tema, risco, responsável, fila e próxima ação.",
        "prompt": (
            "Atue como triagem geral de Gente e Gestão. Classifique o assunto, identifique o próximo passo, "
            "o responsável provável, o que depende de Anderson e o que pode ser delegado."
        ),
    },
    "Prioridades do Dia": {
        "descricao": "Foco em decidir o que precisa ser tratado hoje, o que pode ser delegado e o que só precisa ser acompanhado.",
        "prompt": (
            "Atue como um chefe de gabinete operacional de RH. Priorize prazo, dinheiro, fechamento de folha, risco legal, "
            "dependência executiva e bloqueios. Diferencie claramente: Crítico, Precisa de mim, Delegável, Aguardando terceiro e Informativo."
        ),
    },
    "Folha de Pagamento": {
        "descricao": "Prioriza competência, corte, evento, valor, retroatividade, encargos e risco de fechamento.",
        "prompt": (
            "Atue como especialista de Payroll/DP. Procure competência, data de corte, evento de folha, base de cálculo, "
            "valor, retroatividade, impacto financeiro, encargos e risco de fechamento. Destaque o que pode afetar pagamento, "
            "desconto, contabilização ou fechamento da competência corrente."
        ),
    },
    "Benefícios": {
        "descricao": "Foco em elegibilidade, movimentação, faturamento, vidas, desconto e fornecedor.",
        "prompt": (
            "Atue como especialista de Benefícios. Identifique elegibilidade, inclusão/exclusão, competência, vidas, faturamento, "
            "desconto em folha, fornecedor, evidência cadastral, divergência e prazo operacional."
        ),
    },
    "Ponto e Jornada": {
        "descricao": "Foco em marcação, ajuste, HE, banco de horas, escala, aprovação e reflexo em folha.",
        "prompt": (
            "Atue como especialista de Ponto/Jornada. Identifique marcações, ajuste, justificativa, hora extra, banco de horas, "
            "escala, aprovação do gestor, período, evidência e possível reflexo em folha."
        ),
    },
    "Desligamentos / Jurídico": {
        "descricao": "Alta cautela: prazo, verbas, documentação, estabilidade, evidência e risco jurídico.",
        "prompt": (
            "Atue em modo de alta cautela para desligamentos e temas jurídicos. Identifique datas, modalidade, aviso, estabilidade, "
            "prazo, verbas, documentos, evidências e risco. Não conclua tese jurídica nem autorize medida sensível sem revisão humana."
        ),
    },
    "Admissão": {
        "descricao": "Foco em prazo de entrada, documentos, cadastro, ASO e impeditivos de admissão.",
        "prompt": (
            "Atue como especialista de admissão. Identifique data de admissão, documentos, cadastro, exame, dados bancários, "
            "benefícios de entrada e qualquer pendência que possa impedir ou atrasar a admissão."
        ),
    },
    "Férias": {
        "descricao": "Foco em período aquisitivo/concessivo, aviso, pagamento e risco de vencimento.",
        "prompt": (
            "Atue como especialista de férias. Identifique período aquisitivo/concessivo, programação, aviso, pagamento, abono, "
            "risco de vencimento e impacto operacional."
        ),
    },
    "Encargos / eSocial": {
        "descricao": "Foco em FGTS, INSS, IRRF, DCTFWeb, eSocial, vencimento e retificação.",
        "prompt": (
            "Atue como especialista de encargos e obrigações acessórias. Identifique competência, base, recolhimento, vencimento, "
            "evento eSocial, DCTFWeb, FGTS Digital e necessidade de retificação."
        ),
    },
}

DESTINO_PADRAO = "Padrão RH"
DESTINOS_COMUNICACAO = {
    "Padrão RH": {
        "descricao": "Resposta profissional, simples, objetiva e orientada ao próximo passo.",
        "prompt": (
            "Escreva em tom profissional de RH, simples e objetivo. Explique o necessário e termine com o próximo passo claro."
        ),
    },
    "Joana": {
        "descricao": "Resumo executivo: conclusão, impacto/risco, recomendação e decisão necessária.",
        "prompt": (
            "Escreva para Joana em formato executivo. Comece pela conclusão, depois impacto/risco, recomendação e, por último, "
            "a decisão ou aprovação necessária. Elimine detalhes operacionais que não mudem a decisão."
        ),
    },
    "Lisiane": {
        "descricao": "Direta e cordial, com contexto, alternativas e próximo passo bem definido.",
        "prompt": (
            "Escreva para Lisiane de forma direta, cordial e contextualizada. Quando houver alternativas, apresente-as objetivamente "
            "e deixe claro qual OK, valor, definição ou responsável é necessário para seguir."
        ),
    },
    "Ricardo Ribeiro": {
        "descricao": "Executiva: fato, decisão solicitada, impacto, prazo e risco, sem presumir alçada.",
        "prompt": (
            "Escreva para Ricardo Ribeiro em formato executivo. Resuma fato, decisão solicitada, impacto financeiro/operacional, "
            "prazo e risco. Não presuma cargo, alçada, autorização ou histórico não informado."
        ),
    },
    "Colaborador": {
        "descricao": "Linguagem acessível, sem jargão de DP, com orientação prática.",
        "prompt": (
            "Escreva para um colaborador. Use linguagem simples e acessível, evite siglas sem explicar e informe claramente o que "
            "aconteceu, o que será feito e se existe alguma ação necessária por parte dele."
        ),
    },
    "Gestor": {
        "descricao": "Orientação objetiva para gestor, com responsabilidade, prazo e ação esperada.",
        "prompt": (
            "Escreva para um gestor. Seja objetivo, informe o contexto mínimo, o impacto para a equipe e exatamente qual ação, "
            "aprovação ou evidência é necessária e até quando."
        ),
    },
    "Fornecedor": {
        "descricao": "Cobrança objetiva com evidência, divergência, prazo e confirmação esperada.",
        "prompt": (
            "Escreva para fornecedor/prestador. Registre a divergência, evidência disponível, impacto, ação esperada e prazo de retorno. "
            "Evite linguagem acusatória quando o fato ainda não estiver comprovado."
        ),
    },
    "Jurídico": {
        "descricao": "Consulta estruturada: fatos, premissas, dúvida, posição preliminar e risco.",
        "prompt": (
            "Escreva como consulta ao Jurídico. Separe fatos confirmados de premissas, formule a dúvida jurídica objetivamente, "
            "registre a posição preliminar de RH quando houver e destaque o risco/decisão que depende de validação."
        ),
    },
}

# aliases de compatibilidade com a V1.2
PERFIL_PADRAO = MODO_PADRAO
PERFIS_AGENTE = MODOS_OPERACIONAIS

# Regras determinísticas locais: servem como primeira camada de triagem e também como
# contexto adicional para a IA. Elas não substituem política interna ou validação humana.
REGRAS_RH = [
    {
        "nome": "Rescisão / prazo legal",
        "termos": ["rescisao", "desligamento", "trct", "aviso previo", "demissao", "art. 477", "artigo 477"],
        "categoria": "Desligamento",
        "urgencia": "Alta",
        "score": 90,
        "acao": "Validar datas, modalidade, verbas, documentos e prazo antes do processamento.",
        "responsavel": "Folha / DP",
        "tipo_decisao": "Validação legal/operacional",
        "revisao_humana": True,
    },
    {
        "nome": "Impacto direto em folha",
        "termos": ["folha", "holerite", "salario", "diferenca salarial", "desconto", "retroativo", "adiantamento", "pagamento incorreto", "nao recebeu"],
        "categoria": "Folha de Pagamento",
        "urgencia": "Alta",
        "score": 85,
        "acao": "Validar competência, evento, base, valor, histórico e impacto na folha corrente.",
        "responsavel": "Folha / DP",
        "tipo_decisao": "Correção de Payroll",
        "revisao_humana": True,
    },
    {
        "nome": "Encargos / tributos",
        "termos": ["fgts", "inss", "irrf", "dctfweb", "esocial", "dar f", "darf", "encargo", "consignado"],
        "categoria": "Encargos",
        "urgencia": "Alta",
        "score": 82,
        "acao": "Validar competência, base, recolhimento, vencimento e eventual retificação.",
        "responsavel": "Folha / Encargos",
        "tipo_decisao": "Recolhimento/retificação",
        "revisao_humana": True,
    },
    {
        "nome": "Ponto / hora extra",
        "termos": ["ponto", "hora extra", "horas extras", "banco de horas", "marcacao", "jornada", "escala"],
        "categoria": "Ponto e Jornada",
        "urgencia": "Média",
        "score": 70,
        "acao": "Validar período, marcações, justificativa, aprovação do gestor e reflexo em folha.",
        "responsavel": "Ponto / DP",
        "tipo_decisao": "Ajuste/aprovação de jornada",
        "revisao_humana": True,
    },
    {
        "nome": "Benefícios / faturamento",
        "termos": ["beneficio", "plano de saude", "odontologico", "sulamerica", "ifood", "vale refeicao", "vale alimentacao", "vale transporte", "faturamento", "vidas"],
        "categoria": "Benefícios",
        "urgencia": "Média",
        "score": 62,
        "acao": "Validar elegibilidade, movimentação, competência, vidas, desconto e faturamento do fornecedor.",
        "responsavel": "Benefícios",
        "tipo_decisao": "Movimentação/cobrança",
        "revisao_humana": False,
    },
    {
        "nome": "Admissão / cadastro",
        "termos": ["admissao", "admissional", "cadastro", "documentos admissionais", "data de admissao"],
        "categoria": "Admissão",
        "urgencia": "Média",
        "score": 58,
        "acao": "Validar data, documentos, cadastro e pendências que impeçam a admissão.",
        "responsavel": "Admissão / DP",
        "tipo_decisao": "Regularização cadastral",
        "revisao_humana": False,
    },
    {
        "nome": "Férias",
        "termos": ["ferias", "abono pecuniario", "recibo de ferias", "periodo aquisitivo", "periodo concessivo"],
        "categoria": "Férias",
        "urgencia": "Média",
        "score": 55,
        "acao": "Validar período aquisitivo/concessivo, datas, aviso, pagamento e impacto operacional.",
        "responsavel": "Férias / DP",
        "tipo_decisao": "Programação/regularização",
        "revisao_humana": False,
    },
    {
        "nome": "Afastamento / dado sensível",
        "termos": ["atestado", "afastamento", "doenca", "internacao", "cid", "inss beneficio", "licenca medica"],
        "categoria": "Afastamentos",
        "urgencia": "Média",
        "score": 65,
        "acao": "Tratar somente os dados necessários, validar prazo/documentação e evitar exposição de informação sensível.",
        "responsavel": "DP / SST",
        "tipo_decisao": "Validação de afastamento",
        "revisao_humana": True,
    },
    {
        "nome": "Jurídico / reclamação",
        "termos": ["juridico", "processo trabalhista", "liminar", "audiencia", "reclamacao", "advogado", "defensoria", "notificacao extrajudicial"],
        "categoria": "Jurídico Trabalhista",
        "urgencia": "Alta",
        "score": 95,
        "acao": "Preservar evidências, identificar prazo e encaminhar para validação jurídica antes de responder conclusivamente.",
        "responsavel": "Jurídico / RH",
        "tipo_decisao": "Validação jurídica",
        "revisao_humana": True,
    },
    {
        "nome": "CCT / acordo coletivo",
        "termos": ["cct", "convencao coletiva", "acordo coletivo", "dissidio", "sindicato", "data base"],
        "categoria": "Relações Trabalhistas / CCT",
        "urgencia": "Alta",
        "score": 88,
        "acao": "Validar vigência, abrangência, cláusula aplicável, data-base e impacto na folha/benefícios.",
        "responsavel": "DP / Jurídico",
        "tipo_decisao": "Aplicação de norma coletiva",
        "revisao_humana": True,
    },
    {
        "nome": "Comissão / prêmio / PLR",
        "termos": ["comissao", "premio", "plr", "ppr", "vcc", "meta", "acelerador"],
        "categoria": "Remuneração Variável",
        "urgencia": "Média",
        "score": 76,
        "acao": "Validar regra, período de apuração, elegibilidade, base, aprovação e reflexo em folha/encargos.",
        "responsavel": "Folha / C&B",
        "tipo_decisao": "Cálculo/aprovação de variável",
        "revisao_humana": True,
    },
    {
        "nome": "Pagamento / fornecedor",
        "termos": ["nota fiscal", "nf", "boleto", "pagamento fornecedor", "vencimento", "faturamento", "cobranca"],
        "categoria": "Pagamentos / Fornecedores",
        "urgencia": "Média",
        "score": 68,
        "acao": "Validar documento, competência, valor, vencimento, centro de custo e responsável pela aprovação.",
        "responsavel": "RH / Pagamentos",
        "tipo_decisao": "Pagamento/aprovação",
        "revisao_humana": False,
    },
]

BASE_DIR = Path(r"C:\_RPA\AppEmailRH")
SAIDA_DIR = BASE_DIR / "saida"
LOGS_DIR = BASE_DIR / "logs"
CONFIG_DIR = BASE_DIR / "config"
CONFIG_FILE = CONFIG_DIR / "config_email_rh.json"

DEFAULT_EXCEL_NAME = "base_atendimento_gente_gestao.xlsx"

CORES = {
    # Paleta Sonova OFC
    "azul": "#0083CA",         # dominante
    "azul_apoio": "#003C64",   # apoio (barra lateral, textos de titulo)
    "verde_apoio": "#005A64",  # apoio
    "azul_claro": "#6EB4DC",   # apoio / botao secundario
    "vinho": "#7D0041",        # acento
    "telha": "#8C321E",        # acento
    "navy": "#003C64",         # compat.: antigo "navy" agora aponta para o apoio Sonova
    "fundo": "#FFFFFF",        # fundo branco
    "fundo_lateral": "#003C64",# barra lateral
    "card": "#FFFFFF",
    "borda": "#CCCCCC",        # borda de card (nunca azul claro)
    "texto": "#333333",
    "texto_sec": "#646464",
    "sucesso": "#2E7D32",      # verde = avancar
    "erro": "#B91C1C",         # vermelho = destrutivo
    "alerta": "#B7791F",
    "auxiliar": "#E0E0E0",     # cinza = auxiliar
    "auxiliar_hover": "#CFCFCF",
    "branco": "#FFFFFF",
}


# Opcoes fixas para os campos de acao (dropdowns do modal de detalhes)
OPCOES_FILA = ["Crítico", "Precisa de mim", "Delegável", "Aguardando terceiro", "Informativo"]
OPCOES_URGENCIA = ["Alta", "Média", "Baixa"]
OPCOES_STATUS = ["Pendente", "Em andamento", "Aguardando terceiro", "Concluído", "Cancelado"]
OPCOES_ESTADO = ["Novo", "Em tratativa", "Aguardando terceiro", "Resolvido"]
OPCOES_SIM_NAO = ["Sim", "Não"]
OPCOES_RESPONSAVEL = [
    "Anderson", "Joana", "Lisiane", "Ricardo Ribeiro",
    "Equipe DP", "Gestor da área", "Fornecedor", "Aguardando terceiro",
]
OPCOES_CATEGORIA = [
    "Folha de Pagamento", "Benefícios", "Ponto e Jornada", "Encargos / eSocial",
    "Admissão", "Desligamento", "Férias", "Remuneração Variável",
    "Relações Trabalhistas / CCT", "Cadastro / Dados", "Fornecedor",
    "Dúvida do Colaborador", "Outros / Informativo",
]
OPCOES_PRAZO = [
    "Hoje", "Esta semana", "Até o fechamento da folha",
    "Próxima competência", "Sem prazo definido",
]
OPCOES_PROXIMO_PASSO = [
    "Responder ao remetente", "Solicitar informações/documentos",
    "Encaminhar ao responsável", "Lançar em folha", "Conferir/validar cálculo",
    "Aguardar retorno de terceiro", "Registrar e arquivar",
    "Levar para decisão da gestão",
]


COLUNAS_BASE = [
    "EntryID", "StoreID", "ConversationID", "ConversationTopic", "Case ID",
    "Data Recebimento", "Hora Recebimento", "Dias Desde Recebimento",
    "Remetente Nome", "Remetente E-mail", "Para", "CC", "Assunto",
    "Corpo do E-mail", "Corpo Limpo", "Texto Truncado", "Tem Anexo",
    "Quantidade de Anexos", "Nomes dos Anexos", "Status Outlook", "Lido",
    "Importância Outlook", "Tamanho do Texto", "Data Extração",
    "Modo Operacional", "Destinatário / Tom",
    "Fila Copilot", "Motivo da Fila", "Resumo Executivo", "Pendência / Decisão",
    "Impacto", "Prazo / Timing", "Próximo Passo", "Pode Delegar", "Estado do Caso",
    "Regras Detectadas", "Score Prioridade", "Tipo de Decisão",
    "Análise ChatGPT", "Urgência Sugerida", "Motivo da Urgência",
    "Categoria Sugerida", "Ação Recomendada", "Responsável Sugerido",
    "Observação RH", "Status Atendimento", "Resposta Sugerida (IA)",
    "E-mails no Caso (sessão)", "Ocorrências Históricas", "Caso Recorrente",
    "Rascunho Outlook Criado",
]


COLUNAS_GLOSSARIO = [
    "Palavra-chave",
    "Categoria",
    "Prioridade",
    "Ação sugerida",
    "Responsável",
    "Observação",
    "Ativo",
]


def garantir_pastas():
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    SAIDA_DIR.mkdir(parents=True, exist_ok=True)
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)


def agora_texto():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def hoje_data():
    return datetime.now().date()


def limpar_texto(texto):
    if texto is None:
        return ""

    texto = str(texto)
    texto = texto.replace("\r", "\n")
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    texto = re.sub(r"[ \t]+", " ", texto)
    texto = texto.replace("\x00", "")
    return texto.strip()


def limpar_texto_uma_linha(texto):
    texto = limpar_texto(texto)
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def limitar_excel(texto, limite=32000):
    texto = "" if texto is None else str(texto)
    if len(texto) <= limite:
        return texto, "Não"
    return texto[:limite] + "\n\n[TEXTO TRUNCADO PARA COMPATIBILIDADE COM EXCEL]", "Sim"


def normalizar_data_outlook(valor):
    if valor is None:
        return None

    try:
        if hasattr(valor, "date"):
            return valor
    except Exception:
        pass

    try:
        return datetime.fromtimestamp(time.mktime(valor.timetuple()))
    except Exception:
        return None


def dias_desde_recebimento(recebido):
    if not recebido:
        return ""

    try:
        data_email = recebido.date()
        return (hoje_data() - data_email).days
    except Exception:
        return ""


def traduzir_importancia(valor):
    try:
        valor = int(valor)
    except Exception:
        return ""

    mapa = {
        0: "Baixa",
        1: "Normal",
        2: "Alta",
    }
    return mapa.get(valor, str(valor))


def carregar_config():
    garantir_pastas()

    config_padrao = {
        "modo_caixa": "principal",
        "caixa_compartilhada": "",
        "pasta_outlook": "Caixa de Entrada",
        "pasta_entry_id": "",
        "pasta_store_id": "",
        "pasta_display": "",
        "quantidade": 200,
        "ultimos_dias": 30,
        "somente_nao_lidos": False,
        "pasta_saida": str(SAIDA_DIR),
        "nome_excel": DEFAULT_EXCEL_NAME,
        "ia_ativa": False,
        "criar_rascunho_outlook": False,
        "modo_operacional": MODO_PADRAO,
        "destino_comunicacao": DESTINO_PADRAO,
        "perfil_agente": PERFIL_PADRAO,
        "ia_modelo": IA_MODELO_PADRAO,
        "ia_api_key": "",
        "ia_workspace_id": "",
        "ia_assinatura": "Atenciosamente,\nGente e Gestão – Sonova",
        "resumo_destinatario_email": "",
        "resumo_folha_auto": True,
        "data_de": "",
        "data_ate": "",
        "agendamento_ativo": False,
        "agendamento_hora": "05:00",
        "alerta_urgente": True,
        "salvar_chave": False,
        "ia_contexto_rh": (
            "Você atende a caixa de Gente e Gestão (RH/DP) da Sonova Brasil. "
            "Os remetentes costumam ser colaboradores e gestores com dúvidas "
            "sobre folha de pagamento, benefícios, férias, admissão, "
            "desligamento, atestados e afastamentos."
        ),
    }

    if not CONFIG_FILE.exists():
        salvar_config(config_padrao)
        return config_padrao

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            config_lido = json.load(f)

        for chave, valor in config_padrao.items():
            config_lido.setdefault(chave, valor)

        return config_lido

    except Exception:
        return config_padrao


def salvar_config(config):
    garantir_pastas()
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=4)


def registrar_log_arquivo(mensagem):
    garantir_pastas()
    caminho_log = LOGS_DIR / "log_execucao.txt"
    with open(caminho_log, "a", encoding="utf-8") as f:
        f.write(f"[{agora_texto()}] {mensagem}\n")


def carregar_glossario(caminho_excel):
    """
    Função reservada para próxima iteração.

    Objetivo futuro:
    Ler a aba GLOSSARIO do Excel e usar as palavras-chave para sugerir
    categoria, prioridade, responsável e ação recomendada.

    Nesta primeira versão, a função não aplica classificação.
    """
    return []


def normalizar_busca(texto):
    texto = "" if texto is None else str(texto)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.lower()
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def assunto_base(assunto):
    txt = normalizar_busca(assunto)
    txt = re.sub(r"^(re|enc|fw|fwd)\s*:\s*", "", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def gerar_case_id(conversation_id="", conversation_topic="", assunto="", remetente_email=""):
    base = (conversation_id or "").strip()
    if not base:
        base = "|".join([
            assunto_base(conversation_topic or assunto),
            normalizar_busca(remetente_email),
        ])
    digest = hashlib.sha1(base.encode("utf-8", errors="ignore")).hexdigest()[:10].upper()
    return f"CASE-{digest}"


def detectar_prazo_textual(texto):
    t = normalizar_busca(texto)
    achados = []
    padroes = [
        r"\bhoje\b", r"\bamanha\b", r"\burgente\b", r"\bdeadline\b",
        r"\bprazo\b", r"\bvencimento\b", r"\bfechamento\b",
        r"\b(?:ate|para)\s+\d{1,2}/\d{1,2}(?:/\d{2,4})?\b",
        r"\b\d{1,2}/\d{1,2}(?:/\d{2,4})?\b",
    ]
    for p in padroes:
        m = re.search(p, t)
        if m:
            achados.append(m.group(0))
    return ", ".join(dict.fromkeys(achados[:4]))


def inferir_fila_local(resultado, texto_email="", dias_desde=None):
    r = resultado or {}
    score = int(r.get("score_prioridade", 20) or 20)
    urg = r.get("urgencia", "Baixa")
    revisao = bool(r.get("precisa_revisao_humana"))
    t = normalizar_busca(texto_email)

    aguardando = any(x in t for x in [
        "aguardando retorno do fornecedor", "aguardando retorno da operadora", "pendente fornecedor",
        "aguardando fornecedor", "aguardando prestador", "dependemos do retorno", "waiting for supplier",
    ])
    informativo = any(x in t for x in ["para conhecimento", "fyi", "somente para ciencia", "apenas para ciencia"])

    if score >= 88 or (urg == "Alta" and revisao):
        return "Crítico", "Prazo/risco elevado ou validação humana obrigatória."
    if aguardando and score < 85:
        return "Aguardando terceiro", "O conteúdo indica dependência de retorno externo/terceiro."
    if revisao or score >= 72:
        return "Precisa de mim", "Exige decisão, validação ou atenção direta antes de seguir."
    if informativo and score < 60:
        return "Informativo", "Mensagem predominantemente informativa, sem ação crítica identificada."
    if r.get("responsavel") and score >= 45:
        return "Delegável", "Há próxima ação operacional e responsável provável identificados."
    return "Informativo", "Nenhum gatilho relevante de ação imediata foi identificado."


def avaliar_regras_locais(texto_email, modo_operacional=MODO_PADRAO, destino_comunicacao=DESTINO_PADRAO):
    texto_norm = normalizar_busca(texto_email)
    encontrados = []

    for regra in REGRAS_RH:
        termos_encontrados = []
        for termo in regra.get("termos", []):
            termo_norm = normalizar_busca(termo)
            if termo_norm and termo_norm in texto_norm:
                termos_encontrados.append(termo)
        if termos_encontrados:
            item = dict(regra)
            item["termos_encontrados"] = termos_encontrados
            encontrados.append(item)

    encontrados.sort(key=lambda x: int(x.get("score", 0)), reverse=True)
    prazo = detectar_prazo_textual(texto_email)

    if not encontrados:
        base = {
            "analise_chatgpt": "", "urgencia": "Baixa",
            "motivo_urgencia": "Nenhuma regra crítica local foi acionada.",
            "categoria": "Não classificado", "acao_recomendada": "Ler o contexto e definir a próxima ação.",
            "responsavel": "", "observacao": "", "status_atendimento": "Pendente",
            "resposta_sugerida": "", "regras_detectadas": "", "score_prioridade": 20,
            "tipo_decisao": "Triagem", "precisa_revisao_humana": False,
            "modo_operacional": modo_operacional, "destino_comunicacao": destino_comunicacao,
            "termos_detectados": "", "prazo": prazo,
        }
    else:
        principal = encontrados[0]
        nomes = [r["nome"] for r in encontrados[:4]]
        termos = []
        for r in encontrados[:4]:
            termos.extend(r.get("termos_encontrados", []))
        revisao = any(bool(r.get("revisao_humana")) for r in encontrados)
        score = max(int(r.get("score", 0)) for r in encontrados)
        if prazo and score < 95:
            score = min(95, score + 5)
        base = {
            "analise_chatgpt": "Regras locais: " + "; ".join(nomes),
            "urgencia": principal.get("urgencia", "Baixa"),
            "motivo_urgencia": "Regra(s) acionada(s): " + "; ".join(nomes),
            "categoria": principal.get("categoria", "Não classificado"),
            "acao_recomendada": principal.get("acao", ""),
            "responsavel": principal.get("responsavel", ""),
            "observacao": "REVISÃO HUMANA RECOMENDADA pelas regras locais." if revisao else "",
            "status_atendimento": "Pendente", "resposta_sugerida": "",
            "regras_detectadas": "; ".join(nomes), "score_prioridade": score,
            "tipo_decisao": principal.get("tipo_decisao", "Triagem"),
            "precisa_revisao_humana": revisao,
            "modo_operacional": modo_operacional, "destino_comunicacao": destino_comunicacao,
            "termos_detectados": ", ".join(dict.fromkeys(termos)), "prazo": prazo,
        }

    fila, motivo = inferir_fila_local(base, texto_email=texto_email)
    base.update({
        "fila_copilot": fila,
        "motivo_fila": motivo,
        "resumo_executivo": base.get("analise_chatgpt", "") or base.get("acao_recomendada", ""),
        "pendencia_decisao": base.get("tipo_decisao", "Triagem"),
        "impacto": "A validar",
        "proximo_passo": base.get("acao_recomendada", ""),
        "pode_delegar": "Sim" if fila == "Delegável" else "Não",
        "estado_caso": "Novo/Pendente",
    })
    return base


def classificar_por_glossario(texto_email, glossario, modo_operacional=MODO_PADRAO, destino_comunicacao=DESTINO_PADRAO):
    return avaliar_regras_locais(texto_email, modo_operacional=modo_operacional, destino_comunicacao=destino_comunicacao)


def consolidar_classificacao(resultado_local, resultado_ia, texto_email=""):
    """Combina regras + IA. Alertas locais nunca são apagados pela IA."""
    local = resultado_local or {}
    ia = resultado_ia or {}
    ordem = {"Baixa": 1, "Média": 2, "Alta": 3}
    urg_local = local.get("urgencia", "Baixa")
    urg_ia = ia.get("urgencia", "Baixa")
    urgencia = urg_ia if ordem.get(urg_ia, 0) >= ordem.get(urg_local, 0) else urg_local

    categoria_ia = (ia.get("categoria") or "").strip()
    categoria = categoria_ia if categoria_ia and categoria_ia != "Não classificado" else local.get("categoria", "Não classificado")

    observacoes = [str(x).strip() for x in [local.get("observacao", ""), ia.get("observacao", "")] if str(x).strip()]
    revisao = bool(local.get("precisa_revisao_humana")) or bool(ia.get("precisa_revisao_humana"))
    if revisao and not any("REVISÃO HUMANA" in x for x in observacoes):
        observacoes.insert(0, "REVISÃO HUMANA RECOMENDADA.")

    score_local = int(local.get("score_prioridade", 20) or 20)
    score_ia = ia.get("score_prioridade")
    try:
        score = max(score_local, int(score_ia)) if score_ia not in (None, "") else score_local
    except Exception:
        score = score_local

    fila_ia = ia.get("fila_copilot")
    filas_validas = {"Crítico", "Precisa de mim", "Delegável", "Aguardando terceiro", "Informativo"}
    fila_local, motivo_local = inferir_fila_local({**local, "score_prioridade": score, "urgencia": urgencia, "precisa_revisao_humana": revisao}, texto_email)
    fila = fila_ia if fila_ia in filas_validas else fila_local

    # A IA não pode rebaixar um crítico local automaticamente.
    if fila_local == "Crítico" and fila != "Crítico":
        fila = "Crítico"

    return {
        "analise_chatgpt": ia.get("analise_chatgpt") or local.get("analise_chatgpt", ""),
        "urgencia": urgencia,
        "motivo_urgencia": ia.get("motivo_urgencia") or local.get("motivo_urgencia", ""),
        "categoria": categoria,
        "acao_recomendada": ia.get("acao_recomendada") or local.get("acao_recomendada", ""),
        "responsavel": ia.get("responsavel") or local.get("responsavel", ""),
        "observacao": " ".join(observacoes),
        "status_atendimento": ia.get("status_atendimento") or local.get("status_atendimento", "Pendente"),
        "resposta_sugerida": ia.get("resposta_sugerida", ""),
        "regras_detectadas": local.get("regras_detectadas", ""),
        "score_prioridade": score,
        "tipo_decisao": ia.get("tipo_decisao") or local.get("tipo_decisao", "Triagem"),
        "precisa_revisao_humana": revisao,
        "modo_operacional": local.get("modo_operacional", MODO_PADRAO),
        "destino_comunicacao": local.get("destino_comunicacao", DESTINO_PADRAO),
        "fila_copilot": fila,
        "motivo_fila": ia.get("motivo_fila") or motivo_local,
        "resumo_executivo": ia.get("resumo_executivo") or ia.get("analise_chatgpt") or local.get("resumo_executivo", ""),
        "pendencia_decisao": ia.get("pendencia_decisao") or local.get("pendencia_decisao", ""),
        "impacto": ia.get("impacto") or local.get("impacto", "A validar"),
        "prazo": ia.get("prazo") or local.get("prazo", ""),
        "proximo_passo": ia.get("proximo_passo") or ia.get("acao_recomendada") or local.get("proximo_passo", ""),
        "pode_delegar": ia.get("pode_delegar") or ("Sim" if fila == "Delegável" else "Não"),
        "estado_caso": ia.get("estado_caso") or local.get("estado_caso", "Novo/Pendente"),
    }


# ===========================================================================
# IA (Claude / Anthropic) - analise e geracao de resposta
# ===========================================================================
# ===========================================================================
# IA (Claude / Anthropic) - analise e geracao de resposta
# ===========================================================================

def _extrair_json(texto):
    """
    Extrai o primeiro objeto JSON de um texto, tolerando cercas ```json
    e eventual preambulo do modelo. Retorna dict (ou {} em falha).
    """
    if not texto:
        return {}

    limpo = str(texto).strip()
    limpo = limpo.replace("```json", "").replace("```", "").strip()

    try:
        return json.loads(limpo)
    except Exception:
        pass

    inicio = limpo.find("{")
    fim = limpo.rfind("}")
    if inicio != -1 and fim != -1 and fim > inicio:
        trecho = limpo[inicio:fim + 1]
        try:
            return json.loads(trecho)
        except Exception:
            return {}

    return {}


def analisar_email_com_ia(assunto, corpo, remetente_nome, remetente_email, config_ia,
                           para="", cc="", resultado_regras=None, log_callback=None):
    """Envia o e-mail ao Claude e retorna classificação, fila Copilot e resposta sugerida."""
    fallback = {
        "analise_chatgpt": "", "urgencia": "Baixa", "motivo_urgencia": "",
        "categoria": "Não classificado", "acao_recomendada": "", "responsavel": "",
        "observacao": "", "status_atendimento": "Pendente", "resposta_sugerida": "",
        "precisa_revisao_humana": False, "fila_copilot": "", "motivo_fila": "",
        "resumo_executivo": "", "pendencia_decisao": "", "impacto": "", "prazo": "",
        "proximo_passo": "", "pode_delegar": "", "estado_caso": "", "tipo_decisao": "",
        "score_prioridade": "",
    }

    api_key = (config_ia.get("ia_api_key") or "").strip() or os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        fallback["analise_chatgpt"] = "[IA ligada, mas nenhuma chave da API foi informada]"
        return fallback

    modelo = (config_ia.get("ia_modelo") or IA_MODELO_PADRAO).strip()
    contexto_rh = (config_ia.get("ia_contexto_rh") or "").strip()
    workspace_id = (config_ia.get("ia_workspace_id") or "").strip() or os.environ.get("ANTHROPIC_WORKSPACE_ID", "").strip()
    modo = config_ia.get("modo_operacional", MODO_PADRAO)
    destino = config_ia.get("destino_comunicacao", DESTINO_PADRAO)
    modo_cfg = MODOS_OPERACIONAIS.get(modo, MODOS_OPERACIONAIS[MODO_PADRAO])
    destino_cfg = DESTINOS_COMUNICACAO.get(destino, DESTINOS_COMUNICACAO[DESTINO_PADRAO])
    corpo_curto = (corpo or "")[:IA_LIMITE_CORPO]
    regras = resultado_regras or {}

    system_prompt = (
        "Você é o Agente IA Outlook RH Copilot da área de Gente e Gestão/DP. Trabalha em português do Brasil. "
        + contexto_rh + " "
        + "MODO OPERACIONAL: " + modo + ". " + modo_cfg.get("prompt", "") + " "
        + "DESTINATÁRIO/TOM: " + destino + ". " + destino_cfg.get("prompt", "") + " "
        + "Seu objetivo não é apenas responder e-mail: é ajudar a decidir o trabalho do dia. "
        + "Classifique cada mensagem em exatamente uma fila: Crítico, Precisa de mim, Delegável, Aguardando terceiro ou Informativo. "
        + "Crítico = risco legal, prazo muito próximo, fechamento/pagamento ou impacto relevante. "
        + "Precisa de mim = decisão, validação, aprovação ou julgamento de RH. "
        + "Delegável = ação operacional clara que pode ser executada por outro responsável. "
        + "Aguardando terceiro = não há próxima ação útil até retorno externo. Informativo = ciência, sem ação material. "
        + "NÃO invente política, prazo, valor, alçada, fato ou contexto. Diferencie fato confirmado de inferência. "
        + "Marque precisa_revisao_humana=true em risco jurídico, dado sensível, reclamação, rescisão, ponto/hora extra, "
          "pagamento/desconto, encargo ou qualquer medida sensível. "
        + "score_prioridade deve ser inteiro de 0 a 100. "
        + "Responda EXCLUSIVAMENTE com JSON válido contendo: analise, urgencia, motivo_urgencia, categoria, acao_recomendada, "
          "responsavel, observacao, status_atendimento, resposta_sugerida, precisa_revisao_humana, fila_copilot, motivo_fila, "
          "resumo_executivo, pendencia_decisao, impacto, prazo, proximo_passo, pode_delegar, estado_caso, tipo_decisao, score_prioridade. "
        + "urgencia: Alta, Média ou Baixa. pode_delegar: Sim ou Não. resposta_sugerida sem assinatura."
    )

    contexto_regras = (
        f"Regras locais: {regras.get('regras_detectadas', '') or 'nenhuma'}\n"
        f"Score local: {regras.get('score_prioridade', 20)}\n"
        f"Fila local: {regras.get('fila_copilot', '')}\n"
        f"Categoria local: {regras.get('categoria', 'Não classificado')}\n"
        f"Ação local: {regras.get('acao_recomendada', '')}\n"
        f"Tipo de decisão local: {regras.get('tipo_decisao', 'Triagem')}\n"
        f"Prazo textual detectado: {regras.get('prazo', '')}"
    )

    user_content = (
        f"Remetente: {remetente_nome} <{remetente_email}>\nPara: {para}\nCC: {cc}\nAssunto: {assunto}\n\n"
        f"{contexto_regras}\n\nCorpo do e-mail:\n{corpo_curto}"
    )

    try:
        from anthropic import Anthropic
    except Exception:
        fallback["analise_chatgpt"] = "[Biblioteca 'anthropic' não instalada. Rode: pip install anthropic]"
        if log_callback:
            log_callback("Biblioteca 'anthropic' não encontrada. Instale com: pip install anthropic")
        return fallback

    try:
        headers = {"anthropic-workspace-id": workspace_id} if workspace_id else {}
        client = Anthropic(api_key=api_key, default_headers=headers) if headers else Anthropic(api_key=api_key)
        resposta = client.messages.create(
            model=modelo, max_tokens=1900, system=system_prompt,
            messages=[{"role": "user", "content": user_content}],
        )
        texto_final = "\n".join(getattr(b, "text", "") for b in resposta.content if getattr(b, "text", None)).strip()
        dados = _extrair_json(texto_final)
        if not dados:
            fallback["analise_chatgpt"] = "[IA respondeu em formato inesperado]"
            fallback["resposta_sugerida"] = texto_final
            return fallback

        obs = str(dados.get("observacao", "") or "")
        revisao = bool(dados.get("precisa_revisao_humana", False))
        if revisao and "REVISÃO HUMANA" not in obs:
            obs = ("REVISÃO HUMANA RECOMENDADA. " + obs).strip()

        return {
            "analise_chatgpt": str(dados.get("analise", "") or ""),
            "urgencia": str(dados.get("urgencia", "") or "Baixa"),
            "motivo_urgencia": str(dados.get("motivo_urgencia", "") or ""),
            "categoria": str(dados.get("categoria", "") or "Não classificado"),
            "acao_recomendada": str(dados.get("acao_recomendada", "") or ""),
            "responsavel": str(dados.get("responsavel", "") or ""),
            "observacao": obs,
            "status_atendimento": str(dados.get("status_atendimento", "") or "Pendente"),
            "resposta_sugerida": str(dados.get("resposta_sugerida", "") or ""),
            "precisa_revisao_humana": revisao,
            "fila_copilot": str(dados.get("fila_copilot", "") or ""),
            "motivo_fila": str(dados.get("motivo_fila", "") or ""),
            "resumo_executivo": str(dados.get("resumo_executivo", "") or ""),
            "pendencia_decisao": str(dados.get("pendencia_decisao", "") or ""),
            "impacto": str(dados.get("impacto", "") or ""),
            "prazo": str(dados.get("prazo", "") or ""),
            "proximo_passo": str(dados.get("proximo_passo", "") or ""),
            "pode_delegar": str(dados.get("pode_delegar", "") or ""),
            "estado_caso": str(dados.get("estado_caso", "") or ""),
            "tipo_decisao": str(dados.get("tipo_decisao", "") or ""),
            "score_prioridade": dados.get("score_prioridade", ""),
        }
    except Exception as e:
        if log_callback:
            log_callback(f"Falha na chamada à IA: {e}")
        fallback["analise_chatgpt"] = f"[Erro na IA: {e}]"
        return fallback


def criar_rascunho_reply(mail, corpo_resposta, assinatura, log_callback=None):
    """
    Cria um rascunho de RESPOSTA no Outlook e o salva na pasta Rascunhos.
    NUNCA envia. O usuário revisa e clica em enviar manualmente.
    Retorna True se o rascunho foi salvo, False caso contrário.
    """
    corpo_resposta = (corpo_resposta or "").strip()
    if not corpo_resposta:
        return False

    try:
        reply = mail.Reply()  # cria a resposta (somente remetente), sem enviar

        try:
            corpo_original = reply.Body or ""
        except Exception:
            corpo_original = ""

        novo_corpo = corpo_resposta
        if assinatura:
            novo_corpo += "\n\n" + assinatura
        novo_corpo += "\n\n" + corpo_original

        reply.Body = novo_corpo
        reply.Save()  # salva em Rascunhos; nao dispara envio
        return True

    except Exception as e:
        if log_callback:
            log_callback(f"Falha ao criar rascunho no Outlook: {e}")
        return False


def obter_item_outlook_por_ids(entry_id, store_id=""):
    outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
    if store_id:
        return outlook.GetItemFromID(entry_id, store_id)
    return outlook.GetItemFromID(entry_id)


def abrir_item_outlook(entry_id, store_id=""):
    pythoncom.CoInitialize()
    try:
        item = obter_item_outlook_por_ids(entry_id, store_id)
        item.Display()
        return True
    finally:
        pythoncom.CoUninitialize()


def get_smtp_address(mail):
    """
    Tenta obter o e-mail SMTP real do remetente.
    Funciona melhor em contas Exchange/Outlook corporativas.
    """
    try:
        if getattr(mail, "SenderEmailType", "") == "EX":
            try:
                exchange_user = mail.Sender.GetExchangeUser()
                if exchange_user:
                    smtp = exchange_user.PrimarySmtpAddress
                    if smtp:
                        return smtp
            except Exception:
                pass

            try:
                PR_SMTP_ADDRESS = "http://schemas.microsoft.com/mapi/proptag/0x39FE001E"
                smtp = mail.Sender.PropertyAccessor.GetProperty(PR_SMTP_ADDRESS)
                if smtp:
                    return smtp
            except Exception:
                pass

        return getattr(mail, "SenderEmailAddress", "") or ""

    except Exception:
        return ""


def obter_pasta_por_nome(pasta_raiz, nome_pasta):
    """
    Procura uma pasta pelo nome dentro da árvore do Outlook.
    Se nome_pasta for vazio ou Caixa de Entrada, usa a própria pasta recebida.
    """
    nome_pasta = (nome_pasta or "").strip()

    if nome_pasta == "":
        return pasta_raiz

    # Tolerancia: se o usuario digitou um caminho (ex.: "conta\Caixa de Entrada\Sub"),
    # usa apenas o ultimo segmento como nome da pasta a procurar.
    if "\\" in nome_pasta or "/" in nome_pasta:
        partes = [p.strip() for p in re.split(r"[\\/]", nome_pasta) if p.strip()]
        if partes:
            nome_pasta = partes[-1]

    nomes_padrao_inbox = {
        "caixa de entrada",
        "inbox",
        "entrada",
    }

    if nome_pasta.lower() in nomes_padrao_inbox:
        return pasta_raiz

    def buscar_recursivo(pasta):
        try:
            if pasta.Name.strip().lower() == nome_pasta.lower():
                return pasta
        except Exception:
            pass

        try:
            for subpasta in pasta.Folders:
                achou = buscar_recursivo(subpasta)
                if achou is not None:
                    return achou
        except Exception:
            pass

        return None

    encontrado = buscar_recursivo(pasta_raiz)

    if encontrado is None:
        raise Exception(f"Pasta não encontrada no Outlook: {nome_pasta}")

    return encontrado


def obter_pasta_outlook(modo_caixa, caixa_compartilhada, pasta_outlook,
                        pasta_entry_id="", pasta_store_id="", log_callback=None):
    outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")

    # 1) Se o usuario escolheu uma pasta pelo seletor do Outlook, ela tem prioridade.
    pasta_entry_id = (pasta_entry_id or "").strip()
    pasta_store_id = (pasta_store_id or "").strip()
    if pasta_entry_id and pasta_store_id:
        try:
            pasta_final = outlook.GetFolderFromID(pasta_entry_id, pasta_store_id)
            if log_callback:
                try:
                    log_callback(f"Pasta selecionada: {pasta_final.FolderPath}")
                except Exception:
                    log_callback("Pasta selecionada pelo seletor do Outlook.")
            return pasta_final
        except Exception as e:
            if log_callback:
                log_callback(f"Não foi possível abrir a pasta escolhida ({e}). Usando os campos de origem.")

    modo_caixa = (modo_caixa or "principal").strip().lower()
    caixa_compartilhada = (caixa_compartilhada or "").strip()
    pasta_outlook = (pasta_outlook or "Caixa de Entrada").strip()

    if modo_caixa == "compartilhada":
        if not caixa_compartilhada:
            raise Exception("Informe o nome ou e-mail da caixa compartilhada.")

        if log_callback:
            log_callback(f"Tentando acessar caixa compartilhada: {caixa_compartilhada}")

        recipient = outlook.CreateRecipient(caixa_compartilhada)
        recipient.Resolve()

        if not recipient.Resolved:
            raise Exception(f"Não foi possível resolver a caixa compartilhada: {caixa_compartilhada}")

        inbox = outlook.GetSharedDefaultFolder(recipient, 6)

    else:
        if log_callback:
            log_callback("Acessando Caixa de Entrada principal do Outlook.")
        inbox = outlook.GetDefaultFolder(6)

    pasta_final = obter_pasta_por_nome(inbox, pasta_outlook)

    if log_callback:
        try:
            log_callback(f"Pasta selecionada: {pasta_final.Name}")
        except Exception:
            log_callback("Pasta selecionada com sucesso.")

    return pasta_final


def extrair_dados_email(mail, config_ia=None, log_callback=None):
    def attr(nome, padrao=""):
        try:
            return getattr(mail, nome, padrao) or padrao
        except Exception:
            return padrao

    entry_id = attr("EntryID")
    try:
        store_id = getattr(getattr(mail, "Parent", None), "StoreID", "") or ""
    except Exception:
        store_id = ""
    conversation_id = attr("ConversationID")
    conversation_topic = attr("ConversationTopic")
    recebido = normalizar_data_outlook(attr("ReceivedTime", None))

    data_recebimento = recebido.strftime("%d/%m/%Y") if recebido else ""
    hora_recebimento = recebido.strftime("%H:%M:%S") if recebido else ""
    remetente_nome = attr("SenderName")
    remetente_email = get_smtp_address(mail)
    para, cc, assunto, corpo = attr("To"), attr("CC"), attr("Subject"), attr("Body")
    corpo = limpar_texto(corpo)
    corpo_limpo = limpar_texto_uma_linha(corpo)
    corpo_excel, truncado = limitar_excel(corpo)
    corpo_limpo_excel, truncado_limpo = limitar_excel(corpo_limpo)
    texto_truncado = "Sim" if truncado == "Sim" or truncado_limpo == "Sim" else "Não"

    nomes_anexos = []
    try:
        qtd_anexos = int(mail.Attachments.Count)
        for i in range(1, qtd_anexos + 1):
            try:
                nomes_anexos.append(mail.Attachments.Item(i).FileName)
            except Exception:
                nomes_anexos.append("Anexo não identificado")
    except Exception:
        qtd_anexos = 0
    tem_anexo = "Sim" if qtd_anexos > 0 else "Não"

    try:
        lido = "Não" if bool(getattr(mail, "UnRead", False)) else "Sim"
    except Exception:
        lido = ""
    importancia = traduzir_importancia(attr("Importance", ""))
    tamanho_texto = len(corpo)

    modo = (config_ia or {}).get("modo_operacional", MODO_PADRAO)
    destino = (config_ia or {}).get("destino_comunicacao", DESTINO_PADRAO)
    texto_regras = f"{remetente_nome} {remetente_email} {para} {cc} {assunto}\n{corpo}"
    local = avaliar_regras_locais(texto_regras, modo_operacional=modo, destino_comunicacao=destino)

    if config_ia and config_ia.get("ia_ativa"):
        ia = analisar_email_com_ia(
            assunto=assunto, corpo=corpo, remetente_nome=remetente_nome,
            remetente_email=remetente_email, para=para, cc=cc,
            config_ia=config_ia, resultado_regras=local, log_callback=log_callback,
        )
        classificacao = consolidar_classificacao(local, ia, texto_email=texto_regras)
    else:
        classificacao = local

    case_id = gerar_case_id(conversation_id, conversation_topic, assunto, remetente_email)

    return {
        "EntryID": entry_id, "StoreID": store_id, "ConversationID": conversation_id,
        "ConversationTopic": conversation_topic, "Case ID": case_id,
        "Data Recebimento": data_recebimento, "Hora Recebimento": hora_recebimento,
        "Dias Desde Recebimento": dias_desde_recebimento(recebido),
        "Remetente Nome": remetente_nome, "Remetente E-mail": remetente_email,
        "Para": para, "CC": cc, "Assunto": assunto,
        "Corpo do E-mail": corpo_excel, "Corpo Limpo": corpo_limpo_excel,
        "Texto Truncado": texto_truncado, "Tem Anexo": tem_anexo,
        "Quantidade de Anexos": qtd_anexos, "Nomes dos Anexos": " | ".join(nomes_anexos),
        "Status Outlook": "Não lido" if lido == "Não" else "Lido", "Lido": lido,
        "Importância Outlook": importancia, "Tamanho do Texto": tamanho_texto,
        "Data Extração": agora_texto(), "Modo Operacional": modo, "Destinatário / Tom": destino,
        "Fila Copilot": classificacao.get("fila_copilot", "Informativo"),
        "Motivo da Fila": classificacao.get("motivo_fila", ""),
        "Resumo Executivo": classificacao.get("resumo_executivo", ""),
        "Pendência / Decisão": classificacao.get("pendencia_decisao", ""),
        "Impacto": classificacao.get("impacto", ""), "Prazo / Timing": classificacao.get("prazo", ""),
        "Próximo Passo": classificacao.get("proximo_passo", ""),
        "Pode Delegar": classificacao.get("pode_delegar", "Não"),
        "Estado do Caso": classificacao.get("estado_caso", "Novo/Pendente"),
        "Regras Detectadas": classificacao.get("regras_detectadas", ""),
        "Score Prioridade": classificacao.get("score_prioridade", 20),
        "Tipo de Decisão": classificacao.get("tipo_decisao", "Triagem"),
        "Análise ChatGPT": classificacao.get("analise_chatgpt", ""),
        "Urgência Sugerida": classificacao.get("urgencia", ""),
        "Motivo da Urgência": classificacao.get("motivo_urgencia", ""),
        "Categoria Sugerida": classificacao.get("categoria", "Não classificado"),
        "Ação Recomendada": classificacao.get("acao_recomendada", ""),
        "Responsável Sugerido": classificacao.get("responsavel", ""),
        "Observação RH": classificacao.get("observacao", ""),
        "Status Atendimento": classificacao.get("status_atendimento", "Pendente"),
        "Resposta Sugerida (IA)": classificacao.get("resposta_sugerida", ""),
        "E-mails no Caso (sessão)": 1, "Ocorrências Históricas": 1, "Caso Recorrente": "Não",
        "Rascunho Outlook Criado": "Não",
    }


def parse_data_br(txt):
    """Converte 'dd/mm/aaaa' (ou variações) em datetime. Retorna None se vazio/ inválido."""
    txt = (txt or "").strip()
    if not txt:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(txt, fmt)
        except Exception:
            continue
    return None


def _dentro_do_intervalo(recebido, dt_de, dt_ate):
    """True se a data recebida está dentro de [dt_de, dt_ate] (limites opcionais)."""
    if recebido is None:
        return True
    try:
        d = recebido.date()
    except Exception:
        return True
    if dt_de and d < dt_de.date():
        return False
    if dt_ate and d > dt_ate.date():
        return False
    return True


def listar_cabecalhos_outlook(modo_caixa, caixa_compartilhada, pasta_outlook,
                              pasta_entry_id="", pasta_store_id="",
                              data_de="", data_ate="", ultimos_dias=0,
                              somente_nao_lidos=False, limite=1000, log_callback=None):
    """
    Lê SOMENTE cabeçalhos (data, remetente, assunto, lido) — sem IA e sem
    extração pesada. Serve para o usuário escolher, com filtro, o que analisar.
    Itens vêm ordenados do mais recente para o mais antigo; como a lista é
    ordenada, quando encontramos um item mais antigo que 'data_de' paramos.
    """
    pythoncom.CoInitialize()
    try:
        pasta = obter_pasta_outlook(
            modo_caixa=modo_caixa, caixa_compartilhada=caixa_compartilhada,
            pasta_outlook=pasta_outlook, pasta_entry_id=pasta_entry_id,
            pasta_store_id=pasta_store_id, log_callback=log_callback,
        )
        items = pasta.Items
        items.Sort("[ReceivedTime]", True)

        dt_de = parse_data_br(data_de)
        dt_ate = parse_data_br(data_ate)
        limite = int(limite) if limite else 1000

        limite_dias = None
        try:
            ultimos_dias = int(ultimos_dias)
        except Exception:
            ultimos_dias = 0

        cabecalhos = []
        for item in items:
            if len(cabecalhos) >= limite:
                break
            try:
                if getattr(item, "Class", None) != 43:
                    continue
                try:
                    recebido = normalizar_data_outlook(getattr(item, "ReceivedTime", None))
                except Exception:
                    recebido = None

                # Ordenado desc: se já passou do início do intervalo, pode parar.
                if dt_de and recebido is not None:
                    try:
                        if recebido.date() < dt_de.date():
                            break
                    except Exception:
                        pass

                if not _dentro_do_intervalo(recebido, dt_de, dt_ate):
                    continue

                # Se não há intervalo, ainda respeita 'últimos dias' como antes.
                if not dt_de and not dt_ate and ultimos_dias > 0 and recebido is not None:
                    dias = dias_desde_recebimento(recebido)
                    if isinstance(dias, int) and dias > ultimos_dias:
                        break

                if somente_nao_lidos:
                    try:
                        if not bool(getattr(item, "UnRead", False)):
                            continue
                    except Exception:
                        continue

                try:
                    remetente = str(getattr(item, "SenderName", "") or "")
                except Exception:
                    remetente = ""
                try:
                    email_rem = get_smtp_address(item)
                except Exception:
                    email_rem = ""
                try:
                    assunto = str(getattr(item, "Subject", "") or "")
                except Exception:
                    assunto = ""
                try:
                    lido = "Não" if bool(getattr(item, "UnRead", False)) else "Sim"
                except Exception:
                    lido = ""

                data_txt = ""
                hora_txt = ""
                if recebido is not None:
                    try:
                        data_txt = recebido.strftime("%d/%m/%Y")
                        hora_txt = recebido.strftime("%H:%M")
                    except Exception:
                        pass

                cabecalhos.append({
                    "EntryID": getattr(item, "EntryID", "") or "",
                    "StoreID": pasta_store_id or "",
                    "Data": data_txt, "Hora": hora_txt,
                    "Remetente Nome": remetente, "Remetente E-mail": email_rem,
                    "Assunto": assunto, "Lido": lido,
                })
            except Exception:
                continue

        if log_callback:
            log_callback(f"Cabeçalhos listados (sem IA): {len(cabecalhos)}.")
        return cabecalhos
    finally:
        pythoncom.CoUninitialize()


def ler_emails_outlook(
    modo_caixa,
    caixa_compartilhada,
    pasta_outlook,
    quantidade,
    ultimos_dias,
    somente_nao_lidos,
    config_ia=None,
    pasta_entry_id="",
    pasta_store_id="",
    data_de="",
    data_ate="",
    progress_callback=None,
    log_callback=None,
):
    pythoncom.CoInitialize()

    try:
        pasta = obter_pasta_outlook(
            modo_caixa=modo_caixa,
            caixa_compartilhada=caixa_compartilhada,
            pasta_outlook=pasta_outlook,
            pasta_entry_id=pasta_entry_id,
            pasta_store_id=pasta_store_id,
            log_callback=log_callback,
        )

        items = pasta.Items
        items.Sort("[ReceivedTime]", True)

        total_outlook = items.Count

        if log_callback:
            log_callback(f"Total aparente de itens na pasta: {total_outlook}")

        registros = []
        erros = []

        quantidade = int(quantidade)
        ultimos_dias = int(ultimos_dias)

        dt_de = parse_data_br(data_de)
        dt_ate = parse_data_br(data_ate)
        usar_intervalo = bool(dt_de or dt_ate)

        limite_data = None
        if not usar_intervalo and ultimos_dias > 0:
            limite_data = hoje_data()
            if log_callback:
                log_callback(f"Filtro de período ativado: últimos {ultimos_dias} dias.")
        if usar_intervalo and log_callback:
            log_callback(f"Filtro por intervalo: {data_de or 'início'} até {data_ate or 'hoje'}.")

        analisados = 0
        coletados = 0

        for item in items:
            if coletados >= quantidade:
                break

            analisados += 1

            try:
                item_class = getattr(item, "Class", None)
                if item_class != 43:
                    continue

                try:
                    recebido = normalizar_data_outlook(getattr(item, "ReceivedTime", None))
                except Exception:
                    recebido = None

                if usar_intervalo:
                    # Ordenado desc: mais antigo que o início => pode parar.
                    if dt_de and recebido is not None:
                        try:
                            if recebido.date() < dt_de.date():
                                break
                        except Exception:
                            pass
                    if not _dentro_do_intervalo(recebido, dt_de, dt_ate):
                        continue
                elif ultimos_dias > 0 and recebido:
                    dias = dias_desde_recebimento(recebido)
                    if isinstance(dias, int) and dias > ultimos_dias:
                        continue

                if somente_nao_lidos:
                    try:
                        if not bool(getattr(item, "UnRead", False)):
                            continue
                    except Exception:
                        continue

                dados = extrair_dados_email(item, config_ia=config_ia, log_callback=log_callback)
                # V2: nenhum rascunho é criado durante a varredura. A Mesa Copilot exige seleção humana.

                registros.append(dados)
                coletados += 1

                if progress_callback:
                    progress_callback(coletados, quantidade)

                if log_callback and coletados % 10 == 0:
                    log_callback(f"E-mails coletados: {coletados}")

            except Exception as e:
                erros.append(str(e))
                continue

        if log_callback:
            log_callback(f"Leitura finalizada. Itens analisados: {analisados}. E-mails coletados: {coletados}.")

        return registros, erros

    finally:
        pythoncom.CoUninitialize()


def enriquecer_casos_sessao(registros):
    grupos = defaultdict(list)
    for r in registros:
        grupos[r.get("Case ID", "")].append(r)
    for case_id, itens in grupos.items():
        qtd = len(itens)
        for r in itens:
            r["E-mails no Caso (sessão)"] = qtd
    return registros


def carregar_historico_casos():
    caminho = CONFIG_DIR / "historico_casos.json"
    if not caminho.exists():
        return {}
    try:
        return json.loads(caminho.read_text(encoding="utf-8"))
    except Exception:
        return {}


def atualizar_historico_casos(registros):
    caminho = CONFIG_DIR / "historico_casos.json"
    hist = carregar_historico_casos()
    for r in registros:
        cid = r.get("Case ID", "")
        if not cid:
            continue
        item = hist.setdefault(cid, {"entry_ids": [], "primeira_vez": r.get("Data Recebimento", "")})
        ids = item.setdefault("entry_ids", [])
        eid = r.get("EntryID", "")
        if eid and eid not in ids:
            ids.append(eid)
            if len(ids) > 100:
                del ids[:-100]
        item.update({
            "ultima_vez": r.get("Data Recebimento", ""),
            "assunto": r.get("ConversationTopic") or r.get("Assunto", ""),
            "categoria": r.get("Categoria Sugerida", ""),
            "fila": r.get("Fila Copilot", ""),
            "estado": r.get("Estado do Caso", ""),
        })
        r["Ocorrências Históricas"] = len(ids) or 1
        r["Caso Recorrente"] = "Sim" if len(ids) > 1 else "Não"
    try:
        caminho.write_text(json.dumps(hist, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass
    return registros


def ordenar_registros_copilot(registros):
    ordem = {"Crítico": 0, "Precisa de mim": 1, "Delegável": 2, "Aguardando terceiro": 3, "Informativo": 4}
    return sorted(registros, key=lambda r: (ordem.get(r.get("Fila Copilot"), 9), -int(r.get("Score Prioridade", 0) or 0), r.get("Data Recebimento", "")))



def _data_hora_registro(registro):
    try:
        data_txt = str(registro.get("Data Recebimento", "") or "").strip()
        hora_txt = str(registro.get("Hora Recebimento", "") or "00:00:00").strip() or "00:00:00"
        return datetime.strptime(f"{data_txt} {hora_txt}", "%d/%m/%Y %H:%M:%S")
    except Exception:
        return datetime.min


def consolidar_registros_por_caso(registros):
    """Consolida uma thread/caso em uma única linha para o resumo executivo."""
    grupos = defaultdict(list)
    for r in registros or []:
        chave = r.get("Case ID") or r.get("ConversationID") or r.get("EntryID") or str(id(r))
        grupos[chave].append(r)

    ordem_fila = {"Crítico": 5, "Precisa de mim": 4, "Delegável": 3, "Aguardando terceiro": 2, "Informativo": 1}
    saida = []
    for _, itens in grupos.items():
        mais_recente = max(itens, key=_data_hora_registro)
        mais_critico = max(
            itens,
            key=lambda x: (ordem_fila.get(x.get("Fila Copilot"), 0), int(x.get("Score Prioridade", 0) or 0), _data_hora_registro(x)),
        )
        consolidado = dict(mais_recente)
        consolidado["Fila Copilot"] = mais_critico.get("Fila Copilot", mais_recente.get("Fila Copilot", "Informativo"))
        consolidado["Score Prioridade"] = max(int(x.get("Score Prioridade", 0) or 0) for x in itens)
        consolidado["Urgência Sugerida"] = "Alta" if any(x.get("Urgência Sugerida") == "Alta" for x in itens) else mais_recente.get("Urgência Sugerida", "")
        consolidado["E-mails no Caso (sessão)"] = len(itens)
        # Se o e-mail mais recente vier com campos vazios, aproveita a análise mais crítica.
        for campo in [
            "Resumo Executivo", "Pendência / Decisão", "Impacto", "Prazo / Timing",
            "Próximo Passo", "Responsável Sugerido", "Categoria Sugerida", "Regras Detectadas"
        ]:
            if not str(consolidado.get(campo, "") or "").strip():
                consolidado[campo] = mais_critico.get(campo, "")
        saida.append(consolidado)
    return ordenar_registros_copilot(saida)


def registro_impacta_folha(registro):
    categoria = normalizar_busca(registro.get("Categoria Sugerida", ""))
    regras = normalizar_busca(registro.get("Regras Detectadas", ""))
    texto = normalizar_busca(" ".join(str(registro.get(c, "") or "") for c in [
        "Assunto", "Resumo Executivo", "Impacto", "Pendência / Decisão", "Categoria Sugerida", "Regras Detectadas", "Corpo Limpo"
    ]))

    categorias_diretas = [
        "folha de pagamento", "encargos", "remuneracao variavel", "desligamento",
        "relacoes trabalhistas / cct"
    ]
    if any(c in categoria for c in categorias_diretas):
        return True

    gatilhos = [
        "impacto direto em folha", "folha", "payroll", "salario", "holerite", "retroativo",
        "desconto", "adiantamento", "rescisao", "trct", "fgts", "inss", "irrf", "dctfweb",
        "esocial", "consignado", "cct", "convencao coletiva", "dissidio", "comissao", "premio",
        "plr", "ppr", "vcc", "hora extra", "banco de horas", "reflexo em folha"
    ]
    return any(g in regras or g in texto for g in gatilhos)


def registro_e_urgente(registro):
    fila = registro.get("Fila Copilot", "")
    urg = registro.get("Urgência Sugerida", "")
    try:
        score = int(registro.get("Score Prioridade", 0) or 0)
    except Exception:
        score = 0
    return fila == "Crítico" or urg == "Alta" or score >= 82


def selecionar_casos_folha_urgencias(registros):
    casos = consolidar_registros_por_caso(registros)
    return [r for r in casos if registro_impacta_folha(r) or registro_e_urgente(r)]


def _texto_caso_resumo(registro):
    assunto = str(registro.get("ConversationTopic") or registro.get("Assunto") or "Sem assunto").strip()
    resumo = str(registro.get("Resumo Executivo", "") or registro.get("Análise ChatGPT", "") or "").strip()
    impacto = str(registro.get("Impacto", "") or "A validar").strip()
    prazo = str(registro.get("Prazo / Timing", "") or "").strip()
    proximo = str(registro.get("Próximo Passo", "") or registro.get("Ação Recomendada", "") or "").strip()
    responsavel = str(registro.get("Responsável Sugerido", "") or "").strip()
    fila = str(registro.get("Fila Copilot", "") or "Informativo")
    score = str(registro.get("Score Prioridade", "") or "")
    partes = [f"- [{fila}{' | score ' + score if score else ''}] {assunto}"]
    if resumo:
        partes.append(f"  Resumo: {resumo}")
    if impacto:
        partes.append(f"  Impacto: {impacto}")
    if prazo:
        partes.append(f"  Prazo: {prazo}")
    if proximo:
        partes.append(f"  Próximo passo: {proximo}")
    if responsavel:
        partes.append(f"  Responsável: {responsavel}")
    return "\n".join(partes)


def montar_email_resumo_deterministico(registros, destino_comunicacao=DESTINO_PADRAO, origem="automatico"):
    casos = consolidar_registros_por_caso(registros)
    impactos_folha = [r for r in casos if registro_impacta_folha(r)]
    urgencias_adicionais = [r for r in casos if registro_e_urgente(r) and not registro_impacta_folha(r)]
    outros = [r for r in casos if r not in impactos_folha and r not in urgencias_adicionais]

    data_txt = datetime.now().strftime("%d/%m/%Y")
    if origem == "automatico":
        assunto = f"[Copilot RH] Impactos em Folha e urgências - {data_txt}"
    else:
        assunto = f"[Copilot RH] Resumo dos casos selecionados - {data_txt}"

    nome = destino_comunicacao if destino_comunicacao in {"Joana", "Lisiane", "Ricardo Ribeiro"} else ""
    linhas = []
    if nome:
        linhas.append(f"{nome},")
        linhas.append("")
    linhas.append("Segue um resumo consolidado dos temas que exigem atenção, considerando cada conversa como um único caso.")
    linhas.append("")
    linhas.append(f"Casos consolidados: {len(casos)} | Impactos em Folha: {len(impactos_folha)} | Urgências adicionais: {len(urgencias_adicionais)}")

    if impactos_folha:
        linhas.extend(["", "IMPACTOS EM FOLHA", ""])
        for r in impactos_folha:
            linhas.append(_texto_caso_resumo(r))
            linhas.append("")

    if urgencias_adicionais:
        linhas.extend(["URGÊNCIAS ADICIONAIS", ""])
        for r in urgencias_adicionais:
            linhas.append(_texto_caso_resumo(r))
            linhas.append("")

    if origem != "automatico" and outros:
        linhas.extend(["OUTROS CASOS SELECIONADOS", ""])
        for r in outros:
            linhas.append(_texto_caso_resumo(r))
            linhas.append("")

    linhas.append("Pontos acima devem ser revisados antes de qualquer decisão ou envio externo.")
    return assunto, "\n".join(linhas).strip()


def gerar_email_resumo_com_ia(registros, config_ia, destino_comunicacao=DESTINO_PADRAO, origem="automatico", log_callback=None):
    """Gera um único e-mail de resumo. Em falha/IA desligada, usa o resumo determinístico."""
    assunto_fallback, corpo_fallback = montar_email_resumo_deterministico(registros, destino_comunicacao, origem)
    if not config_ia or not config_ia.get("ia_ativa"):
        return assunto_fallback, corpo_fallback

    api_key = (config_ia.get("ia_api_key") or "").strip() or os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return assunto_fallback, corpo_fallback

    casos = consolidar_registros_por_caso(registros)[:35]
    blocos = []
    total_chars = 0
    for r in casos:
        corpo_trecho = str(r.get("Corpo Limpo", "") or "")[:650]
        bloco = (
            f"CASO {r.get('Case ID','')}\n"
            f"Fila: {r.get('Fila Copilot','')} | Score: {r.get('Score Prioridade','')} | Urgência: {r.get('Urgência Sugerida','')}\n"
            f"Categoria: {r.get('Categoria Sugerida','')}\n"
            f"Assunto: {r.get('ConversationTopic') or r.get('Assunto','')}\n"
            f"Resumo já apurado: {r.get('Resumo Executivo','')}\n"
            f"Impacto: {r.get('Impacto','')}\nPrazo: {r.get('Prazo / Timing','')}\n"
            f"Pendência/decisão: {r.get('Pendência / Decisão','')}\nPróximo passo: {r.get('Próximo Passo','')}\n"
            f"Responsável: {r.get('Responsável Sugerido','')}\nTrecho do e-mail mais recente: {corpo_trecho}\n"
        )
        if total_chars + len(bloco) > 22000:
            break
        blocos.append(bloco)
        total_chars += len(bloco)

    destino_prompt = DESTINOS_COMUNICACAO.get(destino_comunicacao, DESTINOS_COMUNICACAO[DESTINO_PADRAO]).get("prompt", "")
    system = (
        "Você prepara um ÚNICO e-mail-resumo executivo de RH/Payroll a partir de vários casos já lidos pelo Copilot. "
        "Use exclusivamente os fatos fornecidos. Não invente valores, prazos, decisões, causas, políticas ou conclusões. "
        "Agrupe por impacto em Folha e urgências adicionais, elimine duplicidades da mesma conversa e priorize o que muda decisão, prazo, pagamento ou risco. "
        "O texto deve ser curto o suficiente para leitura executiva, mas manter próximo passo, responsável e prazo quando existirem. "
        + destino_prompt + " "
        "Retorne exclusivamente JSON válido com as chaves assunto e corpo. Não inclua assinatura no corpo."
    )
    user = (
        f"Tipo de resumo: {'Folha + urgências automáticas' if origem == 'automatico' else 'casos selecionados pelo usuário'}\n"
        f"Destinatário/tom: {destino_comunicacao}\n\nCASOS:\n" + "\n---\n".join(blocos)
    )

    try:
        from anthropic import Anthropic
        headers = {}
        workspace_id = (config_ia.get("ia_workspace_id") or "").strip() or os.environ.get("ANTHROPIC_WORKSPACE_ID", "").strip()
        if workspace_id:
            headers["anthropic-workspace-id"] = workspace_id
        client = Anthropic(api_key=api_key, default_headers=headers) if headers else Anthropic(api_key=api_key)
        resp = client.messages.create(
            model=config_ia.get("ia_modelo") or IA_MODELO_PADRAO,
            max_tokens=2400,
            system=system,
            messages=[{"role": "user", "content": user}],
        )
        texto = "\n".join(getattr(x, "text", "") for x in resp.content if getattr(x, "text", "")).strip()
        dados = _extrair_json(texto)
        assunto = str(dados.get("assunto", "") or "").strip() if dados else ""
        corpo = str(dados.get("corpo", "") or "").strip() if dados else ""
        if assunto and corpo:
            return assunto, corpo
    except Exception as e:
        if log_callback:
            log_callback(f"Resumo consolidado: falha na IA; usando versão local. {e}")
    return assunto_fallback, corpo_fallback


def criar_rascunho_novo_outlook(destinatario, assunto, corpo, assinatura="", abrir=True, log_callback=None):
    """Cria um único MailItem em Rascunhos. Nunca envia automaticamente."""
    corpo = (corpo or "").strip()
    assunto = (assunto or "Resumo RH").strip()
    if not corpo:
        return False
    try:
        outlook = win32.Dispatch("Outlook.Application")
        draft = outlook.CreateItem(0)  # olMailItem
        if (destinatario or "").strip():
            draft.To = destinatario.strip()
        draft.Subject = assunto
        texto = corpo
        if (assinatura or "").strip():
            texto += "\n\n" + assinatura.strip()
        draft.Body = texto
        draft.Save()
        if abrir:
            draft.Display()
        return True
    except Exception as e:
        if log_callback:
            log_callback(f"Falha ao criar rascunho consolidado: {e}")
        return False


def aplicar_estilo_planilha(ws, tipo="base"):
    navy = CORES["navy"].replace("#", "")
    branco = "FFFFFF"
    borda_cor = "D9DDE3"
    fill_header = PatternFill("solid", fgColor=navy)
    font_header = Font(color=branco, bold=True)
    border = Border(
        left=Side(style="thin", color=borda_cor), right=Side(style="thin", color=borda_cor),
        top=Side(style="thin", color=borda_cor), bottom=Side(style="thin", color=borda_cor),
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = fill_header; cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx in range(1, ws.max_column + 1):
        letra = get_column_letter(idx)
        titulo = str(ws.cell(1, idx).value or "")
        if titulo in {"Corpo do E-mail", "Corpo Limpo", "Resposta Sugerida (IA)"}:
            largura = 70
        elif titulo in {"Assunto", "Resumo Executivo", "Próximo Passo", "Ação Recomendada", "Observação RH"}:
            largura = 45
        elif titulo in {"EntryID", "StoreID", "ConversationID"}:
            largura = 22
        else:
            largura = min(max(len(titulo) + 4, 14), 30)
        ws.column_dimensions[letra].width = largura
    ws.row_dimensions[1].height = 32
    if tipo == "base":
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 70


def criar_aba_base(wb, registros):
    ws = wb.active
    ws.title = "BASE_EMAILS"
    ws.append(COLUNAS_BASE)
    for registro in registros:
        ws.append([registro.get(coluna, "") for coluna in COLUNAS_BASE])
    aplicar_estilo_planilha(ws, "base")


def criar_aba_prioridades(wb, registros):
    ws = wb.create_sheet("COPILOT_PRIORIDADES")
    colunas = ["Fila Copilot", "Score Prioridade", "Case ID", "Data Recebimento", "Remetente Nome", "Assunto",
               "Resumo Executivo", "Pendência / Decisão", "Impacto", "Prazo / Timing", "Próximo Passo",
               "Responsável Sugerido", "Pode Delegar", "Estado do Caso", "Rascunho Outlook Criado"]
    ws.append(colunas)
    for r in ordenar_registros_copilot(registros):
        ws.append([r.get(c, "") for c in colunas])
    aplicar_estilo_planilha(ws, "prioridades")


def criar_aba_casos(wb, registros):
    ws = wb.create_sheet("CASOS")
    cols = ["Case ID", "E-mails na sessão", "Último assunto", "Categoria", "Fila mais crítica", "Maior score",
            "Responsável", "Estado do Caso", "Ocorrências Históricas", "Caso Recorrente"]
    ws.append(cols)
    grupos = defaultdict(list)
    for r in registros:
        grupos[r.get("Case ID", "")].append(r)
    ordem = {"Crítico": 5, "Precisa de mim": 4, "Delegável": 3, "Aguardando terceiro": 2, "Informativo": 1}
    for cid, itens in grupos.items():
        principal = max(itens, key=lambda x: (ordem.get(x.get("Fila Copilot"), 0), int(x.get("Score Prioridade", 0) or 0)))
        ws.append([
            cid, len(itens), principal.get("ConversationTopic") or principal.get("Assunto", ""),
            principal.get("Categoria Sugerida", ""), principal.get("Fila Copilot", ""),
            max(int(x.get("Score Prioridade", 0) or 0) for x in itens),
            principal.get("Responsável Sugerido", ""), principal.get("Estado do Caso", ""),
            max(int(x.get("Ocorrências Históricas", 1) or 1) for x in itens),
            "Sim" if any(x.get("Caso Recorrente") == "Sim" for x in itens) else "Não",
        ])
    aplicar_estilo_planilha(ws, "casos")


def criar_aba_glossario(wb):
    ws = wb.create_sheet("GLOSSARIO")
    ws.append(COLUNAS_GLOSSARIO)
    ws.append(["", "", "", "", "", "Use esta aba para futuras regras validadas pela operação.", "Sim"])
    aplicar_estilo_planilha(ws, "glossario")


def criar_aba_resumo(wb, registros):
    ws = wb.create_sheet("RESUMO")
    filas = Counter(r.get("Fila Copilot", "Informativo") for r in registros)
    categorias = Counter(r.get("Categoria Sugerida", "Não classificado") for r in registros)
    total = len(registros)
    linhas = [
        ["Indicador", "Valor"], ["Total de e-mails analisados", total],
        ["Críticos", filas.get("Crítico", 0)], ["Precisa de mim", filas.get("Precisa de mim", 0)],
        ["Delegáveis", filas.get("Delegável", 0)], ["Aguardando terceiro", filas.get("Aguardando terceiro", 0)],
        ["Informativos", filas.get("Informativo", 0)],
        ["Com resposta sugerida", sum(1 for r in registros if str(r.get("Resposta Sugerida (IA)", "")).strip())],
        ["Casos únicos", len({r.get("Case ID") for r in registros if r.get("Case ID")})],
        ["Data da extração", agora_texto()], ["Aplicativo", f"{APP_NAME} v{APP_VERSION}"],
    ]
    for k, v in categorias.most_common(10):
        linhas.append([f"Categoria: {k}", v])
    for linha in linhas:
        ws.append(linha)
    aplicar_estilo_planilha(ws, "resumo")


def gerar_excel(registros, caminho_saida):
    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    if caminho_saida.exists():
        try:
            os.rename(caminho_saida, caminho_saida)
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            caminho_saida = caminho_saida.with_name(f"{caminho_saida.stem}_{timestamp}{caminho_saida.suffix}")
    wb = Workbook()
    criar_aba_base(wb, registros)
    criar_aba_prioridades(wb, registros)
    criar_aba_casos(wb, registros)
    criar_aba_glossario(wb)
    criar_aba_resumo(wb, registros)
    wb.save(caminho_saida)
    return caminho_saida


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        garantir_pastas()

        self.config_app = carregar_config()
        self.processo_rodando = False
        self.caminho_excel_gerado = None
        self.registros_sessao = []
        self.registro_em_edicao = None

        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.title(APP_NAME)
        self.geometry("1240x880")
        self.minsize(1120, 760)
        self.configure(fg_color=CORES["fundo"])

        self.criar_variaveis()
        self.criar_layout()
        self.carregar_variaveis_config()
        self.iniciar_agendador()

        # Se iniciado pela pasta Inicializar do Windows: vai para a bandeja e roda.
        self._modo_startup = ("--startup" in sys.argv) or ("--auto" in sys.argv)
        if self._modo_startup:
            self.after(60000, self._boot_autorun)  # aguarda ~60s o sistema/Outlook subirem

    def criar_variaveis(self):
        self.var_modo_caixa = ctk.StringVar(value="principal")
        self.var_caixa_compartilhada = ctk.StringVar(value="")
        self.var_pasta_outlook = ctk.StringVar(value="Caixa de Entrada")
        self.var_pasta_entry_id = ctk.StringVar(value="")
        self.var_pasta_store_id = ctk.StringVar(value="")
        self.var_pasta_display = ctk.StringVar(value="")
        self.var_quantidade = ctk.StringVar(value="200")
        self.var_ultimos_dias = ctk.StringVar(value="30")
        self.var_data_de = ctk.StringVar(value="")
        self.var_data_ate = ctk.StringVar(value="")
        self.var_somente_nao_lidos = ctk.BooleanVar(value=False)
        self.var_pasta_saida = ctk.StringVar(value=str(SAIDA_DIR))
        self.var_nome_excel = ctk.StringVar(value=DEFAULT_EXCEL_NAME)
        self.var_ia_ativa = ctk.BooleanVar(value=False)
        self.var_criar_rascunho = ctk.BooleanVar(value=False)
        self.var_resumo_folha_auto = ctk.BooleanVar(value=True)
        self.var_modo_operacional = ctk.StringVar(value=MODO_PADRAO)
        self.var_destino_comunicacao = ctk.StringVar(value=DESTINO_PADRAO)
        self.var_comando_rapido = ctk.StringVar(value="")
        self.var_filtro_fila = ctk.StringVar(value="Todas")
        self.var_busca_copilot = ctk.StringVar(value="")
        self.var_ia_api_key = ctk.StringVar(value="")
        self.var_ia_workspace_id = ctk.StringVar(value="")
        self.var_resumo_destinatario_email = ctk.StringVar(value="")
        self.var_agendamento_ativo = ctk.BooleanVar(value=False)
        self.var_agendamento_hora = ctk.StringVar(value="05:00")
        self.var_alerta_urgente = ctk.BooleanVar(value=True)
        self.var_salvar_chave = ctk.BooleanVar(value=False)
        self.var_iniciar_com_windows = ctk.BooleanVar(value=False)
        self.var_ia_modelo_rotulo = ctk.StringVar(value=self._rotulo_por_modelo(IA_MODELO_PADRAO))
        self._tray_icon = None
        self._agendador_ativo = False
        self._ultima_exec_agendada = None

    def _rotulo_por_modelo(self, modelo_id):
        for rotulo, mid in IA_MODELOS.items():
            if mid == modelo_id:
                return rotulo
        return next(iter(IA_MODELOS))

    def _modelo_por_rotulo(self, rotulo):
        return IA_MODELOS.get(rotulo, IA_MODELO_PADRAO)

    def carregar_variaveis_config(self):
        c = self.config_app
        self.var_modo_caixa.set(c.get("modo_caixa", "principal"))
        self.var_caixa_compartilhada.set(c.get("caixa_compartilhada", ""))
        self.var_pasta_outlook.set(c.get("pasta_outlook", "Caixa de Entrada"))
        self.var_pasta_entry_id.set(c.get("pasta_entry_id", "")); self.var_pasta_store_id.set(c.get("pasta_store_id", ""))
        self.var_pasta_display.set(c.get("pasta_display", "")); self.var_quantidade.set(str(c.get("quantidade", 200)))
        self.var_ultimos_dias.set(str(c.get("ultimos_dias", 30))); self.var_somente_nao_lidos.set(bool(c.get("somente_nao_lidos", False)))
        self.var_data_de.set(c.get("data_de", "")); self.var_data_ate.set(c.get("data_ate", ""))
        self.var_pasta_saida.set(c.get("pasta_saida", str(SAIDA_DIR))); self.var_nome_excel.set(c.get("nome_excel", DEFAULT_EXCEL_NAME))
        self.var_ia_ativa.set(bool(c.get("ia_ativa", False))); self.var_criar_rascunho.set(bool(c.get("criar_rascunho_outlook", False)))
        self.var_resumo_folha_auto.set(bool(c.get("resumo_folha_auto", True)))
        modo = c.get("modo_operacional", c.get("perfil_agente", MODO_PADRAO))
        self.var_modo_operacional.set(modo if modo in MODOS_OPERACIONAIS else MODO_PADRAO)
        destino = c.get("destino_comunicacao", DESTINO_PADRAO)
        self.var_destino_comunicacao.set(destino if destino in DESTINOS_COMUNICACAO else DESTINO_PADRAO)
        self.var_ia_api_key.set(c.get("ia_api_key", "")); self.var_ia_workspace_id.set(c.get("ia_workspace_id", ""))
        self.var_resumo_destinatario_email.set(c.get("resumo_destinatario_email", ""))
        self.var_agendamento_ativo.set(bool(c.get("agendamento_ativo", False)))
        self.var_agendamento_hora.set(c.get("agendamento_hora", "05:00"))
        self.var_alerta_urgente.set(bool(c.get("alerta_urgente", True)))
        self.var_salvar_chave.set(bool(c.get("salvar_chave", False)))
        try:
            self.var_iniciar_com_windows.set(os.path.exists(self._caminho_atalho_startup()))
        except Exception:
            self.var_iniciar_com_windows.set(False)
        self.var_ia_modelo_rotulo.set(self._rotulo_por_modelo(c.get("ia_modelo", IA_MODELO_PADRAO)))
        self.ao_mudar_modo(); self.ao_mudar_destino()

    def obter_config_tela(self):
        return {
            "modo_caixa": self.var_modo_caixa.get(), "caixa_compartilhada": self.var_caixa_compartilhada.get().strip(),
            "pasta_outlook": self.var_pasta_outlook.get().strip(), "pasta_entry_id": self.var_pasta_entry_id.get().strip(),
            "pasta_store_id": self.var_pasta_store_id.get().strip(), "pasta_display": self.var_pasta_display.get().strip(),
            "quantidade": int(self.var_quantidade.get()), "ultimos_dias": int(self.var_ultimos_dias.get()),
            "data_de": self.var_data_de.get().strip(), "data_ate": self.var_data_ate.get().strip(),
            "somente_nao_lidos": bool(self.var_somente_nao_lidos.get()), "pasta_saida": self.var_pasta_saida.get().strip(),
            "nome_excel": self.var_nome_excel.get().strip(), "ia_ativa": bool(self.var_ia_ativa.get()),
            "criar_rascunho_outlook": bool(self.var_criar_rascunho.get()),
            "resumo_folha_auto": bool(self.var_resumo_folha_auto.get()),
            "modo_operacional": self.var_modo_operacional.get(), "destino_comunicacao": self.var_destino_comunicacao.get(),
            "perfil_agente": self.var_modo_operacional.get(),
            "ia_modelo": self._modelo_por_rotulo(self.var_ia_modelo_rotulo.get()),
            "ia_api_key": self.var_ia_api_key.get().strip(), "ia_workspace_id": self.var_ia_workspace_id.get().strip(),
            "resumo_destinatario_email": self.var_resumo_destinatario_email.get().strip(),
            "agendamento_ativo": bool(self.var_agendamento_ativo.get()),
            "agendamento_hora": self.var_agendamento_hora.get().strip() or "05:00",
            "alerta_urgente": bool(self.var_alerta_urgente.get()),
            "salvar_chave": bool(self.var_salvar_chave.get()),
            "ia_assinatura": self.config_app.get("ia_assinatura", "Atenciosamente,\nGente e Gestão – Sonova"),
            "ia_contexto_rh": self.config_app.get("ia_contexto_rh", ""),
        }

    def _descricao_modo(self):
        return MODOS_OPERACIONAIS.get(self.var_modo_operacional.get(), MODOS_OPERACIONAIS[MODO_PADRAO]).get("descricao", "")

    def _descricao_destino(self):
        return DESTINOS_COMUNICACAO.get(self.var_destino_comunicacao.get(), DESTINOS_COMUNICACAO[DESTINO_PADRAO]).get("descricao", "")

    def ao_mudar_modo(self, _valor=None):
        try: self.lbl_modo_info.configure(text=self._descricao_modo())
        except Exception: pass

    def ao_mudar_destino(self, _valor=None):
        try: self.lbl_destino_info.configure(text=self._descricao_destino())
        except Exception: pass

    def interpretar_comando_copilot(self, texto):
        comando = normalizar_busca(texto)
        resultado = {}
        aliases_modo = [
            (["prioridades", "o que preciso fazer", "meu dia", "prioridade do dia"], "Prioridades do Dia"),
            (["folha", "payroll"], "Folha de Pagamento"), (["beneficios", "beneficio"], "Benefícios"),
            (["ponto", "jornada"], "Ponto e Jornada"), (["desligamento", "rescisao", "juridico trabalhista"], "Desligamentos / Jurídico"),
            (["admissao"], "Admissão"), (["ferias"], "Férias"), (["encargos", "esocial", "fgts", "dctfweb"], "Encargos / eSocial"),
            (["geral rh", "geral"], "Geral RH"),
        ]
        for termos, valor in aliases_modo:
            if any(t in comando for t in termos): resultado["modo"] = valor; break
        aliases_destino = [(["joana"], "Joana"), (["lisiane", "lisi"], "Lisiane"), (["ricardo ribeiro", "ricardo"], "Ricardo Ribeiro"),
                           (["colaborador", "funcionario"], "Colaborador"), (["gestor", "gerente"], "Gestor"),
                           (["fornecedor", "prestador"], "Fornecedor"), (["consulta juridica", "para juridico", "juridico"], "Jurídico")]
        for termos, valor in aliases_destino:
            if any(t in comando for t in termos): resultado["destino"] = valor; break
        if "nao lidos" in comando or "nao lido" in comando: resultado["nao_lidos"] = True
        if "todos os emails" in comando or "lidos e nao lidos" in comando: resultado["nao_lidos"] = False
        m = re.search(r"ultimos?\s+(\d+)\s+dias?", comando)
        if m: resultado["dias"] = m.group(1)
        m = re.search(r"(?:ler|varrer|buscar|extrair)?\s*(\d+)\s+(?:emails|e-mails)", comando)
        if m: resultado["quantidade"] = m.group(1)
        if "sem ia" in comando or "ia desligada" in comando: resultado["ia"] = False
        elif "com ia" in comando or "analisar" in comando or "responder" in comando: resultado["ia"] = True
        if "rascunho" in comando and "sem rascunho" not in comando: resultado["rascunho"] = True
        if "sem rascunho" in comando: resultado["rascunho"] = False
        mapa_filas = {"critico":"Crítico", "precisa de mim":"Precisa de mim", "delegavel":"Delegável",
                      "aguardando terceiro":"Aguardando terceiro", "informativo":"Informativo"}
        for termo, valor in mapa_filas.items():
            if termo in comando:
                resultado["filtro_fila"] = valor
                break
        resultado["executar"] = any(x in comando for x in ["executar", "varrer agora", "rodar agora", "extrair agora", "analisar agora"])
        return resultado

    def aplicar_comando_rapido(self):
        original = self.var_comando_rapido.get().strip()
        if not original: return
        r = self.interpretar_comando_copilot(original)
        if r.get("modo"): self.var_modo_operacional.set(r["modo"])
        if r.get("destino"): self.var_destino_comunicacao.set(r["destino"])
        if "nao_lidos" in r: self.var_somente_nao_lidos.set(r["nao_lidos"])
        if r.get("dias"): self.var_ultimos_dias.set(r["dias"])
        if r.get("quantidade"): self.var_quantidade.set(r["quantidade"])
        if "ia" in r: self.var_ia_ativa.set(r["ia"])
        if "rascunho" in r: self.var_criar_rascunho.set(r["rascunho"]); self.var_ia_ativa.set(True if r["rascunho"] else self.var_ia_ativa.get())
        if r.get("filtro_fila"): self.var_filtro_fila.set(r["filtro_fila"])
        self.ao_mudar_modo(); self.ao_mudar_destino()
        self.log(f"Comando Copilot aplicado: {original}")
        self.log(f"Modo: {self.var_modo_operacional.get()} | Destino: {self.var_destino_comunicacao.get()} | dias: {self.var_ultimos_dias.get()}")
        if r.get("executar"): self.after(150, self.iniciar_extracao)

    def criar_layout(self):
        self.grid_columnconfigure(0, weight=0)   # barra lateral (largura fixa)
        self.grid_columnconfigure(1, weight=1)   # conteudo
        self.grid_rowconfigure(0, weight=1)

        self.views = {}
        self.nav_botoes = {}

        self.criar_sidebar()

        direita = ctk.CTkFrame(self, fg_color=CORES["fundo"], corner_radius=0)
        direita.grid(row=0, column=1, sticky="nsew")
        direita.grid_columnconfigure(0, weight=1)
        direita.grid_rowconfigure(1, weight=1)

        self.criar_header(direita)
        self.criar_area_conteudo(direita)
        self.criar_rodape(direita)

        self.mostrar_view("copilot")

    def criar_sidebar(self):
        barra = ctk.CTkFrame(self, fg_color=CORES["fundo_lateral"], corner_radius=0, width=232)
        barra.grid(row=0, column=0, sticky="nsew")
        barra.grid_propagate(False)
        barra.grid_columnconfigure(0, weight=1)

        marca = ctk.CTkLabel(
            barra,
            text="Sonova",
            font=("Segoe UI", 22, "bold"),
            text_color=CORES["branco"],
            anchor="w",
        )
        marca.grid(row=0, column=0, padx=22, pady=(26, 0), sticky="ew")

        sub = ctk.CTkLabel(
            barra,
            text="Gente e Gestão",
            font=("Segoe UI", 12),
            text_color=CORES["azul_claro"],
            anchor="w",
        )
        sub.grid(row=1, column=0, padx=22, pady=(0, 26), sticky="ew")

        itens = [
            ("copilot", "Copilot do Dia"),
            ("config", "Origem e Saída"),
            ("ia", "Agente e Comunicação"),
            ("exec", "Execução"),
        ]

        linha = 2
        for chave, texto in itens:
            btn = ctk.CTkButton(
                barra,
                text=texto,
                anchor="w",
                command=lambda c=chave: self.mostrar_view(c),
                height=44,
                corner_radius=8,
                fg_color="transparent",
                hover_color=CORES["azul"],
                text_color=CORES["branco"],
                font=("Segoe UI", 14),
            )
            btn.grid(row=linha, column=0, padx=12, pady=4, sticky="ew")
            self.nav_botoes[chave] = btn
            linha += 1

        barra.grid_rowconfigure(linha, weight=1)

        versao = ctk.CTkLabel(
            barra,
            text=f"Leitor Outlook RH · v{APP_VERSION}",
            font=("Segoe UI", 11),
            text_color=CORES["azul_claro"],
            anchor="w",
            justify="left",
            wraplength=190,
        )
        versao.grid(row=linha + 1, column=0, padx=22, pady=(0, 16), sticky="sw")

    def mostrar_view(self, chave):
        for k, frame in self.views.items():
            if k == chave:
                frame.grid()
            else:
                frame.grid_remove()

        for k, btn in self.nav_botoes.items():
            if k == chave:
                btn.configure(fg_color=CORES["azul"])
            else:
                btn.configure(fg_color="transparent")

    def criar_header(self, parent):
        header = ctk.CTkFrame(parent, fg_color=CORES["azul"], corner_radius=0, height=92)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        titulo = ctk.CTkLabel(
            header,
            text="Agente IA Outlook RH Copilot",
            font=("Segoe UI", 26, "bold"),
            text_color=CORES["branco"],
            anchor="w",
        )
        titulo.grid(row=0, column=0, padx=28, pady=(18, 0), sticky="ew")

        subtitulo = ctk.CTkLabel(
            header,
            text="Varre a caixa, organiza prioridades, agrupa casos e prepara respostas. Rascunhos só são criados após sua seleção.",
            font=("Segoe UI", 14),
            text_color=CORES["branco"],
            anchor="w",
        )
        subtitulo.grid(row=1, column=0, padx=28, pady=(2, 18), sticky="ew")

    def _nova_view_scroll(self, container):
        view = ctk.CTkScrollableFrame(
            container,
            fg_color=CORES["fundo"],
            scrollbar_button_color=CORES["azul_claro"],
            scrollbar_button_hover_color=CORES["azul"],
        )
        view.grid(row=0, column=0, sticky="nsew", padx=18, pady=18)
        return view

    def _novo_card(self, parent, coluna=0, colspan=1, padx=0):
        card = ctk.CTkFrame(
            parent,
            fg_color=CORES["card"],
            border_color=CORES["borda"],
            border_width=1,
            corner_radius=14,
        )
        card.grid(row=0, column=coluna, columnspan=colspan, sticky="nsew", padx=padx)
        card.grid_columnconfigure(1, weight=1)
        return card

    def criar_area_conteudo(self, parent):
        container = ctk.CTkFrame(parent, fg_color=CORES["fundo"], corner_radius=0)
        container.grid(row=1, column=0, sticky="nsew")
        container.grid_columnconfigure(0, weight=1); container.grid_rowconfigure(0, weight=1)

        view_copilot = ctk.CTkFrame(container, fg_color=CORES["fundo"], corner_radius=0)
        view_copilot.grid(row=0, column=0, sticky="nsew", padx=18, pady=18)
        view_copilot.grid_columnconfigure(0, weight=1); view_copilot.grid_rowconfigure(2, weight=1)
        self.views["copilot"] = view_copilot
        self.criar_view_copilot(view_copilot)

        view_cfg = self._nova_view_scroll(container); view_cfg.grid_columnconfigure(0, weight=1); view_cfg.grid_columnconfigure(1, weight=1)
        self.views["config"] = view_cfg
        card_origem = self._novo_card(view_cfg, coluna=0, padx=(0, 10)); card_saida = self._novo_card(view_cfg, coluna=1, padx=(10, 0))
        self.criar_card_origem(card_origem); self.criar_card_saida(card_saida)
        card_agenda = ctk.CTkFrame(view_cfg, fg_color=CORES["card"], border_color=CORES["borda"], border_width=1, corner_radius=14)
        card_agenda.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(20, 0)); card_agenda.grid_columnconfigure(1, weight=1)
        self.criar_card_agendamento(card_agenda)

        view_ia = self._nova_view_scroll(container); view_ia.grid_columnconfigure(0, weight=1); self.views["ia"] = view_ia
        card_ia = self._novo_card(view_ia, coluna=0); self.criar_card_ia(card_ia)

        view_exec = ctk.CTkFrame(container, fg_color=CORES["fundo"], corner_radius=0)
        view_exec.grid(row=0, column=0, sticky="nsew", padx=18, pady=18); view_exec.grid_columnconfigure(0, weight=1); view_exec.grid_rowconfigure(0, weight=1)
        self.views["exec"] = view_exec; self.criar_area_log(view_exec)

    def criar_card_ia(self, parent):
        self.criar_titulo_card(parent, "Agente e Comunicação")
        chk_ia = ctk.CTkCheckBox(parent, text="Usar IA para análise, fila Copilot e resposta sugerida", variable=self.var_ia_ativa,
                                 fg_color=CORES["azul"], hover_color=CORES["azul_apoio"], text_color=CORES["texto"])
        chk_ia.grid(row=1, column=0, columnspan=3, padx=18, pady=(4, 4), sticky="w")
        frame_chks = ctk.CTkFrame(parent, fg_color="transparent")
        frame_chks.grid(row=2, column=0, columnspan=3, padx=18, pady=(0, 8), sticky="w")
        chk_rascunho = ctk.CTkCheckBox(frame_chks, text="Permitir criação de rascunhos após revisão na Mesa Copilot", variable=self.var_criar_rascunho,
                                       fg_color=CORES["azul"], hover_color=CORES["azul_apoio"], text_color=CORES["texto"])
        chk_rascunho.grid(row=0, column=0, sticky="w", pady=(0, 4))
        chk_resumo_auto = ctk.CTkCheckBox(frame_chks, text="Ao terminar a varredura, já preparar o e-mail único de Folha (onde atuar)", variable=self.var_resumo_folha_auto,
                                          fg_color=CORES["azul"], hover_color=CORES["azul_apoio"], text_color=CORES["texto"])
        chk_resumo_auto.grid(row=1, column=0, sticky="w")

        self.criar_label(parent, "Modo operacional", 3)
        ctk.CTkOptionMenu(parent, variable=self.var_modo_operacional, values=list(MODOS_OPERACIONAIS.keys()), command=self.ao_mudar_modo,
                          fg_color=CORES["azul"], button_color=CORES["azul_apoio"], button_hover_color=CORES["navy"], text_color=CORES["branco"]).grid(
                          row=3, column=1, columnspan=2, padx=(4,18), pady=7, sticky="ew")
        self.lbl_modo_info = ctk.CTkLabel(parent, text=self._descricao_modo(), font=("Segoe UI",12), text_color=CORES["texto_sec"],
                                          justify="left", wraplength=980, anchor="w")
        self.lbl_modo_info.grid(row=4, column=0, columnspan=3, padx=18, pady=(0,8), sticky="ew")

        self.criar_label(parent, "Destinatário / tom", 5)
        ctk.CTkOptionMenu(parent, variable=self.var_destino_comunicacao, values=list(DESTINOS_COMUNICACAO.keys()), command=self.ao_mudar_destino,
                          fg_color=CORES["azul"], button_color=CORES["azul_apoio"], button_hover_color=CORES["navy"], text_color=CORES["branco"]).grid(
                          row=5, column=1, columnspan=2, padx=(4,18), pady=7, sticky="ew")
        self.lbl_destino_info = ctk.CTkLabel(parent, text=self._descricao_destino(), font=("Segoe UI",12), text_color=CORES["texto_sec"],
                                             justify="left", wraplength=980, anchor="w")
        self.lbl_destino_info.grid(row=6, column=0, columnspan=3, padx=18, pady=(0,8), sticky="ew")

        self.criar_label(parent, "Comando Copilot", 7)
        frame_cmd = ctk.CTkFrame(parent, fg_color="transparent"); frame_cmd.grid(row=7, column=1, columnspan=2, padx=(4,18), pady=7, sticky="ew"); frame_cmd.grid_columnconfigure(0, weight=1)
        entry_cmd = ctk.CTkEntry(frame_cmd, textvariable=self.var_comando_rapido, height=36,
                                 placeholder_text="Ex.: Folha para Joana, não lidos, últimos 3 dias, analisar agora",
                                 fg_color=CORES["branco"], border_color=CORES["borda"], text_color=CORES["texto"])
        entry_cmd.grid(row=0, column=0, padx=(0,8), sticky="ew"); entry_cmd.bind("<Return>", lambda _e: self.aplicar_comando_rapido())
        ctk.CTkButton(frame_cmd, text="Aplicar", command=self.aplicar_comando_rapido, width=90, height=36,
                      fg_color=CORES["azul_apoio"], hover_color=CORES["navy"], text_color=CORES["branco"]).grid(row=0,column=1)

        self.criar_label(parent, "Modelo", 8)
        ctk.CTkOptionMenu(parent, variable=self.var_ia_modelo_rotulo, values=list(IA_MODELOS.keys()), fg_color=CORES["azul"],
                          button_color=CORES["azul_apoio"], button_hover_color=CORES["navy"], text_color=CORES["branco"]).grid(
                          row=8, column=1, columnspan=2, padx=(4,18), pady=7, sticky="ew")
        self.criar_label(parent, "Chave da API", 9)
        ctk.CTkEntry(parent, textvariable=self.var_ia_api_key, placeholder_text="ANTHROPIC_API_KEY (preferir variável de ambiente)", height=34,
                     show="•", fg_color=CORES["branco"], border_color=CORES["borda"], text_color=CORES["texto"]).grid(row=9,column=1,columnspan=2,padx=(4,18),pady=7,sticky="ew")
        self.criar_label(parent, "Workspace ID", 10)
        ctk.CTkEntry(parent, textvariable=self.var_ia_workspace_id, placeholder_text="wrkspc_... (se exigido)", height=34,
                     fg_color=CORES["branco"], border_color=CORES["borda"], text_color=CORES["texto"]).grid(row=10,column=1,columnspan=2,padx=(4,18),pady=7,sticky="ew")
        self.criar_label(parent, "E-mail do resumo", 11)
        ctk.CTkEntry(parent, textvariable=self.var_resumo_destinatario_email,
                     placeholder_text="Opcional — deixe vazio para preencher no Outlook", height=34,
                     fg_color=CORES["branco"], border_color=CORES["borda"], text_color=CORES["texto"]).grid(
                     row=11,column=1,columnspan=2,padx=(4,18),pady=7,sticky="ew")
        ctk.CTkLabel(parent, text=("Modo operacional define COMO o agente pensa; Destinatário/Tom define COMO ele escreve. "
                                  "O comando é independente da interface e pode receber no futuro texto transcrito por voz. "
                                  "A varredura nunca cria respostas automaticamente. O resumo consolidado também é salvo apenas como rascunho."), font=("Segoe UI",12), text_color=CORES["texto_sec"],
                     justify="left", wraplength=980, anchor="w").grid(row=12,column=0,columnspan=3,padx=18,pady=(6,16),sticky="ew")

    def criar_view_copilot(self, parent):
        topo = ctk.CTkFrame(parent, fg_color="transparent"); topo.grid(row=0,column=0,sticky="ew",pady=(0,10)); topo.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(topo, text="Mesa Copilot — prioridades e casos", font=("Segoe UI",20,"bold"), text_color=CORES["navy"], anchor="w").grid(row=0,column=0,sticky="w")
        ctk.CTkButton(topo, text="Varrer caixa", command=self.iniciar_extracao, height=38, fg_color=CORES["sucesso"], hover_color="#256628").grid(row=0,column=1,padx=(8,0))

        self.frame_metricas = ctk.CTkFrame(parent, fg_color=CORES["card"], border_color=CORES["borda"], border_width=1, corner_radius=12)
        self.frame_metricas.grid(row=1,column=0,sticky="ew",pady=(0,10));
        self.metric_labels = {}
        nomes = ["Crítico","Precisa de mim","Delegável","Aguardando terceiro","Informativo"]
        for i,nome in enumerate(nomes):
            self.frame_metricas.grid_columnconfigure(i, weight=1)
            box = ctk.CTkFrame(self.frame_metricas, fg_color="transparent"); box.grid(row=0,column=i,padx=8,pady=10,sticky="ew")
            ctk.CTkLabel(box,text=nome,font=("Segoe UI",11),text_color=CORES["texto_sec"]).pack()
            lbl = ctk.CTkLabel(box,text="0",font=("Segoe UI",22,"bold"),text_color=CORES["navy"]); lbl.pack(); self.metric_labels[nome]=lbl
        self.lbl_brief_copilot = ctk.CTkLabel(self.frame_metricas, text="Nenhuma varredura realizada nesta sessão.",
                                              font=("Segoe UI",12), text_color=CORES["texto"], justify="left",
                                              anchor="w", wraplength=1050)
        self.lbl_brief_copilot.grid(row=1,column=0,columnspan=5,padx=14,pady=(0,12),sticky="ew")

        corpo = ctk.CTkFrame(parent, fg_color=CORES["card"], border_color=CORES["borda"], border_width=1, corner_radius=12)
        corpo.grid(row=2,column=0,sticky="nsew"); corpo.grid_columnconfigure(0, weight=1); corpo.grid_rowconfigure(1, weight=1)
        filtros = ctk.CTkFrame(corpo, fg_color="transparent"); filtros.grid(row=0,column=0,sticky="ew",padx=12,pady=10); filtros.grid_columnconfigure(2,weight=1)
        ctk.CTkLabel(filtros,text="Fila",text_color=CORES["texto"]).grid(row=0,column=0,padx=(0,6))
        ctk.CTkOptionMenu(filtros, variable=self.var_filtro_fila, values=["Todas","Crítico","Precisa de mim","Delegável","Aguardando terceiro","Informativo"],
                          command=lambda _v:self.atualizar_mesa_copilot(), width=170).grid(row=0,column=1,padx=(0,10))
        busca = ctk.CTkEntry(filtros,textvariable=self.var_busca_copilot,placeholder_text="Buscar remetente, assunto, categoria, caso...",height=34)
        busca.grid(row=0,column=2,sticky="ew"); busca.bind("<KeyRelease>", lambda _e:self.atualizar_mesa_copilot())

        tree_frame = ctk.CTkFrame(corpo, fg_color="transparent"); tree_frame.grid(row=1,column=0,sticky="nsew",padx=12,pady=(0,8)); tree_frame.grid_columnconfigure(0,weight=1); tree_frame.grid_rowconfigure(0,weight=1)
        cols = ("fila","score","case","data","remetente","assunto","categoria","responsavel")
        self.tree_copilot = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="extended")
        titulos = {"fila":"Fila","score":"Score","case":"Caso","data":"Data","remetente":"Remetente","assunto":"Assunto","categoria":"Categoria","responsavel":"Responsável"}
        larg = {"fila":135,"score":60,"case":105,"data":85,"remetente":180,"assunto":300,"categoria":170,"responsavel":150}
        for c in cols: self.tree_copilot.heading(c,text=titulos[c]); self.tree_copilot.column(c,width=larg[c],minwidth=50,stretch=(c in {"assunto","remetente"}))
        self.tree_copilot.grid(row=0,column=0,sticky="nsew"); sb=ttk.Scrollbar(tree_frame,orient="vertical",command=self.tree_copilot.yview); sb.grid(row=0,column=1,sticky="ns"); self.tree_copilot.configure(yscrollcommand=sb.set)
        self.tree_copilot.bind("<Double-1>", lambda _e:self.ver_detalhes_selecionado())

        acoes = ctk.CTkFrame(corpo, fg_color="transparent"); acoes.grid(row=2,column=0,sticky="ew",padx=12,pady=(0,6))
        ctk.CTkButton(acoes,text="Ver detalhes / editar resposta",command=self.ver_detalhes_selecionado,height=36,fg_color=CORES["azul_apoio"]).pack(side="left",padx=(0,8))
        ctk.CTkButton(acoes,text="Abrir e-mail no Outlook",command=self.abrir_email_selecionado,height=36,fg_color=CORES["auxiliar"],hover_color=CORES["auxiliar_hover"],text_color=CORES["texto"]).pack(side="left",padx=(0,8))
        ctk.CTkButton(acoes,text="Criar rascunho(s) selecionado(s)",command=self.criar_rascunhos_selecionados,height=36,fg_color=CORES["sucesso"],hover_color="#256628").pack(side="right")

        acoes_resumo = ctk.CTkFrame(corpo, fg_color="transparent"); acoes_resumo.grid(row=3,column=0,sticky="ew",padx=12,pady=(0,12))
        ctk.CTkLabel(acoes_resumo,text="Rascunho único:",text_color=CORES["texto_sec"]).pack(side="left",padx=(0,8))
        ctk.CTkButton(acoes_resumo,text="Resumo dos selecionados",command=self.preparar_resumo_selecionados,height=36,
                      fg_color=CORES["azul_apoio"],hover_color=CORES["navy"]).pack(side="left",padx=(0,8))
        ctk.CTkButton(acoes_resumo,text="Resumo Folha + Urgências",command=self.preparar_resumo_folha_urgencias,height=36,
                      fg_color=CORES["vinho"],hover_color=CORES["telha"],text_color=CORES["branco"]).pack(side="right")

    def atualizar_mesa_copilot(self):
        if not hasattr(self, "tree_copilot"): return
        for item in self.tree_copilot.get_children(): self.tree_copilot.delete(item)
        filtro = self.var_filtro_fila.get(); busca = normalizar_busca(self.var_busca_copilot.get())
        cont = Counter(r.get("Fila Copilot","Informativo") for r in self.registros_sessao)
        for nome,lbl in self.metric_labels.items(): lbl.configure(text=str(cont.get(nome,0)))
        if self.registros_sessao:
            top = ordenar_registros_copilot(self.registros_sessao)[:5]
            linhas = []
            for posicao, reg in enumerate(top, 1):
                assunto = str(reg.get("Assunto", ""))[:80]
                passo = str(reg.get("Próximo Passo", ""))[:110]
                linhas.append(f"{posicao}. [{reg.get('Fila Copilot','')}] {assunto}" + (f" — {passo}" if passo else ""))
            casos = len({r.get("Case ID") for r in self.registros_sessao if r.get("Case ID")})
            texto_brief = f"Leitura executiva: {len(self.registros_sessao)} e-mails em {casos} caso(s). Prioridades no topo:\n" + "\n".join(linhas)
        else:
            texto_brief = "Nenhuma varredura realizada nesta sessão."
        try: self.lbl_brief_copilot.configure(text=texto_brief)
        except Exception: pass
        ordem_fila = {"Crítico":0, "Precisa de mim":1, "Delegável":2, "Aguardando terceiro":3, "Informativo":4}
        pares = list(enumerate(self.registros_sessao))
        pares.sort(key=lambda ir: (ordem_fila.get(ir[1].get("Fila Copilot"),9), -int(ir[1].get("Score Prioridade",0) or 0)))
        for idx,r in pares:
            if filtro != "Todas" and r.get("Fila Copilot") != filtro: continue
            alvo = normalizar_busca(" ".join(str(r.get(k,"")) for k in ["Remetente Nome","Remetente E-mail","Assunto","Categoria Sugerida","Case ID","Resumo Executivo"]))
            if busca and busca not in alvo: continue
            iid = str(idx)
            self.tree_copilot.insert("", "end", iid=iid, values=(r.get("Fila Copilot"), r.get("Score Prioridade"), r.get("Case ID"), r.get("Data Recebimento"),
                                                                    r.get("Remetente Nome"), r.get("Assunto"), r.get("Categoria Sugerida"), r.get("Responsável Sugerido")))

    def _indices_selecionados(self):
        return [int(i) for i in self.tree_copilot.selection() if str(i).isdigit()]

    def ver_detalhes_selecionado(self):
        inds = self._indices_selecionados()
        if not inds: messagebox.showinfo("Copilot", "Selecione um e-mail na mesa."); return
        idx = inds[0]; r = self.registros_sessao[idx]
        win = ctk.CTkToplevel(self); win.title(f"Copilot — {r.get('Case ID','')}"); win.geometry("940x860")
        win.grid_columnconfigure(0, weight=1); win.grid_rowconfigure(4, weight=1)

        cab = (f"De: {r.get('Remetente Nome')} <{r.get('Remetente E-mail')}>\nAssunto: {r.get('Assunto')}\n"
               f"Caso: {r.get('Case ID')} | Score {r.get('Score Prioridade')} | e-mails na sessão: {r.get('E-mails no Caso (sessão)')}")
        ctk.CTkLabel(win, text=cab, justify="left", anchor="w", font=("Segoe UI",13,"bold"),
                     text_color=CORES["navy"], wraplength=890).grid(row=0, column=0, padx=18, pady=(18,8), sticky="ew")

        info = (f"RESUMO: {r.get('Resumo Executivo','') or r.get('Análise ChatGPT','')}\n\n"
                f"PENDÊNCIA/DECISÃO: {r.get('Pendência / Decisão','')}\nIMPACTO: {r.get('Impacto','')}")
        ctk.CTkLabel(win, text=info, justify="left", anchor="w", text_color=CORES["texto"],
                     wraplength=890).grid(row=1, column=0, padx=18, pady=(0,8), sticky="ew")

        # ---- Campos de ação (dropdowns editáveis) ----
        acao = ctk.CTkFrame(win, fg_color=CORES["card"], border_color=CORES["borda"], border_width=1, corner_radius=12)
        acao.grid(row=2, column=0, padx=18, pady=(0,10), sticky="ew")
        acao.grid_columnconfigure(1, weight=1); acao.grid_columnconfigure(3, weight=1)
        ctk.CTkLabel(acao, text="Campos de ação", font=("Segoe UI",13,"bold"), text_color=CORES["navy"],
                     anchor="w").grid(row=0, column=0, columnspan=4, padx=12, pady=(10,6), sticky="w")

        def _combo(rr, cc, rotulo, valores, atual):
            ctk.CTkLabel(acao, text=rotulo, text_color=CORES["texto"], anchor="w").grid(
                row=rr, column=cc, padx=(12,6), pady=6, sticky="w")
            var = ctk.StringVar(value=str(atual or ""))
            ctk.CTkComboBox(acao, values=valores, variable=var, state="normal",
                            fg_color=CORES["branco"], border_color=CORES["borda"], text_color=CORES["texto"],
                            button_color=CORES["azul"], button_hover_color=CORES["azul_apoio"],
                            dropdown_fg_color=CORES["branco"], dropdown_hover_color=CORES["azul_claro"]).grid(
                            row=rr, column=cc+1, padx=(0,12), pady=6, sticky="ew")
            return var

        v_fila   = _combo(1, 0, "Fila",        OPCOES_FILA,        r.get("Fila Copilot"))
        v_urg    = _combo(1, 2, "Urgência",    OPCOES_URGENCIA,    r.get("Urgência Sugerida"))
        v_cat    = _combo(2, 0, "Categoria",   OPCOES_CATEGORIA,   r.get("Categoria Sugerida"))
        v_resp   = _combo(2, 2, "Responsável", OPCOES_RESPONSAVEL, r.get("Responsável Sugerido"))
        v_stat   = _combo(3, 0, "Status",      OPCOES_STATUS,      r.get("Status Atendimento"))
        v_prazo  = _combo(3, 2, "Prazo",       OPCOES_PRAZO,       r.get("Prazo / Timing"))
        v_estado = _combo(4, 0, "Estado do caso", OPCOES_ESTADO,   r.get("Estado do Caso"))
        v_deleg  = _combo(4, 2, "Pode delegar",   OPCOES_SIM_NAO,  r.get("Pode Delegar"))

        ctk.CTkLabel(acao, text="Próximo passo", text_color=CORES["texto"], anchor="w").grid(
            row=5, column=0, padx=(12,6), pady=(6,12), sticky="w")
        v_passo = ctk.StringVar(value=str(r.get("Próximo Passo","") or r.get("Ação Recomendada","") or ""))
        ctk.CTkComboBox(acao, values=OPCOES_PROXIMO_PASSO, variable=v_passo, state="normal",
                        fg_color=CORES["branco"], border_color=CORES["borda"], text_color=CORES["texto"],
                        button_color=CORES["azul"], button_hover_color=CORES["azul_apoio"],
                        dropdown_fg_color=CORES["branco"], dropdown_hover_color=CORES["azul_claro"]).grid(
                        row=5, column=1, columnspan=3, padx=(0,12), pady=(6,12), sticky="ew")

        # ---- Resposta sugerida (texto livre) ----
        ctk.CTkLabel(win, text="Resposta sugerida (edite se quiser)", text_color=CORES["texto"],
                     anchor="w").grid(row=3, column=0, padx=18, pady=(4,2), sticky="ew")
        txt = ctk.CTkTextbox(win, wrap="word"); txt.grid(row=4, column=0, padx=18, pady=(0,8), sticky="nsew")
        txt.insert("1.0", r.get("Resposta Sugerida (IA)", ""))

        def salvar():
            reg = self.registros_sessao[idx]
            reg["Fila Copilot"]          = v_fila.get().strip() or reg.get("Fila Copilot", "")
            reg["Urgência Sugerida"]     = v_urg.get().strip()
            reg["Categoria Sugerida"]    = v_cat.get().strip()
            reg["Responsável Sugerido"]  = v_resp.get().strip()
            reg["Status Atendimento"]    = v_stat.get().strip()
            reg["Prazo / Timing"]        = v_prazo.get().strip()
            reg["Estado do Caso"]        = v_estado.get().strip()
            reg["Pode Delegar"]          = v_deleg.get().strip()
            reg["Próximo Passo"]         = v_passo.get().strip()
            reg["Resposta Sugerida (IA)"] = txt.get("1.0", "end").strip()
            self.atualizar_mesa_copilot()
            messagebox.showinfo("Copilot", "Caso atualizado na sessão. A Mesa e o próximo Excel/refletem as mudanças.", parent=win)

        botoes = ctk.CTkFrame(win, fg_color="transparent"); botoes.grid(row=5, column=0, padx=18, pady=(0,18), sticky="ew")
        ctk.CTkButton(botoes, text="Salvar edição", command=salvar, fg_color=CORES["azul"], hover_color=CORES["azul_apoio"]).pack(side="left")
        ctk.CTkButton(botoes, text="Abrir no Outlook", command=lambda:self._abrir_registro(r), fg_color=CORES["azul_apoio"], hover_color=CORES["navy"]).pack(side="left", padx=8)

    def _abrir_registro(self, r):
        try: abrir_item_outlook(r.get("EntryID",""), r.get("StoreID",""))
        except Exception as e: messagebox.showerror("Outlook",f"Não foi possível abrir o e-mail.\n\n{e}")

    def abrir_email_selecionado(self):
        inds=self._indices_selecionados()
        if not inds: messagebox.showinfo("Copilot","Selecione um e-mail."); return
        self._abrir_registro(self.registros_sessao[inds[0]])

    def criar_rascunhos_selecionados(self):
        inds = self._indices_selecionados()
        if not inds: messagebox.showinfo("Copilot","Selecione um ou mais e-mails."); return
        if not self.var_criar_rascunho.get():
            messagebox.showinfo("Copilot","Habilite 'Permitir criação de rascunhos após revisão' em Agente e Comunicação."); return
        sem_resposta=[i for i in inds if not self.registros_sessao[i].get("Resposta Sugerida (IA)","").strip()]
        if sem_resposta:
            messagebox.showwarning("Copilot",f"{len(sem_resposta)} item(ns) não possuem resposta sugerida. Ative a IA e varra novamente ou edite a resposta nos detalhes."); return
        if not messagebox.askyesno("Criar rascunhos", f"Criar {len(inds)} rascunho(s) no Outlook?\n\nNada será enviado automaticamente."):
            return
        assinatura = self.obter_config_tela().get("ia_assinatura", "")
        def worker():
            pythoncom.CoInitialize(); ok=0; falhas=[]
            try:
                for i in inds:
                    r=self.registros_sessao[i]
                    try:
                        mail=obter_item_outlook_por_ids(r.get("EntryID",""),r.get("StoreID",""))
                        sucesso=criar_rascunho_reply(mail,r.get("Resposta Sugerida (IA)",""),assinatura)
                        r["Rascunho Outlook Criado"]="Sim" if sucesso else "Falha"
                        ok += 1 if sucesso else 0
                    except Exception as e: r["Rascunho Outlook Criado"]="Falha"; falhas.append(str(e))
            finally:
                pythoncom.CoUninitialize()
            if self.caminho_excel_gerado:
                try:
                    self.caminho_excel_gerado = gerar_excel(self.registros_sessao, self.caminho_excel_gerado)
                except Exception as e:
                    falhas.append(f"Excel não atualizado: {e}")
            self.after(0,self.atualizar_mesa_copilot); self.after(0,messagebox.showinfo,"Copilot",f"Rascunhos criados: {ok}. Falhas: {len(falhas)}.")
        threading.Thread(target=worker,daemon=True).start()

    def _config_ia_resumo(self):
        c = self.obter_config_tela()
        return {
            "ia_ativa": c.get("ia_ativa", False),
            "ia_modelo": c.get("ia_modelo", IA_MODELO_PADRAO),
            "ia_api_key": c.get("ia_api_key", ""),
            "ia_workspace_id": c.get("ia_workspace_id", ""),
        }

    def preparar_resumo_folha_urgencias(self):
        if not self.registros_sessao:
            messagebox.showinfo("Resumo consolidado", "Faça uma varredura da caixa primeiro.")
            return
        registros = selecionar_casos_folha_urgencias(self.registros_sessao)
        if not registros:
            messagebox.showinfo("Resumo consolidado", "Nenhum caso de impacto em Folha ou alta urgência foi identificado nesta varredura.")
            return
        self._gerar_preview_resumo_async(registros, origem="automatico")

    def preparar_resumo_selecionados(self):
        inds = self._indices_selecionados()
        if not inds:
            messagebox.showinfo("Resumo consolidado", "Selecione um ou mais e-mails/casos na Mesa Copilot.")
            return
        registros = [self.registros_sessao[i] for i in inds]
        self._gerar_preview_resumo_async(registros, origem="selecionados")

    def _gerar_preview_resumo_async(self, registros, origem="automatico"):
        destino = self.var_destino_comunicacao.get()
        config_ia = self._config_ia_resumo()
        self.log(f"Preparando rascunho único: {len(registros)} registro(s), origem={origem}, destino/tom={destino}.")

        def worker():
            assunto, corpo = gerar_email_resumo_com_ia(
                registros=registros,
                config_ia=config_ia,
                destino_comunicacao=destino,
                origem=origem,
                log_callback=lambda msg: self.after(0, self.log, msg),
            )
            casos = consolidar_registros_por_caso(registros)
            self.after(0, self.abrir_preview_resumo, assunto, corpo, len(casos), origem)

        threading.Thread(target=worker, daemon=True).start()

    def abrir_preview_resumo(self, assunto, corpo, qtd_casos, origem):
        win = ctk.CTkToplevel(self)
        win.title("Copilot — rascunho consolidado")
        win.geometry("980x780")
        win.minsize(760, 620)
        win.grid_columnconfigure(0, weight=1)
        win.grid_rowconfigure(4, weight=1)

        titulo = "Resumo automático de Folha + Urgências" if origem == "automatico" else "Resumo dos casos selecionados"
        ctk.CTkLabel(win, text=titulo, font=("Segoe UI", 19, "bold"), text_color=CORES["navy"], anchor="w").grid(
            row=0, column=0, padx=18, pady=(18, 4), sticky="ew"
        )
        ctk.CTkLabel(
            win,
            text=f"{qtd_casos} caso(s) consolidado(s). Revise destinatário, assunto e texto antes de salvar em Rascunhos.",
            font=("Segoe UI", 12), text_color=CORES["texto_sec"], anchor="w"
        ).grid(row=1, column=0, padx=18, pady=(0, 10), sticky="ew")

        meta = ctk.CTkFrame(win, fg_color="transparent")
        meta.grid(row=2, column=0, padx=18, pady=4, sticky="ew")
        meta.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(meta, text="Para", text_color=CORES["texto"]).grid(row=0, column=0, padx=(0, 8), pady=5, sticky="w")
        var_para = ctk.StringVar(value=self.var_resumo_destinatario_email.get().strip())
        ctk.CTkEntry(meta, textvariable=var_para, placeholder_text="Opcional — pode preencher depois no Outlook").grid(
            row=0, column=1, pady=5, sticky="ew"
        )
        ctk.CTkLabel(meta, text="Assunto", text_color=CORES["texto"]).grid(row=1, column=0, padx=(0, 8), pady=5, sticky="w")
        var_assunto = ctk.StringVar(value=assunto)
        ctk.CTkEntry(meta, textvariable=var_assunto).grid(row=1, column=1, pady=5, sticky="ew")

        ctk.CTkLabel(win, text="Corpo do resumo", text_color=CORES["texto"], anchor="w").grid(
            row=3, column=0, padx=18, pady=(8, 2), sticky="ew"
        )
        txt = ctk.CTkTextbox(win, wrap="word", font=("Segoe UI", 12))
        txt.grid(row=4, column=0, padx=18, pady=(0, 10), sticky="nsew")
        txt.insert("1.0", corpo)

        botoes = ctk.CTkFrame(win, fg_color="transparent")
        botoes.grid(row=5, column=0, padx=18, pady=(0, 18), sticky="ew")

        def criar():
            if not self.var_criar_rascunho.get():
                messagebox.showinfo(
                    "Resumo consolidado",
                    "Habilite 'Permitir criação de rascunhos após revisão' em Agente e Comunicação.",
                    parent=win,
                )
                return
            para = var_para.get().strip()
            assunto_final = var_assunto.get().strip()
            corpo_final = txt.get("1.0", "end").strip()
            if not assunto_final or not corpo_final:
                messagebox.showwarning("Resumo consolidado", "Assunto e corpo não podem ficar vazios.", parent=win)
                return
            self.var_resumo_destinatario_email.set(para)
            assinatura = self.obter_config_tela().get("ia_assinatura", "")

            def salvar_worker():
                pythoncom.CoInitialize()
                try:
                    ok = criar_rascunho_novo_outlook(
                        destinatario=para,
                        assunto=assunto_final,
                        corpo=corpo_final,
                        assinatura=assinatura,
                        abrir=True,
                        log_callback=lambda msg: self.after(0, self.log, msg),
                    )
                finally:
                    pythoncom.CoUninitialize()
                if ok:
                    self.after(0, self.log, f"Rascunho consolidado criado: {assunto_final}")
                    self.after(0, messagebox.showinfo, "Resumo consolidado", "Um único rascunho foi salvo e aberto no Outlook. Nada foi enviado.")
                else:
                    self.after(0, messagebox.showerror, "Resumo consolidado", "Não foi possível criar o rascunho no Outlook.")

            threading.Thread(target=salvar_worker, daemon=True).start()

        ctk.CTkButton(botoes, text="Criar único rascunho no Outlook", command=criar, height=40,
                      fg_color=CORES["sucesso"], hover_color="#256628").pack(side="right")
        ctk.CTkButton(botoes, text="Fechar", command=win.destroy, height=40,
                      fg_color=CORES["auxiliar"], hover_color=CORES["auxiliar_hover"], text_color=CORES["texto"]).pack(side="right", padx=(0, 8))

    def criar_titulo_card(self, parent, texto):
        label = ctk.CTkLabel(
            parent,
            text=texto,
            font=("Segoe UI", 17, "bold"),
            text_color=CORES["navy"],
            anchor="w",
        )
        label.grid(row=0, column=0, columnspan=3, padx=18, pady=(16, 10), sticky="ew")

    def criar_label(self, parent, texto, row):
        label = ctk.CTkLabel(
            parent,
            text=texto,
            font=("Segoe UI", 13),
            text_color=CORES["texto"],
            anchor="w",
        )
        label.grid(row=row, column=0, padx=(18, 8), pady=7, sticky="w")
        return label

    def criar_entry(self, parent, variable, row, placeholder=""):
        entry = ctk.CTkEntry(
            parent,
            textvariable=variable,
            placeholder_text=placeholder,
            height=34,
            fg_color=CORES["branco"],
            border_color=CORES["borda"],
            text_color=CORES["texto"],
        )
        entry.grid(row=row, column=1, padx=(4, 18), pady=7, sticky="ew")
        return entry

    def criar_card_origem(self, parent):
        self.criar_titulo_card(parent, "Origem dos e-mails")

        radio_principal = ctk.CTkRadioButton(
            parent,
            text="Caixa principal",
            variable=self.var_modo_caixa,
            value="principal",
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        radio_principal.grid(row=1, column=0, padx=18, pady=5, sticky="w")

        radio_compartilhada = ctk.CTkRadioButton(
            parent,
            text="Caixa compartilhada",
            variable=self.var_modo_caixa,
            value="compartilhada",
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        radio_compartilhada.grid(row=1, column=1, padx=8, pady=5, sticky="w")

        self.criar_label(parent, "Caixa compartilhada", 2)
        self.criar_entry(
            parent,
            self.var_caixa_compartilhada,
            2,
            "Exemplo: genteegestao@empresa.com.br",
        )

        self.criar_label(parent, "Pasta Outlook", 3)
        self.criar_entry(
            parent,
            self.var_pasta_outlook,
            3,
            "Caixa de Entrada ou nome de subpasta",
        )

        frame_sel = ctk.CTkFrame(parent, fg_color="transparent")
        frame_sel.grid(row=4, column=0, columnspan=3, padx=18, pady=(2, 2), sticky="ew")

        btn_escolher = ctk.CTkButton(
            frame_sel,
            text="Escolher pasta do Outlook...",
            command=self.escolher_pasta_outlook,
            height=34,
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["branco"],
            font=("Segoe UI", 13, "bold"),
        )
        btn_escolher.grid(row=0, column=0, padx=(0, 8), sticky="w")

        btn_limpar_sel = ctk.CTkButton(
            frame_sel,
            text="Limpar seleção",
            command=self.limpar_pasta_escolhida,
            height=34,
            width=120,
            fg_color=CORES["borda"],
            hover_color="#C4C9D1",
            text_color=CORES["texto"],
            font=("Segoe UI", 13),
        )
        btn_limpar_sel.grid(row=0, column=1, sticky="w")

        self.lbl_pasta_escolhida = ctk.CTkLabel(
            parent,
            text=self._texto_pasta_escolhida(),
            font=("Segoe UI", 12),
            text_color=CORES["texto_sec"],
            justify="left",
            wraplength=480,
            anchor="w",
        )
        self.lbl_pasta_escolhida.grid(row=5, column=0, columnspan=3, padx=18, pady=(0, 6), sticky="ew")

        self.criar_label(parent, "Quantidade máxima", 6)
        self.criar_entry(parent, self.var_quantidade, 6, "200")

        self.criar_label(parent, "Últimos dias", 7)
        self.criar_entry(parent, self.var_ultimos_dias, 7, "30. Use 0 para não filtrar")

        self.criar_label(parent, "De (data)", 8)
        self.criar_entry(parent, self.var_data_de, 8, "dd/mm/aaaa (opcional)")

        self.criar_label(parent, "Até (data)", 9)
        self.criar_entry(parent, self.var_data_ate, 9, "dd/mm/aaaa (opcional)")

        chk = ctk.CTkCheckBox(
            parent,
            text="Ler somente e-mails não lidos",
            variable=self.var_somente_nao_lidos,
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["texto"],
        )
        chk.grid(row=10, column=0, columnspan=2, padx=18, pady=(8, 6), sticky="w")

        ctk.CTkButton(
            parent,
            text="Listar e escolher e-mails (sem IA)",
            command=self.iniciar_listagem,
            height=40,
            fg_color=CORES["azul"],
            hover_color=CORES["azul_apoio"],
            text_color=CORES["branco"],
            font=("Segoe UI", 13, "bold"),
        ).grid(row=11, column=0, columnspan=3, padx=18, pady=(4, 4), sticky="ew")

        ctk.CTkLabel(
            parent,
            text=("Preencha De/Até para ler só um intervalo. Se preenchidos, o intervalo tem "
                  "prioridade sobre 'Últimos dias'. Use 'Listar e escolher' para marcar, com filtro, "
                  "apenas os e-mails que a IA vai analisar — mais rápido e mais barato."),
            font=("Segoe UI", 12), text_color=CORES["texto_sec"],
            justify="left", wraplength=480, anchor="w",
        ).grid(row=12, column=0, columnspan=3, padx=18, pady=(0, 14), sticky="ew")

    def _texto_pasta_escolhida(self):
        display = self.var_pasta_display.get().strip()
        if display:
            return f"Pasta selecionada: {display}  (tem prioridade sobre os campos acima)"
        return "Nenhuma pasta selecionada pelo seletor — usando os campos de origem acima."

    def escolher_pasta_outlook(self):
        try:
            try:
                pythoncom.CoInitialize()
            except Exception:
                pass

            outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
            folder = outlook.PickFolder()

            if folder is None:
                return  # usuário cancelou

            entry_id = getattr(folder, "EntryID", "") or ""
            store_id = getattr(folder, "StoreID", "") or ""
            try:
                display = folder.FolderPath
            except Exception:
                display = getattr(folder, "Name", "Pasta selecionada")

            self.var_pasta_entry_id.set(entry_id)
            self.var_pasta_store_id.set(store_id)
            self.var_pasta_display.set(display)
            self.lbl_pasta_escolhida.configure(text=self._texto_pasta_escolhida())
            self.log(f"Pasta escolhida no seletor: {display}")

        except Exception as e:
            messagebox.showerror(
                "Escolher pasta",
                f"Não foi possível abrir o seletor de pastas do Outlook.\n\n{e}",
            )

    def limpar_pasta_escolhida(self):
        self.var_pasta_entry_id.set("")
        self.var_pasta_store_id.set("")
        self.var_pasta_display.set("")
        self.lbl_pasta_escolhida.configure(text=self._texto_pasta_escolhida())
        self.log("Seleção de pasta limpa. Usando os campos de origem.")

    def criar_card_saida(self, parent):
        self.criar_titulo_card(parent, "Saída em Excel")

        self.criar_label(parent, "Pasta de saída", 1)

        frame_saida = ctk.CTkFrame(parent, fg_color="transparent")
        frame_saida.grid(row=1, column=1, padx=(4, 18), pady=7, sticky="ew")
        frame_saida.grid_columnconfigure(0, weight=1)

        entry_saida = ctk.CTkEntry(
            frame_saida,
            textvariable=self.var_pasta_saida,
            height=34,
            fg_color=CORES["branco"],
            border_color=CORES["borda"],
            text_color=CORES["texto"],
        )
        entry_saida.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        btn_pasta = ctk.CTkButton(
            frame_saida,
            text="Selecionar",
            command=self.selecionar_pasta_saida,
            width=95,
            height=34,
            fg_color=CORES["azul_apoio"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
        )
        btn_pasta.grid(row=0, column=1)

        self.criar_label(parent, "Nome do Excel", 2)
        self.criar_entry(parent, self.var_nome_excel, 2, DEFAULT_EXCEL_NAME)

        info = ctk.CTkLabel(
            parent,
            text="O Excel será gerado com BASE_EMAILS, COPILOT_PRIORIDADES, CASOS, GLOSSARIO e RESUMO. Na Mesa Copilot você também pode gerar um único rascunho consolidando impactos em Folha e urgências.",
            font=("Segoe UI", 13),
            text_color=CORES["texto_sec"],
            justify="left",
            wraplength=480,
            anchor="w",
        )
        info.grid(row=3, column=0, columnspan=2, padx=18, pady=(8, 12), sticky="ew")

        botoes = ctk.CTkFrame(parent, fg_color="transparent")
        botoes.grid(row=4, column=0, columnspan=2, padx=18, pady=(4, 16), sticky="ew")
        botoes.grid_columnconfigure(0, weight=1)
        botoes.grid_columnconfigure(1, weight=1)

        self.btn_extrair = ctk.CTkButton(
            botoes,
            text="Extrair e-mails para Excel",
            command=self.iniciar_extracao,
            height=42,
            fg_color=CORES["azul"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
            font=("Segoe UI", 14, "bold"),
        )
        self.btn_extrair.grid(row=0, column=0, padx=(0, 8), sticky="ew")

        self.btn_abrir = ctk.CTkButton(
            botoes,
            text="Abrir pasta de saída",
            command=self.abrir_pasta_saida,
            height=42,
            fg_color=CORES["azul_apoio"],
            hover_color=CORES["navy"],
            text_color=CORES["branco"],
            font=("Segoe UI", 14, "bold"),
        )
        self.btn_abrir.grid(row=0, column=1, padx=(8, 0), sticky="ew")

    def criar_card_agendamento(self, parent):
        self.criar_titulo_card(parent, "Agendamento e segundo plano")

        chk_ag = ctk.CTkCheckBox(parent, text="Executar automaticamente todo dia no horário", variable=self.var_agendamento_ativo,
                                 fg_color=CORES["azul"], hover_color=CORES["azul_apoio"], text_color=CORES["texto"])
        chk_ag.grid(row=1, column=0, columnspan=3, padx=18, pady=(4, 4), sticky="w")

        self.criar_label(parent, "Horário (HH:MM)", 2)
        ctk.CTkEntry(parent, textvariable=self.var_agendamento_hora, placeholder_text="05:00", height=34, width=120,
                     fg_color=CORES["branco"], border_color=CORES["borda"], text_color=CORES["texto"]).grid(
                     row=2, column=1, padx=(4, 18), pady=7, sticky="w")

        chk_al = ctk.CTkCheckBox(parent, text="Alertar no Outlook quando houver algo que pode ser urgente", variable=self.var_alerta_urgente,
                                 fg_color=CORES["azul"], hover_color=CORES["azul_apoio"], text_color=CORES["texto"])
        chk_al.grid(row=3, column=0, columnspan=3, padx=18, pady=(4, 4), sticky="w")

        chk_key = ctk.CTkCheckBox(parent, text="Salvar a chave da API no aplicativo (necessário para a execução automática)", variable=self.var_salvar_chave,
                                  fg_color=CORES["azul"], hover_color=CORES["azul_apoio"], text_color=CORES["texto"])
        chk_key.grid(row=4, column=0, columnspan=3, padx=18, pady=(4, 4), sticky="w")

        chk_win = ctk.CTkCheckBox(parent, text="Iniciar com o Windows e ler ao ligar o computador", variable=self.var_iniciar_com_windows,
                                  command=self._toggle_startup,
                                  fg_color=CORES["azul"], hover_color=CORES["azul_apoio"], text_color=CORES["texto"])
        chk_win.grid(row=5, column=0, columnspan=3, padx=18, pady=(4, 8), sticky="w")

        botoes = ctk.CTkFrame(parent, fg_color="transparent")
        botoes.grid(row=6, column=0, columnspan=3, padx=18, pady=(0, 8), sticky="w")
        ctk.CTkButton(botoes, text="Minimizar para a bandeja (segundo plano)", command=self.minimizar_para_bandeja, height=38,
                      fg_color=CORES["azul"], hover_color=CORES["azul_apoio"], text_color=CORES["branco"]).pack(side="left", padx=(0, 8))
        ctk.CTkButton(botoes, text="Rodar agora (não lidos)", command=self.executar_agendado, height=38,
                      fg_color=CORES["azul_apoio"], hover_color=CORES["navy"], text_color=CORES["branco"]).pack(side="left")

        ctk.CTkLabel(parent, text=(
            "Com 'Iniciar com o Windows', ao ligar/entrar o app abre na bandeja e já lê os NÃO LIDOS "
            "(monta o rascunho único de Folha e, se houver urgência, abre um alerta de alta importância "
            "no Outlook — nada é enviado). O horário fixo é opcional e só dispara com o app aberto e a "
            "máquina ligada. A chave, se salva, fica no arquivo de configuração — restrinja o acesso a "
            "essa pasta ou prefira a variável de ambiente ANTHROPIC_API_KEY."),
            font=("Segoe UI", 12), text_color=CORES["texto_sec"], justify="left", wraplength=980, anchor="w").grid(
            row=7, column=0, columnspan=3, padx=18, pady=(2, 16), sticky="ew")

    def criar_area_log(self, parent):
        card = ctk.CTkFrame(
            parent,
            fg_color=CORES["card"],
            border_color=CORES["borda"],
            border_width=1,
            corner_radius=14,
        )
        card.grid(row=0, column=0, sticky="nsew")
        card.grid_columnconfigure(0, weight=1)
        card.grid_rowconfigure(2, weight=1)

        header_exec = ctk.CTkFrame(card, fg_color="transparent")
        header_exec.grid(row=0, column=0, padx=18, pady=(16, 8), sticky="ew")
        header_exec.grid_columnconfigure(0, weight=1)

        titulo = ctk.CTkLabel(
            header_exec,
            text="Execução",
            font=("Segoe UI", 17, "bold"),
            text_color=CORES["navy"],
            anchor="w",
        )
        titulo.grid(row=0, column=0, sticky="w")

        self.btn_executar = ctk.CTkButton(
            header_exec,
            text="Executar extração",
            command=self.iniciar_extracao,
            height=44,
            width=230,
            fg_color=CORES["sucesso"],
            hover_color="#256628",
            text_color=CORES["branco"],
            font=("Segoe UI", 15, "bold"),
        )
        self.btn_executar.grid(row=0, column=1, sticky="e")

        self.progress = ctk.CTkProgressBar(
            card,
            height=14,
            progress_color=CORES["azul"],
            fg_color=CORES["borda"],
        )
        self.progress.grid(row=1, column=0, padx=18, pady=(0, 10), sticky="ew")
        self.progress.set(0)

        self.txt_log = ctk.CTkTextbox(
            card,
            height=280,
            fg_color="#0F172A",
            text_color="#E5E7EB",
            border_color=CORES["borda"],
            border_width=1,
            font=("Consolas", 12),
        )
        self.txt_log.grid(row=2, column=0, padx=18, pady=(0, 18), sticky="nsew")

        self.log("Aplicativo iniciado.")
        self.log("Modo seguro: nenhum e-mail é movido, excluído, marcado ou enviado.")
        self.log("V2.1: a varredura nunca cria rascunhos automaticamente. A Mesa Copilot permite respostas individuais ou um único resumo Folha + Urgências.")

    def criar_rodape(self, parent):
        rodape = ctk.CTkFrame(parent, fg_color=CORES["fundo"], corner_radius=0, height=36)
        rodape.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 8))
        rodape.grid_columnconfigure(0, weight=1)

        label = ctk.CTkLabel(
            rodape,
            text="Anderson Marinho | Igarapé Digital",
            font=("Segoe UI", 12),
            text_color=CORES["texto_sec"],
            anchor="e",
        )
        label.grid(row=0, column=0, sticky="e")

    def log(self, mensagem):
        mensagem = str(mensagem)
        linha = f"[{agora_texto()}] {mensagem}\n"

        try:
            self.txt_log.insert("end", linha)
            self.txt_log.see("end")
        except Exception:
            pass

        try:
            registrar_log_arquivo(mensagem)
        except Exception:
            pass

    def selecionar_pasta_saida(self):
        pasta = filedialog.askdirectory(
            title="Selecione a pasta de saída",
            initialdir=self.var_pasta_saida.get() or str(SAIDA_DIR),
        )
        if pasta:
            self.var_pasta_saida.set(pasta)

    def abrir_pasta_saida(self):
        pasta = self.var_pasta_saida.get().strip()
        if not pasta:
            pasta = str(SAIDA_DIR)

        Path(pasta).mkdir(parents=True, exist_ok=True)
        os.startfile(pasta)

    def validar_campos(self):
        try:
            quantidade = int(self.var_quantidade.get())
            if quantidade <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Validação", "Informe uma quantidade válida maior que zero.")
            return False

        if not self.var_ultimos_dias.get().strip():
            self.var_ultimos_dias.set("0")
        try:
            dias = int(self.var_ultimos_dias.get())
            if dias < 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Validação", "Informe os últimos dias como número. Use 0 para não filtrar.")
            return False

        de_txt = self.var_data_de.get().strip(); ate_txt = self.var_data_ate.get().strip()
        dt_de = parse_data_br(de_txt); dt_ate = parse_data_br(ate_txt)
        if de_txt and dt_de is None:
            messagebox.showerror("Validação", "Data 'De' inválida. Use o formato dd/mm/aaaa."); return False
        if ate_txt and dt_ate is None:
            messagebox.showerror("Validação", "Data 'Até' inválida. Use o formato dd/mm/aaaa."); return False
        if dt_de and dt_ate and dt_de.date() > dt_ate.date():
            messagebox.showerror("Validação", "A data 'De' não pode ser maior que a data 'Até'."); return False

        if self.var_modo_caixa.get() == "compartilhada" and not self.var_caixa_compartilhada.get().strip():
            messagebox.showerror("Validação", "Informe o nome ou e-mail da caixa compartilhada.")
            return False

        pasta_saida = self.var_pasta_saida.get().strip()
        if not pasta_saida:
            messagebox.showerror("Validação", "Informe a pasta de saída.")
            return False

        nome_excel = self.var_nome_excel.get().strip()
        if not nome_excel:
            messagebox.showerror("Validação", "Informe o nome do arquivo Excel.")
            return False

        if not nome_excel.lower().endswith(".xlsx"):
            self.var_nome_excel.set(nome_excel + ".xlsx")

        if self.var_ia_ativa.get():
            chave = self.var_ia_api_key.get().strip() or os.environ.get("ANTHROPIC_API_KEY", "").strip()
            if not chave:
                messagebox.showerror(
                    "Assistente de IA",
                    "A IA está ativada, mas nenhuma chave da API foi encontrada.\n\n"
                    "Informe a chave no campo 'Chave da API' ou defina a variável "
                    "de ambiente ANTHROPIC_API_KEY.",
                )
                return False

        return True

    def travar_interface(self, travar=True):
        estado = "disabled" if travar else "normal"

        try:
            self.btn_extrair.configure(state=estado)
        except Exception:
            pass

        try:
            self.btn_executar.configure(state=estado)
        except Exception:
            pass

        self.processo_rodando = travar

    def atualizar_progresso(self, atual, total):
        try:
            valor = min(max(atual / total, 0), 1)
            self.progress.set(valor)
        except Exception:
            pass

    def iniciar_extracao(self):
        if self.processo_rodando:
            messagebox.showinfo("Execução", "Já existe uma extração em andamento.")
            return

        if not self.validar_campos():
            return

        config = self.obter_config_tela()
        self._persistir_config(config)

        self.travar_interface(True)
        self.progress.set(0)
        self.log("Iniciando extração dos e-mails.")

        thread = threading.Thread(target=self.executar_extracao_thread, args=(config,), daemon=True)
        thread.start()

    def _persistir_config(self, config=None):
        """Salva a config. A chave só é gravada em disco se 'Salvar chave' estiver ativo."""
        try:
            config = dict(config or self.obter_config_tela())
            if not config.get("salvar_chave"):
                config["ia_api_key"] = ""  # não grava a chave em texto puro
            salvar_config(config)
        except Exception:
            pass

    def executar_extracao_thread(self, config):
        try:
            config_ia = {
                "ia_ativa": config.get("ia_ativa", False),
                "criar_rascunho_outlook": False,  # nunca durante a varredura
                "ia_modelo": config.get("ia_modelo", IA_MODELO_PADRAO),
                "ia_api_key": config.get("ia_api_key", ""), "ia_workspace_id": config.get("ia_workspace_id", ""),
                "ia_assinatura": config.get("ia_assinatura", ""), "ia_contexto_rh": config.get("ia_contexto_rh", ""),
                "modo_operacional": config.get("modo_operacional", MODO_PADRAO),
                "destino_comunicacao": config.get("destino_comunicacao", DESTINO_PADRAO),
            }
            self.after(0,self.log,f"Modo operacional: {config_ia['modo_operacional']} | Destino/Tom: {config_ia['destino_comunicacao']}")
            if config_ia["ia_ativa"]: self.after(0,self.log,f"IA ativada. Modelo: {config_ia['ia_modelo']}.")
            else: self.after(0,self.log,"IA desligada. Regras locais e filas Copilot continuarão funcionando.")

            registros, erros = ler_emails_outlook(
                modo_caixa=config["modo_caixa"], caixa_compartilhada=config["caixa_compartilhada"], pasta_outlook=config["pasta_outlook"],
                quantidade=config["quantidade"], ultimos_dias=config["ultimos_dias"], somente_nao_lidos=config["somente_nao_lidos"],
                config_ia=config_ia, pasta_entry_id=config.get("pasta_entry_id", ""), pasta_store_id=config.get("pasta_store_id", ""),
                data_de=config.get("data_de", ""), data_ate=config.get("data_ate", ""),
                progress_callback=lambda atual,total:self.after(0,self.atualizar_progresso,atual,total),
                log_callback=lambda msg:self.after(0,self.log,msg),
            )
            if not registros:
                self.after(0,self.log,"Nenhum e-mail foi extraído com os filtros informados.")
                self.after(0,messagebox.showwarning,"Resultado","Nenhum e-mail foi extraído com os filtros informados."); return

            self._pos_processar_registros(registros, erros, config)
        except Exception as e:
            erro=traceback.format_exc(); self.after(0,self.log,f"Erro na execução: {e}"); registrar_log_arquivo(erro)
            self.after(0,messagebox.showerror,"Erro",f"Ocorreu um erro na execução:\n\n{e}")
        finally:
            self.after(0,self.travar_interface,False)

    def _pos_processar_registros(self, registros, erros, config):
        """Etapas comuns após obter os registros: casos, histórico, Mesa, Excel e resumo de Folha."""
        enriquecer_casos_sessao(registros); atualizar_historico_casos(registros)
        self.registros_sessao = registros
        self.after(0,self.atualizar_mesa_copilot)
        self.after(0,self.mostrar_view,"copilot")

        pasta_saida=Path(config["pasta_saida"]); nome_excel=config["nome_excel"]
        if not nome_excel.lower().endswith(".xlsx"): nome_excel += ".xlsx"
        caminho_saida=pasta_saida/nome_excel
        self.after(0,self.log,"Gerando Excel de apoio do Copilot.")
        caminho_gerado=gerar_excel(registros,caminho_saida); self.caminho_excel_gerado=caminho_gerado
        self.after(0,self.progress.set,1); self.after(0,self.log,f"Excel gerado: {caminho_gerado}")
        cont=Counter(r.get("Fila Copilot") for r in registros)
        resumo=(f"Análise concluída.\n\nCríticos: {cont.get('Crítico',0)}\nPrecisa de mim: {cont.get('Precisa de mim',0)}\n"
                f"Delegáveis: {cont.get('Delegável',0)}\nAguardando terceiro: {cont.get('Aguardando terceiro',0)}\nInformativos: {cont.get('Informativo',0)}\n\n"
                "A Mesa Copilot foi atualizada. Selecione os casos que realmente quer responder.")
        if erros: self.after(0,self.log,f"Itens ignorados por erro: {len(erros)}")
        self.after(0,messagebox.showinfo,"Copilot",resumo)

        if config.get("resumo_folha_auto", True):
            self.after(0,self.log,"Preparando automaticamente o e-mail único de Folha (onde atuar).")
            self.after(0,self.preparar_resumo_folha_urgencias)

    # ------------------------------------------------------------------
    # Fluxo: listar cabeçalhos (sem IA) -> escolher com checkbox -> analisar
    # ------------------------------------------------------------------
    def iniciar_listagem(self):
        if self.processo_rodando:
            messagebox.showinfo("Listagem", "Já existe um processo em andamento."); return
        if not self.validar_campos():
            return
        config = self.obter_config_tela()
        self.travar_interface(True); self.progress.set(0)
        self.log("Listando cabeçalhos (sem IA) para seleção.")
        threading.Thread(target=self._listar_thread, args=(config,), daemon=True).start()

    def _listar_thread(self, config):
        try:
            cabecalhos = listar_cabecalhos_outlook(
                modo_caixa=config["modo_caixa"], caixa_compartilhada=config["caixa_compartilhada"],
                pasta_outlook=config["pasta_outlook"], pasta_entry_id=config.get("pasta_entry_id",""),
                pasta_store_id=config.get("pasta_store_id",""), data_de=config.get("data_de",""),
                data_ate=config.get("data_ate",""), ultimos_dias=config.get("ultimos_dias",0),
                somente_nao_lidos=config.get("somente_nao_lidos", False),
                limite=max(int(config.get("quantidade",200)), 1000),
                log_callback=lambda msg:self.after(0,self.log,msg),
            )
            if not cabecalhos:
                self.after(0,messagebox.showwarning,"Listagem","Nenhum e-mail no intervalo/filtro informado.")
                return
            self.after(0, self.abrir_selecao_emails, cabecalhos, config)
        except Exception as e:
            erro=traceback.format_exc(); self.after(0,self.log,f"Erro na listagem: {e}"); registrar_log_arquivo(erro)
            self.after(0,messagebox.showerror,"Listagem",f"Não foi possível listar os e-mails.\n\n{e}")
        finally:
            self.after(0,self.travar_interface,False)

    def abrir_selecao_emails(self, cabecalhos, config):
        win = ctk.CTkToplevel(self); win.title("Escolher e-mails para analisar"); win.geometry("1040x760")
        win.grid_columnconfigure(0, weight=1); win.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(win, text=f"{len(cabecalhos)} e-mail(s) no intervalo. Marque os que a IA deve analisar.",
                     font=("Segoe UI",15,"bold"), text_color=CORES["navy"], anchor="w").grid(row=0,column=0,padx=16,pady=(16,4),sticky="ew")

        topo = ctk.CTkFrame(win, fg_color="transparent"); topo.grid(row=1,column=0,padx=16,pady=6,sticky="ew"); topo.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(topo, text="Filtro", text_color=CORES["texto"]).grid(row=0,column=0,padx=(0,8))
        var_filtro = ctk.StringVar(value="")
        ent = ctk.CTkEntry(topo, textvariable=var_filtro, placeholder_text="Filtrar por remetente, e-mail ou assunto...", height=34)
        ent.grid(row=0,column=1,sticky="ew")

        tree_frame = ctk.CTkFrame(win, fg_color=CORES["card"], border_color=CORES["borda"], border_width=1, corner_radius=10)
        tree_frame.grid(row=2,column=0,padx=16,pady=6,sticky="nsew"); tree_frame.grid_columnconfigure(0,weight=1); tree_frame.grid_rowconfigure(0,weight=1)
        cols = ("sel","data","hora","remetente","email","assunto","lido")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="none")
        titulos = {"sel":"✓","data":"Data","hora":"Hora","remetente":"Remetente","email":"E-mail","assunto":"Assunto","lido":"Lido"}
        larg = {"sel":36,"data":90,"hora":60,"remetente":170,"email":210,"assunto":330,"lido":50}
        for c in cols:
            tree.heading(c, text=titulos[c])
            tree.column(c, width=larg[c], minwidth=32, anchor=("center" if c in {"sel","lido","hora"} else "w"),
                        stretch=(c in {"assunto","email"}))
        tree.grid(row=0,column=0,sticky="nsew")
        sb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview); sb.grid(row=0,column=1,sticky="ns"); tree.configure(yscrollcommand=sb.set)

        # Estado de seleção por EntryID
        marcados = set(h["EntryID"] for h in cabecalhos)  # começa tudo marcado
        por_id = {h["EntryID"]: h for h in cabecalhos}

        def _visiveis():
            f = normalizar_busca(var_filtro.get())
            if not f:
                return cabecalhos
            out = []
            for h in cabecalhos:
                alvo = normalizar_busca(" ".join([h.get("Remetente Nome",""), h.get("Remetente E-mail",""), h.get("Assunto","")]))
                if f in alvo:
                    out.append(h)
            return out

        def repovoar():
            tree.delete(*tree.get_children())
            for h in _visiveis():
                eid = h["EntryID"]
                marca = "☑" if eid in marcados else "☐"
                tree.insert("", "end", iid=eid, values=(marca, h.get("Data",""), h.get("Hora",""),
                            h.get("Remetente Nome",""), h.get("Remetente E-mail",""), h.get("Assunto",""), h.get("Lido","")))
            atualizar_contador()

        def atualizar_contador():
            lbl_cont.configure(text=f"Marcados: {len(marcados)} de {len(cabecalhos)}")

        def toggle(event):
            iid = tree.identify_row(event.y)
            if not iid:
                return
            if iid in marcados: marcados.discard(iid)
            else: marcados.add(iid)
            tree.set(iid, "sel", "☑" if iid in marcados else "☐")
            atualizar_contador()

        tree.bind("<Button-1>", toggle)
        ent.bind("<KeyRelease>", lambda _e: repovoar())

        rodape = ctk.CTkFrame(win, fg_color="transparent"); rodape.grid(row=3,column=0,padx=16,pady=(6,14),sticky="ew"); rodape.grid_columnconfigure(1, weight=1)
        def marcar_todos():
            for h in _visiveis(): marcados.add(h["EntryID"])
            repovoar()
        def desmarcar_todos():
            for h in _visiveis(): marcados.discard(h["EntryID"])
            repovoar()
        ctk.CTkButton(rodape, text="Marcar todos", command=marcar_todos, height=36, width=120,
                      fg_color=CORES["azul_apoio"], hover_color=CORES["navy"]).grid(row=0,column=0,padx=(0,8))
        lbl_cont = ctk.CTkLabel(rodape, text="", text_color=CORES["texto_sec"]); lbl_cont.grid(row=0,column=1,sticky="w")
        ctk.CTkButton(rodape, text="Desmarcar todos", command=desmarcar_todos, height=36, width=130,
                      fg_color=CORES["auxiliar"], hover_color=CORES["auxiliar_hover"], text_color=CORES["texto"]).grid(row=0,column=2,padx=(0,8))

        def analisar():
            escolhidos = [por_id[e] for e in marcados if e in por_id]
            if not escolhidos:
                messagebox.showinfo("Seleção", "Marque ao menos um e-mail.", parent=win); return
            if not messagebox.askyesno("Analisar", f"Analisar {len(escolhidos)} e-mail(s) com a IA?\n\nCada e-mail é uma chamada.", parent=win):
                return
            win.destroy()
            self.travar_interface(True); self.progress.set(0)
            self.log(f"Analisando {len(escolhidos)} e-mail(s) selecionado(s).")
            threading.Thread(target=self._analisar_selecionados_thread, args=(escolhidos, config), daemon=True).start()

        ctk.CTkButton(rodape, text="Analisar selecionados (IA)", command=analisar, height=36,
                      fg_color=CORES["sucesso"], hover_color="#256628").grid(row=0,column=3)

        repovoar()

    def _analisar_selecionados_thread(self, escolhidos, config):
        pythoncom.CoInitialize()
        try:
            config_ia = {
                "ia_ativa": config.get("ia_ativa", False), "criar_rascunho_outlook": False,
                "ia_modelo": config.get("ia_modelo", IA_MODELO_PADRAO),
                "ia_api_key": config.get("ia_api_key", ""), "ia_workspace_id": config.get("ia_workspace_id", ""),
                "ia_assinatura": config.get("ia_assinatura", ""), "ia_contexto_rh": config.get("ia_contexto_rh", ""),
                "modo_operacional": config.get("modo_operacional", MODO_PADRAO),
                "destino_comunicacao": config.get("destino_comunicacao", DESTINO_PADRAO),
            }
            registros = []; erros = []; total = len(escolhidos)
            for i, h in enumerate(escolhidos, start=1):
                try:
                    mail = obter_item_outlook_por_ids(h.get("EntryID",""), h.get("StoreID",""))
                    if mail is None:
                        erros.append(f"Item não encontrado: {h.get('Assunto','')}"); continue
                    dados = extrair_dados_email(mail, config_ia=config_ia, log_callback=lambda msg:self.after(0,self.log,msg))
                    registros.append(dados)
                except Exception as e:
                    erros.append(str(e))
                self.after(0, self.atualizar_progresso, i, total)
            if not registros:
                self.after(0,messagebox.showwarning,"Análise","Nenhum e-mail selecionado pôde ser lido.")
                return
            self._pos_processar_registros(registros, erros, config)
        except Exception as e:
            erro=traceback.format_exc(); self.after(0,self.log,f"Erro na análise dos selecionados: {e}"); registrar_log_arquivo(erro)
            self.after(0,messagebox.showerror,"Erro",f"Ocorreu um erro:\n\n{e}")
        finally:
            pythoncom.CoUninitialize()
            self.after(0,self.travar_interface,False)

    # ==================================================================
    # Agendamento, execução em segundo plano e alerta de urgentes
    # ==================================================================
    def iniciar_agendador(self):
        """Thread leve que dispara a execução agendada uma vez por dia no horário."""
        self._agendador_ativo = True
        self._ultima_exec_agendada = None

        def loop():
            while getattr(self, "_agendador_ativo", False):
                try:
                    if self.var_agendamento_ativo.get():
                        alvo = (self.var_agendamento_hora.get() or "").strip()
                        if re.match(r"^\d{1,2}:\d{2}$", alvo):
                            hh, mm = alvo.split(":")
                            agora = datetime.now()
                            if agora.hour == int(hh) and agora.minute == int(mm):
                                hoje = agora.date()
                                if self._ultima_exec_agendada != hoje and not self.processo_rodando:
                                    self._ultima_exec_agendada = hoje
                                    self.after(0, self.executar_agendado)
                except Exception:
                    pass
                time.sleep(20)

        threading.Thread(target=loop, daemon=True).start()

    def executar_agendado(self):
        """Dispara a leitura automática dos NÃO LIDOS, de forma silenciosa."""
        if self.processo_rodando:
            self.log("Agendado: já há um processo em andamento; execução ignorada.")
            return
        config = self.obter_config_tela()
        config["somente_nao_lidos"] = True     # tudo que não foi lido
        config["data_de"] = ""; config["data_ate"] = ""  # não limitar por intervalo no agendado
        self._persistir_config(config)
        self.travar_interface(True); self.progress.set(0)
        self.log("Execução agendada: lendo e-mails não lidos.")
        threading.Thread(target=self._executar_agendado_thread, args=(config,), daemon=True).start()

    def _executar_agendado_thread(self, config):
        try:
            config_ia = {
                "ia_ativa": config.get("ia_ativa", False), "criar_rascunho_outlook": False,
                "ia_modelo": config.get("ia_modelo", IA_MODELO_PADRAO),
                "ia_api_key": config.get("ia_api_key", ""), "ia_workspace_id": config.get("ia_workspace_id", ""),
                "ia_assinatura": config.get("ia_assinatura", ""), "ia_contexto_rh": config.get("ia_contexto_rh", ""),
                "modo_operacional": config.get("modo_operacional", MODO_PADRAO),
                "destino_comunicacao": config.get("destino_comunicacao", DESTINO_PADRAO),
            }
            registros, erros = ler_emails_outlook(
                modo_caixa=config["modo_caixa"], caixa_compartilhada=config["caixa_compartilhada"],
                pasta_outlook=config["pasta_outlook"], quantidade=config["quantidade"], ultimos_dias=0,
                somente_nao_lidos=True, config_ia=config_ia,
                pasta_entry_id=config.get("pasta_entry_id",""), pasta_store_id=config.get("pasta_store_id",""),
                progress_callback=lambda a,t:self.after(0,self.atualizar_progresso,a,t),
                log_callback=lambda msg:self.after(0,self.log,msg),
            )
            if not registros:
                self.after(0,self.log,"Agendado: nenhum e-mail não lido novo."); return

            enriquecer_casos_sessao(registros); atualizar_historico_casos(registros)
            self.registros_sessao = registros
            self.after(0, self.atualizar_mesa_copilot)

            # Excel de apoio
            try:
                pasta_saida=Path(config["pasta_saida"]); nome_excel=config["nome_excel"]
                if not nome_excel.lower().endswith(".xlsx"): nome_excel += ".xlsx"
                caminho_gerado=gerar_excel(registros, pasta_saida/nome_excel); self.caminho_excel_gerado=caminho_gerado
                self.after(0,self.log,f"Agendado: Excel gerado ({caminho_gerado}).")
            except Exception as e:
                self.after(0,self.log,f"Agendado: falha ao gerar Excel: {e}")

            casos_folha = selecionar_casos_folha_urgencias(registros)
            urgentes = [r for r in consolidar_registros_por_caso(registros) if registro_e_urgente(r)]

            # Rascunho único de Folha + alerta de urgentes (num único bloco COM)
            pythoncom.CoInitialize()
            try:
                if casos_folha:
                    assunto, corpo = gerar_email_resumo_com_ia(
                        registros=casos_folha, config_ia=config_ia,
                        destino_comunicacao=config.get("destino_comunicacao", DESTINO_PADRAO),
                        origem="automatico", log_callback=lambda msg:self.after(0,self.log,msg),
                    )
                    criar_rascunho_novo_outlook(
                        destinatario=config.get("resumo_destinatario_email",""),
                        assunto=assunto, corpo=corpo, assinatura=config.get("ia_assinatura",""),
                        abrir=False, log_callback=lambda msg:self.after(0,self.log,msg),
                    )
                    self.after(0,self.log,"Agendado: rascunho único de Folha salvo em Rascunhos.")
                if urgentes and config.get("alerta_urgente", True):
                    self._alertar_urgentes_outlook(urgentes, config)
            finally:
                pythoncom.CoUninitialize()

            if urgentes and config.get("alerta_urgente", True):
                self._notificar_bandeja("RH Copilot", f"{len(urgentes)} caso(s) podem ser urgentes. Veja o alerta no Outlook.")
            self.after(0,self.log,f"Agendado concluído: {len(registros)} lido(s), {len(urgentes)} possível(is) urgente(s).")
        except Exception as e:
            erro=traceback.format_exc(); self.after(0,self.log,f"Erro no agendado: {e}"); registrar_log_arquivo(erro)
        finally:
            self.after(0,self.travar_interface,False)

    def _alertar_urgentes_outlook(self, urgentes, config):
        """Cria um rascunho de ALERTA de alta importância no Outlook (não envia)."""
        try:
            linhas = [f"{len(urgentes)} caso(s) podem ser urgentes na caixa de Gente e Gestão:", ""]
            for r in urgentes[:25]:
                assunto = r.get("ConversationTopic") or r.get("Assunto") or "Sem assunto"
                linhas.append(f"- [{r.get('Fila Copilot','')}] {assunto} — {r.get('Remetente Nome','')}"
                              + (f" | prazo: {r.get('Prazo / Timing','')}" if r.get('Prazo / Timing') else ""))
            corpo = "\n".join(linhas)
            assunto_alerta = f"⚠ URGENTE — {len(urgentes)} caso(s) em RH ({datetime.now():%d/%m %H:%M})"
            outlook = win32.Dispatch("Outlook.Application")
            m = outlook.CreateItem(0)
            m.Subject = assunto_alerta
            m.Body = corpo
            try: m.Importance = 2  # olImportanceHigh
            except Exception: pass
            dest = (config.get("resumo_destinatario_email","") or "").strip()
            if dest: m.To = dest
            m.Save(); m.Display()
            self.after(0,self.log,"Alerta de urgentes aberto no Outlook (alta importância, não enviado).")
        except Exception as e:
            self.after(0,self.log,f"Falha ao criar alerta no Outlook: {e}")

    # ----- Bandeja do Windows (segundo plano) -----
    def _iniciar_tray(self):
        if getattr(self, "_tray_icon", None):
            return True
        try:
            import pystray
            from PIL import Image, ImageDraw
        except Exception:
            self.log("Bandeja indisponível: instale com  python -m pip install --user pystray pillow.")
            return False
        img = Image.new("RGB", (64, 64), "#0083CA")
        d = ImageDraw.Draw(img); d.rectangle([18, 18, 46, 46], fill="white")

        def _abrir(icon=None, item=None): self.after(0, self._restaurar_da_bandeja)
        def _rodar(icon=None, item=None): self.after(0, self.executar_agendado)
        def _sair(icon=None, item=None):
            self._agendador_ativo = False
            try: self._tray_icon.stop()
            except Exception: pass
            self._tray_icon = None
            self.after(0, self.destroy)

        menu = pystray.Menu(
            pystray.MenuItem("Abrir", _abrir, default=True),
            pystray.MenuItem("Rodar agora (não lidos)", _rodar),
            pystray.MenuItem("Sair", _sair),
        )
        self._tray_icon = pystray.Icon("rhcopilot", img, "RH Copilot", menu)
        threading.Thread(target=self._tray_icon.run, daemon=True).start()
        return True

    def minimizar_para_bandeja(self):
        self._persistir_config()
        if self._iniciar_tray():
            self.withdraw()
            self.log("Rodando em segundo plano na bandeja. O agendamento continua ativo. Clique no ícone para reabrir.")
        else:
            self.iconify()
            self.log("Minimizado. Para o ícone na bandeja, instale pystray e pillow.")

    def _restaurar_da_bandeja(self):
        try:
            self.deiconify(); self.state("normal"); self.lift(); self.focus_force()
        except Exception:
            pass

    def _notificar_bandeja(self, titulo, mensagem):
        ic = getattr(self, "_tray_icon", None)
        if ic:
            try:
                ic.notify(mensagem, titulo)
            except Exception:
                pass

    # ----- Iniciar com o Windows (pasta Inicializar) -----
    def _pasta_startup(self):
        appdata = os.environ.get("APPDATA", "")
        return os.path.join(appdata, "Microsoft", "Windows", "Start Menu", "Programs", "Startup")

    def _caminho_atalho_startup(self):
        return os.path.join(self._pasta_startup(), "RH Copilot.lnk")

    def _alvo_e_args_startup(self):
        """Descobre o executável e os argumentos para o atalho, com ou sem cx_Freeze."""
        script = os.path.abspath(sys.argv[0])
        if getattr(sys, "frozen", False):
            # App compilado: o próprio executável.
            return sys.executable, "--startup", os.path.dirname(sys.executable)
        # Rodando como script: usa pythonw.exe (sem console) quando existir.
        interp = sys.executable
        if interp.lower().endswith("python.exe"):
            cand = interp[:-len("python.exe")] + "pythonw.exe"
            if os.path.exists(cand):
                interp = cand
        args = f'"{script}" --startup'
        return interp, args, os.path.dirname(script)

    def criar_atalho_startup(self):
        try:
            pasta = self._pasta_startup()
            os.makedirs(pasta, exist_ok=True)
            alvo, args, workdir = self._alvo_e_args_startup()
            shell = win32.Dispatch("WScript.Shell")
            atalho = shell.CreateShortCut(self._caminho_atalho_startup())
            atalho.TargetPath = alvo
            atalho.Arguments = args
            atalho.WorkingDirectory = workdir
            atalho.WindowStyle = 7  # minimizado
            atalho.Description = "RH Copilot - início automático com o Windows"
            try:
                atalho.IconLocation = alvo
            except Exception:
                pass
            atalho.save()
            return True
        except Exception as e:
            self.log(f"Falha ao criar o atalho de inicialização: {e}")
            return False

    def remover_atalho_startup(self):
        try:
            caminho = self._caminho_atalho_startup()
            if os.path.exists(caminho):
                os.remove(caminho)
            return True
        except Exception as e:
            self.log(f"Falha ao remover o atalho de inicialização: {e}")
            return False

    def _toggle_startup(self):
        if self.var_iniciar_com_windows.get():
            ok = self.criar_atalho_startup()
            if ok:
                self.log("Início com o Windows ativado. Ao ligar/entrar, o app abre na bandeja e lê os não lidos.")
            else:
                self.var_iniciar_com_windows.set(False)
        else:
            if self.remover_atalho_startup():
                self.log("Início com o Windows desativado.")

    def _boot_autorun(self):
        """Chamado quando o app é iniciado pela pasta Inicializar do Windows."""
        self.log("Início automático com o Windows: indo para a bandeja e lendo os não lidos.")
        self.minimizar_para_bandeja()
        self.executar_agendado()


def main():
    garantir_pastas()
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()